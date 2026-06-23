import openpyxl
import sys

try:
    wb = openpyxl.load_workbook("Расчет сырья Тектум. Основа..xlsx", data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"Sheet: {sheet_name}")
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            print(row)
        print("---")
except Exception as e:
    print(f"Error: {e}")
