import openpyxl
import io
from datetime import datetime, date
from sqlalchemy.orm import Session
import models

def import_aci_excel_data(file_source, db: Session) -> dict:
    """
    Импортирует данные напрямую из Excel-файла рапорта АЦИ (лист "рапорт").
    file_source может быть путем к файлу (str) или файловым объектом (BytesIO/BinaryIO).
    """
    # Загружаем книгу
    if isinstance(file_source, (str, bytes)):
        wb = openpyxl.load_workbook(file_source, data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_source.read()) if hasattr(file_source, 'read') else file_source, data_only=True)
        
    if "рапорт" not in wb.sheetnames:
        raise ValueError("В Excel-файле отсутствует лист 'рапорт'")
        
    sheet = wb["рапорт"]
    
    # 1. Первым проходом собираем всех мастеров
    masters_cache = {}
    db_masters = db.query(models.Master).all()
    for m in db_masters:
        masters_cache[m.name.strip().lower()] = m.id
        
    # Считываем строки, начиная с 3-й
    rows_data = []
    current_date = datetime.now().date()
    
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
            
        date_val = row[0]
        if isinstance(date_val, datetime):
            row_date = date_val.date()
        elif isinstance(date_val, date):
            row_date = date_val
        else:
            try:
                row_date = datetime.strptime(str(date_val).split()[0], "%Y-%m-%d").date()
            except ValueError:
                try:
                    row_date = datetime.strptime(str(date_val).split()[0], "%d.%m.%Y").date()
                except ValueError:
                    continue
                    
        # Игнорируем будущие даты при парсинге
        if row_date > current_date:
            continue
            
        master_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        product_name = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        
        # Пропускаем пустые строки без мастера и продукта
        if not master_name and not product_name:
            continue
            
        rows_data.append((row_date, row))
        
        # Добавляем мастера, если его нет
        if master_name:
            m_key = master_name.lower()
            if m_key not in masters_cache:
                m = models.Master(name=master_name, pin="0000", role="master")
                db.add(m)
                db.commit()
                db.refresh(m)
                masters_cache[m_key] = m.id
                
    # 2. Вторым проходом импортируем смены, отчеты ЛФМ и партии
    shifts_cache = {}
    imported_shifts_count = 0
    imported_batches_count = 0
    imported_lfm_count = 0
    
    for row_date, row in rows_data:
        # Линия
        raw_line = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ""
        if "хоргос" in raw_line:
            db_line = "Линия 1"
        elif "шанхай" in raw_line:
            db_line = "Линия 2"
        elif "1" in raw_line:
            db_line = "Линия 1"
        elif "2" in raw_line:
            db_line = "Линия 2"
        else:
            db_line = "Линия 1"
            
        # Смена (День / Ночь)
        raw_shift = str(row[3]).strip().upper() if len(row) > 3 and row[3] else ""
        if "Д" in raw_shift:
            db_shift_name = "День"
        elif "Н" in raw_shift:
            db_shift_name = "Ночь"
        else:
            db_shift_name = "День"
            
        shift_key = (row_date, db_shift_name, db_line)
        
        # Получаем или создаем смену
        if shift_key not in shifts_cache:
            master_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            master_id = masters_cache.get(master_name.lower()) if master_name else None
            
            shift = db.query(models.Shift).filter_by(date=row_date, shift_name=db_shift_name, line=db_line).first()
            if not shift:
                shift = models.Shift(
                    date=row_date,
                    shift_name=db_shift_name,
                    line=db_line,
                    master_id=master_id,
                    status="closed"
                )
                db.add(shift)
                db.commit()
                db.refresh(shift)
                imported_shifts_count += 1
                
            shifts_cache[shift_key] = {
                "id": shift.id,
                "model": shift,
                "data_filled": False
            }
            
        shift_info = shifts_cache[shift_key]
        shift_model = shift_info["model"]
        shift_id = shift_info["id"]
        
        # Заполняем показатели расхода сырья ЗО один раз для смены
        if not shift_info["data_filled"]:
            batches_count = row[6] if len(row) > 6 and row[6] else 0
            if batches_count or (len(row) > 18 and row[18]): # receipt_cement_total
                shift_model.zo_batches = int(batches_count)
                
                # Цемент по силосам (в кг)
                shift_model.zo_cement_silo1 = float(row[14] or 0.0)
                shift_model.zo_cement_silo2 = float(row[15] or 0.0)
                shift_model.zo_cement_silo3 = float(row[16] or 0.0)
                shift_model.zo_cement_silo4 = float(row[17] or 0.0)
                shift_model.zo_cement = (shift_model.zo_cement_silo1 + shift_model.zo_cement_silo2 + 
                                         shift_model.zo_cement_silo3 + shift_model.zo_cement_silo4)
                
                # Сливы смеси
                shift_model.zo_asb_drain = float(row[19] or 0.0) if len(row) > 19 else 0.0
                shift_model.zo_cem_drain = float(row[20] or 0.0) if len(row) > 20 else 0.0
                
                # Приход / Расход сырья
                shift_model.receipt_laprol = float(row[21] or 0.0) if len(row) > 21 else 0.0
                shift_model.zo_laprol = float(row[22] or 0.0) if len(row) > 22 else 0.0
                
                shift_model.receipt_cellulose = float(row[23] or 0.0) if len(row) > 23 else 0.0
                shift_model.zo_cellulose = float(row[24] or 0.0) if len(row) > 24 else 0.0
                
                shift_model.receipt_fiberglass = float(row[25] or 0.0) if len(row) > 25 else 0.0
                shift_model.zo_fiberglass = float(row[26] or 0.0) if len(row) > 26 else 0.0
                
                shift_model.receipt_crushed_slate = float(row[27] or 0.0) if len(row) > 27 else 0.0
                shift_model.zo_crushed_slate = float(row[28] or 0.0) if len(row) > 28 else 0.0
                
                shift_model.zo_asbozurit = float(row[29] or 0.0) if len(row) > 29 else 0.0
                shift_model.zo_submitted = True
                
                db.commit()
                shift_info["data_filled"] = True
                
        # Добавление партии
        batch_num = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        product_name = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        
        if batch_num and product_name:
            existing_batch = db.query(models.Batch).filter_by(shift_id=shift_id, batch_number=batch_num).first()
            transferred = int(row[9] or 0) if len(row) > 9 and row[9] is not None else 0
            
            if not existing_batch:
                batch = models.Batch(
                    shift_id=shift_id,
                    batch_number=batch_num,
                    product_name=product_name,
                    status="qcd_checked",
                    stacked_stacks=0,
                    ds_condition=transferred,
                    qcd_condition=transferred,
                    qcd_first_grade=int(row[10] or 0) if len(row) > 10 and row[10] is not None else 0,
                    qcd_defect=int(row[11] or 0) if len(row) > 11 and row[11] is not None else 0
                )
                db.add(batch)
                imported_batches_count += 1
            else:
                existing_batch.ds_condition = transferred
                existing_batch.qcd_condition = transferred
                existing_batch.qcd_first_grade = int(row[10] or 0) if len(row) > 10 and row[10] is not None else 0
                existing_batch.qcd_defect = int(row[11] or 0) if len(row) > 11 and row[11] is not None else 0
                
        # Добавление отчета ЛФМ
        if product_name and (len(row) > 7 and row[7] is not None):
            existing_lfm = db.query(models.LFMReport).filter_by(shift_id=shift_id, product_name=product_name).first()
            lfm_sheets_val = int(row[7] or 0)
            lfm_wind_resets_val = int(row[12] or 0) if len(row) > 12 and row[12] is not None else 0
            
            if not existing_lfm:
                lfm = models.LFMReport(
                    shift_id=shift_id,
                    product_name=product_name,
                    lfm_sheets=lfm_sheets_val,
                    lfm_wind_resets=lfm_wind_resets_val,
                    formed_1st_grade=int(row[10] or 0) if len(row) > 10 and row[10] is not None else 0,
                    formed_defect=int(row[11] or 0) if len(row) > 11 and row[11] is not None else 0,
                    transferred_to_warehouse=transferred
                )
                db.add(lfm)
                imported_lfm_count += 1
            else:
                existing_lfm.lfm_sheets = lfm_sheets_val
                existing_lfm.lfm_wind_resets = lfm_wind_resets_val
                existing_lfm.transferred_to_warehouse = transferred
                
    db.commit()
    return {
        "status": "success",
        "shifts": imported_shifts_count,
        "batches": imported_batches_count,
        "lfm_reports": imported_lfm_count
    }
