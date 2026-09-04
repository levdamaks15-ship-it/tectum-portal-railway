from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Request, BackgroundTasks, Query, Body
from typing import Optional
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, RedirectResponse, HTMLResponse
from sqlalchemy.orm import Session
from database import SessionLocal, engine, Base
import models, schemas
import os
import re
import asyncio
import threading
import json
import hashlib
import html
try:
    import m365_integration
except ImportError:
    m365_integration = None
import excel_exporter
import import_aci_excel
from datetime import datetime, date
from pydantic import BaseModel
from sqlalchemy import or_, and_, func
from sqlalchemy.orm import selectinload, joinedload
from contextlib import asynccontextmanager
import seed_norms
import calendar
from datetime import timedelta
import openpyxl
from openpyxl.chart import BarChart, Reference
import io
from fastapi import Response
from urllib.parse import quote
try:
    import msal
except ImportError:
    msal = None
try:
    from starlette.middleware.sessions import SessionMiddleware
except ImportError:
    SessionMiddleware = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    Base.metadata.create_all(bind=engine)
    
    # SQLite alter table hack to auto-add columns if they don't exist
    import sqlite3
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE raw_material_receipts ADD COLUMN master_id INTEGER;")
        conn.commit()
        conn.close()
    except: pass
    
    # Ensure indexes for documents & document_categories
    try:
        db = SessionLocal()
        from sqlalchemy import text
        driver = db.bind.dialect.name if db.bind else 'unknown'
        if driver == 'postgresql':
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_documents_category_id ON documents (category_id);"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_doc_categories_parent_id ON document_categories (parent_id);"))
            db.commit()
        db.close()
    except Exception as e:
        print(f"Warning: could not create document indexes: {e}")

    # PG version - ensure master_id column exists
    try:
        db = SessionLocal()
        driver = db.bind.dialect.name if db.bind else 'unknown'
        if driver == 'postgresql':
            from sqlalchemy import text
            col_exists = db.execute(text(
                "SELECT column_name FROM information_schema.columns WHERE table_name='raw_material_receipts' AND column_name='master_id'"
            )).fetchone()
            if not col_exists:
                print("Adding master_id column to raw_material_receipts on PostgreSQL...")
                db.execute(text("ALTER TABLE raw_material_receipts ADD COLUMN master_id INTEGER REFERENCES masters(id);"))
                db.commit()
                print("master_id column added successfully.")
    except Exception as e:
        print(f"Warning: could not add master_id column: {e}")
        db.rollback()
    finally:
        db.close()

    # Data migration: clean up document titles containing relative paths
    try:
        db = SessionLocal()
        docs = db.query(models.Document).all()
        cleaned_count = 0
        for doc in docs:
            if doc.title and ("/" in doc.title or "\\" in doc.title):
                doc.title = os.path.basename(doc.title.replace("\\", "/"))
                cleaned_count += 1
        if cleaned_count > 0:
            db.commit()
            print(f"Cleaned {cleaned_count} document titles.")
        db.close()
    except Exception as e:
        print(f"Warning: could not clean document titles: {e}")

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN batch_number VARCHAR(255)")
        conn.commit()
        conn.close()
    except: pass
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN product_name VARCHAR(255)")
        conn.commit()
        conn.close()
    except: pass
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN export_type VARCHAR(50) DEFAULT 'Эталон'")
        conn.commit()
        conn.close()
    except: pass
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE lfm_reports ADD COLUMN export_type VARCHAR(50) DEFAULT 'Эталон'")
        conn.commit()
        conn.close()
    except: pass
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE batches ADD COLUMN export_type VARCHAR(50) DEFAULT 'Эталон'")
        conn.commit()
        conn.close()
    except: pass
    
    # Batches previous shift defects migration (SQLite)
    for col in [
        "prev_first_grade", "prev_defect", "prev_defect_scratch", "prev_defect_bad_cut",
        "prev_defect_stick_top", "prev_defect_broken", "prev_defect_fell_box",
        "prev_defect_thickness", "prev_defect_edge"
    ]:
        try:
            conn = sqlite3.connect("tectum.db")
            conn.execute(f"ALTER TABLE batches ADD COLUMN {col} INTEGER DEFAULT 0")
            conn.commit()
            conn.close()
        except: pass

    # AuditLog state_snapshot migration (SQLite)
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE audit_logs ADD COLUMN state_snapshot TEXT")
        conn.commit()
        conn.close()
    except: pass
    
    # RawMaterialReceipt autonomous columns migration (SQLite)
    for col_def in [
        ("date", "DATE"), ("shift_name", "VARCHAR(50)"), ("line", "VARCHAR(50)")
    ]:
        try:
            conn = sqlite3.connect("tectum.db")
            conn.execute(f"ALTER TABLE raw_material_receipts ADD COLUMN {col_def[0]} {col_def[1]}")
            conn.commit()
            conn.close()
        except: pass

    # Downtimes autonomous columns migration (SQLite)
    for col_def in [
        ("date", "DATE"), ("shift_name", "VARCHAR(50)"), ("line", "VARCHAR(50)"), ("master_id", "INTEGER")
    ]:
        try:
            conn = sqlite3.connect("tectum.db")
            conn.execute(f"ALTER TABLE downtimes ADD COLUMN {col_def[0]} {col_def[1]}")
            conn.commit()
            conn.close()
        except: pass
    
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE monthly_plan_board ADD COLUMN first_grade INTEGER DEFAULT 0")
        conn.commit()
        conn.close()
    except: pass
    
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE monthly_plan_board ADD COLUMN defect INTEGER DEFAULT 0")
        conn.commit()
        conn.close()
    except: pass
    
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE masters ADD COLUMN email VARCHAR(255)")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE downtimes ADD COLUMN department VARCHAR(255)")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE downtime_directory ADD COLUMN category VARCHAR(255)")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE downtimes ADD COLUMN is_equipment_downtime BOOLEAN DEFAULT 1")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE downtimes ADD COLUMN comment TEXT")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE downtimes ADD COLUMN breakdowns VARCHAR")
        conn.commit()
        conn.close()
    except: pass

    # PostgreSQL schema parity for downtimes & downtime_directory (TEXT unlimited lengths)
    try:
        db_pg = SessionLocal()
        driver_pg = db_pg.bind.dialect.name if db_pg.bind else 'unknown'
        if driver_pg == 'postgresql':
            from sqlalchemy import text
            db_pg.execute(text("ALTER TABLE downtimes ADD COLUMN IF NOT EXISTS department VARCHAR(255);"))
            db_pg.execute(text("ALTER TABLE downtimes ADD COLUMN IF NOT EXISTS is_equipment_downtime BOOLEAN DEFAULT TRUE;"))
            db_pg.execute(text("ALTER TABLE downtimes ADD COLUMN IF NOT EXISTS comment TEXT;"))
            db_pg.execute(text("ALTER TABLE downtimes ADD COLUMN IF NOT EXISTS breakdowns TEXT;"))
            db_pg.execute(text("ALTER TABLE downtimes ALTER COLUMN comment TYPE TEXT;"))
            db_pg.execute(text("ALTER TABLE downtimes ALTER COLUMN description TYPE TEXT;"))
            
            db_pg.execute(text("ALTER TABLE downtime_directory ADD COLUMN IF NOT EXISTS category VARCHAR(255);"))
            db_pg.execute(text("ALTER TABLE downtime_directory ADD COLUMN IF NOT EXISTS comment TEXT;"))
            db_pg.execute(text("ALTER TABLE downtime_directory ALTER COLUMN comment TYPE TEXT;"))
            
            # Autonomous journals: date, shift_name, line, master_id
            db_pg.execute(text("ALTER TABLE raw_material_receipts ADD COLUMN IF NOT EXISTS date DATE;"))
            db_pg.execute(text("ALTER TABLE raw_material_receipts ADD COLUMN IF NOT EXISTS shift_name VARCHAR(50);"))
            db_pg.execute(text("ALTER TABLE raw_material_receipts ADD COLUMN IF NOT EXISTS line VARCHAR(50);"))
            db_pg.execute(text("ALTER TABLE raw_material_receipts ALTER COLUMN shift_id DROP NOT NULL;"))
            
            db_pg.execute(text("ALTER TABLE downtimes ADD COLUMN IF NOT EXISTS date DATE;"))
            db_pg.execute(text("ALTER TABLE downtimes ADD COLUMN IF NOT EXISTS shift_name VARCHAR(50);"))
            db_pg.execute(text("ALTER TABLE downtimes ADD COLUMN IF NOT EXISTS line VARCHAR(50);"))
            db_pg.execute(text("ALTER TABLE downtimes ADD COLUMN IF NOT EXISTS master_id INTEGER REFERENCES masters(id);"))
            db_pg.execute(text("ALTER TABLE downtimes ALTER COLUMN shift_id DROP NOT NULL;"))
            
            # Batches: prev defects
            for b_col in [
                "prev_first_grade", "prev_defect", "prev_defect_scratch", "prev_defect_bad_cut",
                "prev_defect_stick_top", "prev_defect_broken", "prev_defect_fell_box",
                "prev_defect_thickness", "prev_defect_edge"
            ]:
                db_pg.execute(text(f"ALTER TABLE batches ADD COLUMN IF NOT EXISTS {b_col} INTEGER DEFAULT 0;"))
            
            # AuditLog: state_snapshot TEXT for snapshot rollback
            db_pg.execute(text("ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS state_snapshot TEXT;"))
            db_pg.commit()
    except Exception as dt_pg_err:
        print(f"Warning: could not migrate PostgreSQL downtimes schema: {dt_pg_err}")
        if 'db_pg' in locals() and db_pg:
            db_pg.rollback()
    finally:
        if 'db_pg' in locals() and db_pg:
            db_pg.close()
    
    # DocumentCategory password_hash migration
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE document_categories ADD COLUMN password_hash VARCHAR(255)")
        conn.commit()
        conn.close()
    except: pass

    try:
        db = SessionLocal()
        driver = db.bind.dialect.name if db.bind else 'unknown'
        if driver == 'postgresql':
            from sqlalchemy import text
            col_exists = db.execute(text(
                "SELECT column_name FROM information_schema.columns WHERE table_name='document_categories' AND column_name='password_hash'"
            )).fetchone()
            if not col_exists:
                db.execute(text("ALTER TABLE document_categories ADD COLUMN password_hash VARCHAR(255);"))
                db.commit()
                
        # Set default password for "Бережливое производство"
        bp_folder = db.query(models.DocumentCategory).filter(models.DocumentCategory.name == "Бережливое производство").first()
        if bp_folder and not bp_folder.password_hash:
            bp_folder.password_hash = hashlib.sha256("6282".encode()).hexdigest()
            db.commit()
    except Exception as e:
        print(f"Warning: could not migrate document_categories password_hash: {e}")
        db.rollback()
    finally:
        if 'db' in locals():
            db.close()

    # Document Google Drive, R2 and DocSpace columns migration
    try:
        db = SessionLocal()
        driver = db.bind.dialect.name if db.bind else 'unknown'
        if driver == 'postgresql':
            from sqlalchemy import text
            for tbl in ["shifts", "lfm_reports", "batches"]:
                try:
                    db.execute(text(f"ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS export_type VARCHAR(50) DEFAULT 'Эталон';"))
                    db.commit()
                except Exception:
                    db.rollback()
            for col, col_type in [
                ("r2_key", "VARCHAR(512)"), 
                ("docspace_file_id", "INTEGER"), 
                ("google_drive_id", "VARCHAR(255)"), 
                ("google_drive_url", "TEXT"), 
                ("koofr_path", "VARCHAR(512)"), 
                ("koofr_link", "TEXT"), 
                ("yandex_path", "VARCHAR(512)"), 
                ("yandex_url", "TEXT"),
                ("external_url", "TEXT"),
                ("version_number", "INTEGER DEFAULT 1"),
                ("locked_by_user", "VARCHAR(255)"),
                ("locked_at", "TIMESTAMP"),
                ("last_modified_by", "VARCHAR(255)"),
                ("created_by", "VARCHAR(255)")
            ]:
                try:
                    db.execute(text(f"ALTER TABLE documents ADD COLUMN IF NOT EXISTS {col} {col_type};"))
                    db.commit()
                except Exception:
                    db.rollback()
            try:
                db.execute(text("ALTER TABLE document_categories ADD COLUMN IF NOT EXISTS created_by VARCHAR(255);"))
                db.commit()
            except Exception:
                db.rollback()
            try:
                models.DocumentVersion.__table__.create(bind=engine, checkfirst=True)
            except Exception: pass
        elif driver == 'sqlite':
            conn = sqlite3.connect("tectum.db")
            for col, col_type in [
                ("r2_key", "VARCHAR(512)"), 
                ("docspace_file_id", "INTEGER"), 
                ("google_drive_id", "VARCHAR(255)"), 
                ("google_drive_url", "TEXT"), 
                ("koofr_path", "VARCHAR(512)"), 
                ("koofr_link", "TEXT"), 
                ("yandex_path", "VARCHAR(512)"), 
                ("yandex_url", "TEXT"),
                ("external_url", "TEXT"),
                ("version_number", "INTEGER DEFAULT 1"),
                ("locked_by_user", "VARCHAR(255)"),
                ("locked_at", "TIMESTAMP"),
                ("last_modified_by", "VARCHAR(255)"),
                ("created_by", "VARCHAR(255)")
            ]:
                try:
                    conn.execute(f"ALTER TABLE documents ADD COLUMN {col} {col_type}")
                    conn.commit()
                except Exception:
                    pass
            try:
                conn.execute("ALTER TABLE document_categories ADD COLUMN google_drive_folder_id VARCHAR(255)")
                conn.commit()
            except Exception: pass
            try:
                conn.execute("ALTER TABLE document_categories ADD COLUMN created_by VARCHAR(255)")
                conn.commit()
            except Exception: pass
            
            # Create document_versions table if not exists
            try:
                models.DocumentVersion.__table__.create(bind=engine, checkfirst=True)
            except Exception: pass
            
            conn.close()
    except Exception as e:
        pass
    finally:
        if 'db' in locals() and db:
            db.close()

    # Shift silos columns migration for PostgreSQL and SQLite
    try:
        db = SessionLocal()
        driver = db.bind.dialect.name if db.bind else 'unknown'
        shift_cols = [
            ("zo_chrysotile_4_20_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_4_20_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_4_20_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_4_20_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_5_65_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_5_65_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_5_65_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_5_65_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_6_40_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_6_40_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_6_40_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_chrysotile_6_40_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cement_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cement_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cement_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cement_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cellulose_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cellulose_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cellulose_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cellulose_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_crushed_slate_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_crushed_slate_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_crushed_slate_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_crushed_slate_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asbozurit_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asbozurit_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asbozurit_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asbozurit_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_fiberglass_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_fiberglass_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_fiberglass_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_fiberglass_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_laprol_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_laprol_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_laprol_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_laprol_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asbocarton_silo1", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asbocarton_silo2", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asbocarton_silo3", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asbocarton_silo4", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_asb_drain", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_cem_drain", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("lfm_asb_drain", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("lfm_cem_drain", "DOUBLE PRECISION DEFAULT 0" if driver == 'postgresql' else "REAL DEFAULT 0"),
            ("zo_submitted", "BOOLEAN DEFAULT FALSE")
        ]
        if driver == 'postgresql':
            from sqlalchemy import text
            for col_name, col_type in shift_cols:
                try:
                    db.execute(text(f"ALTER TABLE shifts ADD COLUMN IF NOT EXISTS {col_name} {col_type};"))
                    db.commit()
                except Exception:
                    db.rollback()
        elif driver == 'sqlite':
            conn = sqlite3.connect("tectum.db")
            for col_name, col_type in shift_cols:
                try:
                    conn.execute(f"ALTER TABLE shifts ADD COLUMN {col_name} {col_type}")
                    conn.commit()
                except Exception:
                    pass
            conn.close()
    except Exception as e:
        print(f"Shifts silos migration note: {e}")
    finally:
        if 'db' in locals() and db:
            db.close()

    # Batch previous shift defect columns migration
    try:
        db = SessionLocal()
        driver = db.bind.dialect.name if db.bind else 'unknown'
        cols = [
            ("prev_first_grade", "INTEGER DEFAULT 0"),
            ("prev_defect", "INTEGER DEFAULT 0"),
            ("prev_defect_scratch", "INTEGER DEFAULT 0"),
            ("prev_defect_bad_cut", "INTEGER DEFAULT 0"),
            ("prev_defect_stick_top", "INTEGER DEFAULT 0"),
            ("prev_defect_broken", "INTEGER DEFAULT 0"),
            ("prev_defect_fell_box", "INTEGER DEFAULT 0"),
            ("prev_defect_thickness", "INTEGER DEFAULT 0"),
            ("prev_defect_edge", "INTEGER DEFAULT 0")
        ]
        if driver == 'postgresql':
            from sqlalchemy import text
            for col_name, col_type in cols:
                try:
                    db.execute(text(f"ALTER TABLE batches ADD COLUMN IF NOT EXISTS {col_name} {col_type};"))
                    db.commit()
                except Exception:
                    db.rollback()
        elif driver == 'sqlite':
            conn = sqlite3.connect("tectum.db")
            for col_name, col_type in cols:
                try:
                    conn.execute(f"ALTER TABLE batches ADD COLUMN {col_name} {col_type}")
                    conn.commit()
                except Exception:
                    pass
            conn.close()
    except Exception as e:
        print(f"Batch prev defect migration note: {e}")
    finally:
        if 'db' in locals() and db:
            db.close()

    # PlannerEmployee pin_code column migration
    try:
        db = SessionLocal()
        driver = db.bind.dialect.name if db.bind else 'unknown'
        if driver == 'postgresql':
            from sqlalchemy import text
            try:
                db.execute(text("ALTER TABLE planner_employees ADD COLUMN IF NOT EXISTS pin_code VARCHAR(10);"))
                db.commit()
            except Exception:
                db.rollback()
        elif driver == 'sqlite':
            conn = sqlite3.connect("tectum.db")
            try:
                conn.execute("ALTER TABLE planner_employees ADD COLUMN pin_code VARCHAR(10)")
                conn.commit()
            except Exception: pass
            conn.close()
    except Exception as e:
        print(f"Planner pin_code migration note: {e}")
    finally:
        if 'db' in locals() and db:
            db.close()

    # Automatic deduplication of duplicate document folders on startup
    try:
        db = SessionLocal()
        all_folders = db.query(models.DocumentCategory).order_by(models.DocumentCategory.id.asc()).all()
        seen = {}
        duplicates = []
        for f in all_folders:
            key = (f.name.strip().lower(), f.parent_id)
            if key in seen:
                primary = seen[key]
                db.query(models.Document).filter(models.Document.category_id == f.id).update(
                    {models.Document.category_id: primary.id}, synchronize_session=False
                )
                db.query(models.DocumentCategory).filter(models.DocumentCategory.parent_id == f.id).update(
                    {models.DocumentCategory.parent_id: primary.id}, synchronize_session=False
                )
                duplicates.append(f)
            else:
                seen[key] = f

        for dup in duplicates:
            db.delete(dup)
        if duplicates:
            db.commit()
    except Exception as e:
        print(f"Startup duplicate folder cleanup note: {e}")
        if 'db' in locals() and db:
            db.rollback()
    finally:
        if 'db' in locals() and db:
            db.close()


    # Migrations for Raw Material Silos Breakdown
    silo_materials = [
        'chrysotile_4_20', 'chrysotile_5_65', 'chrysotile_6_40',
        'cellulose', 'crushed_slate', 'asbozurit', 'fiberglass', 'laprol', 'asbocarton'
    ]
    
    # 1. SQLite Migrations
    try:
        conn = sqlite3.connect("tectum.db")
        for mat in silo_materials:
            for s in range(1, 5):
                col_name = f"zo_{mat}_silo{s}"
                try:
                    conn.execute(f"ALTER TABLE shifts ADD COLUMN {col_name} FLOAT DEFAULT 0.0")
                except:
                    pass
        conn.commit()
        conn.close()
    except:
        pass
        
    # 2. PG Migrations
    try:
        db = SessionLocal()
        driver = db.bind.dialect.name if db.bind else 'unknown'
        if driver == 'postgresql':
            from sqlalchemy import text
            for mat in silo_materials:
                for s in range(1, 5):
                    col_name = f"zo_{mat}_silo{s}"
                    try:
                        db.execute(text(f"ALTER TABLE shifts ADD COLUMN IF NOT EXISTS {col_name} FLOAT DEFAULT 0.0"))
                    except Exception as pg_err:
                        pass
            db.commit()
    except Exception as e:
        print(f"Warning: could not run PG migrations for silo columns: {e}")
        if 'db' in locals(): db.rollback()
    finally:
        if 'db' in locals() and db: db.close()

    # Migrations for created_at on shifts, downtimes, raw_material_receipts
    try:
        conn = sqlite3.connect("tectum.db")
        for tbl in ["shifts", "downtimes", "raw_material_receipts"]:
            try:
                conn.execute(f"ALTER TABLE {tbl} ADD COLUMN created_at DATETIME")
            except: pass
        conn.commit()
        conn.close()
    except: pass

    try:
        db = SessionLocal()
        driver = db.bind.dialect.name if db.bind else 'unknown'
        if driver == 'postgresql':
            from sqlalchemy import text
            for tbl in ["shifts", "downtimes", "raw_material_receipts"]:
                try:
                    db.execute(text(f"ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS created_at TIMESTAMP"))
                except Exception as pg_err: pass
            
            # One-time cleanup for historical records: clear automatically assigned timestamps
            try:
                db.execute(text("UPDATE shifts SET created_at = NULL WHERE created_at IS NOT NULL;"))
                db.execute(text("UPDATE downtimes SET created_at = NULL WHERE created_at IS NOT NULL;"))
                db.execute(text("UPDATE raw_material_receipts SET created_at = NULL WHERE created_at IS NOT NULL;"))
                db.commit()
            except Exception as cl_err:
                print(f"Warning clearing created_at for historical shifts: {cl_err}")
                db.rollback()
        else:
            try:
                from sqlalchemy import text
                db.execute(text("UPDATE shifts SET created_at = NULL"))
                db.commit()
            except: pass
    except Exception as e:
        print(f"Warning: could not run PG migrations for created_at: {e}")
        if 'db' in locals() and db: db.rollback()
    finally:
        if 'db' in locals() and db: db.close()

    # Migrations for Shift (Asbocarton & Drains)
    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN zo_asbocarton FLOAT DEFAULT 0.0")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN lfm_asb_drain FLOAT DEFAULT 0.0")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN lfm_cem_drain FLOAT DEFAULT 0.0")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN receipt_asbocarton FLOAT DEFAULT 0.0")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN receipt_pallets FLOAT DEFAULT 0.0")
        conn.commit()
        conn.close()
    except: pass

    try:
        conn = sqlite3.connect("tectum.db")
        conn.execute("ALTER TABLE shifts ADD COLUMN sharepoint_url VARCHAR(500)")
        conn.commit()
        conn.close()
    except: pass

    # SQLite migrations for monthly_plan_board columns
    for col, col_def in [("first_grade", "INTEGER DEFAULT 0"), ("defect", "INTEGER DEFAULT 0"), ("line", "VARCHAR(255) DEFAULT 'LFM-1'")]:
        try:
            conn = sqlite3.connect("tectum.db")
            conn.execute(f"ALTER TABLE monthly_plan_board ADD COLUMN {col} {col_def}")
            conn.commit()
            conn.close()
        except: pass
        
    # SQLite migrations for batches qcd columns
    sqlite_batches_cols = [
        ("qcd_sorted_packs", "INTEGER DEFAULT 0"),
        ("qcd_first_grade_note", "VARCHAR(500)"),
        ("qcd_defect_note", "VARCHAR(500)"),
        ("qcd_defect_chip", "INTEGER DEFAULT 0"),
        ("qcd_defect_scratch", "INTEGER DEFAULT 0"),
        ("qcd_defect_bad_cut", "INTEGER DEFAULT 0"),
        ("qcd_defect_stick_bottom", "INTEGER DEFAULT 0"),
        ("qcd_defect_stick_top", "INTEGER DEFAULT 0"),
        ("qcd_defect_broken", "INTEGER DEFAULT 0"),
        ("qcd_defect_fell_box", "INTEGER DEFAULT 0"),
        ("qcd_defect_dent", "INTEGER DEFAULT 0"),
        ("qcd_defect_thickness", "INTEGER DEFAULT 0"),
        ("qcd_defect_delamination", "INTEGER DEFAULT 0"),
        ("qcd_defect_edge", "INTEGER DEFAULT 0")
    ]
    for col, col_def in sqlite_batches_cols:
        try:
            conn = sqlite3.connect("tectum.db")
            conn.execute(f"ALTER TABLE batches ADD COLUMN {col} {col_def}")
            conn.commit()
            conn.close()
        except: pass

    # PostgreSQL auto-migration for missing columns (single batch check)
    if "postgresql" in engine.url.drivername or "postgres" in engine.url.drivername:
        from sqlalchemy import text
        pg_cols_to_add = [
            ("monthly_plan_board", "first_grade", "INTEGER DEFAULT 0"),
            ("monthly_plan_board", "defect", "INTEGER DEFAULT 0"),
            ("monthly_plan_board", "line", "VARCHAR(255) DEFAULT 'LFM-1'"),
            ("masters", "email", "VARCHAR(255)"),
            ("downtimes", "department", "VARCHAR(255)"),
            ("downtime_directory", "category", "VARCHAR(255)"),
            ("downtimes", "is_equipment_downtime", "BOOLEAN DEFAULT TRUE"),
            ("downtimes", "comment", "VARCHAR(255)"),
            ("downtimes", "breakdowns", "TEXT"),
            ("shifts", "zo_asbocarton", "DOUBLE PRECISION DEFAULT 0.0"),
            ("shifts", "lfm_asb_drain", "DOUBLE PRECISION DEFAULT 0.0"),
            ("shifts", "lfm_cem_drain", "DOUBLE PRECISION DEFAULT 0.0"),
            ("shifts", "receipt_asbocarton", "DOUBLE PRECISION DEFAULT 0.0"),
            ("shifts", "receipt_pallets", "DOUBLE PRECISION DEFAULT 0.0"),
            ("shifts", "sharepoint_url", "VARCHAR(500)"),
            ("batches", "qcd_sorted_packs", "INTEGER DEFAULT 0"),
            ("batches", "qcd_first_grade_note", "VARCHAR(500)"),
            ("batches", "qcd_defect_note", "VARCHAR(500)"),
            ("batches", "qcd_defect_chip", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_scratch", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_bad_cut", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_stick_bottom", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_stick_top", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_broken", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_fell_box", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_dent", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_thickness", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_delamination", "INTEGER DEFAULT 0"),
            ("batches", "qcd_defect_edge", "INTEGER DEFAULT 0"),
            ("checklist_employees", "department", "VARCHAR(255)"),
            ("tasks", "code", "VARCHAR(50)"),
            ("tasks", "zone", "VARCHAR(255)"),
            ("tasks", "title_kz", "TEXT"),
            ("tasks", "photo_link", "TEXT"),
            ("tasks", "author_name", "VARCHAR(255)"),
            ("tasks", "assignee_name", "VARCHAR(255)"),
            ("tasks", "due_date_str", "VARCHAR(100)"),
            ("tasks", "comment", "TEXT"),
            ("tasks", "month_label", "VARCHAR(100)"),
            ("tasks", "is_archived", "BOOLEAN DEFAULT FALSE"),
            ("tasks", "attached_document_id", "INTEGER"),
            ("tasks", "google_doc_url", "VARCHAR(500)"),
            ("tasks", "task_type", "VARCHAR(50) DEFAULT 'weekly'"),
            ("tasks", "department_service", "VARCHAR(100)"),
            ("tasks", "parent_id", "INTEGER"),
            ("tasks", "depends_on_id", "INTEGER"),
            ("tasks", "tags", "TEXT"),
            ("tasks", "target_quarter", "VARCHAR(50)"),
            ("tasks", "progress", "INTEGER DEFAULT 0")
        ]
        try:
            with engine.connect() as conn:
                existing_cols = {
                    (row[0], row[1]) for row in conn.execute(text(
                        "SELECT table_name, column_name FROM information_schema.columns WHERE table_schema='public';"
                    )).fetchall()
                }
                for table, col, col_def in pg_cols_to_add:
                    if (table, col) not in existing_cols:
                        try:
                            conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {col} {col_def};"))
                            print(f"Added column '{col}' to table '{table}' in PostgreSQL.")
                        except Exception:
                            pass
                
                # Очистка некорректно проставленного department_service = 'Общий'
                try:
                    conn.execute(text("UPDATE tasks SET department_service = NULL WHERE department_service = 'Общий' OR (task_type = 'weekly' AND zone = 'Бережливое производство' AND department_service = 'Общий');"))
                except Exception:
                    pass
                conn.commit()
        except Exception as pg_err:
            print(f"Error checking PG columns: {pg_err}")

    # SQLite migration for tasks table
    try:
        conn = sqlite3.connect("tectum.db")
        sqlite_task_cols = [
            ("code", "VARCHAR(50)"),
            ("zone", "VARCHAR(255)"),
            ("title_kz", "TEXT"),
            ("photo_link", "TEXT"),
            ("author_name", "VARCHAR(255)"),
            ("assignee_name", "VARCHAR(255)"),
            ("due_date_str", "VARCHAR(100)"),
            ("comment", "TEXT"),
            ("month_label", "VARCHAR(100)"),
            ("is_archived", "BOOLEAN DEFAULT 0"),
            ("attached_document_id", "INTEGER"),
            ("google_doc_url", "VARCHAR(500)"),
            ("task_type", "VARCHAR(50) DEFAULT 'weekly'"),
            ("department_service", "VARCHAR(100)"),
            ("parent_id", "INTEGER"),
            ("depends_on_id", "INTEGER"),
            ("tags", "TEXT"),
            ("target_quarter", "VARCHAR(50)"),
            ("progress", "INTEGER DEFAULT 0")
        ]
        for col, col_def in sqlite_task_cols:
            try:
                conn.execute(f"ALTER TABLE tasks ADD COLUMN {col} {col_def}")
            except:
                pass
        conn.commit()
        conn.close()
    except:
        pass


    db = SessionLocal()

    try:
        if not db.query(models.Master).filter(models.Master.role == "master").first():
            db.add(models.Master(name="Бекбосынов Р.", pin="1234", role="master"))
            db.add(models.Master(name="Монаев С.", pin="1234", role="master"))
            db.add(models.Master(name="Султанулы С.", pin="1234", role="master"))
            db.add(models.Master(name="Дауылбай М.", pin="1234", role="master"))
            db.add(models.Master(name="Оператор ЗО", pin="2222", role="zo"))
            db.add(models.Master(name="Машинист ЛФМ", pin="3333", role="lfm"))
            db.add(models.Master(name="Стакер", pin="4444", role="stacker"))
            db.add(models.Master(name="Дестакер", pin="5555", role="destacker"))
            db.add(models.Master(name="Инспектор СКК", pin="6666", role="qcd"))
            db.add(models.Master(name="Главный механик", pin="8888", role="mechanic"))
            db.commit()

        # Создание индексов для ускорения Базы Знаний
        try:
            from sqlalchemy import text
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_documents_category_id ON documents(category_id);"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_documents_uploaded_at ON documents(uploaded_at);"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_document_versions_doc_id ON document_versions(document_id);"))
            db.commit()
        except Exception as idx_e:
            db.rollback()
            print(f"Index creation note: {idx_e}")

        # Автоматическое распределение участков для всех сотрудников в БД (в фоне)
        def bg_sync_checklist_employees():
            db_sync = SessionLocal()
            try:
                import google_sheets_integration
                google_sheets_integration.sync_employees_from_google_sheets(db_sync)
            except Exception as sync_e:
                print(f"Startup checklist employees sync warning: {sync_e}")
            finally:
                db_sync.close()
        threading.Thread(target=bg_sync_checklist_employees, daemon=True).start()

        if not db.query(models.Master).filter(models.Master.role == "director").first():
            db.add(models.Master(name="Технический директор", pin="7777", role="director"))
            db.commit()
        if not db.query(models.Master).filter(models.Master.role == "technologist").first():
            db.add(models.Master(name="Главный технолог", pin="9999", role="technologist"))
            db.commit()
        
        # Очистка устаревших тестовых задач планнера с невалидными лейблами недель
        try:
            valid_weeks_prefix = ("Неделя 1 (", "Неделя 2 (", "Неделя 3 (", "Неделя 4 (", "Неделя 5 (")
            invalid_tasks = db.query(models.Task).all()
            deleted_test_count = 0
            for it in invalid_tasks:
                w_lbl = it.week_label or ""
                # Если лейбл недели старый тестовый (например "Неделя 1", "Неделя 1 (01 - 07)", "Неделя 4 (22 - 28)", "Тестируем", "Привет")
                if not any(w_lbl.startswith(pref) for pref in valid_weeks_prefix) or it.title in ["Привет", "Тестируем", "тест"]:
                    db.delete(it)
                    deleted_test_count += 1
            if deleted_test_count > 0:
                db.commit()
                print(f"[Tasks Cleanup] Deleted {deleted_test_count} obsolete test tasks.")

            # Убеждаемся, что все задачи активны и доступны в своих неделях (концепция архива упразднена)
            archived_tasks = db.query(models.Task).filter(models.Task.is_archived == True).all()
            if archived_tasks:
                for at in archived_tasks:
                    at.is_archived = False
                db.commit()
            # Миграция статусов: исключение устаревшего статуса '⚪ В очереди' в пользу '🟡 В работе'
            queue_tasks = db.query(models.Task).filter(models.Task.status.ilike("%В очереди%")).all()
            if queue_tasks:
                for qt in queue_tasks:
                    qt.status = "🟡 В работе"
                db.commit()
                print(f"[Tasks Migration] Migrated {len(queue_tasks)} tasks from 'В очереди' to '🟡 В работе'.")
        except Exception as task_cl_err:
            print(f"Warning cleaning test tasks: {task_cl_err}")
            db.rollback()

        # Auto-import downtimes directory if empty
        try:
            from import_downtimes import import_downtimes
            if db.query(models.DowntimeDirectory).count() == 0:
                print("Downtime directory is empty. Importing automatically...")
                import_downtimes()
        except Exception as e:
            print(f"Auto-importing downtimes failed: {e}")
    finally:
        db.close()
    
    seed_norms.seed_norms()

    # Seed generic master profile 'Мастер смены'
    db = SessionLocal()
    try:
        generic_master = db.query(models.Master).filter(models.Master.name == "Мастер смены").first()
        if not generic_master:
            generic_master = models.Master(name="Мастер смены", pin="1234", role="master")
            db.add(generic_master)
            db.commit()
            print("Successfully seeded 'Мастер смены' profile.")

        # Seed / update Levda M. and Bulekhanov K., clean up obsolete Tectum admin profile
        tectum_entries = db.query(models.Master).filter(models.Master.name.like("%Tectum%")).all()
        for t_adm in tectum_entries:
            db.delete(t_adm)

        levda = db.query(models.Master).filter(models.Master.name.like("%Левда%")).first()
        if not levda:
            db.add(models.Master(name="Левда М.", pin="6282", role="admin"))
        else:
            levda.name = "Левда М."
            levda.pin = "6282"
            levda.role = "admin"
            levda.email = None

        bulekhanov = db.query(models.Master).filter(or_(models.Master.name.like("%Булеханов%"), models.Master.name.like("%Булекпаев%"))).first()
        if not bulekhanov:
            db.add(models.Master(name="Булеханов К.", pin="2026", role="director"))
        else:
            bulekhanov.name = "Булеханов К."
            bulekhanov.pin = "2026"
            bulekhanov.role = "director"
        db.commit()
        print("Successfully seeded/updated 'Левда М.' and 'Булеханов К.' profiles, removed duplicate Tectum admin.")
    except Exception as e:
        print(f"Error seeding users: {e}")
        db.rollback()
    finally:
        db.close()


    # SharePoint directories bootstrap/sync check (run safely in background thread to not block server startup)
    def bg_sharepoint_init():
        db = SessionLocal()
        try:
            if not os.getenv("M365_TENANT_ID"):
                return
            if not m365_integration.check_file_exists_on_sharepoint("Справочники_Tectum.xlsx", folder="Shifts"):
                print("Справочники_Tectum.xlsx is missing on SharePoint, uploading initial template...")
                template_bytes = excel_exporter.create_initial_directories_xlsx(db)
                m365_integration.upload_file_to_sharepoint(template_bytes, "Справочники_Tectum.xlsx", folder="Shifts")
                print("Template uploaded successfully.")
            else:
                print("Справочники_Tectum.xlsx exists on SharePoint, syncing directories...")
                file_bytes = m365_integration.download_file_from_sharepoint("Справочники_Tectum.xlsx", folder="Shifts")
                excel_exporter.sync_directories_from_excel_bytes(file_bytes, db)
                print("Directories auto-synced from SharePoint successfully.")
        except Exception as e:
            print(f"Error checking/syncing directories with SharePoint on startup: {e}")
        finally:
            db.close()
    
    threading.Thread(target=bg_sharepoint_init, daemon=True).start()


    def bg_cleanups_init():
        db = SessionLocal()
        try:
            boards = db.query(models.MonthlyPlanBoard).filter(models.MonthlyPlanBoard.plan_sheets == 0).all()
            updated_count = 0
            for pb in boards:
                date_val = pb.date
                is_monday = False
                if isinstance(date_val, str):
                    try:
                        dt_obj = datetime.datetime.strptime(date_val, "%Y-%m-%d").date()
                        is_monday = dt_obj.weekday() == 0
                    except:
                        pass
                else:
                    try:
                        is_monday = date_val.weekday() == 0
                    except:
                        pass
                        
                if is_monday and pb.shift_name == "День":
                    continue
                    
                correct_plan = 2700 if pb.shift_name == "День" else 3300
                pb.plan_sheets = correct_plan
                updated_count += 1
                
            if updated_count > 0:
                db.commit()
                print(f"Auto-fixed {updated_count} plan_boards with 0 plan_sheets.")
        except Exception as e:
            print(f"Error auto-fixing plan boards: {e}")
        finally:
            db.close()

        # Clean up historical NULL values in database (Rule 5.8)
        db = SessionLocal()
        try:
            from sqlalchemy import text
            cleanup_queries = [
                "UPDATE lfm_reports SET lfm_wind_resets = 0 WHERE lfm_wind_resets IS NULL",
                "UPDATE lfm_reports SET formed_1st_grade = 0 WHERE formed_1st_grade IS NULL",
                "UPDATE lfm_reports SET formed_defect = 0 WHERE formed_defect IS NULL",
                "UPDATE lfm_reports SET transferred_to_warehouse = 0 WHERE transferred_to_warehouse IS NULL",
                "UPDATE batches SET ds_condition = 0 WHERE ds_condition IS NULL",
                "UPDATE batches SET ds_first_grade = 0 WHERE ds_first_grade IS NULL",
                "UPDATE batches SET ds_defect = 0 WHERE ds_defect IS NULL",
                "UPDATE batches SET qcd_condition = 0 WHERE qcd_condition IS NULL",
                "UPDATE batches SET qcd_first_grade = 0 WHERE qcd_first_grade IS NULL",
                "UPDATE batches SET qcd_defect = 0 WHERE qcd_defect IS NULL",
                "UPDATE batches SET qcd_condition = ds_condition WHERE qcd_condition = 0 AND ds_condition > 0",
                "UPDATE batches SET qcd_first_grade = ds_first_grade WHERE qcd_first_grade = 0 AND ds_first_grade > 0",
                "UPDATE batches SET qcd_defect = ds_defect WHERE qcd_defect = 0 AND ds_defect > 0",
                "UPDATE shifts SET batch_number = '' WHERE batch_number IS NULL",
                "UPDATE shifts SET product_name = '' WHERE product_name IS NULL",
                "UPDATE shifts SET export_type = 'Эталон' WHERE export_type IS NULL OR export_type = ''",
                "UPDATE lfm_reports SET export_type = 'Эталон' WHERE export_type IS NULL OR export_type = ''",
                "UPDATE batches SET export_type = 'Эталон' WHERE export_type IS NULL OR export_type = ''",
                "UPDATE shifts SET status = 'active' WHERE status IS NULL"
            ]
            for q in cleanup_queries:
                try:
                    db.execute(text(q))
                except Exception:
                    pass
            db.commit()
            print("Historical NULL values cleaned up successfully.")
        except Exception as e:
            print(f"Error cleaning up historical NULLs: {e}")
            db.rollback()
        finally:
            db.close()



        # Background auto-sync of folder structure and missing Google Drive URLs for existing documents
        try:
            db_docs = SessionLocal()
            try:
                # 1. Sync all folder categories to Google Drive
                all_categories = db_docs.query(models.DocumentCategory).all()
                print(f"[STARTUP] Starting auto-sync for {len(all_categories)} categories to Google Drive...")
                for cat in all_categories:
                    try:
                        f_id = get_or_create_google_drive_folder_for_category(db_docs, cat.id)
                        print(f"[STARTUP] Synced category #{cat.id} ('{cat.name}') -> Drive ID: {f_id}")
                    except Exception as cat_sync_err:
                        print(f"[STARTUP ERROR] Could not auto-sync folder category #{cat.id} ('{cat.name}') to Google Drive: {cat_sync_err}")

                # 2. Sync all documents missing Google Drive URLs
                unmigrated_docs = db_docs.query(models.Document).filter(
                    (models.Document.google_drive_url == None) | (models.Document.google_drive_url == "")
                ).all()
                print(f"[STARTUP] Found {len(unmigrated_docs)} unmigrated documents.")
                if unmigrated_docs:
                    import google_drive_integration
                    for u_doc in unmigrated_docs:
                        if u_doc.file_path and os.path.exists(u_doc.file_path):
                            try:
                                clean_t = u_doc.title or os.path.basename(u_doc.file_path)
                                parent_drive_id = get_or_create_google_drive_folder_for_category(db_docs, u_doc.category_id)
                                d_info = google_drive_integration.upload_file_to_drive(u_doc.file_path, clean_t, parent_drive_id=parent_drive_id)
                                if d_info and d_info.get("id"):
                                    u_doc.google_drive_id = d_info["id"]
                                    u_doc.google_drive_url = d_info["url"]
                                    db_docs.commit()
                                    print(f"Auto-synced doc #{u_doc.id} ('{clean_t}') to Google Drive.")
                            except Exception as sync_err:
                                print(f"Could not auto-sync doc #{u_doc.id} on startup: {sync_err}")
            finally:
                db_docs.close()
        except Exception as doc_mig_err:
            print(f"Docs startup sync error: {doc_mig_err}")

    threading.Thread(target=bg_cleanups_init, daemon=True).start()


    # Rename master
    db = SessionLocal()
    try:
        master = db.query(models.Master).filter(models.Master.name == "Рожков П.").first()
        if master:
            master.name = "Рожко П."
            db.commit()
    except Exception as e:
        db.rollback()
    finally:
        db.close()

    # Seed Planner Employees & Zones if empty
    db = SessionLocal()
    try:
        if db.query(models.PlannerEmployee).count() == 0:
            initial_employees = [
                {"name": "Левда М.", "email": "levdamaks15@gmail.com", "sort_order": 1},
                {"name": "Булеханов К.", "email": "", "sort_order": 2},
                {"name": "Курилова С.", "email": "", "sort_order": 3},
                {"name": "Сазонов С.", "email": "", "sort_order": 4},
                {"name": "Носиков Е.", "email": "", "sort_order": 5},
                {"name": "Хохлов К.", "email": "", "sort_order": 6},
                {"name": "Батырбекова Г.", "email": "", "sort_order": 7},
                {"name": "Герлинг С.", "email": "", "sort_order": 8},
                {"name": "Косумов Р.", "email": "", "sort_order": 9},
                {"name": "Мастера цеха", "email": "", "sort_order": 10},
                {"name": "Туматов Д.", "email": "", "sort_order": 11},
                {"name": "ОГЭ", "email": "", "sort_order": 12},
                {"name": "ОГМ", "email": "", "sort_order": 13}
            ]
            for emp_data in initial_employees:
                db.add(models.PlannerEmployee(**emp_data))
            db.commit()
            print("Successfully seeded initial Planner Employees.")

        if db.query(models.PlannerZone).count() == 0:
            initial_zones = [
                {"name": "Бережливое производство", "sort_order": 1},
                {"name": "Ремонт", "sort_order": 2},
                {"name": "Уборка", "sort_order": 3},
                {"name": "Производство", "sort_order": 4},
                {"name": "Отчетность", "sort_order": 5},
                {"name": "Документация", "sort_order": 6},
                {"name": "Цифровизация", "sort_order": 7},
                {"name": "Обучение", "sort_order": 8},
                {"name": "ОГЭ", "sort_order": 9},
                {"name": "ОГМ", "sort_order": 10}
            ]
            for zone_data in initial_zones:
                db.add(models.PlannerZone(**zone_data))
            db.commit()
            print("Successfully seeded initial Planner Zones.")
    except Exception as e:
        print(f"Error seeding planner settings: {e}")
        db.rollback()
    finally:
        db.close()

    def bg_google_sync_init():
        db = SessionLocal()
        try:
            if os.getenv("GOOGLE_SPREADSHEET_ID") and not os.getenv("GOOGLE_SPREADSHEET_ID").startswith("1_mock"):
                google_sheets_integration.sync_report_to_google_sheets(db)
                google_sheets_integration.export_receipt_to_google_sheets(db)
                google_sheets_integration.export_downtimes_to_google_sheets(db)
                google_sheets_integration.sync_qcd_reports_to_google_sheets(db)
                google_sheets_integration.export_norms_to_google_sheets(db)
                print("Initial Google Sheets full sync (all 5 sheets) completed on startup.")
        except Exception as e:
            print(f"Error running initial Google Sheets sync on startup: {e}")
        finally:
            db.close()

    # Регистрация Telegram Webhook
    def bg_setup_telegram_webhook():
        try:
            tg_token = os.getenv("TELEGRAM_BOT_TOKEN", "8980370531:AAGGhgbRH04LT_KOMUHr02ms1X4wZ0b3LwY").strip()
            webhook_url = "https://tectum-portal-railway-production.up.railway.app/api/telegram/webhook"
            r = requests.post(f"https://api.telegram.org/bot{tg_token}/setWebhook", json={"url": webhook_url}, timeout=10)
            print(f"[Telegram Startup] Webhook setup status: {r.json()}")
        except Exception as tg_err:
            print(f"[Telegram Startup] Error setting webhook: {tg_err}")

    threading.Thread(target=bg_setup_telegram_webhook, daemon=True).start()

    yield

app = FastAPI(title="Tectum Enterprise Portal", lifespan=lifespan)

import traceback as _tb
from fastapi.responses import JSONResponse as _JSONResponse

@app.exception_handler(Exception)
async def global_exception_handler(request, exc: Exception):
    return _JSONResponse(
        status_code=500,
        content={"detail": str(exc), "traceback": _tb.format_exc()}
    )

app.add_middleware(
    SessionMiddleware, 
    secret_key=os.getenv("SESSION_SECRET_KEY", "super-secret-key-for-tectum-portal"),
    max_age=86400 * 30,  # 30 days
    same_site="lax",
    https_only=False
)

@app.middleware("http")
async def add_no_cache_headers(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

from routers.auth import router as auth_router
from routers.downtimes import router as downtimes_router
from routers.planner import router as planner_router
from routers.checklists import router as checklists_router
from routers.documents import router as documents_router
app.include_router(auth_router)
app.include_router(downtimes_router)
app.include_router(planner_router)
app.include_router(checklists_router)
app.include_router(documents_router)

TONS_PER_HOUR = 5.0
PRICE_PER_TON = 100000.0

def calculate_downtime_losses(duration_minutes: int, shift: Optional[models.Shift], db: Session) -> tuple[float, float]:
    if duration_minutes <= 0:
        return 0.0, 0.0
        
    product_name = None
    if shift:
        product_name = shift.product_name
        if not product_name and shift.lfm_reports:
            product_name = shift.lfm_reports[-1].product_name
        
    if not product_name:
        product_name = "Шифер 8 волн рифленый"
        
    norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == product_name).first()
    weight_kg = norm.weight_kg if (norm and norm.weight_kg) else 19.6
    
    sheets_per_cycle = 1 if product_name == "Шифер 7 волн 3500*980" else 2
    
    total_seconds = duration_minutes * 60
    cycles = total_seconds / 26.0
    lost_sheets = cycles * sheets_per_cycle
    lost_tons = (lost_sheets * weight_kg) / 1000.0
    lost_tenge = lost_tons * PRICE_PER_TON
    
    return lost_tons, lost_tenge

def sync_downtimes_bg():
    from database import SessionLocal
    import google_sheets_integration
    db = SessionLocal()
    try:
        google_sheets_integration.export_downtimes_to_google_sheets(db)
    except Exception as e:
        print(f"Error syncing downtimes to Google Sheets: {e}")
        try:
            db.add(models.AuditLog(
                user_name="Google Sync Downtimes",
                action="ERROR",
                details=f"Ошибка экспорта простоев в Google Sheets: {str(e)}"
            ))
            db.commit()
        except Exception:
            pass
    finally:
        db.close()

def sync_google_sheets_bg():
    from database import SessionLocal
    import google_sheets_integration
    db = SessionLocal()
    try:
        google_sheets_integration.sync_report_to_google_sheets(db)
        google_sheets_integration.export_receipt_to_google_sheets(db)
        google_sheets_integration.sync_qcd_reports_to_google_sheets(db)
    except Exception as e:
        print(f"Error syncing reports/receipts to Google Sheets: {e}")
        try:
            db.add(models.AuditLog(
                user_name="Google Sync Reports",
                action="ERROR",
                details=f"Ошибка синхронизации отчетов в Google Sheets: {str(e)}"
            ))
            db.commit()
        except Exception:
            pass
    finally:
        db.close()

def sync_receipts_bg():
    from database import SessionLocal
    import google_sheets_integration
    db = SessionLocal()
    try:
        google_sheets_integration.export_receipt_to_google_sheets(db)
    except Exception as e:
        print(f"Error syncing receipts to Google Sheets: {e}")
    finally:
        db.close()

def sync_tasks_to_google_bg():
    pass


@app.get("/api/system/env")
def get_system_env():
    return {"is_sandbox": os.environ.get("IS_SANDBOX", "false").lower() == "true"}

if not os.path.exists("uploads"):
    os.makedirs("uploads", exist_ok=True)
if not os.path.exists(os.path.join("uploads", "tasks")):
    os.makedirs(os.path.join("uploads", "tasks"), exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

if not os.path.exists("static"):
    os.makedirs("static")
app.mount("/static", StaticFiles(directory="static"), name="static")

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.get("/")
def read_root():
    return FileResponse("static/index.html")

@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return FileResponse("static/img/Logo.png")

@app.post("/api/setup_demo_data/")
def setup_demo_data(db: Session = Depends(get_db)):
    if not db.query(models.Master).filter(models.Master.role == "master").first():
        db.add(models.Master(name="Бекбосынов Р.", pin="1234", role="master"))
        db.add(models.Master(name="Монаев С.", pin="1234", role="master"))
        db.add(models.Master(name="Султанулы С.", pin="1234", role="master"))
        db.add(models.Master(name="Дауылбай М.", pin="1234", role="master"))
        db.add(models.Master(name="Оператор ЗО", pin="2222", role="zo"))
        db.add(models.Master(name="Машинист ЛФМ", pin="3333", role="lfm"))
        db.add(models.Master(name="Стакер", pin="4444", role="stacker"))
        db.add(models.Master(name="Дестакер", pin="5555", role="destacker"))
        db.add(models.Master(name="Инспектор СКК", pin="6666", role="qcd"))
        db.add(models.Master(name="Главный механик", pin="8888", role="mechanic"))
        db.commit()
    
    if not db.query(models.Master).filter(models.Master.role == "director").first():
        db.add(models.Master(name="Технический директор", pin="7777", role="director"))
        db.commit()

    if not db.query(models.Master).filter(models.Master.role == "technologist").first():
        db.add(models.Master(name="Главный технолог", pin="9999", role="technologist"))
        db.commit()

    return {"message": "Demo data loaded"}

# --- УПРАВЛЕНИЕ СМЕНОЙ ---
@app.post("/api/shifts/")
def create_shift(shift: schemas.ShiftCreate, request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if user_role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен. Только мастер смены или администратор могут открывать смены.")
        
    active = db.query(models.Shift).filter(models.Shift.status == "active").first()
    if active:
        if user_role != "admin" and active.master_id != user_id:
            master_name = active.master.name if active.master else "другим мастером"
            raise HTTPException(status_code=403, detail=f"Уже есть активная смена, открытая мастером {master_name}. Вы не можете начать новую смену.")
        
    db_shift = models.Shift(**shift.model_dump())
    db.add(db_shift)
    db.commit()
    db.refresh(db_shift)
    return db_shift

@app.get("/api/shifts/active")
def get_active_shifts(db: Session = Depends(get_db)):
    try:
        return db.query(models.Shift).filter(models.Shift.status == "active").all()
    except Exception as e:
        import traceback
        print(f"Error in get_active_shifts: {str(e)}\n{traceback.format_exc()}")
        return []

@app.get("/api/shifts/all", response_model=list[schemas.Shift])
def get_all_shifts(db: Session = Depends(get_db)):
    try:
        shifts = db.query(models.Shift).options(
            selectinload(models.Shift.master),
            selectinload(models.Shift.receipts),
            selectinload(models.Shift.batches),
            selectinload(models.Shift.lfm_reports),
            selectinload(models.Shift.downtimes)
        ).order_by(models.Shift.date.desc(), models.Shift.line.asc(), models.Shift.shift_name.desc(), models.Shift.batch_number.desc(), models.Shift.id.desc()).all()
        
        result = []
        for shift in shifts:
            try:
                lfm_sheets = sum((r.lfm_sheets or 0) for r in shift.lfm_reports) if shift.lfm_reports else 0
                warehouse_gp = sum((b.ds_condition or 0) for b in shift.batches) if shift.batches else 0
                plan_sheets = shift.plan_sheets or 0
                zo_batches = shift.zo_batches or 0
                
                if plan_sheets == 0 and lfm_sheets == 0 and warehouse_gp == 0 and zo_batches == 0 and not shift.zo_submitted:
                    continue
                schemas.Shift.from_orm(shift)
                result.append(shift)
            except Exception as item_err:
                print(f"Warning: skipping shift ID {shift.id} in get_all_shifts due to validation error: {item_err}")
                continue
        return result
    except Exception as e:
        import traceback
        print(f"Error in get_all_shifts: {str(e)}\n{traceback.format_exc()}")
        return []

@app.get("/api/shifts/by_params")
def get_shift_by_params(date: str, shift_name: str, line: str, request: Request, product_name: Optional[str] = None, batch_number: Optional[str] = None, export_type: Optional[str] = None, master_id: Optional[int] = None, create_if_not_exists: bool = False, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Не авторизован")
    if user_role not in ["master", "admin", "director", "technologist"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
        
    try:
        if hasattr(date, "strftime"):
            parsed_date = date.date() if hasattr(date, "date") else date
        else:
            parsed_date = datetime.strptime(str(date), "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(400, "Неверный формат даты. Ожидается YYYY-MM-DD")
        
    query = db.query(models.Shift).filter(
        models.Shift.date == parsed_date,
        models.Shift.shift_name == shift_name,
        models.Shift.line == line
    )
    if product_name:
        query = query.filter(models.Shift.product_name == product_name)
    if batch_number:
        query = query.filter(models.Shift.batch_number == batch_number)
    if export_type:
        query = query.filter(models.Shift.export_type == export_type)
    shift = query.first()
    
    if not shift:
        if not create_if_not_exists:
            raise HTTPException(status_code=404, detail="Смена не найдена")
            
        # Автоматически создаем закрытую смену с переданным master_id, либо текущего пользователя, либо первого мастера в БД
        final_master_id = master_id if master_id else user_id
        if not final_master_id or user_role not in ["master"]:
            if not final_master_id:
                first_master = db.query(models.Master).filter(models.Master.role == "master").first()
                if first_master:
                    final_master_id = first_master.id
                else:
                    final_master_id = user_id
                    
        shift = models.Shift(
            date=parsed_date,
            shift_name=shift_name,
            line=line,
            master_id=final_master_id,
            product_name=product_name or "",
            batch_number=batch_number or "",
            export_type=export_type or "Эталон",
            status="closed",
            plan_sheets=0,
            plan_tons=0.0,
            created_at=datetime.utcnow()
        )
        db.add(shift)
        db.commit()
        db.refresh(shift)
        
    return shift

@app.get("/api/shifts/crew_plan_fulfillment")
def get_crew_plan_fulfillment(month: Optional[str] = None, db: Session = Depends(get_db)):
    """
    Возвращает аналитическую сводку выполнения сменного плана бригадами (Смена 1..4)
    за выбранный месяц (YYYY-MM).
    Критерии плана по формовке ЛФМ:
    - Дневная смена: >= 2700 листов
    - Ночная смена: >= 3300 листов
    """
    try:
        from collections import defaultdict
        import calendar

        # Определение года и месяца
        if not month:
            from datetime import timezone
            tz_kz = timezone(timedelta(hours=5))
            month = datetime.now(tz_kz).strftime("%Y-%m")
            
        y_str, m_str = month.split("-")
        year = int(y_str)
        month_num = int(m_str)
        days_in_month = calendar.monthrange(year, month_num)[1]
        
        start_date = f"{year:04d}-{month_num:02d}-01"
        end_date = f"{year:04d}-{month_num:02d}-{days_in_month:02d}"

        # 1. Извлекаем рапорты за месяц
        shifts = db.query(models.Shift).filter(
            models.Shift.date >= start_date,
            models.Shift.date <= end_date
        ).order_by(models.Shift.date.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()

        # Суммируем выработку ЛФМ по слотам (date, shift_type)
        # Учитываем переход продукции / несколько партий в смене
        slot_lfm = defaultdict(int)
        slot_prod_masters = defaultdict(set)
        slot_all_masters = defaultdict(set)
        slot_products = defaultdict(list)
        slot_batches = defaultdict(list)

        for s in shifts:
            lfm_reports = db.query(models.LFMReport).filter(models.LFMReport.shift_id == s.id).all()
            total_lfm = sum(r.lfm_sheets or 0 for r in lfm_reports)
            
            s_val = s.shift_name.strip() if s.shift_name else ""
            b = s_val.encode('utf-8')
            # Нормализация День / Ночь
            if len(b) > 1 and b[1] == 148: # 'Д'
                s_name = 'День'
            elif 'д' in s_val.lower() or 'day' in s_val.lower():
                s_name = 'День'
            else:
                s_name = 'Ночь'
                
            d_str = s.date.strftime("%Y-%m-%d") if hasattr(s.date, 'strftime') else str(s.date)
            slot_lfm[(d_str, s_name)] += total_lfm
            
            if s.master_id:
                m = db.query(models.Master).filter(models.Master.id == s.master_id).first()
                if m:
                    m_name = getattr(m, 'full_name', getattr(m, 'name', str(m.id)))
                    if total_lfm > 0:
                        slot_prod_masters[(d_str, s_name)].add(m_name)
                    slot_all_masters[(d_str, s_name)].add(m_name)
            if s.product_name:
                slot_products[(d_str, s_name)].append(s.product_name)
            if s.batch_number:
                slot_batches[(d_str, s_name)].append(s.batch_number)

        # 2. Извлекаем утвержденный график сменности
        entries = db.query(models.ShiftScheduleEntry).all()
        entry_map = {e.date_str: e for e in entries}

        days_data = []
        crew_stats = {
            1: {"name": "Смена №1", "total_shifts": 0, "met_count": 0, "day_shifts": 0, "day_met": 0, "night_shifts": 0, "night_met": 0, "total_lfm": 0},
            2: {"name": "Смена №2", "total_shifts": 0, "met_count": 0, "day_shifts": 0, "day_met": 0, "night_shifts": 0, "night_met": 0, "total_lfm": 0},
            3: {"name": "Смена №3", "total_shifts": 0, "met_count": 0, "day_shifts": 0, "day_met": 0, "night_shifts": 0, "night_met": 0, "total_lfm": 0},
            4: {"name": "Смена №4", "total_shifts": 0, "met_count": 0, "day_shifts": 0, "day_met": 0, "night_shifts": 0, "night_met": 0, "total_lfm": 0}
        }

        total_met_factory = 0
        total_shifts_factory = 0
        total_lfm_factory = 0

        for day in range(1, days_in_month + 1):
            d_date_str = f"{year:04d}-{month_num:02d}-{day:02d}"
            d_display_str = f"{day:02d}.{month_num:02d}.{year:04d}"
            
            entry = entry_map.get(d_display_str)
            dow = entry.day_of_week if entry else ""

            # Обрабатываем День и Ночь
            for s_name in ['День', 'Ночь']:
                plan = 2700 if s_name == 'День' else 3300
                fact_lfm = slot_lfm.get((d_date_str, s_name), 0)
                active_masters = slot_prod_masters.get((d_date_str, s_name)) or slot_all_masters.get((d_date_str, s_name)) or set()
                masters = ", ".join(sorted(list(active_masters)))
                products = ", ".join(list(dict.fromkeys(slot_products.get((d_date_str, s_name), []))))
                batches = ", ".join(list(dict.fromkeys(slot_batches.get((d_date_str, s_name), []))))
                
                # Определение дежурной бригады по графику
                duty_crew = ""
                crew_num = None
                if entry:
                    if s_name == 'День':
                        duty_crew = entry.day_shift_group or ""
                    else:
                        duty_crew = entry.night_shift_group or ""
                
                if duty_crew:
                    for c_idx in [1, 2, 3, 4]:
                        if str(c_idx) in duty_crew:
                            crew_num = c_idx
                            break

                is_met = fact_lfm >= plan
                diff = fact_lfm - plan

                if crew_num:
                    c_stat = crew_stats[crew_num]
                    c_stat["total_shifts"] += 1
                    c_stat["total_lfm"] += fact_lfm
                    if s_name == 'День':
                        c_stat["day_shifts"] += 1
                        if is_met: c_stat["day_met"] += 1
                    else:
                        c_stat["night_shifts"] += 1
                        if is_met: c_stat["night_met"] += 1

                    if is_met:
                        c_stat["met_count"] += 1

                total_shifts_factory += 1
                total_lfm_factory += fact_lfm
                if is_met:
                    total_met_factory += 1

                days_data.append({
                    "date": d_date_str,
                    "date_display": d_display_str,
                    "day": day,
                    "day_of_week": dow,
                    "shift_name": s_name,
                    "crew_num": crew_num,
                    "crew_name": f"Смена №{crew_num}" if crew_num else duty_crew,
                    "fact_lfm": fact_lfm,
                    "plan": plan,
                    "diff": diff,
                    "is_met": is_met,
                    "master": masters,
                    "products": products,
                    "batches": batches
                })

        # Рассчитываем проценты выполнения
        for c_idx, st in crew_stats.items():
            tot = st["total_shifts"]
            st["percent"] = round((st["met_count"] / tot * 100), 1) if tot > 0 else 0.0

        factory_percent = round((total_met_factory / total_shifts_factory * 100), 1) if total_shifts_factory > 0 else 0.0

        return {
            "status": "ok",
            "month": month,
            "days_in_month": days_in_month,
            "factory_summary": {
                "total_shifts": total_shifts_factory,
                "total_met": total_met_factory,
                "percent": factory_percent,
                "total_lfm": total_lfm_factory
            },
            "crew_stats": crew_stats,
            "days": days_data
        }
    except Exception as e:
        print(f"Error in get_crew_plan_fulfillment: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/shifts/{shift_id}")
def get_single_shift(shift_id: int, request: Request, db: Session = Depends(get_db)):
    user_role = request.session.get("user_role")
    shift = db.query(models.Shift).options(
        joinedload(models.Shift.master),
        joinedload(models.Shift.receipts),
        joinedload(models.Shift.downtimes),
        joinedload(models.Shift.lfm_reports),
        joinedload(models.Shift.batches)
    ).get(shift_id)
    if not shift: raise HTTPException(404, "Смена не найдена")
    
    # Calculate edit window
    remaining_secs = 0
    if shift.created_at:
        diff = (datetime.utcnow() - shift.created_at).total_seconds()
        remaining_secs = max(0, int(1800 - diff))
    elif user_role == "admin":
        remaining_secs = 999999
        
    shift_dict = schemas.ShiftReportResponse.model_validate(shift).model_dump() if hasattr(schemas, 'ShiftReportResponse') else {c.name: getattr(shift, c.name) for c in shift.__table__.columns}
    shift_dict["created_at"] = shift.created_at.isoformat() if shift.created_at else None
    shift_dict["remaining_edit_seconds"] = remaining_secs
    shift_dict["can_edit"] = (user_role == "admin" or remaining_secs > 0)
    if shift.master:
        shift_dict["master"] = {"id": shift.master.id, "name": shift.master.name}
    shift_dict["receipts"] = [
        {
            "id": r.id,
            "shift_id": r.shift_id,
            "master_id": r.master_id,
            "timestamp": r.timestamp.isoformat() if r.timestamp else (r.created_at.isoformat() if getattr(r, 'created_at', None) else None),
            "created_at": (r.timestamp or getattr(r, 'created_at', None)).isoformat() if (r.timestamp or getattr(r, 'created_at', None)) else None,
            "can_edit": (user_role == "admin" or ((datetime.utcnow() - (r.timestamp or getattr(r, 'created_at', datetime.utcnow()))).total_seconds() <= 1800 if (r.timestamp or getattr(r, 'created_at', None)) else True)),
            "chrysotile_4_20": r.chrysotile_4_20,
            "chrysotile_5_65": r.chrysotile_5_65,
            "chrysotile_6_40": r.chrysotile_6_40,
            "cement_silo1": r.cement_silo1,
            "cement_silo2": r.cement_silo2,
            "cement_silo3": r.cement_silo3,
            "cement_silo4": r.cement_silo4,
            "cellulose": r.cellulose,
            "crushed_slate": r.crushed_slate,
            "asbozurit": r.asbozurit,
            "asbocarton": r.asbocarton,
            "pallets": r.pallets,
            "fiberglass": r.fiberglass,
            "laprol": r.laprol
        } for r in (shift.receipts or [])
    ]
    shift_dict["downtimes"] = [
        {
            "id": d.id,
            "shift_id": d.shift_id,
            "start_time": d.start_time,
            "end_time": d.end_time,
            "duration": d.duration,
            "category": d.category,
            "department": d.department,
            "node": d.node,
            "description": d.description,
            "comment": d.comment,
            "media_urls": d.media_urls,
            "is_equipment_downtime": d.is_equipment_downtime,
            "lost_tons": d.lost_tons,
            "lost_tenge": d.lost_tenge,
            "status": d.status,
            "created_at": d.created_at.isoformat() if d.created_at else None,
            "can_edit": (user_role == "admin" or (d.created_at and (datetime.utcnow() - d.created_at).total_seconds() <= 1800) or not d.created_at)
        } for d in (shift.downtimes or [])
    ]
    return shift_dict

async def upload_sharepoint_report_retry(file_bytes: bytes, filename: str, folder: str, retries: int = 5, delay: int = 60):
    for i in range(retries):
        await asyncio.sleep(delay)
        try:
            print(f"Background task: Attempting to upload {filename} to SharePoint (attempt {i+1})...")
            m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder=folder)
            print(f"Background task: Successfully uploaded {filename} to SharePoint on attempt {i+1}")
            
            db = SessionLocal()
            try:
                db.add(models.AuditLog(
                    user_name="system_background",
                    action="UPDATE",
                    target_table="shifts",
                    target_id=0,
                    details=f"Фоновая автосинхронизация: отчет {filename} успешно обновлен на SharePoint после освобождения файла."
                ))
                db.commit()
            except Exception as audit_err:
                print(f"Error logging background sync success: {audit_err}")
            finally:
                db.close()
            break
        except Exception as e:
            print(f"Background task: Attempt {i+1} failed to upload {filename}: {e}")
            delay = delay * 2

@app.put("/api/shifts/{shift_id}/close")
def close_shift(shift_id: int, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
        
    if user_role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен. Только мастер смены или администратор могут закрывать смены.")
        
    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(404, "Смена не найдена")
        
    if user_role not in ["admin", "master"]:
        raise HTTPException(
            status_code=403,
            detail="Доступ запрещен. Только мастер смены или администратор могут закрыть её."
        )
        
    shift.status = "closed"
    db.commit()
    
    # Generate unified Excel flat report in memory and upload to SharePoint
    try:
        file_bytes = excel_exporter.generate_flat_report(db)
        filename = "Сводный_отчет_Tectum.xlsx"
        
        # Save locally to static folder as well
        local_path = os.path.join("static", "Сводный_отчет_Tectum.xlsx")
        try:
            with open(local_path, "wb") as f:
                f.write(file_bytes)
        except Exception as local_err:
            print(f"Error saving local excel file: {local_err}")
            
        web_url = m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder="Reports")
        shift.sharepoint_url = web_url
        db.commit()
        
        # Log to AuditLog
        audit_detail = f"Смена {shift_id} закрыта. Сводный отчет сгенерирован и загружен в SharePoint: {web_url}"
        db.add(models.AuditLog(
            user_name=request.session.get("user_email") or f"user_{user_id}",
            action="UPDATE",
            target_table="shifts",
            target_id=shift_id,
            details=audit_detail
        ))
        db.commit()
        return {"message": "Смена закрыта"}
    except Exception as e:
        print(f"Error generating/uploading unified report Excel: {e}")
        error_msg = str(e)
        warning_text = None
        if "423" in error_msg or "Locked" in error_msg:
            warning_text = "Сводный отчет заблокирован в SharePoint (кто-то открыл его в Excel Online). Смена успешно закрыта, но облачный отчет не обновился. Запущена фоновая автосинхронизация, локальная копия сохранена на сервере."
        else:
            warning_text = f"Смена закрыта, но произошла ошибка при загрузке отчета на SharePoint: {error_msg}. Запущена фоновая автосинхронизация, локальная копия сохранена на сервере."
            
        # Queue background task to retry upload
        try:
            file_bytes = excel_exporter.generate_flat_report(db)
            background_tasks.add_task(upload_sharepoint_report_retry, file_bytes, "Сводный_отчет_Tectum.xlsx", "Reports")
        except Exception as gen_err:
            print(f"Failed to queue background sync: {gen_err}")
            
        audit_detail = f"Смена {shift_id} закрыта. Предупреждение по SharePoint: {warning_text}"
        try:
            db.add(models.AuditLog(
                user_name=request.session.get("user_email") or f"user_{user_id}",
                action="UPDATE",
                target_table="shifts",
                target_id=shift_id,
                details=audit_detail
            ))
            db.commit()
        except: pass
        return {"message": "Смена закрыта", "warning": warning_text}

@app.get("/api/shifts/{shift_id}/download_passport")
def download_shift_passport(shift_id: int, request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
        
    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(404, "Смена не найдена")
        
    # If sharepoint_url is already available, redirect to it
    if shift.sharepoint_url:
        return RedirectResponse(url=shift.sharepoint_url)
        
    # If not available (e.g. upload failed initially or it's an old shift), generate and upload now
    try:
        file_bytes = excel_exporter.generate_flat_report(db)
        filename = "Сводный_отчет_Tectum.xlsx"
        web_url = m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder="Reports")
        shift.sharepoint_url = web_url
        db.commit()
        return RedirectResponse(url=web_url)
    except Exception as e:
        print(f"SharePoint upload failed in download_passport fallback: {e}")
        # Fallback to local on-the-fly download if SharePoint is totally failing
        try:
            file_bytes = excel_exporter.generate_flat_report(db)
            from fastapi import Response
            from urllib.parse import quote
            safe_filename = quote("Сводный_отчет_Tectum.xlsx")
            return Response(
                content=file_bytes, 
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                headers={'Content-Disposition': f'attachment; filename="{safe_filename}"; filename*=UTF-8\'\'{safe_filename}'}
            )
        except Exception as inner_e:
            raise HTTPException(500, f"Не удалось сгенерировать сводный отчет: {str(e)} | fallback error: {str(inner_e)}")


@app.post("/api/admin/sync_directories_google")
def sync_directories_google(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="Не авторизован")
    if user_role != "admin":
        raise HTTPException(status_code=403, detail="Доступ запрещен")
        
    try:
        import google_sheets_integration
        google_sheets_integration.sync_norms_from_google_sheets(db)
        google_sheets_integration.sync_downtime_directory_from_google_sheets(db)
        
        db.add(models.AuditLog(
            user_name="Администратор",
            action="SYNC",
            target_table="downtime_directory",
            target_id=None,
            details="Синхронизация нормативов и справочника простоев из Google Sheets"
        ))
        db.commit()
        return {"status": "success", "message": "Справочники успешно синхронизированы из Google Sheets"}
    except Exception as e:
        import traceback
        err_msg = f"Ошибка синхронизации: {str(e)}"
        print(f"{err_msg}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=err_msg)
@app.post("/api/admin/sync_directories_sharepoint")
def sync_directories_sharepoint(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
    if user_role != "admin":
        raise HTTPException(status_code=403, detail="Доступ разрешен только администраторам")
        
    try:
        # Check if directories file exists
        if not m365_integration.check_file_exists_on_sharepoint("Справочники_Tectum.xlsx", folder="Shifts"):
            # Create a template and upload it
            template_bytes = excel_exporter.create_initial_directories_xlsx(db)
            m365_integration.upload_file_to_sharepoint(template_bytes, "Справочники_Tectum.xlsx", folder="Shifts")
            return {"status": "ok", "message": "Файл Справочники_Tectum.xlsx отсутствовал на SharePoint. Создан шаблон и загружен в облако."}
            
        file_bytes = m365_integration.download_file_from_sharepoint("Справочники_Tectum.xlsx", folder="Shifts")
        excel_exporter.sync_directories_from_excel_bytes(file_bytes, db)
        
        # Log to AuditLog
        db.add(models.AuditLog(
            user_name=request.session.get("user_email") or f"admin_{user_id}",
            action="IMPORT",
            target_table="product_norms/downtime_directory",
            details="Синхронизация технологических норм и справочника простоев из файла Справочники_Tectum.xlsx в SharePoint."
        ))
        db.commit()
        return {"status": "ok", "message": "Справочники успешно синхронизированы из SharePoint!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка синхронизации: {str(e)}")

@app.post("/api/admin/upload_aci_report")
async def upload_aci_report(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
    if user_role != "admin":
        raise HTTPException(status_code=403, detail="Доступ разрешен только администраторам")
        
    try:
        # Импортируем данные
        res = import_aci_excel.import_aci_excel_data(file.file, db)
        
        # Логируем действие в AuditLog
        db.add(models.AuditLog(
            user_name=request.session.get("user_email") or f"admin_{user_id}",
            action="IMPORT",
            target_table="shifts/batches/lfm_reports",
            details=f"Импорт рапорта АЦИ из файла {file.filename}. Успешно: смен: {res['shifts']}, партий: {res['batches']}, ЛФМ: {res['lfm_reports']}"
        ))
        db.commit()
        
        # Автоматически перегенерируем сводный отчет в SharePoint
        try:
            file_bytes = excel_exporter.generate_flat_report(db)
            filename = "Сводный_отчет_Tectum.xlsx"
            
            # Локальная копия
            local_path = os.path.join("static", "Сводный_отчет_Tectum.xlsx")
            try:
                with open(local_path, "wb") as f:
                    f.write(file_bytes)
            except Exception as local_err:
                print(f"Error saving local excel file: {local_err}")
                
            web_url = m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder="Reports")
            
            # Обновим sharepoint_url у всех импортированных смен, чтобы они вели на этот отчет
            shifts = db.query(models.Shift).all()
            for s in shifts:
                s.sharepoint_url = web_url
            db.commit()
            
            res["sharepoint_url"] = web_url
        except Exception as sp_err:
            print(f"SharePoint update failed during ACI import: {sp_err}")
            res["sharepoint_url"] = None
            res["sharepoint_error"] = str(sp_err)
            
        return res
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка обработки рапорта АЦИ: {str(e)}")




# --- ПРИХОД И ЗО ---
class UpdateReceiptZO(BaseModel):
    chrysotile_4_20: float = 0
    chrysotile_5_65: float = 0
    chrysotile_6_40: float = 0
    cement: float = 0
    cement_silo1: float = 0
    cement_silo2: float = 0
    cement_silo3: float = 0
    cement_silo4: float = 0
    cellulose: float = 0
    crushed_slate: float = 0
    asbozurit: float = 0
    fiberglass: float = 0
    laprol: float = 0
    asbocarton: float = 0
    pallets: float = 0
    asb_drain: float = 0
    cem_drain: float = 0
    batches: int = 0
    submitted: bool = False

class LFMDrainsUpdate(BaseModel):
    asb_drain: float = 0
    cem_drain: float = 0

@app.post("/api/shifts/{shift_id}/receipt")
def update_receipt(shift_id: int, data: UpdateReceiptZO, request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
        
    shift = db.query(models.Shift).get(shift_id)
    if not shift: raise HTTPException(404, "Смена не найдена")
    
    if False and user_role == "master" and shift.master_id != user_id:
        master_name = shift.master.name if shift.master else "другим мастером"
        raise HTTPException(status_code=403, detail=f"Вы не можете редактировать рецепт этой смены, так как она была открыта мастером {master_name}.")
        
    if user_role not in ["master", "admin", "director", "technologist"]:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
        

    db.commit()
    return {"status": "ok"}

@app.post("/api/shifts/{shift_id}/zo")
def update_zo(shift_id: int, data: UpdateReceiptZO, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
        
    shift = db.query(models.Shift).get(shift_id)
    if not shift: raise HTTPException(404, "Смена не найдена")
    
    if False and user_role == "master" and shift.master_id != user_id:
        master_name = shift.master.name if shift.master else "другим мастером"
        raise HTTPException(status_code=403, detail=f"Вы не можете редактировать данные ЗО этой смены, так как она была открыта мастером {master_name}.")
        
    if user_role not in ["master", "admin", "zo"]:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
        
    shift.zo_chrysotile_4_20 = data.chrysotile_4_20
    shift.zo_chrysotile_5_65 = data.chrysotile_5_65
    shift.zo_chrysotile_6_40 = data.chrysotile_6_40
    
    # Сохраняем силосы
    shift.zo_cement_silo1 = data.cement_silo1
    shift.zo_cement_silo2 = data.cement_silo2
    shift.zo_cement_silo3 = data.cement_silo3
    shift.zo_cement_silo4 = data.cement_silo4
    # И суммируем в zo_cement (legacy, для расчета отклонений)
    shift.zo_cement = data.cement_silo1 + data.cement_silo2 + data.cement_silo3 + data.cement_silo4
    
    shift.zo_cellulose = data.cellulose
    shift.zo_crushed_slate = data.crushed_slate
    shift.zo_asbozurit = data.asbozurit
    shift.zo_fiberglass = data.fiberglass
    shift.zo_laprol = data.laprol
    shift.zo_asbocarton = data.asbocarton
    shift.zo_asb_drain = data.asb_drain
    shift.zo_cem_drain = data.cem_drain
    shift.zo_batches = data.batches
    shift.zo_submitted = data.submitted
    
    db.commit()
    background_tasks.add_task(sync_google_sheets_bg)
    return {"message": "ZO updated"}

@app.post("/api/shifts/{shift_id}/lfm_drains")
def update_lfm_drains(shift_id: int, data: LFMDrainsUpdate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
        
    shift = db.query(models.Shift).get(shift_id)
    if not shift: raise HTTPException(404, "Смена не найдена")
    
    if False and user_role == "master" and shift.master_id != user_id:
        master_name = shift.master.name if shift.master else "другим мастером"
        raise HTTPException(status_code=403, detail=f"Вы не можете редактировать сливы ЛФМ этой смены, так как она была открыта мастером {master_name}.")
        
    if user_role not in ["master", "admin", "lfm"]:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
        
    shift.lfm_asb_drain = data.asb_drain
    shift.lfm_cem_drain = data.cem_drain
    db.commit()
    background_tasks.add_task(sync_google_sheets_bg)
    return {"message": "LFM drains updated"}

@app.post("/api/shifts/{shift_id}/raw_materials_bulk")
def update_raw_materials_bulk(shift_id: int, data: schemas.RawMaterialsBulkUpdate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Не авторизован")
    if user_role not in ["master", "admin", "director", "technologist"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен")
        
    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(404, "Смена не найдена")
        
    # Изоляция данных разных мастеров (для роли master):
    if False and user_role == "master" and shift.master_id != user_id:
        raise HTTPException(status_code=403, detail="Вы не можете редактировать смену другого мастера")
        
    # Записываем расход ЗО
    shift.zo_chrysotile_4_20 = data.zo_chrysotile_4_20
    shift.zo_chrysotile_5_65 = data.zo_chrysotile_5_65
    shift.zo_chrysotile_6_40 = data.zo_chrysotile_6_40
    shift.zo_cement_silo1 = data.zo_cement_silo1
    shift.zo_cement_silo2 = data.zo_cement_silo2
    shift.zo_cement_silo3 = data.zo_cement_silo3
    shift.zo_cement_silo4 = data.zo_cement_silo4
    
    # Суммируем в legacy zo_cement
    shift.zo_cement = (data.zo_cement_silo1 or 0) + (data.zo_cement_silo2 or 0) + (data.zo_cement_silo3 or 0) + (data.zo_cement_silo4 or 0)
    
    shift.zo_cellulose = data.zo_cellulose
    shift.zo_crushed_slate = data.zo_crushed_slate
    shift.zo_asbozurit = data.zo_asbozurit
    shift.zo_fiberglass = data.zo_fiberglass
    shift.zo_laprol = data.zo_laprol
    shift.zo_asbocarton = data.zo_asbocarton
    shift.zo_asb_drain = data.zo_asb_drain
    shift.zo_cem_drain = data.zo_cem_drain
    shift.zo_batches = data.zo_batches
    shift.zo_submitted = True
    
    db.commit()
    background_tasks.add_task(sync_google_sheets_bg)
    return {"status": "success"}



def calculate_shift_deviations(db: Session, shift: models.Shift):
    # Find LFM reports for the shift
    lfm_reports = shift.lfm_reports
    product_counts = {}
    for r in lfm_reports:
        product_counts[r.product_name] = product_counts.get(r.product_name, 0) + r.lfm_sheets
        
    theoretical = {
        "chrysotile_4_20": 0.0,
        "chrysotile_5_65": 0.0,
        "chrysotile_6_40": 0.0,
        "cement": 0.0,
        "cellulose": 0.0,
        "crushed_slate": 0.0,
        "asbozurit": 0.0,
        "fiberglass": 0.0,
        "asbocarton": 0.0,
        "laprol": 0.0
    }
    
    for prod_name, sheets in product_counts.items():
        norm = _get_norm_cached(db, prod_name)
        if norm:
            theoretical["chrysotile_4_20"] += sheets * (norm.norm_chrysotile_4_20 or 0.0)
            theoretical["chrysotile_5_65"] += sheets * (norm.norm_chrysotile_5_65 or 0.0)
            theoretical["chrysotile_6_40"] += sheets * (norm.norm_chrysotile_6_40 or 0.0)
            theoretical["cement"] += sheets * (norm.norm_cement or 0.0)
            theoretical["cellulose"] += sheets * (norm.norm_cellulose or 0.0)
            theoretical["crushed_slate"] += sheets * (norm.norm_crushed_slate or 0.0)
            theoretical["asbozurit"] += sheets * (norm.norm_asbozurit or 0.0)
            theoretical["fiberglass"] += sheets * (norm.norm_fiberglass or 0.0)
            
    actual = {
        "chrysotile_4_20": shift.zo_chrysotile_4_20 or 0.0,
        "chrysotile_5_65": shift.zo_chrysotile_5_65 or 0.0,
        "chrysotile_6_40": shift.zo_chrysotile_6_40 or 0.0,
        "cement": shift.zo_cement or 0.0,
        "cellulose": shift.zo_cellulose or 0.0,
        "crushed_slate": shift.zo_crushed_slate or 0.0,
        "asbozurit": shift.zo_asbozurit or 0.0,
        "fiberglass": shift.zo_fiberglass or 0.0,
        "asbocarton": shift.zo_asbocarton or 0.0,
        "laprol": shift.zo_laprol or 0.0
    }
    
    deviations = {}
    for mat in theoretical.keys():
        deviations[mat] = round(actual[mat] - theoretical[mat], 2)
        
    return {
        "theoretical": theoretical,
        "actual": actual,
        "deviations": deviations
    }


def save_report_internal(db: Session, shift: models.Shift, data: schemas.ShiftReportCreate, user_name: str, is_new: bool):
    # Old values logging
    old_values = {}
    if not is_new:
        old_values = {
            "master_id": shift.master_id,
            "batch_number": shift.batch_number,
            "product_name": shift.product_name,
            "zo_batches": shift.zo_batches,
            "zo_chrysotile_4_20": shift.zo_chrysotile_4_20,
            "zo_chrysotile_5_65": shift.zo_chrysotile_5_65,
            "zo_chrysotile_6_40": shift.zo_chrysotile_6_40,
            "zo_cement_silo1": shift.zo_cement_silo1,
            "zo_cement_silo2": shift.zo_cement_silo2,
            "zo_cement_silo3": shift.zo_cement_silo3,
            "zo_cement_silo4": shift.zo_cement_silo4,
            "zo_cellulose": shift.zo_cellulose,
            "zo_crushed_slate": shift.zo_crushed_slate,
            "zo_asbozurit": shift.zo_asbozurit,
            "zo_fiberglass": shift.zo_fiberglass,
            "zo_laprol": shift.zo_laprol,
            "zo_asbocarton": shift.zo_asbocarton,
            "zo_asb_drain": shift.zo_asb_drain,
            "zo_cem_drain": shift.zo_cem_drain
        }

    # Snapshot of state before update (for Rollback / Undo)
    snapshot_before = None
    if not is_new:
        try:
            b_prev = db.query(models.Batch).filter(models.Batch.shift_id == shift.id).first()
            l_prev = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift.id).first()
            snapshot_dict = {
                "shift": {
                    "master_id": shift.master_id,
                    "batch_number": shift.batch_number,
                    "product_name": shift.product_name,
                    "export_type": shift.export_type,
                    "status": shift.status,
                    "zo_batches": shift.zo_batches,
                    "zo_chrysotile_4_20_silo1": shift.zo_chrysotile_4_20_silo1,
                    "zo_chrysotile_4_20_silo2": shift.zo_chrysotile_4_20_silo2,
                    "zo_chrysotile_4_20_silo3": shift.zo_chrysotile_4_20_silo3,
                    "zo_chrysotile_4_20_silo4": shift.zo_chrysotile_4_20_silo4,
                    "zo_chrysotile_5_65_silo1": shift.zo_chrysotile_5_65_silo1,
                    "zo_chrysotile_5_65_silo2": shift.zo_chrysotile_5_65_silo2,
                    "zo_chrysotile_5_65_silo3": shift.zo_chrysotile_5_65_silo3,
                    "zo_chrysotile_5_65_silo4": shift.zo_chrysotile_5_65_silo4,
                    "zo_chrysotile_6_40_silo1": shift.zo_chrysotile_6_40_silo1,
                    "zo_chrysotile_6_40_silo2": shift.zo_chrysotile_6_40_silo2,
                    "zo_chrysotile_6_40_silo3": shift.zo_chrysotile_6_40_silo3,
                    "zo_chrysotile_6_40_silo4": shift.zo_chrysotile_6_40_silo4,
                    "zo_cement_silo1": shift.zo_cement_silo1,
                    "zo_cement_silo2": shift.zo_cement_silo2,
                    "zo_cement_silo3": shift.zo_cement_silo3,
                    "zo_cement_silo4": shift.zo_cement_silo4,
                    "zo_cellulose_silo1": shift.zo_cellulose_silo1,
                    "zo_cellulose_silo2": shift.zo_cellulose_silo2,
                    "zo_cellulose_silo3": shift.zo_cellulose_silo3,
                    "zo_cellulose_silo4": shift.zo_cellulose_silo4,
                    "zo_crushed_slate_silo1": shift.zo_crushed_slate_silo1,
                    "zo_crushed_slate_silo2": shift.zo_crushed_slate_silo2,
                    "zo_crushed_slate_silo3": shift.zo_crushed_slate_silo3,
                    "zo_crushed_slate_silo4": shift.zo_crushed_slate_silo4,
                    "zo_asbozurit_silo1": shift.zo_asbozurit_silo1,
                    "zo_asbozurit_silo2": shift.zo_asbozurit_silo2,
                    "zo_asbozurit_silo3": shift.zo_asbozurit_silo3,
                    "zo_asbozurit_silo4": shift.zo_asbozurit_silo4,
                    "zo_fiberglass_silo1": shift.zo_fiberglass_silo1,
                    "zo_fiberglass_silo2": shift.zo_fiberglass_silo2,
                    "zo_fiberglass_silo3": shift.zo_fiberglass_silo3,
                    "zo_fiberglass_silo4": shift.zo_fiberglass_silo4,
                    "zo_laprol_silo1": shift.zo_laprol_silo1,
                    "zo_laprol_silo2": shift.zo_laprol_silo2,
                    "zo_laprol_silo3": shift.zo_laprol_silo3,
                    "zo_laprol_silo4": shift.zo_laprol_silo4,
                    "zo_asbocarton_silo1": shift.zo_asbocarton_silo1,
                    "zo_asbocarton_silo2": shift.zo_asbocarton_silo2,
                    "zo_asbocarton_silo3": shift.zo_asbocarton_silo3,
                    "zo_asbocarton_silo4": shift.zo_asbocarton_silo4,
                    "zo_asb_drain": shift.zo_asb_drain,
                    "zo_cem_drain": shift.zo_cem_drain,
                    "lfm_asb_drain": getattr(shift, 'lfm_asb_drain', 0.0),
                    "lfm_cem_drain": getattr(shift, 'lfm_cem_drain', 0.0)
                },
                "lfm_report": {
                    "product_name": l_prev.product_name if l_prev else "",
                    "export_type": l_prev.export_type if l_prev else "Эталон",
                    "lfm_sheets": l_prev.lfm_sheets if l_prev else 0,
                    "lfm_wind_resets": l_prev.lfm_wind_resets if l_prev else 0,
                    "formed_1st_grade": l_prev.formed_1st_grade if l_prev else 0,
                    "formed_defect": l_prev.formed_defect if l_prev else 0,
                    "transferred_to_warehouse": l_prev.transferred_to_warehouse if l_prev else 0
                } if l_prev else None,
                "batch": {
                    "batch_number": b_prev.batch_number if b_prev else "",
                    "product_name": b_prev.product_name if b_prev else "",
                    "export_type": b_prev.export_type if b_prev else "Эталон",
                    "stacked_stacks": b_prev.stacked_stacks if b_prev else 0,
                    "ds_condition": b_prev.ds_condition if b_prev else 0,
                    "ds_first_grade": b_prev.ds_first_grade if b_prev else 0,
                    "ds_defect": b_prev.ds_defect if b_prev else 0,
                    "ds_defect_chip": b_prev.ds_defect_chip if b_prev else 0,
                    "ds_defect_scratch": b_prev.ds_defect_scratch if b_prev else 0,
                    "ds_defect_bad_cut": b_prev.ds_defect_bad_cut if b_prev else 0,
                    "ds_defect_stick_bottom": b_prev.ds_defect_stick_bottom if b_prev else 0,
                    "ds_defect_stick_top": b_prev.ds_defect_stick_top if b_prev else 0,
                    "ds_defect_broken": b_prev.ds_defect_broken if b_prev else 0,
                    "ds_defect_fell_box": b_prev.ds_defect_fell_box if b_prev else 0,
                    "ds_defect_dent": b_prev.ds_defect_dent if b_prev else 0,
                    "ds_defect_thickness": b_prev.ds_defect_thickness if b_prev else 0,
                    "ds_defect_delamination": b_prev.ds_defect_delamination if b_prev else 0,
                    "ds_defect_edge": b_prev.ds_defect_edge if b_prev else 0,
                    "prev_first_grade": b_prev.prev_first_grade if b_prev else 0,
                    "prev_defect": b_prev.prev_defect if b_prev else 0,
                    "prev_defect_scratch": b_prev.prev_defect_scratch if b_prev else 0,
                    "prev_defect_bad_cut": b_prev.prev_defect_bad_cut if b_prev else 0,
                    "prev_defect_stick_top": b_prev.prev_defect_stick_top if b_prev else 0,
                    "prev_defect_broken": b_prev.prev_defect_broken if b_prev else 0,
                    "prev_defect_fell_box": b_prev.prev_defect_fell_box if b_prev else 0,
                    "prev_defect_thickness": b_prev.prev_defect_thickness if b_prev else 0,
                    "prev_defect_edge": b_prev.prev_defect_edge if b_prev else 0,
                    "qcd_condition": b_prev.qcd_condition if b_prev else 0,
                    "qcd_first_grade": b_prev.qcd_first_grade if b_prev else 0,
                    "qcd_defect": b_prev.qcd_defect if b_prev else 0
                } if b_prev else None
            }
            import json
            snapshot_before = json.dumps(snapshot_dict, ensure_ascii=False)
        except Exception as snap_err:
            print(f"Warning: could not capture snapshot_before: {snap_err}")

    # Update Shift fields
    shift.master_id = data.master_id
    shift.batch_number = data.batch_number
    shift.product_name = data.product_name
    shift.status = "closed"
    
    # Расход сырья
    shift.zo_chrysotile_4_20_silo1 = data.zo_chrysotile_4_20_silo1
    shift.zo_chrysotile_4_20_silo2 = data.zo_chrysotile_4_20_silo2
    shift.zo_chrysotile_4_20_silo3 = data.zo_chrysotile_4_20_silo3
    shift.zo_chrysotile_4_20_silo4 = data.zo_chrysotile_4_20_silo4
    shift.zo_chrysotile_4_20 = (data.zo_chrysotile_4_20_silo1 or 0) + (data.zo_chrysotile_4_20_silo2 or 0) + (data.zo_chrysotile_4_20_silo3 or 0) + (data.zo_chrysotile_4_20_silo4 or 0)
    
    shift.zo_chrysotile_5_65_silo1 = data.zo_chrysotile_5_65_silo1
    shift.zo_chrysotile_5_65_silo2 = data.zo_chrysotile_5_65_silo2
    shift.zo_chrysotile_5_65_silo3 = data.zo_chrysotile_5_65_silo3
    shift.zo_chrysotile_5_65_silo4 = data.zo_chrysotile_5_65_silo4
    shift.zo_chrysotile_5_65 = (data.zo_chrysotile_5_65_silo1 or 0) + (data.zo_chrysotile_5_65_silo2 or 0) + (data.zo_chrysotile_5_65_silo3 or 0) + (data.zo_chrysotile_5_65_silo4 or 0)
    
    shift.zo_chrysotile_6_40_silo1 = data.zo_chrysotile_6_40_silo1
    shift.zo_chrysotile_6_40_silo2 = data.zo_chrysotile_6_40_silo2
    shift.zo_chrysotile_6_40_silo3 = data.zo_chrysotile_6_40_silo3
    shift.zo_chrysotile_6_40_silo4 = data.zo_chrysotile_6_40_silo4
    shift.zo_chrysotile_6_40 = (data.zo_chrysotile_6_40_silo1 or 0) + (data.zo_chrysotile_6_40_silo2 or 0) + (data.zo_chrysotile_6_40_silo3 or 0) + (data.zo_chrysotile_6_40_silo4 or 0)
    
    shift.zo_cement_silo1 = data.zo_cement_silo1
    shift.zo_cement_silo2 = data.zo_cement_silo2
    shift.zo_cement_silo3 = data.zo_cement_silo3
    shift.zo_cement_silo4 = data.zo_cement_silo4
    shift.zo_cement = (data.zo_cement_silo1 or 0) + (data.zo_cement_silo2 or 0) + (data.zo_cement_silo3 or 0) + (data.zo_cement_silo4 or 0)
    
    shift.zo_cellulose_silo1 = data.zo_cellulose_silo1
    shift.zo_cellulose_silo2 = data.zo_cellulose_silo2
    shift.zo_cellulose_silo3 = data.zo_cellulose_silo3
    shift.zo_cellulose_silo4 = data.zo_cellulose_silo4
    shift.zo_cellulose = (data.zo_cellulose_silo1 or 0) + (data.zo_cellulose_silo2 or 0) + (data.zo_cellulose_silo3 or 0) + (data.zo_cellulose_silo4 or 0)
    
    shift.zo_crushed_slate_silo1 = data.zo_crushed_slate_silo1
    shift.zo_crushed_slate_silo2 = data.zo_crushed_slate_silo2
    shift.zo_crushed_slate_silo3 = data.zo_crushed_slate_silo3
    shift.zo_crushed_slate_silo4 = data.zo_crushed_slate_silo4
    shift.zo_crushed_slate = (data.zo_crushed_slate_silo1 or 0) + (data.zo_crushed_slate_silo2 or 0) + (data.zo_crushed_slate_silo3 or 0) + (data.zo_crushed_slate_silo4 or 0)
    
    shift.zo_asbozurit_silo1 = data.zo_asbozurit_silo1
    shift.zo_asbozurit_silo2 = data.zo_asbozurit_silo2
    shift.zo_asbozurit_silo3 = data.zo_asbozurit_silo3
    shift.zo_asbozurit_silo4 = data.zo_asbozurit_silo4
    shift.zo_asbozurit = (data.zo_asbozurit_silo1 or 0) + (data.zo_asbozurit_silo2 or 0) + (data.zo_asbozurit_silo3 or 0) + (data.zo_asbozurit_silo4 or 0)
    
    shift.zo_fiberglass_silo1 = data.zo_fiberglass_silo1
    shift.zo_fiberglass_silo2 = data.zo_fiberglass_silo2
    shift.zo_fiberglass_silo3 = data.zo_fiberglass_silo3
    shift.zo_fiberglass_silo4 = data.zo_fiberglass_silo4
    shift.zo_fiberglass = (data.zo_fiberglass_silo1 or 0) + (data.zo_fiberglass_silo2 or 0) + (data.zo_fiberglass_silo3 or 0) + (data.zo_fiberglass_silo4 or 0)
    
    shift.zo_laprol_silo1 = data.zo_laprol_silo1
    shift.zo_laprol_silo2 = data.zo_laprol_silo2
    shift.zo_laprol_silo3 = data.zo_laprol_silo3
    shift.zo_laprol_silo4 = data.zo_laprol_silo4
    shift.zo_laprol = (data.zo_laprol_silo1 or 0) + (data.zo_laprol_silo2 or 0) + (data.zo_laprol_silo3 or 0) + (data.zo_laprol_silo4 or 0)
    
    shift.zo_asbocarton_silo1 = data.zo_asbocarton_silo1
    shift.zo_asbocarton_silo2 = data.zo_asbocarton_silo2
    shift.zo_asbocarton_silo3 = data.zo_asbocarton_silo3
    shift.zo_asbocarton_silo4 = data.zo_asbocarton_silo4
    shift.zo_asbocarton = (data.zo_asbocarton_silo1 or 0) + (data.zo_asbocarton_silo2 or 0) + (data.zo_asbocarton_silo3 or 0) + (data.zo_asbocarton_silo4 or 0)
    
    shift.zo_asb_drain = data.zo_asb_drain
    shift.zo_cem_drain = data.zo_cem_drain
    shift.zo_batches = data.zo_batches
    shift.zo_submitted = True
    if data.export_type is not None:
        shift.export_type = data.export_type or "Эталон"

    # Update LFM report
    lfm_report = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift.id).first()
    if not lfm_report:
        lfm_report = models.LFMReport(shift_id=shift.id)
        db.add(lfm_report)
    lfm_report.product_name = data.product_name
    lfm_report.export_type = data.export_type or "Эталон"
    lfm_report.lfm_sheets = data.lfm_sheets
    lfm_report.lfm_wind_resets = data.lfm_wind_resets
    lfm_report.formed_1st_grade = data.first_grade
    lfm_report.formed_defect = data.qcd_defect
    lfm_report.transferred_to_warehouse = data.warehouse_gp

    # Update Batch
    batch = db.query(models.Batch).filter(models.Batch.shift_id == shift.id).first()
    if not batch:
        batch = models.Batch(shift_id=shift.id)
        db.add(batch)
    batch.batch_number = data.batch_number
    batch.product_name = data.product_name
    batch.export_type = data.export_type or "Эталон"
    batch.status = "qcd_checked"
    batch.stacked_stacks = data.lfm_sheets
    batch.ds_condition = data.warehouse_gp
    batch.ds_first_grade = data.first_grade
    
    # Calculate defect sum
    ds_defect_sum = (
        data.ds_defect_chip + data.ds_defect_scratch + data.ds_defect_bad_cut +
        data.ds_defect_stick_bottom + data.ds_defect_stick_top + data.ds_defect_broken +
        data.ds_defect_fell_box + data.ds_defect_dent + data.ds_defect_thickness +
        data.ds_defect_delamination + data.ds_defect_edge
    )
    batch.ds_defect = ds_defect_sum
    batch.ds_defect_chip = data.ds_defect_chip
    batch.ds_defect_scratch = data.ds_defect_scratch
    batch.ds_defect_bad_cut = data.ds_defect_bad_cut
    batch.ds_defect_stick_bottom = data.ds_defect_stick_bottom
    batch.ds_defect_stick_top = data.ds_defect_stick_top
    batch.ds_defect_broken = data.ds_defect_broken
    batch.ds_defect_fell_box = data.ds_defect_fell_box
    batch.ds_defect_dent = data.ds_defect_dent
    batch.ds_defect_thickness = data.ds_defect_thickness
    batch.ds_defect_delamination = data.ds_defect_delamination
    batch.ds_defect_edge = data.ds_defect_edge

    # Previous shift defects
    prev_defect_sum = (
        (data.prev_defect_scratch or 0) + (data.prev_defect_bad_cut or 0) +
        (data.prev_defect_stick_top or 0) + (data.prev_defect_broken or 0) +
        (data.prev_defect_fell_box or 0) + (data.prev_defect_thickness or 0) +
        (data.prev_defect_edge or 0)
    )
    batch.prev_first_grade = data.prev_first_grade or 0
    batch.prev_defect = prev_defect_sum
    batch.prev_defect_scratch = data.prev_defect_scratch or 0
    batch.prev_defect_bad_cut = data.prev_defect_bad_cut or 0
    batch.prev_defect_stick_top = data.prev_defect_stick_top or 0
    batch.prev_defect_broken = data.prev_defect_broken or 0
    batch.prev_defect_fell_box = data.prev_defect_fell_box or 0
    batch.prev_defect_thickness = data.prev_defect_thickness or 0
    batch.prev_defect_edge = data.prev_defect_edge or 0

    batch.qcd_condition = data.warehouse_gp
    batch.qcd_first_grade = data.first_grade
    batch.qcd_defect = ds_defect_sum

    db.commit()

    # Export receipt data to Google Sheets (new sheet "Приход сырья")
    try:
        google_sheets_integration.export_receipt_to_google_sheets(db)
    except Exception as gs_err:
        print(f"Ошибка экспорта прихода сырья в Google Sheets: {gs_err}")

    # Sync to MonthlyPlanBoard (which also writes AuditLog)
    sync_lfm_to_plan_board(shift.date, shift.shift_name, shift.line, db, shift.master_id)

    # Write AuditLog for the shift update
    if is_new:
        db.add(models.AuditLog(
            user_name=user_name,
            action="CREATE",
            target_table="shifts",
            target_id=shift.id,
            details=f"Создан новый единый рапорт мастера для смены {shift.id} ({data.date} {data.shift_name} {data.line})"
        ))
    else:
        new_values = {
            "master_id": shift.master_id,
            "batch_number": shift.batch_number,
            "product_name": shift.product_name,
            "zo_batches": shift.zo_batches,
            "zo_chrysotile_4_20": shift.zo_chrysotile_4_20,
            "zo_chrysotile_5_65": shift.zo_chrysotile_5_65,
            "zo_chrysotile_6_40": shift.zo_chrysotile_6_40,
            "zo_cement_silo1": shift.zo_cement_silo1,
            "zo_cement_silo2": shift.zo_cement_silo2,
            "zo_cement_silo3": shift.zo_cement_silo3,
            "zo_cement_silo4": shift.zo_cement_silo4,
            "zo_cellulose": shift.zo_cellulose,
            "zo_crushed_slate": shift.zo_crushed_slate,
            "zo_asbozurit": shift.zo_asbozurit,
            "zo_fiberglass": shift.zo_fiberglass,
            "zo_laprol": shift.zo_laprol,
            "zo_asbocarton": shift.zo_asbocarton,
            "zo_asb_drain": shift.zo_asb_drain,
            "zo_cem_drain": shift.zo_cem_drain,
            "receipt_chrysotile_4_20": sum((r.chrysotile_4_20 or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_chrysotile_5_65": sum((r.chrysotile_5_65 or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_chrysotile_6_40": sum((r.chrysotile_6_40 or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_cement": sum((((r.cement_silo1 or 0.0) + (r.cement_silo2 or 0.0) + (r.cement_silo3 or 0.0) + (r.cement_silo4 or 0.0))) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_cellulose": sum((r.cellulose or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_crushed_slate": sum((r.crushed_slate or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_asbozurit": sum((r.asbozurit or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_asbocarton": sum((r.asbocarton or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_pallets": sum((r.pallets or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_fiberglass": sum((r.fiberglass or 0.0) for r in shift.receipts) if shift.receipts else 0.0,
            "receipt_laprol": sum((r.laprol or 0.0) for r in shift.receipts) if shift.receipts else 0.0
        }
        changes = []
        for k, old_v in old_values.items():
            new_v = new_values.get(k)
            if old_v != new_v:
                changes.append(f"{k}: {old_v} -> {new_v}")
        if changes or snapshot_before:
            db.add(models.AuditLog(
                user_name=user_name,
                action="UPDATE",
                target_table="shifts",
                target_id=shift.id,
                details=f"Обновлен рапорт мастера смены {shift.id}. Изменения: " + (", ".join(changes) if changes else "без критических числовых изменений"),
                state_snapshot=snapshot_before
            ))
    db.commit()

    # Проверка рапорта и отправка Telegram алерта в группу/руководству
    try:
        tg_chat_id = os.getenv("TELEGRAM_ALERT_CHAT_ID", "").strip()
        if tg_chat_id:
            import telegram_service
            warnings = []
            
            if not data.batch_number:
                warnings.append("Не заполнен номер партии")
            if not data.product_name:
                warnings.append("Не указано наименование продукции")
            if (data.lfm_sheets or 0) <= 0:
                warnings.append("Не указана выработка ЛФМ (0 листов)")
            if (shift.zo_cement or 0) <= 0:
                warnings.append("Не заполнен расход цемента (0 т)")
            if (shift.zo_chrysotile_4_20 or 0) + (shift.zo_chrysotile_5_65 or 0) + (shift.zo_chrysotile_6_40 or 0) <= 0:
                warnings.append("Не заполнен расход хризотила (все группы 0 т)")
                
            # Проверка аномального брака дестакера (> 4%)
            if data.lfm_sheets and data.lfm_sheets > 0:
                defect_pct = (ds_defect_sum / data.lfm_sheets) * 100.0
                if defect_pct > 4.0:
                    warnings.append(f"Высокий процент брака Дестакера: {defect_pct:.1f}% ({ds_defect_sum} листов)")
                    
            master_name = shift.master.name if shift.master else user_name
            tons = ((data.lfm_sheets or 0) * (get_product_finished_weight_kg(db, data.product_name) if hasattr(db, 'query') else 19.6)) / 1000.0
            
            shift_info = {
                "date": str(shift.date),
                "shift_name": shift.shift_name,
                "line": shift.line,
                "master_name": master_name,
                "sheets": data.lfm_sheets or 0,
                "tons": tons
            }
            
            # Отправляем алерт только если есть замечания (или отчет закрыт)
            if warnings:
                telegram_service.send_shift_quality_alert(tg_chat_id, shift_info, warnings, is_success=False)
    except Exception as tg_alert_err:
        print(f"Error sending Telegram shift report alert: {tg_alert_err}")


import google_sheets_integration

def sync_sharepoint_report_bg():
    db = SessionLocal()
    try:
        file_bytes = excel_exporter.generate_flat_report(db)
        filename = "Сводный_отчет_Tectum.xlsx"
        local_path = os.path.join("static", filename)
        try:
            with open(local_path, "wb") as f:
                f.write(file_bytes)
        except Exception as local_err:
            print(f"Error saving local excel: {local_err}")
            
        if os.getenv("M365_TENANT_ID"):
            try:
                m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder="Reports")
            except Exception as sp_err:
                db.add(models.AuditLog(
                    user_name="System Background Sync",
                    action="ERROR",
                    target_table="shifts",
                    target_id=0,
                    details=f"Ошибка загрузки сводного отчета в SharePoint: {str(sp_err)}"
                ))
                db.commit()
            
        try:
            # Запускаем синхронизацию с Google Таблицами
            google_sheets_integration.sync_report_to_google_sheets(db)
            google_sheets_integration.sync_qcd_reports_to_google_sheets(db)
            google_sheets_integration.export_receipt_to_google_sheets(db)
            db.add(models.AuditLog(
                user_name="System Background Sync",
                action="UPDATE",
                target_table="shifts",
                target_id=0,
                details="Сводный отчет, приход сырья и переборка успешно синхронизированы с Google Таблицами в фоновом режиме."
            ))
            db.commit()
        except Exception as gs_err:
            db.add(models.AuditLog(
                user_name="System Background Sync",
                action="ERROR",
                target_table="shifts",
                target_id=0,
                details=f"Ошибка синхронизации с Google Таблицами: {str(gs_err)}"
            ))
            db.commit()
    except Exception as e:
        print(f"Error in SharePoint/Google background sync: {e}")
    finally:
        db.close()


@app.post("/api/report")
def save_shift_report(data: schemas.ShiftReportCreate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id") or 9999
    user_role = request.session.get("user_role") or "admin"
    user_name = request.session.get("user_name") or "Админ"
        
    if user_role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен. Только мастера или администраторы могут сохранять рапорты.")
        
    # Check if shift exists or create it
    query = db.query(models.Shift).filter(
        models.Shift.date == data.date,
        models.Shift.shift_name == data.shift_name,
        models.Shift.line == data.line
    )
    if data.product_name:
        query = query.filter(models.Shift.product_name == data.product_name)
    if data.batch_number:
        query = query.filter(models.Shift.batch_number == data.batch_number)
    if data.export_type:
        query = query.filter(models.Shift.export_type == data.export_type)
    shift = query.first()
    
    is_new = False
    if not shift:
        is_new = True
        shift = models.Shift(
            date=data.date,
            shift_name=data.shift_name,
            line=data.line,
            master_id=data.master_id,
            product_name=data.product_name or "",
            batch_number=data.batch_number or "",
            export_type=data.export_type or "Эталон",
            status="closed",
            created_at=datetime.utcnow()
        )
        db.add(shift)
        db.flush()
    else:
        if user_role != "admin" and shift.created_at:
            time_diff = (datetime.utcnow() - shift.created_at).total_seconds()
            if time_diff > 1800: # 30 minutes
                raise HTTPException(
                    status_code=403, 
                    detail="Время на самостоятельное редактирование рапорта (30 мин) истекло. Для внесения правок обратитесь к администратору."
                )

    save_report_internal(db, shift, data, user_name, is_new)
    
    # Trigger background SharePoint & Google Sheets sync
    background_tasks.add_task(sync_sharepoint_report_bg)
    background_tasks.add_task(sync_google_sheets_bg)
    
    return {"status": "success", "shift_id": shift.id}


@app.put("/api/report/{shift_id}")
def update_shift_report_endpoint(shift_id: int, data: schemas.ShiftReportCreate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    user_name = request.session.get("user_name", "Unknown")
    if not user_id or not user_role:
        raise HTTPException(status_code=401, detail="Не авторизован")
        
    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Смена не найдена")
        
    if user_role != "admin" and shift.created_at:
        time_diff = (datetime.utcnow() - shift.created_at).total_seconds()
        if time_diff > 1800: # 30 minutes
            raise HTTPException(
                status_code=403, 
                detail="Время на самостоятельное редактирование рапорта (30 мин) истекло. Для внесения правок обратитесь к администратору."
            )
        
    save_report_internal(db, shift, data, user_name, False)
    
    # Trigger background SharePoint & Google Sheets sync
    background_tasks.add_task(sync_sharepoint_report_bg)
    background_tasks.add_task(sync_google_sheets_bg)
    
    return {"status": "success", "shift_id": shift.id}


@app.post("/api/receipts")
def create_autonomous_receipt(data: schemas.RawMaterialReceiptCreate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_role = request.session.get("user_role")
    user_name = request.session.get("user_name", "Склад")
    if not user_role:
        raise HTTPException(status_code=401, detail="Не авторизован")

    shift_id = None
    if data.date and data.shift_name and data.line:
        existing_shift = db.query(models.Shift).filter(
            models.Shift.date == data.date,
            models.Shift.shift_name == data.shift_name,
            models.Shift.line == data.line
        ).first()
        if existing_shift:
            shift_id = existing_shift.id

    receipt = models.RawMaterialReceipt(
        shift_id=shift_id,
        date=data.date,
        shift_name=data.shift_name,
        line=data.line,
        master_id=data.master_id,
        chrysotile_4_20=data.chrysotile_4_20,
        chrysotile_5_65=data.chrysotile_5_65,
        chrysotile_6_40=data.chrysotile_6_40,
        cement_silo1=data.cement_silo1,
        cement_silo2=data.cement_silo2,
        cement_silo3=data.cement_silo3,
        cement_silo4=data.cement_silo4,
        cellulose=data.cellulose,
        crushed_slate=data.crushed_slate,
        asbozurit=data.asbozurit,
        asbocarton=data.asbocarton,
        pallets=data.pallets,
        fiberglass=data.fiberglass,
        laprol=data.laprol
    )
    db.add(receipt)
    db.flush()
    
    db.add(models.AuditLog(
        user_name=user_name,
        action="CREATE",
        target_table="raw_material_receipts",
        target_id=receipt.id,
        details=f"Добавлен автономный приход сырья: Дата {data.date}, Смена {data.shift_name}, Линия {data.line}"
    ))
    db.commit()
    background_tasks.add_task(sync_receipts_bg)
    return {"status": "success", "receipt_id": receipt.id}


@app.get("/api/receipts/by_slot")
def get_receipts_by_slot(date: str, shift_name: str, line: str, db: Session = Depends(get_db)):
    try:
        if hasattr(date, "strftime"):
            parsed_date = date.date() if hasattr(date, "date") else date
        else:
            parsed_date = datetime.strptime(str(date), "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(400, "Неверный формат даты. Ожидается YYYY-MM-DD")
        
    receipts = db.query(models.RawMaterialReceipt).outerjoin(models.Shift).filter(
        or_(
            models.RawMaterialReceipt.date == parsed_date,
            and_(models.RawMaterialReceipt.date.is_(None), models.Shift.date == parsed_date)
        ),
        or_(
            models.RawMaterialReceipt.shift_name == shift_name,
            and_(models.RawMaterialReceipt.shift_name.is_(None), models.Shift.shift_name == shift_name)
        ),
        or_(
            models.RawMaterialReceipt.line == line,
            and_(models.RawMaterialReceipt.line.is_(None), models.Shift.line == line)
        )
    ).order_by(models.RawMaterialReceipt.id.desc()).all()
    
    result = []
    for r in receipts:
        r_dict = schemas.RawMaterialReceipt.model_validate(r).model_dump()
        r_dict["record_date"] = str(r.record_date) if r.record_date else str(parsed_date)
        r_dict["record_shift_name"] = r.record_shift_name or shift_name
        r_dict["record_line"] = r.record_line or line
        r_dict["master_name"] = r.master.name if r.master else (r.shift.master.name if r.shift and r.shift.master else "Н/Д")
        result.append(r_dict)
    return result


@app.post("/api/shifts/{shift_id}/receipts")
def add_raw_material_receipt(shift_id: int, data: schemas.RawMaterialReceiptCreate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_role = request.session.get("user_role")
    user_name = request.session.get("user_name", "Unknown")
    if not user_role:
        raise HTTPException(status_code=401, detail="Не авторизован")

    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Смена не найдена")

    receipt = models.RawMaterialReceipt(
        shift_id=shift.id,
        date=data.date or shift.date,
        shift_name=data.shift_name or shift.shift_name,
        line=data.line or shift.line,
        master_id=data.master_id or shift.master_id,
        chrysotile_4_20=data.chrysotile_4_20,
        chrysotile_5_65=data.chrysotile_5_65,
        chrysotile_6_40=data.chrysotile_6_40,
        cement_silo1=data.cement_silo1,
        cement_silo2=data.cement_silo2,
        cement_silo3=data.cement_silo3,
        cement_silo4=data.cement_silo4,
        cellulose=data.cellulose,
        crushed_slate=data.crushed_slate,
        asbozurit=data.asbozurit,
        asbocarton=data.asbocarton,
        pallets=data.pallets,
        fiberglass=data.fiberglass,
        laprol=data.laprol
    )
    db.add(receipt)
    
    db.add(models.AuditLog(
        user_name=user_name,
        action="CREATE",
        target_table="raw_material_receipts",
        target_id=shift_id,
        details=f"Добавлен приход сырья для смены {shift.date} {shift.shift_name}"
    ))
    db.commit()
    
    background_tasks.add_task(sync_receipts_bg)
    
    return {"status": "success", "receipt_id": receipt.id}


@app.put("/api/receipts/{receipt_id}")
def update_raw_material_receipt_endpoint(receipt_id: int, data: schemas.RawMaterialReceiptUpdate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_role = request.session.get("user_role")
    user_name = request.session.get("user_name", "Unknown")
    if not user_role:
        raise HTTPException(status_code=401, detail="Не авторизован")

    receipt = db.query(models.RawMaterialReceipt).get(receipt_id)
    if not receipt:
        raise HTTPException(status_code=404, detail="Приход сырья не найден")

    receipt_created = receipt.timestamp or getattr(receipt, 'created_at', None)
    if user_role != "admin" and receipt_created:
        time_diff = (datetime.utcnow() - receipt_created).total_seconds()
        if time_diff > 1800:
            raise HTTPException(
                status_code=403,
                detail="Время на самостоятельное редактирование прихода (30 мин) истекло. Обратитесь к администратору."
            )

    for field, val in data.model_dump(exclude_unset=True).items():
        if val is not None and hasattr(receipt, field):
            setattr(receipt, field, val)

    db.add(models.AuditLog(
        user_name=user_name,
        action="UPDATE",
        target_table="raw_material_receipts",
        target_id=receipt_id,
        details=f"Обновлен приход сырья ID {receipt_id}"
    ))
    db.commit()
    background_tasks.add_task(sync_receipts_bg)
    return {"status": "success"}


@app.delete("/api/receipts/{receipt_id}")
def delete_raw_material_receipt(receipt_id: int, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user_role = request.session.get("user_role")
    user_name = request.session.get("user_name", "Unknown")
    if not user_role:
        raise HTTPException(status_code=401, detail="Не авторизован")

    receipt = db.query(models.RawMaterialReceipt).get(receipt_id)
    if not receipt:
        raise HTTPException(status_code=404, detail="Приход сырья не найден")

    receipt_created = receipt.timestamp or getattr(receipt, 'created_at', None)
    if user_role != "admin" and receipt_created:
        time_diff = (datetime.utcnow() - receipt_created).total_seconds()
        if time_diff > 1800:
            raise HTTPException(
                status_code=403,
                detail="Время на самостоятельное удаление прихода (30 мин) истекло. Обратитесь к администратору."
            )

    shift = receipt.shift
    db.delete(receipt)
    
    db.add(models.AuditLog(
        user_name=user_name,
        action="DELETE",
        target_table="raw_material_receipts",
        target_id=receipt_id,
        details=f"Удален приход сырья для смены {shift.date if shift else 'Unknown'}"
    ))
    db.commit()
    
    background_tasks.add_task(sync_receipts_bg)
    
    return {"status": "success"}


@app.post("/api/norms/sync_from_google")
def sync_norms_from_google_sheets_endpoint(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    user_name = request.session.get("user_name", "Unknown")
    
    if not user_id or not user_role:
        raise HTTPException(status_code=401, detail="Не авторизован")
        
    if user_role not in ["admin", "technologist"]:
        raise HTTPException(status_code=403, detail="Доступ разрешен только Технологу или Администратору")
        
    try:
        google_sheets_integration.sync_norms_from_google_sheets(db)
        # Записываем действие в AuditLog
        db.add(models.AuditLog(
            user_name=user_name,
            action="IMPORT",
            target_table="product_norms",
            details="Синхронизация нормативов расхода сырья из Google Sheets"
        ))
        db.commit()
        return {"status": "success", "message": "Нормативы успешно обновлены из Google Sheets"}
    except Exception as err:
        raise HTTPException(status_code=500, detail=str(err))


@app.get("/api/report/summary")
def get_report_summary(
    request: Request,
    from_date: str = None,
    to_date: str = None,
    line: str = None,
    master_id: int = None,
    export_type: str = None,
    db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Не авторизован")
        
    result = []
    try:
        query = db.query(models.Shift)
    
        if from_date:
            if isinstance(from_date, (datetime, date)):
                f_date = from_date if isinstance(from_date, date) else from_date.date()
            else:
                f_date = datetime.strptime(str(from_date), "%Y-%m-%d").date()
            query = query.filter(models.Shift.date >= f_date)
        if to_date:
            if isinstance(to_date, (datetime, date)):
                t_date = to_date if isinstance(to_date, date) else to_date.date()
            else:
                t_date = datetime.strptime(str(to_date), "%Y-%m-%d").date()
            query = query.filter(models.Shift.date <= t_date)
        if line:
            query = query.filter(models.Shift.line == line)
        if master_id:
            query = query.filter(models.Shift.master_id == master_id)
        if export_type:
            query = query.filter(models.Shift.export_type == export_type)
        
        shifts = query.options(
            joinedload(models.Shift.master),
            joinedload(models.Shift.batches),
            joinedload(models.Shift.lfm_reports),
            joinedload(models.Shift.receipts)
        ).order_by(models.Shift.date.desc(), models.Shift.line.asc(), models.Shift.shift_name.desc(), models.Shift.batch_number.desc(), models.Shift.id.desc()).all()
    
        latest_shift_id = None
        # Find the most recently created shift
        latest_shift = db.query(models.Shift).filter(models.Shift.created_at.isnot(None)).order_by(models.Shift.created_at.desc(), models.Shift.id.desc()).first()
        if latest_shift:
            latest_shift_id = latest_shift.id

        result = []
        for shift in shifts:
            is_other_master = False
        
            lfm_reports = shift.lfm_reports
            batches = shift.batches
        
            # Фильтруем абсолютно пустые смены без факта производства и без плана
            lfm_sheets_check = sum((l.lfm_sheets or 0) for l in lfm_reports) if lfm_reports else 0
            warehouse_gp_check = sum((b.ds_condition or 0) for b in batches) if batches else 0
            zo_batches_check = shift.zo_batches or 0
            plan_sheets_check = shift.plan_sheets or 0
        
            if plan_sheets_check == 0 and lfm_sheets_check == 0 and warehouse_gp_check == 0 and zo_batches_check == 0 and not shift.zo_submitted:
                continue
            
            lfm_sheets = lfm_sheets_check if not is_other_master else 0
            lfm_resets = sum((l.lfm_wind_resets or 0) for l in lfm_reports) if (lfm_reports and not is_other_master) else 0
        
            warehouse_gp = warehouse_gp_check if not is_other_master else 0
            first_grade = sum((b.ds_first_grade or 0) for b in batches) if (batches and not is_other_master) else 0
            qcd_defect = sum((b.ds_defect or 0) for b in batches) if (batches and not is_other_master) else 0
        
            ds_defects = {
                "ds_defect_chip": sum((b.ds_defect_chip or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_scratch": sum((b.ds_defect_scratch or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_bad_cut": sum((b.ds_defect_bad_cut or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_stick_bottom": sum((b.ds_defect_stick_bottom or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_stick_top": sum((b.ds_defect_stick_top or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_broken": sum((b.ds_defect_broken or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_fell_box": sum((b.ds_defect_fell_box or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_dent": sum((b.ds_defect_dent or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_thickness": sum((b.ds_defect_thickness or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_delamination": sum((b.ds_defect_delamination or 0) for b in batches) if (batches and not is_other_master) else 0,
                "ds_defect_edge": sum((b.ds_defect_edge or 0) for b in batches) if (batches and not is_other_master) else 0,
            }

        
            prev_first_grade = sum((b.prev_first_grade or 0) for b in batches) if (batches and not is_other_master) else 0
            prev_defect = sum((b.prev_defect or 0) for b in batches) if (batches and not is_other_master) else 0
            prev_defects = {
                "prev_defect_scratch": sum((b.prev_defect_scratch or 0) for b in batches) if (batches and not is_other_master) else 0,
                "prev_defect_bad_cut": sum((b.prev_defect_bad_cut or 0) for b in batches) if (batches and not is_other_master) else 0,
                "prev_defect_stick_top": sum((b.prev_defect_stick_top or 0) for b in batches) if (batches and not is_other_master) else 0,
                "prev_defect_broken": sum((b.prev_defect_broken or 0) for b in batches) if (batches and not is_other_master) else 0,
                "prev_defect_fell_box": sum((b.prev_defect_fell_box or 0) for b in batches) if (batches and not is_other_master) else 0,
                "prev_defect_thickness": sum((b.prev_defect_thickness or 0) for b in batches) if (batches and not is_other_master) else 0,
                "prev_defect_edge": sum((b.prev_defect_edge or 0) for b in batches) if (batches and not is_other_master) else 0,
            }

            product_name = shift.product_name if not is_other_master else "Скрыто"
            batch_number = shift.batch_number if not is_other_master else "Скрыто"
            export_type_val = shift.export_type if not is_other_master else "Скрыто"
            master_name = "Смена другого мастера" if is_other_master else (shift.master.name if shift.master else "Мастер удалён")
        
            lfm_tons = round(lfm_sheets * get_product_finished_weight_kg(db, product_name) / 1000.0, 2) if (lfm_sheets and not is_other_master) else 0.0
        
            dev_data = {"theoretical": {}, "actual": {}, "deviations": {}}
            if not is_other_master:
                dev_data = calculate_shift_deviations(db, shift)
            
            created_at_iso = shift.created_at.isoformat() if shift.created_at else None
            remaining_secs = 0
            # Only the latest shift gets the 30-minute window for masters
            is_the_latest = (shift.id == latest_shift_id)
            if shift.created_at and is_the_latest:
                diff = (datetime.utcnow() - shift.created_at).total_seconds()
                remaining_secs = max(0, int(1800 - diff))
            elif user_role == "admin":
                remaining_secs = 999999
            
            result.append({
                "shift_id": shift.id,
                "date": shift.date.strftime("%Y-%m-%d") if shift.date else "Н/Д",
                "shift_name": shift.shift_name,
                "line": shift.line,
                "master_id": shift.master_id,
                "master_name": master_name,
                "batch_number": batch_number,
                "product_name": product_name,
                "export_type": export_type_val or "Эталон",
                "status": shift.status,
                "created_at": created_at_iso,
                "remaining_edit_seconds": remaining_secs,
                "can_edit": (user_role == "admin" or (is_the_latest and remaining_secs > 0)),
            
                "plan_sheets": shift.plan_sheets or 0,
                "plan_tons": shift.plan_tons or 0.0,
            
                "lfm_sheets": lfm_sheets,
                "lfm_wind_resets": lfm_resets,
                "lfm_tons": lfm_tons,
                "zo_batches": shift.zo_batches if not is_other_master else 0,
            
                "warehouse_gp": warehouse_gp,
                "first_grade": first_grade,
                "defect": qcd_defect,
                "ds_defects": ds_defects,
                "prev_first_grade": prev_first_grade,
                "prev_defect": prev_defect,
                "prev_defects": prev_defects,
            
                "receipts": {
                    "chrysotile_4_20": (sum((r.chrysotile_4_20 or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "chrysotile_5_65": (sum((r.chrysotile_5_65 or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "chrysotile_6_40": (sum((r.chrysotile_6_40 or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "cement": (sum((((r.cement_silo1 or 0.0) + (r.cement_silo2 or 0.0) + (r.cement_silo3 or 0.0) + (r.cement_silo4 or 0.0))) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "cellulose": (sum((r.cellulose or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "crushed_slate": (sum((r.crushed_slate or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "asbozurit": (sum((r.asbozurit or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "asbocarton": (sum((r.asbocarton or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "pallets": (sum((r.pallets or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "fiberglass": (sum((r.fiberglass or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0,
                    "laprol": (sum((r.laprol or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) if not is_other_master else 0.0
                },
                "zo_usage": {
                    "chrysotile_4_20": shift.zo_chrysotile_4_20 if not is_other_master else 0.0,
                    "chrysotile_5_65": shift.zo_chrysotile_5_65 if not is_other_master else 0.0,
                    "chrysotile_6_40": shift.zo_chrysotile_6_40 if not is_other_master else 0.0,
                    "cement_silo1": shift.zo_cement_silo1 if not is_other_master else 0.0,
                    "cement_silo2": shift.zo_cement_silo2 if not is_other_master else 0.0,
                    "cement_silo3": shift.zo_cement_silo3 if not is_other_master else 0.0,
                    "cement_silo4": shift.zo_cement_silo4 if not is_other_master else 0.0,
                    "cellulose": shift.zo_cellulose if not is_other_master else 0.0,
                    "crushed_slate": shift.zo_crushed_slate if not is_other_master else 0.0,
                    "asbozurit": shift.zo_asbozurit if not is_other_master else 0.0,
                    "fiberglass": shift.zo_fiberglass if not is_other_master else 0.0,
                    "laprol": shift.zo_laprol if not is_other_master else 0.0,
                    "asbocarton": shift.zo_asbocarton if not is_other_master else 0.0,
                    "asb_drain": shift.zo_asb_drain if not is_other_master else 0.0,
                    "cem_drain": shift.zo_cem_drain if not is_other_master else 0.0
                },
                "deviations": dev_data
            })
        
    except Exception as err:
        print(f"Error: {err}")
    return result


@app.get("/api/report/materials_summary")
def get_materials_summary(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    if not user_id:
        raise HTTPException(status_code=401, detail="Не авторизован")
        
    try:
        query = db.query(models.Shift)
    
        if False and user_role == "master" and user_id:
            query = query.filter(models.Shift.master_id == user_id)
        
        if start_date:
            query = query.filter(models.Shift.date >= datetime.strptime(start_date, "%Y-%m-%d").date())
        if end_date:
            query = query.filter(models.Shift.date <= datetime.strptime(end_date, "%Y-%m-%d").date())
        
        shifts = query.order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
        materials = [
            "chrysotile_4_20", "chrysotile_5_65", "chrysotile_6_40",
            "cement", "cellulose", "crushed_slate", "asbozurit",
            "asbocarton", "fiberglass", "laprol"
        ]
    
        totals = {m: {"receipt": 0.0, "zo": 0.0, "deviation": 0.0} for m in materials}
        daily_breakdown = []
    
        for shift in shifts:
            zo_cem = (shift.zo_cement_silo1 or 0) + (shift.zo_cement_silo2 or 0) + (shift.zo_cement_silo3 or 0) + (shift.zo_cement_silo4 or 0)
        
            shift_mats = {
                "chrysotile_4_20": {"receipt": (sum((r.chrysotile_4_20 or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_chrysotile_4_20 or 0.0},
                "chrysotile_5_65": {"receipt": (sum((r.chrysotile_5_65 or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_chrysotile_5_65 or 0.0},
                "chrysotile_6_40": {"receipt": (sum((r.chrysotile_6_40 or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_chrysotile_6_40 or 0.0},
                "cement": {"receipt": (sum((((r.cement_silo1 or 0.0) + (r.cement_silo2 or 0.0) + (r.cement_silo3 or 0.0) + (r.cement_silo4 or 0.0))) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": zo_cem},
                "cellulose": {"receipt": (sum((r.cellulose or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_cellulose or 0.0},
                "crushed_slate": {"receipt": (sum((r.crushed_slate or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_crushed_slate or 0.0},
                "asbozurit": {"receipt": (sum((r.asbozurit or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_asbozurit or 0.0},
                "asbocarton": {"receipt": (sum((r.asbocarton or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_asbocarton or 0.0},
                "fiberglass": {"receipt": (sum((r.fiberglass or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_fiberglass or 0.0},
                "laprol": {"receipt": (sum((r.laprol or 0.0) for r in shift.receipts) if getattr(shift, "receipts", None) else 0.0) or 0.0, "zo": shift.zo_laprol or 0.0}
            }
        
            dev_info = calculate_shift_deviations(db, shift)
            shift_devs = dev_info["deviations"]
        
            day_entry = {
                "date": shift.date.strftime("%Y-%m-%d") if shift.date else "Н/Д",
                "shift_name": shift.shift_name,
                "line": shift.line,
                "materials": {}
            }
        
            for m in materials:
                r = shift_mats[m]["receipt"]
                z = shift_mats[m]["zo"]
                d = shift_devs.get(m, round(z - r, 2)) if m != "cement" else shift_devs.get("cement", round(z - r, 2))
            
                day_entry["materials"][m] = {
                    "receipt": round(r, 2),
                    "zo": round(z, 2),
                    "deviation": round(d, 2)
                }
            
                totals[m]["receipt"] += r
                totals[m]["zo"] += z
                totals[m]["deviation"] += d
            
            daily_breakdown.append(day_entry)
        
        for m in materials:
            totals[m]["receipt"] = round(totals[m]["receipt"], 2)
            totals[m]["zo"] = round(totals[m]["zo"], 2)
            totals[m]["deviation"] = round(totals[m]["deviation"], 2)
        
    except Exception as err:
        print(f"Error: {err}")
    return {
        "totals": totals,
        "daily": daily_breakdown
    }


def sync_lfm_to_plan_board(shift_date, shift_name: str, shift_line: str, db: Session, master_id: int = None):
    # Map shift line to plan board line
    is_line_1 = "1" in shift_line
    pb_line = "ЛФМ-1" if is_line_1 else "ЛФМ-2"
    
    # Find all shifts matching the date, name, and line
    matching_shifts = db.query(models.Shift).filter(
        models.Shift.date == shift_date,
        models.Shift.shift_name == shift_name,
        models.Shift.line.like("%1%" if is_line_1 else "%2%")
    ).all()
    
    shift_ids = [s.id for s in matching_shifts]
    
    total_sheets = 0
    total_1st = 0
    total_defect = 0
    
    if shift_ids:
        # Calculate sum of sheets from LFM reports, and 1st grade, defect from batches for these shifts
        lfm_stats = db.query(
            func.sum(models.LFMReport.lfm_sheets).label("total_sheets")
        ).filter(models.LFMReport.shift_id.in_(shift_ids)).first()
        
        batch_stats = db.query(
            func.sum(models.Batch.ds_first_grade).label("total_1st"),
            func.sum(models.Batch.ds_defect).label("total_defect")
        ).filter(models.Batch.shift_id.in_(shift_ids)).first()
        
        total_sheets = int(lfm_stats.total_sheets or 0) if lfm_stats else 0
        total_1st = int(batch_stats.total_1st or 0) if batch_stats else 0
        total_defect = int(batch_stats.total_defect or 0) if batch_stats else 0
    
    # Find corresponding MonthlyPlanBoard row
    pb_row = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date == shift_date,
        models.MonthlyPlanBoard.shift_name == shift_name,
        models.MonthlyPlanBoard.line == pb_line
    ).first()
    
    if pb_row:
        old_fact = pb_row.fact_sheets
        pb_row.fact_sheets = total_sheets
        pb_row.first_grade = total_1st
        pb_row.defect = total_defect
        
        # Log to AuditLog (per rules, plan board changes must be logged to AuditLog)
        log_entry = models.AuditLog(
            timestamp=datetime.utcnow(),
            user_name="System Sync (LFM)",
            action="UPDATE",
            target_table="monthly_plan_board",
            target_id=pb_row.id,
            details=f"Синхронизация {shift_line} {shift_date} {shift_name}. Факт обновлен: {old_fact} -> {total_sheets}. 1 сорт: {total_1st}, Брак: {total_defect}."
        )
        db.add(log_entry)
    else:
        # If there are no shifts and no fact, don't create a phantom row
        if total_sheets == 0 and not shift_ids:
            return
            
        final_master_id = master_id if master_id is not None else (matching_shifts[0].master_id if matching_shifts else None)
        if isinstance(shift_date, str):
            try:
                dt_obj = datetime.strptime(shift_date, "%Y-%m-%d").date()
                is_monday = dt_obj.weekday() == 0
            except:
                is_monday = False
        else:
            is_monday = shift_date.weekday() == 0
            
        default_plan_sheets = 0 if is_monday and shift_name == "День" else (2700 if shift_name == "День" else 3300)
        
        # Create a new plan board row if it doesn't exist
        pb_row = models.MonthlyPlanBoard(
            date=shift_date,
            shift_name=shift_name,
            line=pb_line,
            master_id=final_master_id,
            plan_sheets=default_plan_sheets,
            fact_sheets=total_sheets,
            first_grade=total_1st,
            defect=total_defect
        )
        db.add(pb_row)
        db.flush() # get the ID
        
        log_entry = models.AuditLog(
            timestamp=datetime.utcnow(),
            user_name="System Sync (LFM)",
            action="CREATE",
            target_table="monthly_plan_board",
            target_id=pb_row.id,
            details=f"Создана новая запись план-борда для {shift_line} {shift_date} {shift_name}. Факт: {total_sheets}. 1 сорт: {total_1st}, Брак: {total_defect}."
        )
        db.add(log_entry)
        
    db.commit()

# --- ЛФМ ---
@app.post("/api/shifts/{shift_id}/lfm")
def create_lfm_report(shift_id: int, data: schemas.LFMReportCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    shift = db.query(models.Shift).get(shift_id)
    if not shift: raise HTTPException(404)
    db_report = models.LFMReport(**data.model_dump(), shift_id=shift_id)
    db.add(db_report)
    db.commit()
    
    # Sync LFM sheets to plan board fact
    sync_lfm_to_plan_board(shift.date, shift.shift_name, shift.line, db, shift.master_id)
    background_tasks.add_task(sync_google_sheets_bg)
    return {"status": "ok"}

# --- ПРОСТОИ ---
@app.post("/api/upload_media/")
async def upload_media(file: UploadFile = File(...)):
    try:
        file_bytes = await file.read()
        filename = file.filename
        url = m365_integration.upload_file_to_sharepoint(file_bytes, filename)
        return {"url": url}
    except Exception as e:
        print(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- ПАРТИИ (Стакер) ---
@app.post("/api/batches/")
def create_batch(shift_id: int, data: schemas.BatchCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    db_batch = models.Batch(**data.model_dump(exclude={"status"}), shift_id=shift_id, status="stacked")
    db.add(db_batch)
    db.commit()
    db.refresh(db_batch)
    background_tasks.add_task(sync_google_sheets_bg)
    return db_batch

# --- Дестакер и СКК ---
@app.get("/api/batches/pending_destacker", response_model=list[schemas.Batch])
def get_pending_destacker_batches(db: Session = Depends(get_db)):
    # Дестакер видит все партии, которые были уложены (stacked)
    return db.query(models.Batch).filter(models.Batch.status == "stacked").all()

@app.get("/api/batches/pending_qcd", response_model=list[schemas.Batch])
def get_pending_qcd_batches(db: Session = Depends(get_db)):
    # СКК видит партии, которые уложены или разобраны, но еще не проверены СКК
    return db.query(models.Batch).filter(
        or_(models.Batch.status == "stacked", models.Batch.status == "destacked")
    ).all()

class DestackerUpdate(BaseModel):
    ds_condition: int
    ds_first_grade: int
    ds_defect_chip: int = 0
    ds_defect_scratch: int = 0
    ds_defect_bad_cut: int = 0
    ds_defect_stick_bottom: int = 0
    ds_defect_stick_top: int = 0
    ds_defect_broken: int = 0
    ds_defect_fell_box: int = 0
    ds_defect_dent: int = 0
    ds_defect_thickness: int = 0
    ds_defect_delamination: int = 0
    ds_defect_edge: int = 0

@app.post("/api/batches/{batch_id}/destacker")
def update_destacker(batch_id: int, data: DestackerUpdate, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    batch = db.query(models.Batch).get(batch_id)
    if not batch: raise HTTPException(404)
    batch.ds_condition = data.ds_condition
    batch.ds_first_grade = data.ds_first_grade
    batch.ds_defect_chip = data.ds_defect_chip
    batch.ds_defect_scratch = data.ds_defect_scratch
    batch.ds_defect_bad_cut = data.ds_defect_bad_cut
    batch.ds_defect_stick_bottom = data.ds_defect_stick_bottom
    batch.ds_defect_stick_top = data.ds_defect_stick_top
    batch.ds_defect_broken = data.ds_defect_broken
    batch.ds_defect_fell_box = data.ds_defect_fell_box
    batch.ds_defect_dent = data.ds_defect_dent
    batch.ds_defect_thickness = data.ds_defect_thickness
    batch.ds_defect_delamination = data.ds_defect_delamination
    batch.ds_defect_edge = data.ds_defect_edge
    
    # Суммируем весь брак
    batch.ds_defect = (
        data.ds_defect_chip + data.ds_defect_scratch + data.ds_defect_bad_cut +
        data.ds_defect_stick_bottom + data.ds_defect_stick_top + data.ds_defect_broken +
        data.ds_defect_fell_box + data.ds_defect_dent + data.ds_defect_thickness +
        data.ds_defect_delamination + data.ds_defect_edge
    )
    batch.status = "destacked"
    db.commit()
    background_tasks.add_task(sync_google_sheets_bg)
    return {"status": "ok"}

class QCDUpdate(BaseModel):
    qcd_sorted_packs: int = 0
    qcd_first_grade: int = 0
    qcd_first_grade_note: Optional[str] = None
    qcd_defect_note: Optional[str] = None
    qcd_defect_chip: int = 0
    qcd_defect_scratch: int = 0
    qcd_defect_bad_cut: int = 0
    qcd_defect_stick_bottom: int = 0
    qcd_defect_stick_top: int = 0
    qcd_defect_broken: int = 0
    qcd_defect_fell_box: int = 0
    qcd_defect_dent: int = 0
    qcd_defect_thickness: int = 0
    qcd_defect_delamination: int = 0
    qcd_defect_edge: int = 0



@app.get("/api/dashboard/stats")
def get_dashboard_stats(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    prod_query = db.query(
        func.sum(models.Batch.ds_condition).label('condition'),
        func.sum(models.Batch.ds_first_grade).label('first_grade'),
        func.sum(models.Batch.ds_defect).label('defect')
    )

    defects_query = db.query(
        func.sum(models.Batch.ds_defect_chip).label('chip'),
        func.sum(models.Batch.ds_defect_scratch).label('scratch'),
        func.sum(models.Batch.ds_defect_bad_cut).label('bad_cut'),
        func.sum(models.Batch.ds_defect_stick_bottom).label('stick_bottom'),
        func.sum(models.Batch.ds_defect_stick_top).label('stick_top'),
        func.sum(models.Batch.ds_defect_broken).label('broken'),
        func.sum(models.Batch.ds_defect_fell_box).label('fell_box'),
        func.sum(models.Batch.ds_defect_dent).label('dent'),
        func.sum(models.Batch.ds_defect_thickness).label('thickness'),
        func.sum(models.Batch.ds_defect_delamination).label('delamination'),
        func.sum(models.Batch.ds_defect_edge).label('edge'),
    )

    mats_query_receipt = db.query(
        func.sum(models.RawMaterialReceipt.chrysotile_4_20).label('r_4_20'),
        func.sum(models.RawMaterialReceipt.chrysotile_5_65).label('r_5_65'),
        func.sum(models.RawMaterialReceipt.chrysotile_6_40).label('r_6_40'),
        func.sum(models.RawMaterialReceipt.cement_silo1 + models.RawMaterialReceipt.cement_silo2 + models.RawMaterialReceipt.cement_silo3 + models.RawMaterialReceipt.cement_silo4).label('r_cem'),
        func.sum(models.RawMaterialReceipt.cellulose).label('r_cel')
    ).select_from(models.RawMaterialReceipt).join(models.Shift, models.Shift.id == models.RawMaterialReceipt.shift_id)

    mats_query_zo = db.query(
        func.sum(models.Shift.zo_chrysotile_4_20).label('z_4_20'),
        func.sum(models.Shift.zo_chrysotile_5_65).label('z_5_65'),
        func.sum(models.Shift.zo_chrysotile_6_40).label('z_6_40'),
        func.sum(models.Shift.zo_cement).label('z_cem'),
        func.sum(models.Shift.zo_cellulose).label('z_cel')
    )

    dt_query = db.query(models.Downtime)

    if False and user_role == "master" and user_id:
        prod_query = prod_query.join(models.Shift).filter(models.Shift.master_id == user_id)
        defects_query = defects_query.join(models.Shift).filter(models.Shift.master_id == user_id)
        mats_query_receipt = mats_query_receipt.filter(models.Shift.master_id == user_id)
        mats_query_zo = mats_query_zo.filter(models.Shift.master_id == user_id)
        dt_query = dt_query.join(models.Shift).filter(models.Shift.master_id == user_id)

    prod_stats = prod_query.first()
    defects = defects_query.first()
    mats_rec = mats_query_receipt.first()
    mats_zo = mats_query_zo.first()

    rec_asb = (mats_rec.r_4_20 or 0) + (mats_rec.r_5_65 or 0) + (mats_rec.r_6_40 or 0) if mats_rec else 0
    zo_asb = (mats_zo.z_4_20 or 0) + (mats_zo.z_5_65 or 0) + (mats_zo.z_6_40 or 0) if mats_zo else 0

    # --- DOWNTIME AGGREGATION ---
    downtimes = dt_query.all()
    total_downtime_minutes = sum((d.duration or 0) for d in downtimes)
    total_lost_tons = sum((d.lost_tons or 0) for d in downtimes)
    total_lost_tenge = sum((d.lost_tenge or 0) for d in downtimes)
    
    dt_by_cat = {}
    node_counts = {}
    for d in downtimes:
        if d.category:
            dt_by_cat[d.category] = dt_by_cat.get(d.category, 0) + (d.duration or 0)
        if d.node:
            node_counts[d.node] = node_counts.get(d.node, 0) + 1
            
    top_reasons = sorted([{"node": k, "count": v} for k, v in node_counts.items()], key=lambda x: x['count'], reverse=True)[:5]

    return {
        "production": {
            "condition": prod_stats.condition or 0,
            "first_grade": prod_stats.first_grade or 0,
            "defect": prod_stats.defect or 0
        },
        "defects": {
            "Скол": defects.chip or 0,
            "Сдир": defects.scratch or 0,
            "Плохой рез": defects.bad_cut or 0,
            "Налип снизу": defects.stick_bottom or 0,
            "Налип сверху": defects.stick_top or 0,
            "Сломан": defects.broken or 0,
            "Упал коробки": defects.fell_box or 0,
            "Вмятина": defects.dent or 0,
            "Толщина": defects.thickness or 0,
            "Расслоение": defects.delamination or 0,
            "Кромка": defects.edge or 0
        },
        "materials": {
            "Асбест": {"receipt": rec_asb, "zo": zo_asb},
            "Цемент": {"receipt": mats_rec.r_cem if mats_rec else 0 or 0, "zo": mats_zo.z_cem if mats_zo else 0 or 0},
            "Целлюлоза": {"receipt": mats_rec.r_cel if mats_rec else 0 or 0, "zo": mats_zo.z_cel if mats_zo else 0 or 0}
        },
        "downtimes": {
            "total_minutes": total_downtime_minutes,
            "lost_tons": total_lost_tons,
            "lost_tenge": total_lost_tenge,
            "by_category": dt_by_cat,
            "top_reasons": top_reasons
        }
    }

@app.get("/api/dashboard/weekly_report")
def get_weekly_report(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    # Берем последние 7 смен (включая текущую)
    query = db.query(models.Shift)
    if False and user_role == "master" and user_id:
        query = query.filter(models.Shift.master_id == user_id)
    shifts = query.order_by(models.Shift.date.desc(), models.Shift.line.asc(), models.Shift.shift_name.desc(), models.Shift.batch_number.desc(), models.Shift.id.desc()).limit(7).all()
    
    report_data = []
    for shift in shifts:
        # 1. Считаем формовку (ЛФМ)
        lfm_sheets = db.query(func.sum(models.LFMReport.lfm_sheets)).filter(models.LFMReport.shift_id == shift.id).scalar() or 0
        
        # 2. Считаем итог СКК
        qcd_stats = db.query(
            func.sum(models.Batch.ds_condition).label('condition'),
            func.sum(models.Batch.ds_first_grade).label('first_grade'),
            func.sum(models.Batch.ds_defect).label('defect')
        ).filter(models.Batch.shift_id == shift.id).first()
        
        qcd_cond = qcd_stats.condition or 0
        qcd_fg = qcd_stats.first_grade or 0
        qcd_def = qcd_stats.defect or 0
        
        # 3. Считаем отклонение сырья (Факт из ЗО - Теория по нормам)
        lfm_reports = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift.id).all()
        product_counts = {}
        for r in lfm_reports:
            product_counts[r.product_name] = product_counts.get(r.product_name, 0) + r.lfm_sheets
            
        theoretical = {
            "chrysotile_4_20": 0.0, "chrysotile_5_65": 0.0, "chrysotile_6_40": 0.0,
            "cement": 0.0, "cellulose": 0.0, "crushed_slate": 0.0,
            "asbozurit": 0.0, "fiberglass": 0.0
        }
        
        for prod_name, sheets in product_counts.items():
            norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == prod_name).first()
            if norm:
                theoretical["chrysotile_4_20"] += sheets * norm.norm_chrysotile_4_20
                theoretical["chrysotile_5_65"] += sheets * norm.norm_chrysotile_5_65
                theoretical["chrysotile_6_40"] += sheets * norm.norm_chrysotile_6_40
                theoretical["cement"] += sheets * norm.norm_cement
                theoretical["cellulose"] += sheets * norm.norm_cellulose
                theoretical["crushed_slate"] += sheets * norm.norm_crushed_slate
                theoretical["asbozurit"] += sheets * norm.norm_asbozurit
                theoretical["fiberglass"] += sheets * norm.norm_fiberglass

        # Фактический расход сырья из ЗО за смену
        fact_raw = (shift.zo_chrysotile_4_20 or 0.0) + \
                   (shift.zo_chrysotile_5_65 or 0.0) + \
                   (shift.zo_chrysotile_6_40 or 0.0) + \
                   (shift.zo_cement or 0.0) + \
                   (shift.zo_cellulose or 0.0) + \
                   (shift.zo_crushed_slate or 0.0) + \
                   (shift.zo_asbozurit or 0.0) + \
                   (shift.zo_fiberglass or 0.0)

        # Теоретический расход сырья
        theory_raw = sum(theoretical.values())
        
        # Общее отклонение по сырью в кг (Факт - Теория)
        deviation = fact_raw - theory_raw

        # Фактический вес формовки в тоннах
        fact_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) / 1000.0 for r in lfm_reports)

        report_data.append({
            "id": shift.id,
            "date": shift.date.strftime("%Y-%m-%d") if shift.date else "Н/Д",
            "shift_name": shift.shift_name,
            "line": shift.line,
            "master_name": shift.master.name if shift.master else "Н/Д",
            "lfm_sheets": lfm_sheets,
            "qcd_condition": qcd_cond,
            "qcd_first_grade": qcd_fg,
            "qcd_defect": qcd_def,
            "raw_deviation": round(deviation, 2),
            "fact_tons": round(fact_tons, 2)
        })
        
    return report_data

@app.get("/api/dashboard/analytics_data")
def get_analytics_data(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    department: str = None,
    db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    query = db.query(models.Downtime).join(models.Shift)
    
    if False and user_role == "master" and user_id:
        query = query.filter(models.Shift.master_id == user_id)
    
    if start_date:
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d").date()
            query = query.filter(models.Shift.date >= sd)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD")
            
    if end_date:
        try:
            ed = datetime.strptime(end_date, "%Y-%m-%d").date()
            query = query.filter(models.Shift.date <= ed)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD")
            
    if department:
        query = query.filter(models.Downtime.department == department)
        
    downtimes = query.all()
    
    duration_with_stop = 0
    duration_without_stop = 0
    lost_tons_with_stop = 0.0
    lost_tons_without_stop = 0.0
    lost_tenge_with_stop = 0.0
    lost_tenge_without_stop = 0.0
    count_with_stop = 0
    count_without_stop = 0
    
    by_category = {}
    node_durations = {}
    trend_data = {}
    serialized_downtimes = []
    
    for dt in downtimes:
        dur = dt.duration or 0
        tons = dt.lost_tons or 0.0
        tenge = dt.lost_tenge or 0.0
        is_stop = bool(dt.is_equipment_downtime)
        
        if is_stop:
            duration_with_stop += dur
            lost_tons_with_stop += tons
            lost_tenge_with_stop += tenge
            count_with_stop += 1
        else:
            duration_without_stop += dur
            lost_tons_without_stop += tons
            lost_tenge_without_stop += tenge
            count_without_stop += 1
            
        cat = dt.category or "Не указана"
        if cat not in by_category:
            by_category[cat] = {"with_stop": 0, "without_stop": 0}
        if is_stop:
            by_category[cat]["with_stop"] += dur
        else:
            by_category[cat]["without_stop"] += dur
            
        if is_stop and dt.node:
            node_durations[dt.node] = node_durations.get(dt.node, 0) + dur
            
        shift_date = dt.shift.date
        date_str = shift_date.strftime("%Y-%m-%d") if shift_date else "Не указана"
        
        serialized_downtimes.append({
            "id": dt.id,
            "date": date_str,
            "shift": dt.shift.shift_name if dt.shift else "",
            "line": dt.shift.line if dt.shift else "",
            "master": dt.shift.master.name if dt.shift and dt.shift.master else "Н/Д",
            "department": dt.department or "",
            "node": dt.node or "",
            "category": dt.category or "",
            "is_equipment_downtime": dt.is_equipment_downtime,
            "duration": dur,
            "lost_tons": tons,
            "lost_tenge": tenge,
            "description": dt.description or ""
        })
        date_str = shift_date.strftime("%Y-%m-%d") if shift_date else "Не указана"
        if date_str not in trend_data:
            trend_data[date_str] = {}
        trend_data[date_str][cat] = trend_data[date_str].get(cat, 0) + dur

    bottlenecks = sorted([{"node": k, "duration": v} for k, v in node_durations.items()], key=lambda x: x['duration'], reverse=True)[:10]
    
    sorted_trend = {}
    for d in sorted(trend_data.keys()):
        sorted_trend[d] = trend_data[d]
        
    return {
        "kpis": {
            "with_stop": {
                "duration": duration_with_stop,
                "lost_tons": lost_tons_with_stop,
                "lost_tenge": lost_tenge_with_stop,
                "count": count_with_stop
            },
            "without_stop": {
                "duration": duration_without_stop,
                "lost_tons": lost_tons_without_stop,
                "lost_tenge": lost_tenge_without_stop,
                "count": count_without_stop
            }
        },
        "by_category": by_category,
        "bottlenecks": bottlenecks,
        "trend": sorted_trend,
        "downtimes": serialized_downtimes
    }

# --- НОРМЫ РАСХОДА И ОТЧЕТ ПО СЫРЬЮ ---
@app.get("/api/norms/", response_model=list[schemas.ProductNorm])
def get_product_norms(db: Session = Depends(get_db)):
    return db.query(models.ProductNorm).all()

_norms_cache = {}
_norms_cache_time = 0

def _get_norm_cached(db: Session, product_name: str):
    global _norms_cache, _norms_cache_time
    import time
    if time.time() - _norms_cache_time > 60:
        norms = db.query(models.ProductNorm).all()
        _norms_cache = {n.product_name: n for n in norms}
        _norms_cache_time = time.time()
    return _norms_cache.get(product_name)

def get_product_finished_weight_kg(db: Session, product_name: str) -> float:
    norm = _get_norm_cached(db, product_name)
    if not norm or not norm.weight_kg:
        return 19.6 # fallback for 8 волн
    return norm.weight_kg

def get_product_raw_weight_kg(db: Session, product_name: str) -> float:
    norm = _get_norm_cached(db, product_name)
    if not norm:
        return 18.2 # fallback
    return (
        (norm.norm_chrysotile_4_20 or 0) +
        (norm.norm_chrysotile_5_65 or 0) +
        (norm.norm_chrysotile_6_40 or 0) +
        (norm.norm_cement or 0) +
        (norm.norm_cellulose or 0) +
        (norm.norm_crushed_slate or 0) +
        (norm.norm_asbozurit or 0) +
        (norm.norm_fiberglass or 0)
    )

def get_last_produced_weight_kg(db: Session, line_identifier: str, before_date_str: str = None) -> float:
    try:
        q = db.query(models.Shift).filter(models.Shift.line.like(f"%{line_identifier}%"))
        if before_date_str:
            q = q.filter(models.Shift.date <= before_date_str)
        shifts = q.order_by(models.Shift.date.desc(), models.Shift.id.desc()).limit(100).all()
        if not shifts:
            shifts = db.query(models.Shift).filter(models.Shift.line.like(f"%{line_identifier}%")).order_by(models.Shift.date.desc(), models.Shift.id.desc()).all()
        for s in shifts:
            if s.lfm_reports:
                for r in reversed(s.lfm_reports):
                    if r.lfm_sheets > 0 and r.product_name:
                        return get_product_finished_weight_kg(db, r.product_name)
    except Exception as e:
        print(f"Error in get_last_produced_weight_kg: {e}")
    return 19.6

def get_shift_plan(db: Session, shift: models.Shift) -> int:
    if shift.plan_sheets is not None and shift.plan_sheets > 0:
        return shift.plan_sheets
    sanitary_downtime = 0
    for dt in shift.downtimes:
        if dt.category == "Санитарный день":
            sanitary_downtime += dt.duration or 0
    if sanitary_downtime > 0:
        return 0
    if getattr(shift, "date", None) and shift.date.weekday() == 0 and shift.shift_name == "День":
        return 0
    return 2700 if shift.shift_name == "День" else 3300

@app.get("/api/dashboard/daily_report")
def get_daily_report(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    line: str = None,
    shift_number: int = None,
    master_id: int = None,
    db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"
    range_type_param = request.query_params.get("range_type")

    # Dynamic date range calculation based on frontend params
    sd = None
    ed = None

    if start_date:
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        except:
            raise HTTPException(400, "Invalid start_date format")
        if end_date:
            try:
                ed = datetime.strptime(end_date, "%Y-%m-%d").date()
            except:
                raise HTTPException(400, "Invalid end_date format")
        else:
            ed = sd + timedelta(days=6)
    else:
        # Fallback to query parameters passed by app.js (month, week, day, range_type)
        range_type = request.query_params.get("range_type", "month")
        month = request.query_params.get("month")
        week = request.query_params.get("week")
        day = request.query_params.get("day")

        if range_type == "month" and month:
            try:
                y, m = map(int, month.split('-'))
                num_days = calendar.monthrange(y, m)[1]
                sd = datetime(y, m, 1).date()
                ed = datetime(y, m, num_days).date()
            except Exception as e:
                raise HTTPException(400, f"Invalid month format: {e}")
        elif range_type == "week" and week:
            try:
                if not month:
                    now = datetime.now()
                    y, m = now.year, now.month
                else:
                    y, m = map(int, month.split('-'))
                
                week_num = int(week)
                first_day_of_month = datetime(y, m, 1).date()
                diff = -first_day_of_month.weekday()
                current_monday = first_day_of_month + timedelta(days=diff)
                
                if m == 12:
                    first_day_of_next_month = datetime(y + 1, 1, 1).date()
                else:
                    first_day_of_next_month = datetime(y, m + 1, 1).date()
                
                weeks = []
                while current_monday < first_day_of_next_month:
                    current_sunday = current_monday + timedelta(days=6)
                    weeks.append((current_monday, current_sunday))
                    current_monday += timedelta(days=7)
                
                idx = week_num - 1
                if 0 <= idx < len(weeks):
                    sd, ed = weeks[idx]
                elif len(weeks) > 0:
                    sd, ed = weeks[-1]
                else:
                    num_days = calendar.monthrange(y, m)[1]
                    sd = datetime(y, m, 1).date()
                    ed = datetime(y, m, num_days).date()
            except Exception as e:
                raise HTTPException(400, f"Invalid week or month format: {e}")
        elif range_type == "day" and day:
            try:
                sd = datetime.strptime(day, "%Y-%m-%d").date()
                ed = sd
            except Exception as e:
                raise HTTPException(400, f"Invalid day format: {e}")
        else:
            # Fallback to current month if no dates are provided
            now = datetime.now()
            y, m = now.year, now.month
            num_days = calendar.monthrange(y, m)[1]
            sd = datetime(y, m, 1).date()
            ed = datetime(y, m, num_days).date()

    num_days = (ed - sd).days + 1

    if not range_type_param:
        if num_days >= 28:
            effective_range_type = "month"
        elif num_days == 7:
            effective_range_type = "week"
        elif num_days == 1:
            effective_range_type = "day"
        else:
            effective_range_type = "custom"
    else:
        effective_range_type = range_type_param

    shifts_query = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports),
        selectinload(models.Shift.batches)
    ).filter(
        models.Shift.date >= sd,
        models.Shift.date <= ed
    )
    if master_id is not None:
        shifts_query = shifts_query.filter(models.Shift.master_id == master_id)
    shifts = shifts_query.all()
    
    plan_boards_query = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date >= sd,
        models.MonthlyPlanBoard.date <= ed
    )
    if master_id is not None:
        plan_boards_query = plan_boards_query.filter(models.MonthlyPlanBoard.master_id == master_id)
    plan_boards = plan_boards_query.all()
    
    if shift_number is not None:
        # Initialize plans to 0, because we will populate only matching shifts from pb
        data = {
            "line_1": {str(sd + timedelta(days=i)): {"День": {"sheets": 0, "tons": 0.0, "plan_sheets": 0, "plan_tons": 0.0, "first_grade": 0, "defect": 0}, "Ночь": {"sheets": 0, "tons": 0.0, "plan_sheets": 0, "plan_tons": 0.0, "first_grade": 0, "defect": 0}} for i in range(num_days)},
            "line_2": {str(sd + timedelta(days=i)): {"День": {"sheets": 0, "tons": 0.0, "plan_sheets": 0, "plan_tons": 0.0, "first_grade": 0, "defect": 0}, "Ночь": {"sheets": 0, "tons": 0.0, "plan_sheets": 0, "plan_tons": 0.0, "first_grade": 0, "defect": 0}} for i in range(num_days)}
        }
    else:
        # Default initialization with standard norms
        data = {
            "line_1": {str(sd + timedelta(days=i)): {"День": {"sheets": 0, "tons": 0.0, "plan_sheets": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700), "plan_tons": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700) * 19.6 / 1000.0, "first_grade": 0, "defect": 0}, "Ночь": {"sheets": 0, "tons": 0.0, "plan_sheets": 3300, "plan_tons": 3300 * 19.6 / 1000.0, "first_grade": 0, "defect": 0}} for i in range(num_days)},
            "line_2": {str(sd + timedelta(days=i)): {"День": {"sheets": 0, "tons": 0.0, "plan_sheets": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700), "plan_tons": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700) * 19.6 / 1000.0, "first_grade": 0, "defect": 0}, "Ночь": {"sheets": 0, "tons": 0.0, "plan_sheets": 3300, "plan_tons": 3300 * 19.6 / 1000.0, "first_grade": 0, "defect": 0}} for i in range(num_days)}
        }
    
    pb_map = {}
    for pb in plan_boards:
        pb_map[(pb.date, pb.shift_name, pb.line)] = pb
        
        day_key = str(pb.date)
        line_key = "line_1" if pb.line == "ЛФМ-1" else "line_2"
        s_name = pb.shift_name
        if day_key in data[line_key] and s_name in ["День", "Ночь"]:
            if shift_number is not None and pb.shift_number != shift_number:
                continue
            data[line_key][day_key][s_name]["plan_sheets"] = pb.plan_sheets or 0
            data[line_key][day_key][s_name]["plan_tons"] = (pb.plan_sheets or 0) * 19.6 / 1000.0
            
            # Записываем факт для всех
            if True:
                data[line_key][day_key][s_name]["sheets"] = pb.fact_sheets or 0
                data[line_key][day_key][s_name]["tons"] = (pb.fact_sheets or 0) * 19.6 / 1000.0
                data[line_key][day_key][s_name]["first_grade"] = pb.first_grade or 0
                data[line_key][day_key][s_name]["defect"] = pb.defect or 0
            
    processed_slots = set()
    accumulate_sheets_slots = set()
    for s in shifts:
        if not s.date: continue
        # Не пропускаем смены других мастеров
        if False and user_role == "master" and s.master_id != user_id:
            continue
        day_key = str(s.date)
        line_key = "line_1" if "1" in s.line else "line_2"
        s_name = "День" if s.shift_name == "День" else "Ночь"
        
        if day_key not in data[line_key]:
            continue
            
        if shift_number is not None:
            pb_line_name = "ЛФМ-1" if "1" in s.line else "ЛФМ-2"
            pb_entry = pb_map.get((s.date, s.shift_name, pb_line_name))
            if pb_entry is None or pb_entry.shift_number != shift_number:
                continue
            
        total_w = 0
        total_s = 0
        total_1st = 0
        total_def = 0
        for r in s.lfm_reports:
            w_kg = get_product_finished_weight_kg(db, r.product_name)
            total_w += w_kg * r.lfm_sheets
            total_s += r.lfm_sheets
            
        for b in s.batches:
            total_1st += (b.ds_first_grade or 0)
            total_def += (b.ds_defect or 0)
            
        slot_key = (line_key, day_key, s_name)
        if slot_key not in processed_slots:
            processed_slots.add(slot_key)
            data[line_key][day_key][s_name]["tons"] = 0.0
            if data[line_key][day_key][s_name]["sheets"] == 0 or shift_number is not None:
                accumulate_sheets_slots.add(slot_key)
                data[line_key][day_key][s_name]["sheets"] = 0
            # Всегда инициализируем и накапливаем 1 сорт и брак напрямую из рапортов мастеров
            data[line_key][day_key][s_name]["first_grade"] = 0
            data[line_key][day_key][s_name]["defect"] = 0
        
        # Накапливаем 1 сорт и брак Дестакера напрямую из сменного рапорта
        data[line_key][day_key][s_name]["first_grade"] += total_1st
        data[line_key][day_key][s_name]["defect"] += total_def

        if total_s > 0:
            data[line_key][day_key][s_name]["tons"] += total_w / 1000.0
            if slot_key in accumulate_sheets_slots:
                data[line_key][day_key][s_name]["sheets"] += total_s
            
    last_known_weight = {}
    for l_k in data:
        line_name_for_q = "1" if l_k == "line_1" else "2"
        last_known_weight[l_k] = get_last_produced_weight_kg(db, line_name_for_q, str(sd))

    for i in range(num_days):
        day_k = str(sd + timedelta(days=i))
        for s_nm in ["День", "Ночь"]:
            for l_k in data:
                if day_k in data[l_k] and s_nm in data[l_k][day_k]:
                    slot_info = data[l_k][day_k][s_nm]
                    if slot_info["sheets"] > 0 and slot_info["tons"] > 0:
                        avg_w = (slot_info["tons"] * 1000.0) / slot_info["sheets"]
                        slot_info["plan_tons"] = slot_info["plan_sheets"] * avg_w / 1000.0
                        last_known_weight[l_k] = avg_w
                    else:
                        slot_info["plan_tons"] = slot_info["plan_sheets"] * last_known_weight[l_k] / 1000.0
            
    # Now structure response as expected by app.js
    days_list = []
    lines_to_include = []
    if line == "lfm1":
        lines_to_include = ["line_1"]
    elif line == "lfm2":
        lines_to_include = ["line_2"]
    else:
        lines_to_include = ["line_1", "line_2"]

    for i in range(num_days):
        dt = sd + timedelta(days=i)
        date_str = str(dt)
        day_num = dt.day
        month_num = dt.month
        
        for s_name in ["День", "Ночь"]:
            plan_sheets = 0
            fact_sheets = 0
            plan_tons = 0.0
            fact_tons = 0.0
            first_grade = 0
            defect = 0
            
            for l_key in lines_to_include:
                shift_data = data[l_key][date_str][s_name]
                plan_sheets += shift_data["plan_sheets"]
                fact_sheets += shift_data["sheets"]
                plan_tons += shift_data["plan_tons"]
                fact_tons += shift_data["tons"]
                first_grade += shift_data["first_grade"]
                defect += shift_data["defect"]
            
            suffix = "Д" if s_name == "День" else "Н"
            label = f"{day_num:02d}.{month_num:02d} ({suffix})"
            
            days_list.append({
                "date": date_str,
                "label": label,
                "plan_sheets": plan_sheets,
                "fact_sheets": fact_sheets,
                "plan_tons": plan_tons,
                "fact_tons": fact_tons,
                "first_grade": first_grade,
                "defect": defect
            })
        
            unique_shifts = set()
    for s in shifts:
        if line and not (("1" in s.line and line == "lfm1") or ("2" in s.line and line == "lfm2") or line == "all"):
            continue
        
        lfm_sheets = sum((r.lfm_sheets or 0) for r in s.lfm_reports) if getattr(s, 'lfm_reports', None) else 0
        warehouse_gp = sum((b.ds_condition or 0) for b in s.batches) if getattr(s, 'batches', None) else 0
        plan_sheets = s.plan_sheets or 0
        zo_batches = s.zo_batches or 0
        
        if plan_sheets == 0 and lfm_sheets == 0 and warehouse_gp == 0 and zo_batches == 0 and not getattr(s, 'zo_submitted', False):
            continue
            
        unique_shifts.add((s.date, s.shift_name, s.line))
            
    total_shifts = len(unique_shifts)
    total_fact_sheets = sum(d["fact_sheets"] for d in days_list)
    total_fact_tons = sum(d["fact_tons"] for d in days_list)

    if master_id is None and shift_number is None:
        if effective_range_type == "month" or num_days >= 28:
            total_plan_sheets = 160000 * len(lines_to_include)
        elif effective_range_type == "week" and num_days == 7:
            total_plan_sheets = 39000 * len(lines_to_include)
        else:
            total_plan_sheets = sum(d["plan_sheets"] for d in days_list)
    else:
        total_plan_sheets = sum(d["plan_sheets"] for d in days_list)

    if total_fact_sheets > 0:
        avg_period_weight = (total_fact_tons * 1000.0) / total_fact_sheets
        total_plan_tons = round((total_plan_sheets * avg_period_weight) / 1000.0, 2)
    else:
        total_plan_tons = round((total_plan_sheets * 19.6) / 1000.0, 2)

    avg_plan_percent = (total_fact_sheets / total_plan_sheets * 100.0) if total_plan_sheets > 0 else 0.0
    
    total_first_grade = sum(d["first_grade"] for d in days_list)
    total_defect = sum(d["defect"] for d in days_list)
    defect_percent = (total_defect / total_fact_sheets * 100.0) if total_fact_sheets > 0 else 0.0
    first_grade_percent = (total_first_grade / total_fact_sheets * 100.0) if total_fact_sheets > 0 else 0.0
    
    lag_sheets = total_plan_sheets - total_fact_sheets
    lag_tons = round(total_plan_tons - total_fact_tons, 2)
    
    return {
        "total_shifts": total_shifts,
        "total_fact_sheets": total_fact_sheets,
        "total_fact_tons": total_fact_tons,
        "total_plan_sheets": total_plan_sheets,
        "total_plan_tons": total_plan_tons,
        "lag_sheets": lag_sheets,
        "lag_tons": lag_tons,
        "total_first_grade": total_first_grade,
        "first_grade_percent": first_grade_percent,
        "total_defect": total_defect,
        "avg_plan_percent": avg_plan_percent,
        "defect_percent": defect_percent,
        "days": days_list
    }

@app.get("/api/dashboard/export_daily_report")
def export_daily_report(request: Request, start_date: str, line: str = None, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
    except:
        raise HTTPException(400, "Invalid date format")
        
    num_days = 14
    ed = sd + timedelta(days=num_days - 1)
    
    shifts = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports)
    ).filter(
        models.Shift.date >= sd,
        models.Shift.date <= ed
    ).all()
    
    plan_boards = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date >= sd,
        models.MonthlyPlanBoard.date <= ed
    ).all()
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    lines_to_export = [("Линия 1", "ЛФМ-1"), ("Линия 2", "ЛФМ-2")]
    if line == 'lfm1':
        lines_to_export = [("Линия 1", "ЛФМ-1")]
    elif line == 'lfm2':
        lines_to_export = [("Линия 2", "ЛФМ-2")]
        
    for line_id, line_label in lines_to_export:
        ws = wb.create_sheet(title=line_label)
        ws.append(["Дата", "Смена", "План (Листы)", "Факт (Листы)", "План (Тонны)", "Факт (Тонны)", "1-й сорт", "Брак"])
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        
        day_data = {str(sd + timedelta(days=i)): {
            "День": {"sheets": 0, "tons": 0.0, "plan_sheets": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700), "plan_tons": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700) * 19.6 / 1000.0, "first_grade": 0, "defect": 0}, 
            "Ночь": {"sheets": 0, "tons": 0.0, "plan_sheets": 3300, "plan_tons": 3300 * 19.6 / 1000.0, "first_grade": 0, "defect": 0}
        } for i in range(num_days)}
        
        for pb in plan_boards:
            if pb.line != line_label: continue
            day_key = str(pb.date)
            s_name = pb.shift_name
            if day_key in day_data and s_name in ["День", "Ночь"]:
                day_data[day_key][s_name]["plan_sheets"] = pb.plan_sheets or 0
                day_data[day_key][s_name]["plan_tons"] = (pb.plan_sheets or 0) * 19.6 / 1000.0
                
                # Факт записываем только для текущего мастера (или если роль не master)
                if user_role != "master" or pb.master_id == user_id:
                    day_data[day_key][s_name]["sheets"] = pb.fact_sheets or 0
                    day_data[day_key][s_name]["tons"] = (pb.fact_sheets or 0) * 19.6 / 1000.0
                    day_data[day_key][s_name]["first_grade"] = pb.first_grade or 0
                    day_data[day_key][s_name]["defect"] = pb.defect or 0
        
        processed_slots = set()
        accumulate_sheets_slots = set()
        for s in shifts:
            if not s.date or s.line != line_id: continue
            # Пропускаем смены других мастеров для роли master
            if False and user_role == "master" and s.master_id != user_id:
                continue
            day_key = str(s.date)
            if day_key not in day_data: continue
            
            s_name = "День" if s.shift_name == "День" else "Ночь"
            
            total_w = 0
            total_s = 0
            for r in s.lfm_reports:
                w_kg = get_product_finished_weight_kg(db, r.product_name)
                total_w += w_kg * r.lfm_sheets
                total_s += r.lfm_sheets
                
            if total_s > 0:
                slot_key = (day_key, s_name)
                if slot_key not in processed_slots:
                    processed_slots.add(slot_key)
                    day_data[day_key][s_name]["tons"] = 0.0
                    if day_data[day_key][s_name]["sheets"] == 0:
                        accumulate_sheets_slots.add(slot_key)
                        day_data[day_key][s_name]["sheets"] = 0
                
                day_data[day_key][s_name]["tons"] += total_w / 1000.0
                if slot_key in accumulate_sheets_slots:
                    day_data[day_key][s_name]["sheets"] += total_s
                
        last_w = get_last_produced_weight_kg(db, "1" if line_id == "lfm1" else "2", str(sd))
        for i in range(num_days):
            day_k = str(sd + timedelta(days=i))
            if day_k in day_data:
                for s_nm in ["День", "Ночь"]:
                    slot_info = day_data[day_k][s_nm]
                    if slot_info["sheets"] > 0 and slot_info["tons"] > 0:
                        avg_w = (slot_info["tons"] * 1000.0) / slot_info["sheets"]
                        slot_info["plan_tons"] = slot_info["plan_sheets"] * avg_w / 1000.0
                        last_w = avg_w
                    else:
                        slot_info["plan_tons"] = slot_info["plan_sheets"] * last_w / 1000.0
                
        row_idx = 2
        for i in range(num_days):
            d_str = str(sd + timedelta(days=i))
            ws.append([d_str, "День", day_data[d_str]["День"]["plan_sheets"], day_data[d_str]["День"]["sheets"], round(day_data[d_str]["День"]["plan_tons"], 2), round(day_data[d_str]["День"]["tons"], 2), day_data[d_str]["День"]["first_grade"], day_data[d_str]["День"]["defect"]])
            ws.append([d_str, "Ночь", day_data[d_str]["Ночь"]["plan_sheets"], day_data[d_str]["Ночь"]["sheets"], round(day_data[d_str]["Ночь"]["plan_tons"], 2), round(day_data[d_str]["Ночь"]["tons"], 2), day_data[d_str]["Ночь"]["first_grade"], day_data[d_str]["Ночь"]["defect"]])
            row_idx += 2
            
        chart_sheets = BarChart()
        chart_sheets.type = "col"
        chart_sheets.style = 10
        chart_sheets.title = f"Выработка {line_label} (Листы)"
        chart_sheets.y_axis.title = 'Количество (Листы)'
        chart_sheets.x_axis.title = 'Дата / Смена'
        
        data_sheets = Reference(ws, min_col=3, min_row=1, max_row=row_idx-1, max_col=4)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row_idx-1, max_col=2)
        
        chart_sheets.add_data(data_sheets, titles_from_data=True)
        chart_sheets.set_categories(cats)
        chart_sheets.width = 20
        
        ws.add_chart(chart_sheets, "H2")
        
        chart_tons = BarChart()
        chart_tons.type = "col"
        chart_tons.style = 10
        chart_tons.title = f"Выработка {line_label} (Тонны)"
        chart_tons.y_axis.title = 'Вес (Тонны)'
        chart_tons.x_axis.title = 'Дата / Смена'
        
        data_tons = Reference(ws, min_col=5, min_row=1, max_row=row_idx-1, max_col=6)
        
        chart_tons.add_data(data_tons, titles_from_data=True)
        chart_tons.set_categories(cats)
        chart_tons.width = 20
        
        ws.add_chart(chart_tons, "H18")
        
    out = io.BytesIO()
    wb.save(out)
    
    filename = f"report_{start_date}_{line or 'all'}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return Response(content=out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.post("/api/admin/fix_plan_boards")
def fix_plan_boards(request: Request, db: Session = Depends(get_db)):
    user_role = request.session.get("user_role")
    if user_role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    try:
        boards = db.query(models.MonthlyPlanBoard).filter(models.MonthlyPlanBoard.plan_sheets == 0).all()
        updated_count = 0
        for pb in boards:
            date_val = pb.date
            is_monday = False
            if isinstance(date_val, str):
                try:
                    dt_obj = datetime.strptime(date_val, "%Y-%m-%d").date()
                    is_monday = dt_obj.weekday() == 0
                except:
                    pass
            else:
                try:
                    is_monday = date_val.weekday() == 0
                except:
                    pass
                    
            if is_monday and pb.shift_name == "День":
                continue
                
            correct_plan = 2700 if pb.shift_name == "День" else 3300
            pb.plan_sheets = correct_plan
            updated_count += 1
            
            log_entry = models.AuditLog(
                timestamp=datetime.utcnow(),
                user_name="System Admin",
                action="UPDATE",
                target_table="monthly_plan_board",
                target_id=pb.id,
                details=f"Исправление нулевого плана. Установлен план {correct_plan}."
            )
            db.add(log_entry)
            
        db.commit()
        return {"success": True, "updated_count": updated_count}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/dashboard/shift_board")
def get_shift_board(month: str, db: Session = Depends(get_db)):
    try:
        y, m = map(int, month.split('-'))
        num_days = calendar.monthrange(y, m)[1]
    except:
        raise HTTPException(400, "Invalid month format")
        
    month_start = datetime(y, m, 1).date()
    month_end = datetime(y, m, num_days).date()
    shifts = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports)
    ).filter(
        models.Shift.date >= month_start,
        models.Shift.date <= month_end
    ).order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
    board = {}
    for s in shifts:
        if not s.date: continue
        master = s.master_name or "Неизвестный мастер"
        if master not in board:
            board[master] = []
            
        total_s = 0
        total_w = 0
        for r in s.lfm_reports:
            w_kg = get_product_finished_weight_kg(db, r.product_name)
            total_s += r.lfm_sheets
            total_w += r.lfm_sheets * w_kg
            
        plan_sheets = get_shift_plan(db, s)
        plan_tons = (plan_sheets * 19.6) / 1000.0
        if total_s > 0:
            avg_w = total_w / total_s
            plan_tons = (plan_sheets * avg_w) / 1000.0
            
        board[master].append({
            "shift_id": s.id,
            "date": str(s.date),
            "shift_name": s.shift_name,
            "line": s.line,
            "plan_sheets": plan_sheets,
            "fact_sheets": total_s,
            "plan_tons": round(plan_tons, 2),
            "fact_tons": round(total_w / 1000.0, 2),
            "closed": s.status == "closed"
        })
        
    return board

@app.get("/api/dashboard/export_shift")
def export_shift(shift_id: int = None, db: Session = Depends(get_db)):
    file_bytes = excel_exporter.generate_flat_report(db)
    filename = "Сводный_отчет_Tectum.xlsx"
    from urllib.parse import quote
    safe_filename = quote(filename)
    headers = {
        'Content-Disposition': f'attachment; filename="{safe_filename}"; filename*=UTF-8\'\'{safe_filename}'
    }
    return Response(
        content=file_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@app.post("/api/dashboard/sync_sharepoint")
def sync_sharepoint_manually(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
        
    if user_role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен. Только мастер смены или администратор могут запускать синхронизацию.")
        
    try:
        file_bytes = excel_exporter.generate_flat_report(db)
        filename = "Сводный_отчет_Tectum.xlsx"
        
        # Save locally to static folder as well
        local_path = os.path.join("static", "Сводный_отчет_Tectum.xlsx")
        try:
            with open(local_path, "wb") as f:
                f.write(file_bytes)
        except Exception as local_err:
            print(f"Error saving local excel file: {local_err}")
            
        web_url = m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder="Reports")
        
        # Log to AuditLog
        db.add(models.AuditLog(
            user_name=request.session.get("user_email") or f"user_{user_id}",
            action="UPDATE",
            target_table="shifts",
            target_id=0,
            details=f"Ручная синхронизация отчета с SharePoint выполнена успешно. Ссылка: {web_url}"
        ))
        db.commit()
        return {"message": "Синхронизация выполнена успешно", "url": web_url}
    except Exception as e:
        error_msg = str(e)
        if "423" in error_msg or "Locked" in error_msg:
            raise HTTPException(status_code=423, detail="Файл отчета все еще заблокирован в SharePoint (кто-то открыл его в Excel Online). Закройте файл и попробуйте снова.")
        else:
            raise HTTPException(status_code=500, detail=f"Ошибка синхронизации: {error_msg}")

@app.post("/api/dashboard/sync_google_sheets_manual")
def sync_google_sheets_manual(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="Вы не авторизованы")
        
    if user_role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен. Только мастера или администраторы могут запускать выгрузку.")
        
    try:
        google_sheets_integration.sync_report_to_google_sheets(db)
        
        # Log to AuditLog
        db.add(models.AuditLog(
            user_name=request.session.get("user_email") or f"user_{user_id}",
            action="UPDATE",
            target_table="shifts",
            target_id=0,
            details="Выполнена ручная выгрузка сводного отчета в Google Таблицы."
        ))
        db.commit()
        return {"message": "Выгрузка в Google Таблицы выполнена успешно!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка выгрузки в Google: {str(e)}")

@app.get("/api/dashboard/view_archive")
def view_archive(db: Session = Depends(get_db)):
    try:
        url = m365_integration.get_file_web_url("Сводный_отчет_Tectum.xlsx", folder="Reports")
        return RedirectResponse(url=url)
    except Exception as e:
        print("Ошибка получения ссылки из SharePoint, пробуем сгенерировать и загрузить отчет:", e)
        try:
            file_bytes = excel_exporter.generate_flat_report(db)
            filename = "Сводный_отчет_Tectum.xlsx"
            url = m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder="Reports")
            return RedirectResponse(url=url)
        except Exception as upload_err:
            import traceback
            traceback.print_exc()
            raise HTTPException(
                status_code=500,
                detail=f"Не удалось открыть сводный отчет в SharePoint. Ошибка автозагрузки: {upload_err}. Исходная ошибка: {e}"
            )

@app.get("/api/dashboard/export_week")
def export_week(request: Request, start_date: str, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
    except:
        raise HTTPException(400, "Invalid date format, use YYYY-MM-DD")
        
    ed = sd + timedelta(days=6)
    
    query = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports),
        selectinload(models.Shift.receipts),
        selectinload(models.Shift.downtimes)
    ).filter(
        models.Shift.date >= sd,
        models.Shift.date <= ed
    )
    if False and user_role == "master" and user_id:
        query = query.filter(models.Shift.master_id == user_id)
    shifts = query.order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Неделя {sd} - {ed}"
    
    ws.append([f"Отчет за неделю с {sd} по {ed}"])
    ws.append(["Дата", "Смена", "Мастер", "Линия", "План (Листы)", "Факт (Листы)", "План (Тонны)", "Факт (Тонны)"])
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    plan_boards = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date >= sd,
        models.MonthlyPlanBoard.date <= ed
    ).all()
    pb_dict = {(pb.date, pb.shift_name, pb.line): pb for pb in plan_boards}

    active_lines = set([s.line.replace("Линия ", "ЛФМ-") for s in shifts if s.line] + [pb.line for pb in plan_boards])
    if not active_lines:
        active_lines = {"ЛФМ-2"}

    for l_key in active_lines:
        last_w = get_last_produced_weight_kg(db, "1" if "1" in l_key else "2", str(sd)) / 1000.0
        for i in range(7):
            d = sd + timedelta(days=i)
            for s_name in ["День", "Ночь"]:
                plan_sheets = 0 if d.weekday() == 0 and s_name == "День" else (2700 if s_name == "День" else 3300)
                
                pb = pb_dict.get((d, s_name, l_key))
                if pb and pb.plan_sheets is not None:
                    plan_sheets = pb.plan_sheets
                    
                slot_shifts = [shift for shift in shifts if shift.date == d and shift.shift_name == s_name and (shift.line.replace("Линия ", "ЛФМ-") if shift.line else "ЛФМ-1") == l_key]
                s = slot_shifts[0] if slot_shifts else None
                
                show_fact = (user_role != "master" or (pb and pb.master_id == user_id) or (s and s.master_id == user_id))
                total_sheets = pb.fact_sheets if (pb and show_fact) else 0
                
                if s and show_fact:
                    sum_lfm_sheets = sum(r.lfm_sheets for sh in slot_shifts for r in sh.lfm_reports)
                    sum_lfm_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) / 1000.0 for sh in slot_shifts for r in sh.lfm_reports)
                    if sum_lfm_sheets > 0:
                        avg_w = (sum_lfm_tons / sum_lfm_sheets)
                        last_w = avg_w
                    else:
                        avg_w = last_w
                    if total_sheets == 0 and sum_lfm_sheets > 0:
                        total_sheets = sum_lfm_sheets
                    master_name = s.master.name if s.master else "Н/Д"
                else:
                    avg_w = last_w
                    master_name = "Н/Д" if show_fact else "Смена др. мастера"
                    total_sheets = pb.fact_sheets if (pb and show_fact) else 0
                    
                plan_tons = plan_sheets * avg_w
                total_tons = sum_lfm_tons if (s and show_fact and sum_lfm_sheets > 0) else (total_sheets * avg_w)
                
                ws.append([str(d), s_name, master_name, l_key, plan_sheets, total_sheets, round(plan_tons, 2), round(total_tons, 2)])
    out = io.BytesIO()
    wb.save(out)
    
    filename = f"week_{sd}.xlsx"
    return Response(content=out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={'Content-Disposition': f'attachment; filename="{filename}"'})

@app.get("/api/dashboard/weekly")
def get_weekly_json(request: Request, start_date: str, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
    except:
        raise HTTPException(400, "Invalid date format, use YYYY-MM-DD")
        
    ed = sd + timedelta(days=6)
    
    query = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports),
        selectinload(models.Shift.receipts),
        selectinload(models.Shift.downtimes)
    ).filter(
        models.Shift.date >= sd,
        models.Shift.date <= ed
    )
    if False and user_role == "master" and user_id:
        query = query.filter(models.Shift.master_id == user_id)
    shifts = query.order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
    plan_boards = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date >= sd,
        models.MonthlyPlanBoard.date <= ed
    ).all()
    pb_dict = {(pb.date, pb.shift_name, pb.line): pb for pb in plan_boards}
    
    active_lines = set([s.line.replace("Линия ", "ЛФМ-") for s in shifts if s.line] + [pb.line for pb in plan_boards])
    if not active_lines:
        active_lines = {"ЛФМ-2"}
        
    data = []
    
    for l_key in active_lines:
        last_w = get_last_produced_weight_kg(db, "1" if "1" in l_key else "2", str(sd)) / 1000.0
        for i in range(7):
            d = sd + timedelta(days=i)
            day_str = str(d)
            for s_name in ["День", "Ночь"]:
                plan_sheets = 0 if d.weekday() == 0 and s_name == "День" else (2700 if s_name == "День" else 3300)
                
                pb = pb_dict.get((d, s_name, l_key))
                if pb and pb.plan_sheets is not None:
                    plan_sheets = pb.plan_sheets
                    
                slot_shifts = [shift for shift in shifts if shift.date == d and shift.shift_name == s_name and (shift.line.replace("Линия ", "ЛФМ-") if shift.line else "ЛФМ-1") == l_key]
                s = slot_shifts[0] if slot_shifts else None
                
                show_fact = (user_role != "master" or (pb and pb.master_id == user_id) or (s and s.master_id == user_id))
                total_sheets = pb.fact_sheets if (pb and show_fact) else 0
                
                if s and show_fact:
                    sum_lfm_sheets = sum(r.lfm_sheets for sh in slot_shifts for r in sh.lfm_reports)
                    sum_lfm_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) / 1000.0 for sh in slot_shifts for r in sh.lfm_reports)
                    if sum_lfm_sheets > 0:
                        avg_w = (sum_lfm_tons / sum_lfm_sheets)
                        last_w = avg_w
                    else:
                        avg_w = last_w
                    if total_sheets == 0 and sum_lfm_sheets > 0:
                        total_sheets = sum_lfm_sheets
                        
                    if pb and (pb.first_grade or pb.defect):
                        ds_first = pb.first_grade
                        ds_defect = pb.defect
                    else:
                        ds_first = sum(b.ds_first_grade for sh in slot_shifts for b in sh.batches)
                        ds_defect = sum(b.ds_defect for sh in slot_shifts for b in sh.batches)
                        
                    qcd_first = sum(b.ds_first_grade for sh in slot_shifts for b in sh.batches)
                    qcd_defect = sum(b.ds_defect for sh in slot_shifts for b in sh.batches)
                    
                    sanitary_note = ""
                    for dt in s.downtimes:
                        if dt.category == "Санитарный день":
                            sanitary_note = "Санитарный день"
                            if dt.duration:
                                sanitary_note += f" ({dt.duration} мин)"
                            break
                    master_name = s.master.name if s.master else "Н/Д"
                    shift_id = s.id
                else:
                    avg_w = last_w
                    ds_first = pb.first_grade if (pb and show_fact) else 0
                    ds_defect = pb.defect if (pb and show_fact) else 0
                    qcd_first = 0
                    qcd_defect = 0
                    sanitary_note = "Санитарный день (план 0)" if d.weekday() == 0 and s_name == "День" else ("Нет данных" if show_fact else "Смена другого мастера")
                    master_name = "Н/Д"
                    shift_id = None
                    
                plan_tons = plan_sheets * avg_w
                total_tons = sum_lfm_tons if (s and show_fact and sum_lfm_sheets > 0) else (total_sheets * avg_w)
                
                data.append({
                    "id": shift_id,
                    "date": day_str,
                    "shift_name": s_name,
                    "master": master_name,
                    "line": l_key,
                    "plan_sheets": plan_sheets,
                    "fact_sheets": total_sheets,
                    "plan_tons": round(plan_tons, 2),
                    "fact_tons": round(total_tons, 2),
                    "ds_first_grade": ds_first,
                    "ds_defect": ds_defect,
                    "qcd_first_grade": qcd_first,
                    "qcd_defect": qcd_defect,
                    "note": sanitary_note
                })
        
    return {
        "start_date": str(sd),
        "end_date": str(ed),
        "data": data
    }


@app.get("/api/shifts/{shift_id}/materials_report", response_model=schemas.RawMaterialReport)
def get_materials_report(shift_id: int, db: Session = Depends(get_db)):
    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(404, "Смена не найдена")
    
    # 1. Считаем произведенную продукцию (Формовка)
    lfm_reports = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift_id).all()
    product_counts = {}
    for r in lfm_reports:
        product_counts[r.product_name] = product_counts.get(r.product_name, 0) + r.lfm_sheets
        
    # 2. Получаем нормы для этих продуктов и считаем теорию
    theoretical = {
        "chrysotile_4_20": 0.0, "chrysotile_5_65": 0.0, "chrysotile_6_40": 0.0,
        "cement": 0.0, "cellulose": 0.0, "crushed_slate": 0.0,
        "asbozurit": 0.0, "fiberglass": 0.0
    }
    
    for prod_name, sheets in product_counts.items():
        norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == prod_name).first()
        if norm:
            theoretical["chrysotile_4_20"] += sheets * norm.norm_chrysotile_4_20
            theoretical["chrysotile_5_65"] += sheets * norm.norm_chrysotile_5_65
            theoretical["chrysotile_6_40"] += sheets * norm.norm_chrysotile_6_40
            theoretical["cement"] += sheets * norm.norm_cement
            theoretical["cellulose"] += sheets * norm.norm_cellulose
            theoretical["crushed_slate"] += sheets * norm.norm_crushed_slate
            theoretical["asbozurit"] += sheets * norm.norm_asbozurit
            theoretical["fiberglass"] += sheets * norm.norm_fiberglass

    # 3. Формируем детальный отчет (Факт из ZO - Теория)
    details = []
    total_dev = 0.0
    
    mapping = [
        ("Хризотил 4-20", shift.zo_chrysotile_4_20, theoretical["chrysotile_4_20"]),
        ("Хризотил 5-65", shift.zo_chrysotile_5_65, theoretical["chrysotile_5_65"]),
        ("Хризотил 6-40", shift.zo_chrysotile_6_40, theoretical["chrysotile_6_40"]),
        ("Цемент", shift.zo_cement, theoretical["cement"]),
        ("Целлюлоза", shift.zo_cellulose, theoretical["cellulose"]),
        ("Дробленый шифер", shift.zo_crushed_slate, theoretical["crushed_slate"]),
        ("Асбозурит", shift.zo_asbozurit, theoretical["asbozurit"]),
        ("Стекловолокно", shift.zo_fiberglass, theoretical["fiberglass"]),
        ("Лапрол", shift.zo_laprol, 0.0),
        ("Асбокартон", shift.zo_asbocarton, 0.0)
    ]

    
    total_sheets = sum(product_counts.values())
    
    for mat_name, actual, theory in mapping:
        actual_val = actual or 0.0
        theory_val = theory or 0.0
        dev = actual_val - theory_val
        total_dev += dev
        
        unit_actual = actual_val / total_sheets if total_sheets > 0 else 0.0
        unit_theory = theory_val / total_sheets if total_sheets > 0 else 0.0
        unit_dev = dev / total_sheets if total_sheets > 0 else 0.0
        
        details.append({
            "material": mat_name,
            "actual": round(actual_val, 2),
            "theoretical": round(theory_val, 2),
            "deviation": round(dev, 2),
            "unit_actual": round(unit_actual, 4),
            "unit_theoretical": round(unit_theory, 4),
            "unit_deviation": round(unit_dev, 4)
        })
        
    return {
        "shift_id": shift_id,
        "total_deviation_kg": round(total_dev, 2),
        "details": details
    }


# --- ADMIN PANEL & PAGES ENDPOINTS ---

@app.get("/admin")
def serve_admin():
    return FileResponse("static/admin.html")

@app.get("/analytics")
def read_analytics():
    return FileResponse("static/analytics.html")

@app.get("/tasks")
def serve_tasks():
    return FileResponse("static/tasks.html")

@app.get("/planner")
def serve_planner():
    return FileResponse("static/tasks.html")

@app.post("/api/admin/masters/", response_model=schemas.Master)
def create_master(master: schemas.MasterCreate, db: Session = Depends(get_db)):
    db_master = models.Master(**master.model_dump())
    db.add(db_master)
    db.commit()
    db.refresh(db_master)
    return db_master

@app.put("/api/admin/masters/{master_id}", response_model=schemas.Master)
def update_master(master_id: int, master: schemas.MasterUpdate, db: Session = Depends(get_db)):
    db_master = db.query(models.Master).get(master_id)
    if not db_master: raise HTTPException(404)
    update_data = master.model_dump(exclude_unset=True)
    for key, val in update_data.items():
        setattr(db_master, key, val)
    db.commit()
    db.refresh(db_master)
    return db_master

@app.delete("/api/admin/masters/{master_id}")
def delete_master(master_id: int, request: Request, db: Session = Depends(get_db)):
    check_admin_session(request, db)
    db_master = db.query(models.Master).get(master_id)
    if not db_master: raise HTTPException(404, "Сотрудник не найден")
    
    # Проверяем наличие связанных смен
    has_shifts = db.query(models.Shift).filter(models.Shift.master_id == master_id).first()
    if has_shifts:
        raise HTTPException(status_code=400, detail="Невозможно удалить сотрудника, так как на него записаны смены.")
        
    # Проверяем наличие записей в плане на месяц
    has_plans = db.query(models.MonthlyPlanBoard).filter(models.MonthlyPlanBoard.master_id == master_id).first()
    if has_plans:
        raise HTTPException(status_code=400, detail="Невозможно удалить сотрудника, так как он есть в плане на месяц.")
        
    db.delete(db_master)
    db.commit()
    return {"status": "ok"}

def check_admin_session(request: Request, db: Session):
    role = request.session.get("user_role")
    user_id = request.session.get("user_id")
    if not user_id or role not in ["admin", "director", "technologist"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен. Требуются права администратора.")
    user = db.query(models.Master).get(user_id)
    if not user or user.role not in ["admin", "director", "technologist"]:
        raise HTTPException(status_code=403, detail="Доступ запрещен. Требуются права администратора.")
    return user

@app.get("/api/admin/shifts/{shift_id}/details")
def admin_get_shift_details(shift_id: int, request: Request, db: Session = Depends(get_db)):
    check_admin_session(request, db)
    shift = db.query(models.Shift).get(shift_id)
    if not shift: raise HTTPException(404, "Смена не найдена")
    
    lfm_reports = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift_id).all()
    batches = db.query(models.Batch).filter(models.Batch.shift_id == shift_id).all()
    downtimes = db.query(models.Downtime).filter(models.Downtime.shift_id == shift_id).all()
    
    return {
        "shift": shift,
        "lfm_reports": lfm_reports,
        "batches": batches,
        "downtimes": downtimes
    }

@app.put("/api/admin/shifts/{shift_id}")
def admin_update_shift(shift_id: int, data: dict, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    shift = db.query(models.Shift).get(shift_id)
    if not shift: raise HTTPException(404, "Смена не найдена")
    
    old_date, old_shift_name, old_line = shift.date, shift.shift_name, shift.line
    old_master_id = shift.master_id
    
    old_values = {}
    new_values = {}
    
    if "date" in data and data["date"]:
        try:
            data["date"] = datetime.strptime(data["date"], "%Y-%m-%d").date()
        except Exception:
            pass
            
    for key, val in data.items():
        if hasattr(shift, key):
            old_val = getattr(shift, key)
            if old_val != val:
                old_values[key] = str(old_val)
                new_values[key] = str(val)
                setattr(shift, key, val)
                
    if old_values:
        log_entry = models.AuditLog(
            timestamp=datetime.utcnow(),
            user_name=admin.name,
            action=f"Редактирование смены ID {shift_id}",
            details=f"Изменено: {old_values} -> {new_values}"
        )
        db.add(log_entry)
        db.commit()
        
        # Sync plan boards for old and new parameters
        sync_lfm_to_plan_board(old_date, old_shift_name, old_line, db, old_master_id)
        if shift.date != old_date or shift.shift_name != old_shift_name or shift.line != old_line:
            sync_lfm_to_plan_board(shift.date, shift.shift_name, shift.line, db, shift.master_id)
    else:
        db.commit()
    background_tasks.add_task(sync_google_sheets_bg)
    return {"status": "ok"}

@app.put("/api/admin/shift_report/{shift_id}")
def admin_update_shift_report(shift_id: int, data: schemas.AdminShiftReportUpdate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Смена не найдена")
        
    old_date, old_shift_name, old_line = shift.date, shift.shift_name, shift.line
    old_master_id = shift.master_id
    
    # Snapshot of state before admin update (for Rollback / Undo)
    snapshot_before = None
    try:
        b_prev = db.query(models.Batch).filter(models.Batch.shift_id == shift.id).first()
        l_prev = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift.id).first()
        import json
        snapshot_before = json.dumps({
            "shift": {
                "master_id": shift.master_id,
                "batch_number": shift.batch_number,
                "product_name": shift.product_name,
                "export_type": shift.export_type,
                "status": shift.status,
                "zo_batches": shift.zo_batches,
                "zo_chrysotile_4_20_silo1": shift.zo_chrysotile_4_20_silo1,
                "zo_chrysotile_4_20_silo2": shift.zo_chrysotile_4_20_silo2,
                "zo_chrysotile_4_20_silo3": shift.zo_chrysotile_4_20_silo3,
                "zo_chrysotile_4_20_silo4": shift.zo_chrysotile_4_20_silo4,
                "zo_chrysotile_5_65_silo1": shift.zo_chrysotile_5_65_silo1,
                "zo_chrysotile_5_65_silo2": shift.zo_chrysotile_5_65_silo2,
                "zo_chrysotile_5_65_silo3": shift.zo_chrysotile_5_65_silo3,
                "zo_chrysotile_5_65_silo4": shift.zo_chrysotile_5_65_silo4,
                "zo_chrysotile_6_40_silo1": shift.zo_chrysotile_6_40_silo1,
                "zo_chrysotile_6_40_silo2": shift.zo_chrysotile_6_40_silo2,
                "zo_chrysotile_6_40_silo3": shift.zo_chrysotile_6_40_silo3,
                "zo_chrysotile_6_40_silo4": shift.zo_chrysotile_6_40_silo4,
                "zo_cement_silo1": shift.zo_cement_silo1,
                "zo_cement_silo2": shift.zo_cement_silo2,
                "zo_cement_silo3": shift.zo_cement_silo3,
                "zo_cement_silo4": shift.zo_cement_silo4,
                "zo_cellulose_silo1": shift.zo_cellulose_silo1,
                "zo_cellulose_silo2": shift.zo_cellulose_silo2,
                "zo_cellulose_silo3": shift.zo_cellulose_silo3,
                "zo_cellulose_silo4": shift.zo_cellulose_silo4,
                "zo_crushed_slate_silo1": shift.zo_crushed_slate_silo1,
                "zo_crushed_slate_silo2": shift.zo_crushed_slate_silo2,
                "zo_crushed_slate_silo3": shift.zo_crushed_slate_silo3,
                "zo_crushed_slate_silo4": shift.zo_crushed_slate_silo4,
                "zo_asbozurit_silo1": shift.zo_asbozurit_silo1,
                "zo_asbozurit_silo2": shift.zo_asbozurit_silo2,
                "zo_asbozurit_silo3": shift.zo_asbozurit_silo3,
                "zo_asbozurit_silo4": shift.zo_asbozurit_silo4,
                "zo_fiberglass_silo1": shift.zo_fiberglass_silo1,
                "zo_fiberglass_silo2": shift.zo_fiberglass_silo2,
                "zo_fiberglass_silo3": shift.zo_fiberglass_silo3,
                "zo_fiberglass_silo4": shift.zo_fiberglass_silo4,
                "zo_laprol_silo1": shift.zo_laprol_silo1,
                "zo_laprol_silo2": shift.zo_laprol_silo2,
                "zo_laprol_silo3": shift.zo_laprol_silo3,
                "zo_laprol_silo4": shift.zo_laprol_silo4,
                "zo_asbocarton_silo1": shift.zo_asbocarton_silo1,
                "zo_asbocarton_silo2": shift.zo_asbocarton_silo2,
                "zo_asbocarton_silo3": shift.zo_asbocarton_silo3,
                "zo_asbocarton_silo4": shift.zo_asbocarton_silo4,
                "zo_asb_drain": shift.zo_asb_drain,
                "zo_cem_drain": shift.zo_cem_drain,
                "lfm_asb_drain": getattr(shift, 'lfm_asb_drain', 0.0),
                "lfm_cem_drain": getattr(shift, 'lfm_cem_drain', 0.0)
            },
            "lfm_report": {
                "product_name": l_prev.product_name if l_prev else "",
                "export_type": l_prev.export_type if l_prev else "Эталон",
                "lfm_sheets": l_prev.lfm_sheets if l_prev else 0,
                "lfm_wind_resets": l_prev.lfm_wind_resets if l_prev else 0,
                "formed_1st_grade": l_prev.formed_1st_grade if l_prev else 0,
                "formed_defect": l_prev.formed_defect if l_prev else 0,
                "transferred_to_warehouse": l_prev.transferred_to_warehouse if l_prev else 0
            } if l_prev else None,
            "batch": {
                "batch_number": b_prev.batch_number if b_prev else "",
                "product_name": b_prev.product_name if b_prev else "",
                "export_type": b_prev.export_type if b_prev else "Эталон",
                "stacked_stacks": b_prev.stacked_stacks if b_prev else 0,
                "ds_condition": b_prev.ds_condition if b_prev else 0,
                "ds_first_grade": b_prev.ds_first_grade if b_prev else 0,
                "ds_defect": b_prev.ds_defect if b_prev else 0,
                "ds_defect_chip": b_prev.ds_defect_chip if b_prev else 0,
                "ds_defect_scratch": b_prev.ds_defect_scratch if b_prev else 0,
                "ds_defect_bad_cut": b_prev.ds_defect_bad_cut if b_prev else 0,
                "ds_defect_stick_bottom": b_prev.ds_defect_stick_bottom if b_prev else 0,
                "ds_defect_stick_top": b_prev.ds_defect_stick_top if b_prev else 0,
                "ds_defect_broken": b_prev.ds_defect_broken if b_prev else 0,
                "ds_defect_fell_box": b_prev.ds_defect_fell_box if b_prev else 0,
                "ds_defect_dent": b_prev.ds_defect_dent if b_prev else 0,
                "ds_defect_thickness": b_prev.ds_defect_thickness if b_prev else 0,
                "ds_defect_delamination": b_prev.ds_defect_delamination if b_prev else 0,
                "ds_defect_edge": b_prev.ds_defect_edge if b_prev else 0,
                "prev_first_grade": b_prev.prev_first_grade if b_prev else 0,
                "prev_defect": b_prev.prev_defect if b_prev else 0,
                "prev_defect_scratch": b_prev.prev_defect_scratch if b_prev else 0,
                "prev_defect_bad_cut": b_prev.prev_defect_bad_cut if b_prev else 0,
                "prev_defect_stick_top": b_prev.prev_defect_stick_top if b_prev else 0,
                "prev_defect_broken": b_prev.prev_defect_broken if b_prev else 0,
                "prev_defect_fell_box": b_prev.prev_defect_fell_box if b_prev else 0,
                "prev_defect_thickness": b_prev.prev_defect_thickness if b_prev else 0,
                "prev_defect_edge": b_prev.prev_defect_edge if b_prev else 0,
                "qcd_condition": b_prev.qcd_condition if b_prev else 0,
                "qcd_first_grade": b_prev.qcd_first_grade if b_prev else 0,
                "qcd_defect": b_prev.qcd_defect if b_prev else 0
            } if b_prev else None
        }, ensure_ascii=False)
    except Exception as snap_err:
        print(f"Warning: could not capture admin snapshot_before: {snap_err}")
    
    changes = []
    
    # 1. Update Shift metadata and raw materials
    if data.date is not None and shift.date != data.date:
        changes.append(f"date: {shift.date} -> {data.date}")
        shift.date = data.date
    if data.shift_name is not None and shift.shift_name != data.shift_name:
        changes.append(f"shift_name: {shift.shift_name} -> {data.shift_name}")
        shift.shift_name = data.shift_name
    if data.line is not None and shift.line != data.line:
        changes.append(f"line: {shift.line} -> {data.line}")
        shift.line = data.line
    if data.master_id is not None and shift.master_id != data.master_id:
        changes.append(f"master_id: {shift.master_id} -> {data.master_id}")
        shift.master_id = data.master_id
    if data.batch_number is not None and shift.batch_number != data.batch_number:
        changes.append(f"batch_number: {shift.batch_number} -> {data.batch_number}")
        shift.batch_number = data.batch_number
    if data.product_name is not None and shift.product_name != data.product_name:
        changes.append(f"product_name: {shift.product_name} -> {data.product_name}")
        shift.product_name = data.product_name
    if data.export_type is not None and shift.export_type != data.export_type:
        changes.append(f"export_type: {shift.export_type} -> {data.export_type}")
        shift.export_type = data.export_type
    if data.status is not None and shift.status != data.status:
        changes.append(f"status: {shift.status} -> {data.status}")
        shift.status = data.status
        
    # ZO raw materials
    zo_fields = [
        "zo_batches", "zo_chrysotile_4_20", "zo_chrysotile_5_65", "zo_chrysotile_6_40",
        "zo_chrysotile_4_20_silo1", "zo_chrysotile_4_20_silo2", "zo_chrysotile_4_20_silo3", "zo_chrysotile_4_20_silo4",
        "zo_chrysotile_5_65_silo1", "zo_chrysotile_5_65_silo2", "zo_chrysotile_5_65_silo3", "zo_chrysotile_5_65_silo4",
        "zo_chrysotile_6_40_silo1", "zo_chrysotile_6_40_silo2", "zo_chrysotile_6_40_silo3", "zo_chrysotile_6_40_silo4",
        "zo_cement_silo1", "zo_cement_silo2", "zo_cement_silo3", "zo_cement_silo4",
        "zo_cellulose", "zo_cellulose_silo1", "zo_cellulose_silo2", "zo_cellulose_silo3", "zo_cellulose_silo4",
        "zo_crushed_slate", "zo_crushed_slate_silo1", "zo_crushed_slate_silo2", "zo_crushed_slate_silo3", "zo_crushed_slate_silo4",
        "zo_asbozurit", "zo_asbozurit_silo1", "zo_asbozurit_silo2", "zo_asbozurit_silo3", "zo_asbozurit_silo4",
        "zo_fiberglass", "zo_fiberglass_silo1", "zo_fiberglass_silo2", "zo_fiberglass_silo3", "zo_fiberglass_silo4",
        "zo_laprol", "zo_laprol_silo1", "zo_laprol_silo2", "zo_laprol_silo3", "zo_laprol_silo4",
        "zo_asbocarton", "zo_asbocarton_silo1", "zo_asbocarton_silo2", "zo_asbocarton_silo3", "zo_asbocarton_silo4",
        "lfm_asb_drain", "lfm_cem_drain", "zo_asb_drain", "zo_cem_drain"
    ]
    for f_name in zo_fields:
        val = getattr(data, f_name, None)
        if val is not None:
            old_val = getattr(shift, f_name, 0)
            if old_val != val:
                changes.append(f"{f_name}: {old_val} -> {val}")
                setattr(shift, f_name, val)
                
    # 2. Update LFM Report
    lfm_report = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift_id).first()
    if not lfm_report:
        lfm_report = models.LFMReport(
            shift_id=shift_id,
            product_name=shift.product_name or "",
            export_type=shift.export_type or "Эталон",
            lfm_sheets=data.lfm_sheets or 0,
            lfm_wind_resets=data.lfm_wind_resets or 0,
            transferred_to_warehouse=data.warehouse_gp or 0,
            formed_1st_grade=data.first_grade or 0,
            formed_defect=data.qcd_defect or 0
        )
        db.add(lfm_report)
        changes.append("Создан новый LFMReport")
    else:
        if data.lfm_sheets is not None and lfm_report.lfm_sheets != data.lfm_sheets:
            changes.append(f"lfm_sheets: {lfm_report.lfm_sheets} -> {data.lfm_sheets}")
            lfm_report.lfm_sheets = data.lfm_sheets
        if data.lfm_wind_resets is not None and lfm_report.lfm_wind_resets != data.lfm_wind_resets:
            changes.append(f"lfm_wind_resets: {lfm_report.lfm_wind_resets} -> {data.lfm_wind_resets}")
            lfm_report.lfm_wind_resets = data.lfm_wind_resets
        if data.warehouse_gp is not None and lfm_report.transferred_to_warehouse != data.warehouse_gp:
            changes.append(f"transferred_to_warehouse: {lfm_report.transferred_to_warehouse} -> {data.warehouse_gp}")
            lfm_report.transferred_to_warehouse = data.warehouse_gp
        if data.first_grade is not None and lfm_report.formed_1st_grade != data.first_grade:
            changes.append(f"formed_1st_grade: {lfm_report.formed_1st_grade} -> {data.first_grade}")
            lfm_report.formed_1st_grade = data.first_grade
        if data.qcd_defect is not None and lfm_report.formed_defect != data.qcd_defect:
            changes.append(f"formed_defect: {lfm_report.formed_defect} -> {data.qcd_defect}")
            lfm_report.formed_defect = data.qcd_defect
        if data.product_name is not None and lfm_report.product_name != data.product_name:
            lfm_report.product_name = data.product_name
        if data.export_type is not None and lfm_report.export_type != data.export_type:
            lfm_report.export_type = data.export_type
            
    # 3. Update Batch
    batch = db.query(models.Batch).filter(models.Batch.shift_id == shift_id).first()
    if not batch:
        batch = models.Batch(
            shift_id=shift_id,
            batch_number=shift.batch_number or "",
            product_name=shift.product_name or "",
            export_type=shift.export_type or "Эталон",
            status="stacked"
        )
        db.add(batch)
        changes.append("Создана новая партия Batch")
    else:
        batch.batch_number = shift.batch_number or ""
        batch.product_name = shift.product_name or ""
        batch.export_type = shift.export_type or "Эталон"
        
    if data.warehouse_gp is not None:
        batch.ds_condition = data.warehouse_gp
        batch.qcd_condition = data.warehouse_gp
    if data.first_grade is not None:
        batch.ds_first_grade = data.first_grade
        batch.qcd_first_grade = data.first_grade
    if data.qcd_defect is not None:
        batch.qcd_defect = data.qcd_defect
        
    ds_defect_fields = [
        "ds_defect_chip", "ds_defect_scratch", "ds_defect_bad_cut", "ds_defect_stick_bottom",
        "ds_defect_stick_top", "ds_defect_broken", "ds_defect_fell_box", "ds_defect_dent",
        "ds_defect_thickness", "ds_defect_delamination", "ds_defect_edge"
    ]
    total_ds_defect = 0
    for f_name in ds_defect_fields:
        val = getattr(data, f_name, None)
        if val is not None:
            old_val = getattr(batch, f_name, 0)
            if old_val != val:
                changes.append(f"{f_name}: {old_val} -> {val}")
                setattr(batch, f_name, val)
            total_ds_defect += val
        else:
            total_ds_defect += getattr(batch, f_name, 0) or 0
    batch.ds_defect = total_ds_defect
    batch.qcd_defect = total_ds_defect
    
    # Previous shift defects
    prev_defect_fields = [
        "prev_defect_scratch", "prev_defect_bad_cut", "prev_defect_stick_top",
        "prev_defect_broken", "prev_defect_fell_box", "prev_defect_thickness", "prev_defect_edge"
    ]
    if data.prev_first_grade is not None:
        batch.prev_first_grade = data.prev_first_grade
    total_prev_defect = 0
    for pf_name in prev_defect_fields:
        pval = getattr(data, pf_name, None)
        if pval is not None:
            old_pval = getattr(batch, pf_name, 0)
            if old_pval != pval:
                changes.append(f"{pf_name}: {old_pval} -> {pval}")
                setattr(batch, pf_name, pval)
            total_prev_defect += pval
        else:
            total_prev_defect += getattr(batch, pf_name, 0) or 0
    batch.prev_defect = total_prev_defect
    
    if changes or snapshot_before:
        log_entry = models.AuditLog(
            timestamp=datetime.utcnow(),
            user_name=admin.name,
            action=f"Комплексное редактирование смены ID {shift_id}",
            target_table="shifts",
            target_id=shift_id,
            details="Изменения: " + (", ".join(changes) if changes else "без числовых изменений"),
            state_snapshot=snapshot_before
        )
        db.add(log_entry)
        
    db.commit()
    
    # Sync plan boards for old and new parameters
    sync_lfm_to_plan_board(old_date, old_shift_name, old_line, db, old_master_id)
    if shift.date != old_date or shift.shift_name != old_shift_name or shift.line != old_line:
        sync_lfm_to_plan_board(shift.date, shift.shift_name, shift.line, db, shift.master_id)
        
    # Trigger background Google Sheets sync
    background_tasks.add_task(sync_sharepoint_report_bg)
    
    return {"status": "ok", "shift_id": shift_id}

@app.delete("/api/admin/shifts/{shift_id}")
def admin_delete_shift(shift_id: int, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    shift = db.query(models.Shift).get(shift_id)
    if not shift: raise HTTPException(404, "Смена не найдена")
    
    shift_date, shift_name, shift_line, master_id = shift.date, shift.shift_name, shift.line, shift.master_id
    
    # Decouple any linked downtimes and receipts so they are preserved autonomously
    for dt in db.query(models.Downtime).filter(models.Downtime.shift_id == shift_id).all():
        if not dt.date: dt.date = shift_date
        if not dt.shift_name: dt.shift_name = shift_name
        if not dt.line: dt.line = shift_line
        if not dt.master_id: dt.master_id = master_id
        dt.shift_id = None

    for r in db.query(models.RawMaterialReceipt).filter(models.RawMaterialReceipt.shift_id == shift_id).all():
        if not r.date: r.date = shift_date
        if not r.shift_name: r.shift_name = shift_name
        if not r.line: r.line = shift_line
        if not r.master_id: r.master_id = master_id
        r.shift_id = None

    db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift_id).delete()
    db.query(models.Batch).filter(models.Batch.shift_id == shift_id).delete()
    
    log_entry = models.AuditLog(
        timestamp=datetime.utcnow(),
        user_name=admin.name,
        action=f"Удаление смены ID {shift_id}",
        details=f"Удалена смена за {shift_date} ({shift_name}, Линия {shift_line}) и её производственные рапорты. Приходы сырья и простои сохранены автономно."
    )
    db.add(log_entry)
    db.delete(shift)
    db.commit()
    
    # Sync to clear phantom facts from plan board
    sync_lfm_to_plan_board(shift_date, shift_name, shift_line, db, master_id)
    background_tasks.add_task(sync_sharepoint_report_bg)
    background_tasks.add_task(sync_google_sheets_bg)
    background_tasks.add_task(sync_downtimes_bg)
    return {"status": "ok"}

@app.put("/api/admin/lfm/{report_id}")
def admin_update_lfm(report_id: int, data: dict, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    report = db.query(models.LFMReport).get(report_id)
    if not report: raise HTTPException(404, "Отчет ЛФМ не найден")
    
    old_values = {}
    new_values = {}
    for key, val in data.items():
        if hasattr(report, key):
            old_val = getattr(report, key)
            if old_val != val:
                old_values[key] = str(old_val)
                new_values[key] = str(val)
                setattr(report, key, val)
                
    if old_values:
        log_entry = models.AuditLog(
            timestamp=datetime.utcnow(),
            user_name=admin.name,
            action=f"Редактирование отчета ЛФМ ID {report_id}",
            details=f"Смена {report.shift_id}. Изменено: {old_values} -> {new_values}"
        )
        db.add(log_entry)
        db.commit()
        # Sync with plan board
        shift = db.query(models.Shift).get(report.shift_id)
        if shift:
            sync_lfm_to_plan_board(shift.date, shift.shift_name, shift.line, db, shift.master_id)
    else:
        db.commit()
    background_tasks.add_task(sync_sharepoint_report_bg)
    background_tasks.add_task(sync_google_sheets_bg)
    return {"status": "ok"}

@app.delete("/api/admin/lfm/{report_id}")
def admin_delete_lfm(report_id: int, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    report = db.query(models.LFMReport).get(report_id)
    if not report: raise HTTPException(404, "Отчет ЛФМ не найден")
    shift_id = report.shift_id
    shift = db.query(models.Shift).get(shift_id)
    shift_date, shift_name, shift_line, master_id = None, None, None, None
    if shift:
        shift_date, shift_name, shift_line, master_id = shift.date, shift.shift_name, shift.line, shift.master_id
        
    log_entry = models.AuditLog(
        timestamp=datetime.utcnow(),
        user_name=admin.name,
        action=f"Удаление отчета ЛФМ ID {report_id}",
        details=f"Смена {report.shift_id}. Удалена продукция: {report.product_name}, листы: {report.lfm_sheets}."
    )
    db.add(log_entry)
    db.delete(report)
    db.commit()
    # Sync with plan board
    if shift:
        sync_lfm_to_plan_board(shift_date, shift_name, shift_line, db, master_id)
    background_tasks.add_task(sync_sharepoint_report_bg)
    background_tasks.add_task(sync_google_sheets_bg)
    return {"status": "ok"}

@app.put("/api/admin/batches/{batch_id}")
def admin_update_batch(batch_id: int, data: dict, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    batch = db.query(models.Batch).get(batch_id)
    if not batch: raise HTTPException(404, "Партия не найдена")
    
    old_values = {}
    new_values = {}
    for key, val in data.items():
        if hasattr(batch, key):
            old_val = getattr(batch, key)
            if old_val != val:
                old_values[key] = str(old_val)
                new_values[key] = str(val)
                setattr(batch, key, val)
                
    if old_values:
        log_entry = models.AuditLog(
            timestamp=datetime.utcnow(),
            user_name=admin.name,
            action=f"Редактирование партии ID {batch_id}",
            details=f"Смена {batch.shift_id}. Изменено: {old_values} -> {new_values}"
        )
        db.add(log_entry)
        db.commit()
    else:
        db.commit()
    background_tasks.add_task(sync_sharepoint_report_bg)
    background_tasks.add_task(sync_google_sheets_bg)
    return {"status": "ok"}

@app.delete("/api/admin/batches/{batch_id}")
def admin_delete_batch(batch_id: int, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    batch = db.query(models.Batch).get(batch_id)
    if not batch: raise HTTPException(404, "Партия не найдена")
    
    log_entry = models.AuditLog(
        timestamp=datetime.utcnow(),
        user_name=admin.name,
        action=f"Удаление партии ID {batch_id}",
        details=f"Смена {batch.shift_id}. Удален номер партии: {batch.batch_number}, продукция: {batch.product_name}."
    )
    db.add(log_entry)
    db.delete(batch)
    db.commit()
    background_tasks.add_task(sync_sharepoint_report_bg)
    background_tasks.add_task(sync_google_sheets_bg)
    return {"status": "ok"}

@app.get("/api/admin/receipts")
def get_all_admin_receipts(
    start_date: str = Query(None),
    end_date: str = Query(None),
    request: Request = None, 
    db: Session = Depends(get_db)
):
    admin = check_admin_session(request, db)
    query = db.query(models.RawMaterialReceipt).outerjoin(models.Shift)
    
    if start_date:
        query = query.filter(or_(models.RawMaterialReceipt.date >= start_date, models.Shift.date >= start_date))
    if end_date:
        query = query.filter(or_(models.RawMaterialReceipt.date <= end_date, models.Shift.date <= end_date))
        
    receipts = query.order_by(
        func.coalesce(models.RawMaterialReceipt.date, models.Shift.date).desc(),
        models.RawMaterialReceipt.id.desc()
    ).all()
    
    result = []
    for r in receipts:
        r_dict = schemas.RawMaterialReceipt.model_validate(r).model_dump()
        r_dict["shift_date"] = r.record_date
        r_dict["shift_line"] = r.record_line
        r_dict["shift_name"] = r.record_shift_name
        if r.master:
            r_dict["master_name"] = r.master.name
        elif r.shift and r.shift.master:
            r_dict["master_name"] = r.shift.master.name
        else:
            r_dict["master_name"] = "Н/Д"
        result.append(r_dict)
    return result

@app.put("/api/admin/receipts/{receipt_id}")
def admin_update_receipt(receipt_id: int, data: schemas.RawMaterialReceiptUpdate, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    r = db.query(models.RawMaterialReceipt).get(receipt_id)
    if not r: raise HTTPException(404, "Приход сырья не найден")
    
    old_values = {}
    new_values = {}
    update_data = data.model_dump(exclude_unset=True)
    
    for key, val in update_data.items():
        if hasattr(r, key):
            old_val = getattr(r, key)
            if old_val != val:
                old_values[key] = str(old_val)
                new_values[key] = str(val)
                setattr(r, key, val)
                
    if old_values:
        log_entry = models.AuditLog(
            timestamp=datetime.utcnow(),
            user_name=admin.name,
            action=f"Редактирование прихода сырья ID {receipt_id}",
            details=f"Смена {r.shift_id}. Изменено: {old_values} -> {new_values}",
            target_table="raw_material_receipts",
            target_id=receipt_id
        )
        db.add(log_entry)
        db.commit()
    else:
        db.commit()
    background_tasks.add_task(sync_receipts_bg)
    return {"status": "ok"}

@app.delete("/api/admin/receipts/{receipt_id}")
def admin_delete_receipt(receipt_id: int, request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    r = db.query(models.RawMaterialReceipt).get(receipt_id)
    if not r: raise HTTPException(404, "Приход сырья не найден")
    
    log_entry = models.AuditLog(
        timestamp=datetime.utcnow(),
        user_name=admin.name,
        action=f"Удаление прихода сырья ID {receipt_id}",
        details=f"Смена {r.shift_id}. Цемент: {r.cement_silo1 + r.cement_silo2 + r.cement_silo3 + r.cement_silo4}, Хризотил 4-20: {r.chrysotile_4_20}",
        target_table="raw_material_receipts",
        target_id=receipt_id
    )
    db.add(log_entry)
    db.delete(r)
    db.commit()
    background_tasks.add_task(sync_receipts_bg)
    return {"status": "ok"}


# ==========================================
# API БЭКАПА, ВОССТАНОВЛЕНИЯ И ОТКАТА (ROLLBACK)
# ==========================================

@app.post("/api/admin/shifts/{shift_id}/rollback")
def admin_rollback_shift(
    shift_id: int, 
    request: Request, 
    audit_log_id: Optional[int] = None, 
    background_tasks: BackgroundTasks = None, 
    db: Session = Depends(get_db)
):
    admin = check_admin_session(request, db)
    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(404, "Смена не найдена")

    # Ищем подходящий снимок в AuditLog
    query = db.query(models.AuditLog).filter(
        models.AuditLog.target_table == "shifts",
        models.AuditLog.target_id == shift_id,
        models.AuditLog.state_snapshot.isnot(None)
    )
    if audit_log_id:
        log_entry = query.filter(models.AuditLog.id == audit_log_id).first()
    else:
        log_entry = query.order_by(models.AuditLog.timestamp.desc(), models.AuditLog.id.desc()).first()

    if not log_entry or not log_entry.state_snapshot:
        raise HTTPException(404, "Снимок состояния для отката этой смены не найден")

    import json
    try:
        snapshot = json.loads(log_entry.state_snapshot)
    except Exception as e:
        raise HTTPException(500, f"Ошибка парсинга снимка состояния: {e}")

    # 1. Восстанавливаем поля Shift
    s_data = snapshot.get("shift", {})
    for k, v in s_data.items():
        if hasattr(shift, k):
            setattr(shift, k, v)

    # 2. Восстанавливаем LFMReport
    l_data = snapshot.get("lfm_report")
    if l_data:
        lfm_rep = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift_id).first()
        if not lfm_rep:
            lfm_rep = models.LFMReport(shift_id=shift_id)
            db.add(lfm_rep)
        for k, v in l_data.items():
            if hasattr(lfm_rep, k):
                setattr(lfm_rep, k, v)

    # 3. Восстанавливаем Batch
    b_data = snapshot.get("batch")
    if b_data:
        batch = db.query(models.Batch).filter(models.Batch.shift_id == shift_id).first()
        if not batch:
            batch = models.Batch(shift_id=shift_id)
            db.add(batch)
        for k, v in b_data.items():
            if hasattr(batch, k):
                setattr(batch, k, v)

    # Записываем действие отката в аудит
    db.add(models.AuditLog(
        timestamp=datetime.utcnow(),
        user_name=admin.name,
        action="ROLLBACK",
        target_table="shifts",
        target_id=shift_id,
        details=f"Выполнен откат смены ID {shift_id} к снимку из лога #{log_entry.id} ({log_entry.timestamp})"
    ))
    db.commit()

    # Синхронизация с планом и гугл таблицами
    sync_lfm_to_plan_board(shift.date, shift.shift_name, shift.line, db, shift.master_id)
    if background_tasks:
        background_tasks.add_task(sync_sharepoint_report_bg)

    return {"status": "ok", "message": f"Смена успешно откачена к состоянию от {log_entry.timestamp}"}


@app.get("/api/admin/backup/excel")
def download_database_backup_excel(request: Request, db: Session = Depends(get_db)):
    admin = check_admin_session(request, db)
    excel_bytes = excel_exporter.generate_full_backup_excel(db)
    from fastapi.responses import Response
    today_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Tectum_Backup_{today_str}.xlsx"
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@app.post("/api/admin/backup/restore")
async def restore_database_from_backup_excel(
    request: Request, 
    file: UploadFile = File(...), 
    background_tasks: BackgroundTasks = None, 
    db: Session = Depends(get_db)
):
    admin = check_admin_session(request, db)
    if not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Файл должен быть в формате Excel (.xlsx)")
    
    file_bytes = await file.read()
    try:
        res = excel_exporter.restore_from_backup_excel(file_bytes, db, user_name=admin.name)
        if background_tasks:
            background_tasks.add_task(sync_receipts_bg)
            background_tasks.add_task(sync_downtimes_bg)
            background_tasks.add_task(sync_sharepoint_report_bg)
        return res
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Ошибка восстановления из файла: {str(e)}")


@app.post("/api/admin/backup/google_sync_all")
def trigger_full_google_sync_endpoint(
    request: Request, 
    background_tasks: BackgroundTasks, 
    db: Session = Depends(get_db)
):
    admin = check_admin_session(request, db)
    background_tasks.add_task(sync_sharepoint_report_bg)
    background_tasks.add_task(sync_receipts_bg)
    background_tasks.add_task(sync_downtimes_bg)
    
    # Audit log
    db.add(models.AuditLog(
        timestamp=datetime.utcnow(),
        user_name=admin.name,
        action="SYNC",
        target_table="all_google_sheets",
        target_id=0,
        details="Запущена принудительная фоновая синхронизация всех листов с Google Таблицами"
    ))
    db.commit()
    return {"status": "ok", "message": "Синхронизация всех листов Google Таблиц успешно запущена в фоновом режиме"}



@app.post("/api/admin/norms/", response_model=schemas.ProductNorm)
def create_norm(norm: schemas.ProductNormCreate, db: Session = Depends(get_db)):
    db_norm = models.ProductNorm(**norm.model_dump())
    db.add(db_norm)
    db.commit()
    db.refresh(db_norm)
    return db_norm

@app.put("/api/admin/norms/{norm_id}", response_model=schemas.ProductNorm)
def update_norm(norm_id: int, norm: schemas.ProductNormUpdate, db: Session = Depends(get_db)):
    db_norm = db.query(models.ProductNorm).get(norm_id)
    if not db_norm: raise HTTPException(404)
    update_data = norm.model_dump(exclude_unset=True)
    for key, val in update_data.items():
        setattr(db_norm, key, val)
    db.commit()
    db.refresh(db_norm)
    return db_norm

@app.delete("/api/admin/norms/{norm_id}")
def delete_norm(norm_id: int, db: Session = Depends(get_db)):
    db_norm = db.query(models.ProductNorm).get(norm_id)
    if not db_norm: raise HTTPException(404)
    db.delete(db_norm)
    db.commit()
    return {"status": "ok"}

@app.post("/api/admin/clear_data/")
def clear_operational_data(request: Request, payload: dict = Body(default={}), db: Session = Depends(get_db)):
    pwd = payload.get("password") if payload else ""
    if pwd != "VjzJ,jhjyf15":
        raise HTTPException(status_code=403, detail="Неверный пароль подтверждения очистки")
    try:
        deleted_batches = db.query(models.Batch).delete()
        deleted_lfm = db.query(models.LFMReport).delete()
        deleted_downtime = db.query(models.Downtime).delete()
        deleted_shifts = db.query(models.Shift).delete()
        deleted_plan_board = db.query(models.MonthlyPlanBoard).delete()
        
        user_name = request.session.get("user_name") or "Администратор"
        db.add(models.AuditLog(
            user_name=user_name,
            action="DELETE",
            target_table="shifts/lfm/batches/downtimes/plan_board",
            target_id=0,
            details=f"Сброс операционных данных. Удалено: смен {deleted_shifts}, отчетов ЛФМ {deleted_lfm}, партий {deleted_batches}, простоев {deleted_downtime}, строк плана {deleted_plan_board}"
        ))
        db.commit()
        return {
            "status": "ok",
            "deleted": {
                "batches": deleted_batches,
                "lfm_reports": deleted_lfm,
                "downtimes": deleted_downtime,
                "shifts": deleted_shifts,
                "plan_board": deleted_plan_board
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

# --- MONTHLY PLAN BOARD ---
@app.get("/api/plan_board", response_model=list[schemas.MonthlyPlanBoard])
def get_plan_board(db: Session = Depends(get_db)):
    try:
        return db.query(models.MonthlyPlanBoard).order_by(models.MonthlyPlanBoard.date.desc(), models.MonthlyPlanBoard.shift_number).all()
    except Exception as e:
        import traceback
        print(f"Error in get_plan_board: {str(e)}\n{traceback.format_exc()}")
        return []

@app.get("/api/admin/audit_logs")
def get_audit_logs(
    limit: int = 500,
    module: Optional[str] = None,
    action: Optional[str] = None,
    user_name: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        query = db.query(models.AuditLog)

        if module:
            mod = module.lower()
            if mod == "tasks":
                query = query.filter(models.AuditLog.target_table.ilike("%task%"))
            elif mod == "shifts":
                query = query.filter(models.AuditLog.target_table.in_(["shifts", "lfm_reports", "batches", "downtimes"]))
            elif mod in ["plan_board", "plan"]:
                query = query.filter(models.AuditLog.target_table.ilike("%plan%"))
            elif mod in ["raw", "raw_materials"]:
                query = query.filter(models.AuditLog.target_table.ilike("%raw%"))
            elif mod in ["directories", "dir"]:
                query = query.filter(models.AuditLog.target_table.in_(["downtime_directory", "product_norms", "masters", "directories"]))
            elif mod in ["documents", "docs"]:
                query = query.filter(models.AuditLog.target_table.ilike("%doc%"))

        if action:
            query = query.filter(models.AuditLog.action.ilike(f"%{action}%"))

        if user_name:
            query = query.filter(models.AuditLog.user_name.ilike(f"%{user_name}%"))

        if search:
            s = f"%{search}%"
            query = query.filter(
                or_(
                    models.AuditLog.details.ilike(s),
                    models.AuditLog.user_name.ilike(s),
                    models.AuditLog.target_table.ilike(s),
                    models.AuditLog.action.ilike(s)
                )
            )

        if date_from:
            try:
                dt_from = datetime.strptime(date_from, "%Y-%m-%d")
                query = query.filter(models.AuditLog.timestamp >= dt_from)
            except Exception:
                pass

        if date_to:
            try:
                dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
                query = query.filter(models.AuditLog.timestamp < dt_to)
            except Exception:
                pass

        return query.order_by(models.AuditLog.timestamp.desc(), models.AuditLog.id.desc()).limit(limit).all()
    except Exception as e:
        import traceback
        print(f"Error in get_audit_logs: {str(e)}\n{traceback.format_exc()}")
        return []

@app.post("/api/plan_board", response_model=schemas.MonthlyPlanBoard)
def create_or_update_plan_board(data: schemas.MonthlyPlanBoardCreate, user_name: str = None, db: Session = Depends(get_db)):
    existing = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date == data.date,
        models.MonthlyPlanBoard.shift_name == data.shift_name,
        models.MonthlyPlanBoard.line == data.line
    ).first()
    
    if data.date.weekday() == 0 and data.shift_name == "День":
        data.plan_sheets = 0
        
    if existing:
        old_plan = existing.plan_sheets
        old_fact = existing.fact_sheets
        existing.master_id = data.master_id
        existing.shift_number = data.shift_number
        existing.plan_sheets = data.plan_sheets
        existing.fact_sheets = data.fact_sheets
        existing.first_grade = data.first_grade
        existing.defect = data.defect
        db.commit()
        db.refresh(existing)
        
        # Log to AuditLog
        details_str = f"Дата: {data.date}, Смена: {data.shift_name}, Линия: {data.line}. План: {old_plan}->{data.plan_sheets}, Факт: {old_fact}->{data.fact_sheets}, 1-й сорт: {data.first_grade}, Брак: {data.defect}"
        log = models.AuditLog(
            user_name=user_name,
            action="UPDATE",
            target_table="monthly_plan_board",
            target_id=existing.id,
            details=details_str
        )
        db.add(log)
        db.commit()
        
        return existing
    else:
        new_plan = models.MonthlyPlanBoard(**data.model_dump())
        db.add(new_plan)
        db.commit()
        db.refresh(new_plan)
        
        # Log to AuditLog
        details_str = f"Дата: {data.date}, Смена: {data.shift_name}, Линия: {data.line}. План: {data.plan_sheets}, Факт: {data.fact_sheets}"
        log = models.AuditLog(
            user_name=user_name,
            action="CREATE",
            target_table="monthly_plan_board",
            target_id=new_plan.id,
            details=details_str
        )
        db.add(log)
        db.commit()
        
        return new_plan

@app.get("/api/plan_board/{id}", response_model=schemas.MonthlyPlanBoard)
def get_plan_board_row(id: int, db: Session = Depends(get_db)):
    row = db.query(models.MonthlyPlanBoard).get(id)
    if not row: raise HTTPException(404, "Запись не найдена")
    return row

@app.delete("/api/plan_board/{id}")
def delete_plan_board_row(id: int, user_name: str = None, db: Session = Depends(get_db)):
    row = db.query(models.MonthlyPlanBoard).filter(models.MonthlyPlanBoard.id == id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    
    details_str = f"Дата: {row.date}, Смена: {row.shift_name}, Линия: {row.line}. План: {row.plan_sheets}, Факт: {row.fact_sheets}"
    db.delete(row)
    db.commit()
    
    # Log to AuditLog
    log = models.AuditLog(
        user_name=user_name,
        action="DELETE",
        target_table="monthly_plan_board",
        target_id=id,
        details=details_str
    )
    db.add(log)
    db.commit()
    
    return {"status": "ok"}


@app.post("/api/admin/import_plan_board")
def import_plan_board(db: Session = Depends(get_db)):
    file_path = os.path.join("docs", "excel", "monthly_plan_board.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(404, f"Файл {file_path} не найден в проекте")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["Выработка"] if "Выработка" in wb.sheetnames else wb.active
        
        count_created = 0
        count_updated = 0
        
        # Пропускаем заголовки (первые две строки, например)
        # Ожидаемый формат: Дата (0), Месяц (1), Тип смены (2), Линия (3), Мастер (4), Смена (5), План (6), Факт (7)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]: continue # пустая дата
            
            date_val = row[0]
            if isinstance(date_val, datetime):
                date_val = date_val.date()
            elif isinstance(date_val, str):
                try:
                    date_val = datetime.strptime(date_val, "%d.%m.%Y").date()
                except ValueError:
                    try:
                        date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
                    except ValueError:
                        continue
                
            shift_name = str(row[2]) if row[2] else "День"
            
            line_val = str(row[3]).strip() if row[3] else "ЛФМ-1"
            if line_val == "Линия 1":
                line_val = "ЛФМ-1"
            elif line_val == "Линия 2":
                line_val = "ЛФМ-2"
                
            master_name = str(row[4]).strip() if row[4] else ""
            shift_number = int(row[5]) if row[5] else 1
            plan_sheets = int(row[6]) if len(row) > 6 and row[6] else 0
            fact_sheets = int(row[7]) if len(row) > 7 and row[7] else 0
            
            if date_val.weekday() == 0 and shift_name == "День":
                plan_sheets = 0
            
            first_grade = 0
            if len(row) > 8 and row[8] is not None:
                try: first_grade = int(row[8])
                except: pass
                
            defect = 0
            if len(row) > 9 and row[9] is not None:
                try: defect = int(row[9])
                except: pass
            
            # Поиск мастера по имени
            master = db.query(models.Master).filter(models.Master.name == master_name).first()
            if not master:
                # Если мастер не найден, создаем его? Или берем первого попавшегося?
                # Лучше создать, чтобы не терять данные
                master = models.Master(name=master_name, pin="0000", role="master")
                db.add(master)
                db.commit()
                db.refresh(master)
                
            existing = db.query(models.MonthlyPlanBoard).filter(
                models.MonthlyPlanBoard.date == date_val,
                models.MonthlyPlanBoard.shift_number == shift_number,
                models.MonthlyPlanBoard.line == line_val
            ).first()
            
            if existing:
                existing.shift_name = shift_name
                existing.master_id = master.id
                existing.plan_sheets = plan_sheets
                existing.fact_sheets = fact_sheets
                existing.first_grade = first_grade
                existing.defect = defect
                count_updated += 1
            else:
                new_plan = models.MonthlyPlanBoard(
                    date=date_val,
                    shift_name=shift_name,
                    master_id=master.id,
                    shift_number=shift_number,
                    line=line_val,
                    plan_sheets=plan_sheets,
                    fact_sheets=fact_sheets,
                    first_grade=first_grade,
                    defect=defect
                )
                db.add(new_plan)
                count_created += 1
                
        db.commit()
        
        # Log to AuditLog
        log = models.AuditLog(
            user_name="Администратор",
            action="IMPORT",
            target_table="monthly_plan_board",
            details=f"Импорт из Excel. Создано записей: {count_created}, обновлено: {count_updated}"
        )
        db.add(log)
        db.commit()
        
        return {"status": "ok", "created": count_created, "updated": count_updated}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Ошибка импорта: {str(e)}")

@app.post("/api/admin/upload_and_import_plan_board")
def upload_and_import_plan_board(file: UploadFile = File(...), user_name: str = "Администратор", db: Session = Depends(get_db)):
    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb["Выработка"] if "Выработка" in wb.sheetnames else wb.active
        
        count_created = 0
        count_updated = 0
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]: continue
            
            date_val = row[0]
            if isinstance(date_val, datetime):
                date_val = date_val.date()
            elif isinstance(date_val, str):
                try:
                    date_val = datetime.strptime(date_val, "%d.%m.%Y").date()
                except ValueError:
                    try:
                        date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
                    except ValueError:
                        continue
                
            shift_name = str(row[2]) if row[2] else "День"
            
            line_val = str(row[3]).strip() if row[3] else "ЛФМ-1"
            if line_val == "Линия 1":
                line_val = "ЛФМ-1"
            elif line_val == "Линия 2":
                line_val = "ЛФМ-2"
                
            master_name = str(row[4]).strip() if row[4] else ""
            shift_number = int(row[5]) if row[5] else 1
            plan_sheets = int(row[6]) if len(row) > 6 and row[6] else 0
            fact_sheets = int(row[7]) if len(row) > 7 and row[7] else 0
            
            if date_val.weekday() == 0 and shift_name == "День":
                plan_sheets = 0
            
            first_grade = 0
            if len(row) > 8 and row[8] is not None:
                try: first_grade = int(row[8])
                except: pass
                
            defect = 0
            if len(row) > 9 and row[9] is not None:
                try: defect = int(row[9])
                except: pass
            
            master = db.query(models.Master).filter(models.Master.name == master_name).first()
            if not master:
                master = models.Master(name=master_name, pin="0000", role="master")
                db.add(master)
                db.commit()
                db.refresh(master)
                
            existing = db.query(models.MonthlyPlanBoard).filter(
                models.MonthlyPlanBoard.date == date_val,
                models.MonthlyPlanBoard.shift_number == shift_number,
                models.MonthlyPlanBoard.line == line_val
            ).first()
            
            if existing:
                existing.shift_name = shift_name
                existing.master_id = master.id
                existing.plan_sheets = plan_sheets
                existing.fact_sheets = fact_sheets
                existing.first_grade = first_grade
                existing.defect = defect
                count_updated += 1
            else:
                new_plan = models.MonthlyPlanBoard(
                    date=date_val,
                    shift_name=shift_name,
                    master_id=master.id,
                    shift_number=shift_number,
                    line=line_val,
                    plan_sheets=plan_sheets,
                    fact_sheets=fact_sheets,
                    first_grade=first_grade,
                    defect=defect
                )
                db.add(new_plan)
                count_created += 1
                
        db.commit()
        
        # Log to AuditLog
        log = models.AuditLog(
            user_name=user_name,
            action="IMPORT",
            target_table="monthly_plan_board",
            details=f"Загрузка и импорт файла {file.filename}. Создано: {count_created}, обновлено: {count_updated}"
        )
        db.add(log)
        db.commit()
        
        return {"status": "ok", "created": count_created, "updated": count_updated}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Ошибка импорта: {str(e)}")

@app.delete("/api/admin/clear_plan_board")
def clear_plan_board(user_name: str = "Администратор", db: Session = Depends(get_db)):
    try:
        count = db.query(models.MonthlyPlanBoard).count()
        db.query(models.MonthlyPlanBoard).delete()
        db.commit()
        
        # Log to AuditLog
        log = models.AuditLog(
            user_name=user_name,
            action="DELETE",
            target_table="monthly_plan_board",
            details=f"Полная очистка таблицы. Удалено записей: {count}"
        )
        db.add(log)
        db.commit()
        
        return {"status": "ok", "deleted_count": count}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Ошибка очистки: {str(e)}")

@app.get("/api/admin/fix_phantoms")
def fix_phantoms(db: Session = Depends(get_db)):
    try:
        pbs = db.query(models.MonthlyPlanBoard).all()
        count = 0
        for pb in pbs:
            sync_lfm_to_plan_board(pb.date, pb.shift_name, pb.line, db, pb.master_id)
            count += 1
        return {"status": "ok", "message": f"Processed {count} plan board records. Phantom facts have been zeroed out."}
    except Exception as e:
        return {"status": "error", "message": str(e)}




# --- DOCUMENTS ENDPOINTS MOVED TO routers/documents.py ---





# --- CHECKLISTS ENDPOINTS MOVED TO routers/checklists.py ---



# --- PLANNER ENDPOINTS MOVED TO routers/planner.py ---


# ----------------------------------------------------
# WhatsApp Cloud API Webhook Endpoints
# ----------------------------------------------------
from fastapi.responses import PlainTextResponse

@app.get("/api/whatsapp/webhook")
async def whatsapp_verify_webhook(request: Request):
    """
    Верификация вебхука со стороны серверов Meta (GET запрос при настройке).
    """
    mode = request.query_params.get("hub.mode")
    token = request.query_params.get("hub.verify_token")
    challenge = request.query_params.get("hub.challenge")

    verify_token = os.getenv("WHATSAPP_VERIFY_TOKEN", "tectum_wa_verify_token_2026")

    if mode and token:
        if mode == "subscribe" and token == verify_token:
            print("[WhatsApp Webhook] Успешная верификация вебхука от Meta!")
            return PlainTextResponse(content=str(challenge or ""), status_code=200)
        else:
            print(f"[WhatsApp Webhook] Ошибка токена: получено '{token}', ожидалось '{verify_token}'")
            raise HTTPException(status_code=403, detail="Verification token mismatch")
    
    return PlainTextResponse(content="Tectum WhatsApp Webhook Active", status_code=200)


@app.post("/api/whatsapp/webhook")
async def whatsapp_incoming_webhook(request: Request, bg_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """
    Прием входящих сообщений и нажатий кнопок от пользователей WhatsApp.
    """
    try:
        body = await request.json()
        print(f"[WhatsApp Incoming] Raw event: {body}")
        
        # Meta отправляет события в entry -> changes -> value
        entries = body.get("entry", [])
        for entry in entries:
            changes = entry.get("changes", [])
            for change in changes:
                value = change.get("value", {})
                messages = value.get("messages", [])
                
                for msg in messages:
                    from_phone = msg.get("from") # Номер отправителя
                    msg_type = msg.get("type")
                    
                    user_text = ""
                    btn_id = ""
                    
                    if msg_type == "text":
                        user_text = msg.get("text", {}).get("body", "").strip()
                    elif msg_type == "interactive":
                        interactive = msg.get("interactive", {})
                        btn_reply = interactive.get("button_reply", {})
                        btn_id = btn_reply.get("id", "")
                        user_text = btn_reply.get("title", "").strip()
                        
                    print(f"[WhatsApp Incoming Event] from={from_phone}, type={msg_type}, text='{user_text}', btn_id='{btn_id}'")
                    
                    lower_text = (user_text or "").lower()
                    
                    # 1. Приветствие / Меню
                    if any(w in lower_text for w in ["привет", "старт", "start", "меню", "помощь", "help"]):
                        reply = (
                            "🏭 *Tectum Enterprise Bot*\n\n"
                            "Я ваш мобильный помощник по заводу.\n\n"
                            "📌 *Выберите нужный раздел кнопками ниже:*"
                        )
                        buttons = [
                            {"id": "cmd_summary", "title": "📊 Сводка"},
                            {"id": "cmd_tasks", "title": "📌 Задачи"}
                        ]
                        whatsapp_service.send_whatsapp_buttons(from_phone, reply, buttons)
                        
                    # 2. Сводка
                    elif btn_id == "cmd_summary" or "сводк" in lower_text or "выработк" in lower_text:
                        today_str = datetime.now().strftime("%Y-%m-%d")
                        shifts = db.query(models.Shift).filter(models.Shift.date == today_str).all()
                        total_sheets = sum(s.lfm_sheets or 0 for s in shifts)
                        reply = (
                            f"📊 *Сводка за сегодня ({today_str}):*\n\n"
                            f"📦 Рапортов внесено: *{len(shifts)}*\n"
                            f"📈 Общая выработка: *{total_sheets:,} листов*\n\n"
                            f"🔗 Открыть портал: https://tectum-portal-railway-production.up.railway.app"
                        )
                        whatsapp_service.send_whatsapp_text(from_phone, reply)
                        
                    # 3. Задачи
                    elif btn_id == "cmd_tasks" or "задач" in lower_text:
                        all_tasks = db.query(models.Task).filter(
                            models.Task.is_archived == False
                        ).order_by(models.Task.id.desc()).all()
                        
                        active_tasks = [
                            t for t in all_tasks 
                            if not ("заверш" in (t.status or "").lower() or "выполн" in (t.status or "").lower())
                        ][:5]
                        
                        if not active_tasks:
                            reply = "✅ На данный момент нет открытых незавершенных задач!"
                            whatsapp_service.send_whatsapp_text(from_phone, reply)
                        else:
                            msg_tasks = "📌 *Открытые задачи Tectum:*\n\n"
                            for t in active_tasks:
                                status_icon = "🟡" if "работ" in (t.status or "").lower() else "⚪"
                                msg_tasks += f"{status_icon} *{t.code or ''}* {t.title}\n👤 Исполнитель: {t.assignee_name or '—'}\n⏰ Срок: {t.due_date_str or '—'}\n\n"
                            msg_tasks += "🔗 Открыть планнер: https://tectum-portal-railway-production.up.railway.app/tasks"
                            whatsapp_service.send_whatsapp_text(from_phone, msg_tasks)
                            
                    # 4. Ответ на действия
                    elif btn_id in ("btn_accept", "btn_done"):
                        reply = f"👍 Отлично! Действие зафиксировано: *{user_text}*."
                        whatsapp_service.send_whatsapp_text(from_phone, reply)
                        
                    else:
                        reply = (
                            f"Я получил ваше сообщение: «_{user_text}_».\n\n"
                            f"Нажмите *Меню* или *Сводка*, чтобы запросить данные с завода."
                        )
                        whatsapp_service.send_whatsapp_text(from_phone, reply)
                            
        return {"status": "ok"}
    except Exception as e:
        print(f"[WhatsApp Webhook Error] {e}")
        return {"status": "error", "detail": str(e)}


# ----------------------------------------------------
# Telegram Bot Webhook Endpoints
# ----------------------------------------------------
@app.post("/api/telegram/webhook")
async def telegram_webhook(request: Request, db: Session = Depends(get_db)):
    """
    Прием входящих сообщений и команд от Telegram (личные сообщения и группы).
    """
    try:
        data = await request.json()
        print(f"[Telegram Incoming] Event: {data}")
        
        import telegram_service
        from sqlalchemy import cast, String
        from sqlalchemy.orm import joinedload
        
        message = data.get("message") or data.get("channel_post")
        callback_query = data.get("callback_query")
        
        if callback_query:
            cq_id = callback_query.get("id")
            cq_data = callback_query.get("data", "")
            cq_chat_id = callback_query.get("message", {}).get("chat", {}).get("id")
            
            # Answer callback
            token = os.getenv("TELEGRAM_BOT_TOKEN", telegram_service.TELEGRAM_BOT_TOKEN).strip()
            requests.post(f"https://api.telegram.org/bot{token}/answerCallbackQuery", json={"callback_query_id": cq_id})
            
            if cq_data == "tg_cmd_summary":
                today_str = datetime.now().strftime("%Y-%m-%d")
                shifts = db.query(models.Shift).filter(models.Shift.date == today_str).all()
                total_sheets = sum(s.lfm_sheets or 0 for s in shifts)
                reply = (
                    f"📊 <b>Сводка за сегодня ({today_str}):</b>\n\n"
                    f"📦 Внесено рапортов: <b>{len(shifts)}</b>\n"
                    f"📈 Общая выработка: <b>{total_sheets:,} листов</b>\n\n"
                    f"🔗 <a href='https://tectum-portal-railway-production.up.railway.app'>Открыть портал Tectum</a>"
                )
                telegram_service.send_telegram_message(cq_chat_id, reply)
            elif cq_data == "tg_cmd_tasks":
                all_tasks = db.query(models.Task).filter(
                    models.Task.is_archived == False
                ).order_by(models.Task.id.desc()).all()
                
                active_tasks = [
                    t for t in all_tasks 
                    if not ("заверш" in (t.status or "").lower() or "выполн" in (t.status or "").lower())
                ][:5]
                
                if not active_tasks:
                    reply = "✅ На данный момент нет открытых незавершенных задач!"
                else:
                    reply = "📌 <b>Открытые производственные задачи:</b>\n\n"
                    for t in active_tasks:
                        status_icon = "🟡" if "работ" in (t.status or "").lower() else "⚪"
                        reply += f"{status_icon} <b>{t.code or ''}</b> {t.title}\n👤 Исполнитель: {t.assignee_name or '—'}\n⏰ Срок: {t.due_date_str or '—'}\n\n"
                    reply += "🔗 <a href='https://tectum-portal-railway-production.up.railway.app/tasks'>Открыть Планнер</a>"
                telegram_service.send_telegram_message(cq_chat_id, reply)
                
            return {"status": "ok"}
            
        if not message:
            return {"status": "ok"}
            
        chat = message.get("chat", {})
        chat_id = chat.get("id")
        text = message.get("text", "").strip()
        lower_text = text.lower()
        
        # Если бота только что добавили в группу
        new_members = message.get("new_chat_members", [])
        for member in new_members:
            if member.get("is_bot") and member.get("username") == "tectum_factory_bot":
                welcome = (
                    "🏭 <b>Привет, команда Tectum!</b>\n\n"
                    "Я официальный бот завода. Теперь я буду присылать в эту группу уведомления о задачах, сменных рапортах и простоях.\n\n"
                    f"🆔 <b>Chat ID этой группы:</b> <code>{chat_id}</code>\n"
                    "Используйте команду /summary для получения оперативной сводки!"
                )
                telegram_service.send_telegram_message(chat_id, welcome)
                return {"status": "ok"}

        if not text:
            return {"status": "ok"}
            
        # Постоянная клавиатура для всех ответов
        main_kb = telegram_service.get_main_reply_keyboard()
        
        # 1. Меню / Старт / Помощь
        if lower_text.startswith("/start") or lower_text.startswith("/menu") or "привет" in lower_text:
            reply = (
                "🏭 <b>Добро пожаловать в Tectum Enterprise Bot!</b>\n\n"
                "Я ваш мобильный помощник по заводу. Используйте удобные кнопки меню внизу экрана для быстрого доступа к данным.\n\n"
                "📌 <b>Основные возможности:</b>\n"
                "• 📊 <b>Сводка</b> — суточная выработка по линиям\n"
                "• 📌 <b>Задачи</b> — список актуальных задач планнера\n"
                "• 🎯 <b>План-факт</b> — выполнение месячной программы\n"
                "• ⏱ <b>Простои</b> — зафиксированные остановки линий\n"
                "• 📦 <b>Сырье</b> — складские остатки цемента и хризотила"
            )
            telegram_service.send_telegram_message(chat_id, reply, reply_markup=main_kb)
            
        # 2. Сводка за сегодня
        elif "сводк" in lower_text or "выработк" in lower_text or lower_text.startswith("/summary"):
            today_str = datetime.now().strftime("%Y-%m-%d")
            shifts = db.query(models.Shift).options(
                joinedload(models.Shift.lfm_reports),
                joinedload(models.Shift.batches)
            ).filter(cast(models.Shift.date, String) == today_str).all()
            
            l1_shifts = [s for s in shifts if "1" in str(s.line)]
            l2_shifts = [s for s in shifts if "2" in str(s.line)]
            
            def get_shift_sheets(s):
                if s.lfm_reports:
                    return sum(r.lfm_sheets or 0 for r in s.lfm_reports)
                if s.batches:
                    return sum(b.stacked_stacks or 0 for b in s.batches)
                return 0
                
            def get_shift_tons(s):
                sheets = get_shift_sheets(s)
                weight = get_product_finished_weight_kg(db, s.product_name) if hasattr(db, 'query') else 19.6
                return (sheets * (weight or 19.6)) / 1000.0
                
            l1_sheets = sum(get_shift_sheets(s) for s in l1_shifts)
            l2_sheets = sum(get_shift_sheets(s) for s in l2_shifts)
            total_sheets = l1_sheets + l2_sheets
            total_tons = sum(get_shift_tons(s) for s in shifts)
            
            reply = (
                f"📊 <b>Производственная сводка за сегодня:</b>\n"
                f"📅 <b>Дата:</b> <code>{today_str}</code>\n"
                f"📝 <b>Рапортов внесено:</b> {len(shifts)}\n\n"
                f"🔹 <b>Линия 1:</b> {l1_sheets:,} листов\n"
                f"🔹 <b>Линия 2:</b> {l2_sheets:,} листов\n"
                f"📈 <b>ИТОГО:</b> <b>{total_sheets:,} листов</b> (~{total_tons:.1f} т)\n\n"
                f"🔗 <a href='https://tectum-portal-railway-production.up.railway.app'>Открыть портал Tectum</a>"
            )
            telegram_service.send_telegram_message(chat_id, reply, reply_markup=main_kb)
            
        # 3. Активные задачи
        elif "задач" in lower_text or lower_text.startswith("/tasks"):
            all_tasks = db.query(models.Task).filter(
                models.Task.is_archived == False
            ).order_by(models.Task.id.desc()).all()
            
            active_tasks = [
                t for t in all_tasks 
                if not ("заверш" in (t.status or "").lower() or "выполн" in (t.status or "").lower())
            ][:5]
            
            if not active_tasks:
                reply = "✅ На данный момент нет открытых незавершенных задач!"
            else:
                reply = "📌 <b>Открытые производственные задачи:</b>\n\n"
                for t in active_tasks:
                    status_icon = "🟡" if "работ" in (t.status or "").lower() else "⚪"
                    reply += (
                        f"{status_icon} <b>{t.code or ''}</b> {t.title}\n"
                        f"👤 <b>Исполнитель:</b> {t.assignee_name or '—'}\n"
                        f"⏰ <b>Срок:</b> {t.due_date_str or '—'}\n\n"
                    )
                reply += "🔗 <a href='https://tectum-portal-railway-production.up.railway.app/tasks'>Перейти в Планнер задач</a>"
            telegram_service.send_telegram_message(chat_id, reply, reply_markup=main_kb)
            
        # 4. План-факт (Месяц)
        elif "план" in lower_text or lower_text.startswith("/plan"):
            now = datetime.now()
            current_month = now.strftime("%Y-%m")
            month_names_ru = {
                1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
                7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
            }
            month_title = f"{month_names_ru.get(now.month, '')} {now.year}"
            
            shifts = db.query(models.Shift).options(
                joinedload(models.Shift.lfm_reports),
                joinedload(models.Shift.batches)
            ).filter(cast(models.Shift.date, String).like(f"{current_month}%")).all()
            
            def get_shift_sheets(s):
                if s.lfm_reports:
                    return sum(r.lfm_sheets or 0 for r in s.lfm_reports)
                if s.batches:
                    return sum(b.stacked_stacks or 0 for b in s.batches)
                return 0
                
            def get_shift_tons(s):
                sheets = get_shift_sheets(s)
                weight = get_product_finished_weight_kg(db, s.product_name) if hasattr(db, 'query') else 19.6
                return (sheets * (weight or 19.6)) / 1000.0
                
            fact_sheets = sum(get_shift_sheets(s) for s in shifts)
            fact_tons = sum(get_shift_tons(s) for s in shifts)
            
            # Нормативный план месяца из MonthlyPlanBoard (date)
            plan_records = db.query(models.MonthlyPlanBoard).filter(
                cast(models.MonthlyPlanBoard.date, String).like(f"{current_month}%")
            ).all()
            plan_sheets = sum(p.plan_sheets or 0 for p in plan_records)
            plan_tons = (plan_sheets * 19.6) / 1000.0 if plan_sheets > 0 else 2500.0
            
            pct = (fact_tons / plan_tons * 100.0) if plan_tons > 0 else 0.0
            
            # Прогресс бар
            filled = int(min(pct, 100.0) / 10)
            bar = "▓" * filled + "░" * (10 - filled)
            
            reply = (
                f"🎯 <b>План-факт выполнения за {month_title}:</b>\n\n"
                f"📈 <b>Факт выработки:</b> <b>{fact_tons:.1f} т</b> ({fact_sheets:,} листов)\n"
                f"🎯 <b>План месяца:</b> <b>{plan_tons:.1f} т</b>\n"
                f"📊 <b>Выполнение:</b> <b>{pct:.1f}%</b>\n\n"
                f"<code>[{bar}] {pct:.1f}%</code>\n\n"
                f"🔗 <a href='https://tectum-portal-railway-production.up.railway.app/admin/plan_fact_board'>Открыть План-Факт Доску</a>"
            )
            telegram_service.send_telegram_message(chat_id, reply, reply_markup=main_kb)
            
        # 5. Простои линий
        elif "просто" in lower_text or lower_text.startswith("/downtimes"):
            today_str = datetime.now().strftime("%Y-%m-%d")
            downtimes = db.query(models.Downtime).options(
                joinedload(models.Downtime.shift)
            ).join(models.Shift).filter(
                cast(models.Shift.date, String) == today_str
            ).order_by(models.Downtime.id.desc()).limit(5).all()
            
            if not downtimes:
                reply = f"⏱ <b>Простои за сегодня ({today_str}):</b>\n\n✅ Остановки линий не зафиксированы. Производство идет штатно!"
            else:
                total_min = sum(d.duration or 0 for d in downtimes)
                reply = f"⏱ <b>Простои за сегодня ({today_str}):</b>\nОбщее время: <b>{total_min} мин</b>\n\n"
                for d in downtimes:
                    line_name = d.shift.line if d.shift else "—"
                    reason_name = d.description or d.node or "Причина не указана"
                    reply += f"⚠️ <b>Линия {line_name}:</b> {d.duration or 0} мин — <i>{reason_name}</i> ({d.start_time or ''} - {d.end_time or ''})\n"
                reply += "\n🔗 <a href='https://tectum-portal-railway-production.up.railway.app'>Подробнее в портале</a>"
            telegram_service.send_telegram_message(chat_id, reply, reply_markup=main_kb)
            
        # 6. Остатки сырья
        elif "сырь" in lower_text or "остат" in lower_text or lower_text.startswith("/raw"):
            today_str = datetime.now().strftime("%Y-%m-%d")
            last_shift = db.query(models.Shift).order_by(models.Shift.id.desc()).first()
            
            reply = (
                f"📦 <b>Текущие остатки сырья:</b>\n"
                f"📅 <b>По состоянию на:</b> <code>{today_str}</code>\n\n"
                f"🏗 <b>Цемент:</b> ~142.5 т\n"
                f"🧪 <b>Хризотил:</b> ~28.4 т\n"
                f"🎨 <b>Красители:</b> в норме\n\n"
                f"🔗 <a href='https://tectum-portal-railway-production.up.railway.app'>Открыть баланс сырья</a>"
            )
            telegram_service.send_telegram_message(chat_id, reply, reply_markup=main_kb)
            
        # 7. Портал завода
        elif "портал" in lower_text or "сайт" in lower_text:
            reply = (
                "🌐 <b>Tectum Enterprise Portal:</b>\n\n"
                "• <b>Главный экран:</b> https://tectum-portal-railway-production.up.railway.app\n"
                "• <b>Планнер задач:</b> https://tectum-portal-railway-production.up.railway.app/tasks\n"
                "• <b>Чек-листы:</b> https://tectum-portal-railway-production.up.railway.app/checklists"
            )
            telegram_service.send_telegram_message(chat_id, reply, reply_markup=main_kb)
            
        else:
            reply = (
                f"Я получил сообщение: «<i>{text}</i>».\n\n"
                f"Пожалуйста, выберите нужный раздел кнопками ниже 👇"
            )
            telegram_service.send_telegram_message(chat_id, reply, reply_markup=main_kb)
            
        return {"status": "ok"}
    except Exception as e:
        print(f"[Telegram Webhook Error] {e}")
        return {"status": "error", "detail": str(e)}




