import openpyxl
import os
import sys

# Add current directory to path to allow imports
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from database import SessionLocal, engine
import models

def import_downtimes():
    models.Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    
    # Clean existing
    db.query(models.DowntimeDirectory).delete()
    db.commit()
    
    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs", "excel", "Простои.xlsx")
    if not os.path.exists(excel_path):
        print(f"Error: File not found at {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    last_dept = ""
    last_node = ""
    last_col3 = ""
    
    count = 0
    # Start from row 2 (headers are on row 1)
    for row in range(2, sheet.max_row + 1):
        dept = sheet.cell(row=row, column=1).value
        node = sheet.cell(row=row, column=2).value
        col3 = sheet.cell(row=row, column=3).value
        col4 = sheet.cell(row=row, column=4).value
        
        if dept:
            last_dept = str(dept).strip().replace("\n", ", ")
        if node:
            last_node = str(node).strip().replace("\n", ", ")
        if col3:
            last_col3 = str(col3).strip().replace("\n", ", ")
            
        # Skip empty lines and helper rows
        if not node and not col3 and not col4:
            continue
        if last_node == "Санитарный день":
            continue
            
        dept_val = last_dept
        node_val = last_node
        breakdown_val = ""
        comment_val = None
        
        # Check if it's one of the newer departments with simplified layout
        is_later_dept = any(x in dept_val for x in ["Транспортерная лента", "ВСА", "Бракомешалка", "КВТ", "Дестакер", "Смазчик", "компрессор", "Рекуператор"])
        
        if is_later_dept:
            if col3:
                node_val = last_node
                breakdown_val = last_col3
            else:
                node_val = "Общее"
                breakdown_val = last_node
                
            if col4:
                comment_val = str(col4).strip()
        else:
            if col4:
                col4_str = str(col4).strip()
                if col4_str.lower().startswith("например"):
                    node_val = last_node
                    breakdown_val = last_col3
                    comment_val = col4_str
                else:
                    node_val = f"{last_node} - {last_col3}" if last_col3 else last_node
                    breakdown_val = col4_str
            else:
                node_val = last_node
                breakdown_val = last_col3
                
        if not breakdown_val:
            continue
            
        entry = models.DowntimeDirectory(
            department=dept_val,
            node=node_val,
            breakdown=breakdown_val,
            comment=comment_val
        )
        db.add(entry)
        count += 1
        
    db.commit()
    db.close()
    print(f"Successfully imported {count} downtime directory entries.")

if __name__ == "__main__":
    import_downtimes()
