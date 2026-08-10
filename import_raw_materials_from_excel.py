import os
import openpyxl
from datetime import datetime, date
from database import SessionLocal
import models
from import_aci_excel import safe_float

def import_receipts():
    db = SessionLocal()
    
    file_path = "рапорт_АЦИ 10.06.26.-3.xlsx"
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "сырье" not in wb.sheetnames:
        print("Sheet 'сырье' not found.")
        return
        
    sheet = wb["сырье"]
    
    receipts_added = 0
    
    for i, row in enumerate(sheet.iter_rows(min_row=3, values_only=True)):
        if not row:
            continue
            
        date_val = row[0]
        if not date_val:
            continue
            
        if isinstance(date_val, datetime):
            row_date = date_val.date()
        elif isinstance(date_val, date):
            row_date = date_val
        else:
            # try parsing
            try:
                row_date = datetime.strptime(str(date_val).split()[0], "%Y-%m-%d").date()
            except ValueError:
                continue
                
        # Parse Cement
        c_silo1 = safe_float(row[1])
        c_silo2 = safe_float(row[2])
        c_silo3 = safe_float(row[3])
        c_silo4 = safe_float(row[4])
        
        # Parse Asbestos (assuming col 15 is A5-65, 16 is A4-20, 17 is A6-45, 18 is OM)
        # We also need to be careful, on the screenshot: P=15, Q=16, R=17, S=18
        # index 0=A, 14=O, 15=P(А5-65), 16=Q(А4-20), 17=R(А6-45), 18=S(ОМ)
        a_5_65 = safe_float(row[15])
        a_4_20 = safe_float(row[16])
        a_6_45 = safe_float(row[17])
        a_om = safe_float(row[18])
        a_6_40 = a_6_45 + a_om
        
        total_receipts = c_silo1 + c_silo2 + c_silo3 + c_silo4 + a_5_65 + a_4_20 + a_6_40
        if total_receipts == 0:
            continue
            
        # Try to find a shift for this date to attach the receipt to
        shift = db.query(models.Shift).filter(models.Shift.date == row_date).order_by(models.Shift.id).first()
        shift_id = shift.id if shift else None
        
        # Check if receipt already exists for this date and shift (to avoid duplicates)
        existing = db.query(models.RawMaterialReceipt).filter(
            models.RawMaterialReceipt.shift_id == shift_id
        ).first()
        
        if existing and shift_id is not None:
            # Update existing
            existing.cement_silo1 = (existing.cement_silo1 or 0) + c_silo1
            existing.cement_silo2 = (existing.cement_silo2 or 0) + c_silo2
            existing.cement_silo3 = (existing.cement_silo3 or 0) + c_silo3
            existing.cement_silo4 = (existing.cement_silo4 or 0) + c_silo4
            existing.chrysotile_4_20 = (existing.chrysotile_4_20 or 0) + a_4_20
            existing.chrysotile_5_65 = (existing.chrysotile_5_65 or 0) + a_5_65
            existing.chrysotile_6_40 = (existing.chrysotile_6_40 or 0) + a_6_40
        else:
            # Create new
            new_receipt = models.RawMaterialReceipt(
                shift_id=shift_id,
                cement_silo1=c_silo1,
                cement_silo2=c_silo2,
                cement_silo3=c_silo3,
                cement_silo4=c_silo4,
                chrysotile_4_20=a_4_20,
                chrysotile_5_65=a_5_65,
                chrysotile_6_40=a_6_40,
                # Set others to 0
                cellulose=0.0,
                crushed_slate=0.0,
                asbozurit=0.0,
                asbocarton=0.0,
                fiberglass=0.0,
                laprol=0.0,
                pallets=0
            )
            db.add(new_receipt)
            receipts_added += 1
            
    db.commit()
    print(f"Successfully imported {receipts_added} days of receipts from Excel!")

if __name__ == "__main__":
    import_receipts()
