import openpyxl
import sys

try:
    file_path = "docs/excel/рапорт_АЦИ 10.06.26..xlsx"
    print(f"Reading {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
            # Filter empty rows
            if any(cell is not None for cell in row):
                print(f"Row {idx+1}: {row}")
except Exception as e:
    print(f"Error: {e}")
