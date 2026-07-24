import os
import openpyxl
from datetime import datetime, date
from sqlalchemy.orm import Session
from database import SessionLocal
from models import Master, Shift, LFMReport, Batch, MonthlyPlanBoard, AuditLog
from main import sync_lfm_to_plan_board

def safe_float(val):
    if val is None or val == "":
        return 0.0
    try:
        if isinstance(val, str):
            val = val.replace(" ", "").replace(",", ".")
            # remove % symbol if present
            val = val.replace("%", "")
        return float(val)
    except:
        return 0.0

def safe_int(val):
    if val is None or val == "":
        return 0
    try:
        if isinstance(val, str):
            val = val.replace(" ", "")
            # handle cases like 10.0
            if "." in val:
                return int(float(val))
        return int(val)
    except:
        return 0

def map_line_name(line: str) -> str:
    # DB mostly expects "ЛФМ-1" or "Линия 1" depending on where. 
    # Shift uses "Линия 1", MonthlyPlanBoard uses "ЛФМ-1".
    # We will standardize Shift to "Линия 1", "Линия 2".
    if "1" in str(line):
        return "Линия 1"
    if "2" in str(line) or "шанхай" in str(line).lower():
        return "Линия 2"
    return str(line)

def run_import():
    filepath = r"C:\Users\levda\Downloads\Сводный_отчет_Tectum.xlsx"
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return

    print("Loading Excel file...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "TectumSummaryTable" not in wb.sheetnames:
        print("Error: Sheet 'TectumSummaryTable' not found.")
        return
        
    sheet = wb["TectumSummaryTable"]
    
    # Map headers to indices
    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    
    def get_val(row, col_name):
        try:
            idx = headers.index(col_name)
            return row[idx].value
        except ValueError:
            return None

    db: Session = SessionLocal()
    imported_shifts = 0
    
    try:
        # Iterate over rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            raw_date = get_val(row, "Дата")
            shift_name = get_val(row, "Смена")
            line = get_val(row, "Линия")
            
            if not raw_date or not shift_name or not line:
                # empty row, probably end of file
                continue
                
            # Parse date
            shift_date = None
            if isinstance(raw_date, datetime):
                shift_date = raw_date.date()
            elif isinstance(raw_date, date):
                shift_date = raw_date
            else:
                try:
                    shift_date = datetime.strptime(str(raw_date).strip(), "%d.%m.%Y").date()
                except Exception as e:
                    print(f"Row {row_idx}: Invalid date {raw_date}: {e}")
                    continue
                    
            shift_name = str(shift_name).strip().capitalize()
            line_clean = map_line_name(line)
            
            master_name = str(get_val(row, "Мастер") or "").strip()
            # fuzzy find master
            master = db.query(Master).filter(Master.name.like(f"%{master_name}%")).first()
            master_id = master.id if master else None
            
            if not master_id:
                print(f"Row {row_idx}: Master '{master_name}' not found in DB. Leaving master_id empty.")
                
            product_name = str(get_val(row, "Наименование продукта") or "").strip()
            # Handle standard names
            if "7 волн 3500" in product_name:
                product_name = "Шифер 7 волн 3500*980"
            elif "7 волн" in product_name:
                product_name = "Шифер 7 волн"
            elif "8 волн гладкий" in product_name.lower():
                product_name = "Шифер 8 волн гладкий"
            elif "8 волн" in product_name:
                # Based on the rule, regular '8 волн' or 'пиленый' -> 'рифленый'
                product_name = "Шифер 8 волн рифленый"
                
            batch_number = str(get_val(row, "№ партии") or "").strip()
            zo_batches = safe_int(get_val(row, "Количество замесов"))
            
            # Find existing shift
            shift = db.query(Shift).filter(
                Shift.date == shift_date,
                Shift.shift_name == shift_name,
                Shift.line == line_clean
            ).first()
            
            if not shift:
                shift = Shift(
                    date=shift_date,
                    shift_name=shift_name,
                    line=line_clean
                )
                db.add(shift)
                
            shift.master_id = master_id
            shift.product_name = product_name
            shift.batch_number = batch_number
            shift.zo_batches = zo_batches
            
            # Raw materials (Расход ЗО)
            shift.zo_chrysotile_4_20 = safe_float(get_val(row, "Расход Хризотила 4-20 (кг)"))
            shift.zo_chrysotile_5_65 = safe_float(get_val(row, "Расход Хризотила 5-65 (кг)"))
            shift.zo_chrysotile_6_40 = safe_float(get_val(row, "Расход Хризотила 6-40 (кг)"))
            
            shift.zo_cement_silo1 = safe_float(get_val(row, "Расход Цемента С1 (кг)"))
            shift.zo_cement_silo2 = safe_float(get_val(row, "Расход Цемента С2 (кг)"))
            shift.zo_cement_silo3 = safe_float(get_val(row, "Расход Цемента С3 (кг)"))
            shift.zo_cement_silo4 = safe_float(get_val(row, "Расход Цемента С4 (кг)"))
            shift.zo_cement = safe_float(get_val(row, "Расход Цемента общ. (кг)"))
            
            shift.zo_cellulose = safe_float(get_val(row, "Расход Целлюлозы (кг)"))
            shift.zo_crushed_slate = safe_float(get_val(row, "Расход Дробленого шифера (кг)"))
            shift.zo_asbozurit = safe_float(get_val(row, "Расход Асбозурита (кг)"))
            shift.zo_fiberglass = safe_float(get_val(row, "Расход Стекловолокна (кг)"))
            shift.zo_laprol = safe_float(get_val(row, "Расход Лапрола (кг)"))
            shift.zo_asbocarton = safe_float(get_val(row, "Расход Асбокартона (кг)"))
            
            shift.lfm_asb_drain = safe_float(get_val(row, "Слив асб. (кг)"))
            shift.lfm_cem_drain = safe_float(get_val(row, "Слив цем. (кг)"))
            
            shift.zo_submitted = True
            
            db.flush() # get shift.id
            
            # LFM Report
            lfm = db.query(LFMReport).filter(LFMReport.shift_id == shift.id).first()
            if not lfm:
                lfm = LFMReport(shift_id=shift.id)
                db.add(lfm)
                
            lfm.product_name = product_name
            lfm.lfm_sheets = safe_int(get_val(row, "Формовка (листы)"))
            lfm.lfm_wind_resets = safe_int(get_val(row, "Сбросы наката"))
            lfm.transferred_to_warehouse = safe_int(get_val(row, "Кондиция (на склад)"))
            lfm.formed_1st_grade = safe_int(get_val(row, "1-сорт"))
            lfm.formed_defect = safe_int(get_val(row, "Брак"))
            
            # Batch
            if batch_number:
                b = db.query(Batch).filter(Batch.shift_id == shift.id, Batch.batch_number == batch_number).first()
                if not b:
                    b = Batch(shift_id=shift.id, batch_number=batch_number)
                    db.add(b)
                b.product_name = product_name
                # Usually destacker fields are ds_first_grade / ds_defect.
                # In summary report, 1-сорт / Брак represent final outcome, we can write them here too
                b.ds_first_grade = safe_int(get_val(row, "1-сорт"))
                b.ds_defect = safe_int(get_val(row, "Брак"))
                b.status = "qcd_checked"
            
            db.commit()
            
            # Sync to plan board (Fuzzy matching uses ЛФМ-1/ЛФМ-2 usually)
            pb_line = "ЛФМ-1" if "1" in line_clean else "ЛФМ-2"
            sync_lfm_to_plan_board(shift_date, shift_name, pb_line, db, master_id)
            
            imported_shifts += 1
            if imported_shifts % 10 == 0:
                print(f"Processed {imported_shifts} shifts...")
                
        print(f"Successfully processed {imported_shifts} historical shifts.")

    except Exception as e:
        db.rollback()
        print("An error occurred. Transaction rolled back.")
        import traceback
        traceback.print_exc()
    finally:
        db.close()

if __name__ == "__main__":
    run_import()
