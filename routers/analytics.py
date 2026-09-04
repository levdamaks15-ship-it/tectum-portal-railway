import os
import io
import calendar
from datetime import datetime, date, timedelta
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Request, Response, BackgroundTasks, Query, Body
from fastapi.responses import RedirectResponse
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import or_, and_, func

import models
import schemas
from database import SessionLocal
import excel_exporter
import openpyxl
from routers.common import (
    check_admin_session,
    get_product_finished_weight_kg,
    get_last_produced_weight_kg,
    get_shift_plan
)

try:
    import m365_integration
except ImportError:
    m365_integration = None

try:
    import google_sheets_integration
except ImportError:
    google_sheets_integration = None

router = APIRouter(tags=['analytics'])

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()



@router.get("/api/dashboard/weekly_report")
def get_weekly_report(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    # Р‘РµСЂРµРј РїРѕСЃР»РµРґРЅРёРµ 7 СЃРјРµРЅ (РІРєР»СЋС‡Р°СЏ С‚РµРєСѓС‰СѓСЋ)
    query = db.query(models.Shift)
    if False and user_role == "master" and user_id:
        query = query.filter(models.Shift.master_id == user_id)
    shifts = query.order_by(models.Shift.date.desc(), models.Shift.line.asc(), models.Shift.shift_name.desc(), models.Shift.batch_number.desc(), models.Shift.id.desc()).limit(7).all()
    
    report_data = []
    for shift in shifts:
        # 1. РЎС‡РёС‚Р°РµРј С„РѕСЂРјРѕРІРєСѓ (Р›Р¤Рњ)
        lfm_sheets = db.query(func.sum(models.LFMReport.lfm_sheets)).filter(models.LFMReport.shift_id == shift.id).scalar() or 0
        
        # 2. РЎС‡РёС‚Р°РµРј РёС‚РѕРі РЎРљРљ
        qcd_stats = db.query(
            func.sum(models.Batch.ds_condition).label('condition'),
            func.sum(models.Batch.ds_first_grade).label('first_grade'),
            func.sum(models.Batch.ds_defect).label('defect')
        ).filter(models.Batch.shift_id == shift.id).first()
        
        qcd_cond = qcd_stats.condition or 0
        qcd_fg = qcd_stats.first_grade or 0
        qcd_def = qcd_stats.defect or 0
        
        # 3. РЎС‡РёС‚Р°РµРј РѕС‚РєР»РѕРЅРµРЅРёРµ СЃС‹СЂСЊСЏ (Р¤Р°РєС‚ РёР· Р—Рћ - РўРµРѕСЂРёСЏ РїРѕ РЅРѕСЂРјР°Рј)
        lfm_reports = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift.id).all()
        product_counts = {}
        for r in lfm_reports:
            product_counts[r.product_name] = product_counts.get(r.product_name, 0) + r.lfm_sheets
            
        theoretical = {
            "chrysotile_4_20": 0.0, "chrysotile_5_65": 0.0, "chrysotile_6_40": 0.0,
            "cement": 0.0, "cellulose": 0.0, "crushed_slate": 0.0,
            "asbozurit": 0.0, "fiberglass": 0.0
        }
        
        for prod_name, sheets in product_counts.items():
            norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == prod_name).first()
            if norm:
                theoretical["chrysotile_4_20"] += sheets * norm.norm_chrysotile_4_20
                theoretical["chrysotile_5_65"] += sheets * norm.norm_chrysotile_5_65
                theoretical["chrysotile_6_40"] += sheets * norm.norm_chrysotile_6_40
                theoretical["cement"] += sheets * norm.norm_cement
                theoretical["cellulose"] += sheets * norm.norm_cellulose
                theoretical["crushed_slate"] += sheets * norm.norm_crushed_slate
                theoretical["asbozurit"] += sheets * norm.norm_asbozurit
                theoretical["fiberglass"] += sheets * norm.norm_fiberglass

        # Р¤Р°РєС‚РёС‡РµСЃРєРёР№ СЂР°СЃС…РѕРґ СЃС‹СЂСЊСЏ РёР· Р—Рћ Р·Р° СЃРјРµРЅСѓ
        fact_raw = (shift.zo_chrysotile_4_20 or 0.0) + \
                   (shift.zo_chrysotile_5_65 or 0.0) + \
                   (shift.zo_chrysotile_6_40 or 0.0) + \
                   (shift.zo_cement or 0.0) + \
                   (shift.zo_cellulose or 0.0) + \
                   (shift.zo_crushed_slate or 0.0) + \
                   (shift.zo_asbozurit or 0.0) + \
                   (shift.zo_fiberglass or 0.0)

        # РўРµРѕСЂРµС‚РёС‡РµСЃРєРёР№ СЂР°СЃС…РѕРґ СЃС‹СЂСЊСЏ
        theory_raw = sum(theoretical.values())
        
        # РћР±С‰РµРµ РѕС‚РєР»РѕРЅРµРЅРёРµ РїРѕ СЃС‹СЂСЊСЋ РІ РєРі (Р¤Р°РєС‚ - РўРµРѕСЂРёСЏ)
        deviation = fact_raw - theory_raw

        # Р¤Р°РєС‚РёС‡РµСЃРєРёР№ РІРµСЃ С„РѕСЂРјРѕРІРєРё РІ С‚РѕРЅРЅР°С…
        fact_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) / 1000.0 for r in lfm_reports)

        report_data.append({
            "id": shift.id,
            "date": shift.date.strftime("%Y-%m-%d") if shift.date else "Рќ/Р”",
            "shift_name": shift.shift_name,
            "line": shift.line,
            "master_name": shift.master.name if shift.master else "Рќ/Р”",
            "lfm_sheets": lfm_sheets,
            "qcd_condition": qcd_cond,
            "qcd_first_grade": qcd_fg,
            "qcd_defect": qcd_def,
            "raw_deviation": round(deviation, 2),
            "fact_tons": round(fact_tons, 2)
        })
        
    return report_data

@router.get("/api/dashboard/analytics_data")
def get_analytics_data(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    department: str = None,
    db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    query = db.query(models.Downtime).join(models.Shift)
    
    if False and user_role == "master" and user_id:
        query = query.filter(models.Shift.master_id == user_id)
    
    if start_date:
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d").date()
            query = query.filter(models.Shift.date >= sd)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD")
            
    if end_date:
        try:
            ed = datetime.strptime(end_date, "%Y-%m-%d").date()
            query = query.filter(models.Shift.date <= ed)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD")
            
    if department:
        query = query.filter(models.Downtime.department == department)
        
    downtimes = query.all()
    
    duration_with_stop = 0
    duration_without_stop = 0
    lost_tons_with_stop = 0.0
    lost_tons_without_stop = 0.0
    lost_tenge_with_stop = 0.0
    lost_tenge_without_stop = 0.0
    count_with_stop = 0
    count_without_stop = 0
    
    by_category = {}
    node_durations = {}
    trend_data = {}
    serialized_downtimes = []
    
    for dt in downtimes:
        dur = dt.duration or 0
        tons = dt.lost_tons or 0.0
        tenge = dt.lost_tenge or 0.0
        is_stop = bool(dt.is_equipment_downtime)
        
        if is_stop:
            duration_with_stop += dur
            lost_tons_with_stop += tons
            lost_tenge_with_stop += tenge
            count_with_stop += 1
        else:
            duration_without_stop += dur
            lost_tons_without_stop += tons
            lost_tenge_without_stop += tenge
            count_without_stop += 1
            
        cat = dt.category or "РќРµ СѓРєР°Р·Р°РЅР°"
        if cat not in by_category:
            by_category[cat] = {"with_stop": 0, "without_stop": 0}
        if is_stop:
            by_category[cat]["with_stop"] += dur
        else:
            by_category[cat]["without_stop"] += dur
            
        if is_stop and dt.node:
            node_durations[dt.node] = node_durations.get(dt.node, 0) + dur
            
        shift_date = dt.shift.date
        date_str = shift_date.strftime("%Y-%m-%d") if shift_date else "РќРµ СѓРєР°Р·Р°РЅР°"
        
        serialized_downtimes.append({
            "id": dt.id,
            "date": date_str,
            "shift": dt.shift.shift_name if dt.shift else "",
            "line": dt.shift.line if dt.shift else "",
            "master": dt.shift.master.name if dt.shift and dt.shift.master else "Рќ/Р”",
            "department": dt.department or "",
            "node": dt.node or "",
            "category": dt.category or "",
            "is_equipment_downtime": dt.is_equipment_downtime,
            "duration": dur,
            "lost_tons": tons,
            "lost_tenge": tenge,
            "description": dt.description or ""
        })
        date_str = shift_date.strftime("%Y-%m-%d") if shift_date else "РќРµ СѓРєР°Р·Р°РЅР°"
        if date_str not in trend_data:
            trend_data[date_str] = {}
        trend_data[date_str][cat] = trend_data[date_str].get(cat, 0) + dur

    bottlenecks = sorted([{"node": k, "duration": v} for k, v in node_durations.items()], key=lambda x: x['duration'], reverse=True)[:10]
    
    sorted_trend = {}
    for d in sorted(trend_data.keys()):
        sorted_trend[d] = trend_data[d]
        
    return {
        "kpis": {
            "with_stop": {
                "duration": duration_with_stop,
                "lost_tons": lost_tons_with_stop,
                "lost_tenge": lost_tenge_with_stop,
                "count": count_with_stop
            },
            "without_stop": {
                "duration": duration_without_stop,
                "lost_tons": lost_tons_without_stop,
                "lost_tenge": lost_tenge_without_stop,
                "count": count_without_stop
            }
        },
        "by_category": by_category,
        "bottlenecks": bottlenecks,
        "trend": sorted_trend,
        "downtimes": serialized_downtimes
    }

# --- РќРћР РњР« Р РђРЎРҐРћР”Рђ Р РћРўР§Р•Рў РџРћ РЎР«Р Р¬Р® ---
@router.get("/api/norms/", response_model=list[schemas.ProductNorm])
def get_product_norms(db: Session = Depends(get_db)):
    return db.query(models.ProductNorm).all()

_norms_cache = {}
_norms_cache_time = 0

def _get_norm_cached(db: Session, product_name: str):
    global _norms_cache, _norms_cache_time
    import time
    if time.time() - _norms_cache_time > 60:
        norms = db.query(models.ProductNorm).all()
        _norms_cache = {n.product_name: n for n in norms}
        _norms_cache_time = time.time()
    return _norms_cache.get(product_name)

def get_product_finished_weight_kg(db: Session, product_name: str) -> float:
    norm = _get_norm_cached(db, product_name)
    if not norm or not norm.weight_kg:
        return 19.6 # fallback for 8 РІРѕР»РЅ
    return norm.weight_kg

def get_product_raw_weight_kg(db: Session, product_name: str) -> float:
    norm = _get_norm_cached(db, product_name)
    if not norm:
        return 18.2 # fallback
    return (
        (norm.norm_chrysotile_4_20 or 0) +
        (norm.norm_chrysotile_5_65 or 0) +
        (norm.norm_chrysotile_6_40 or 0) +
        (norm.norm_cement or 0) +
        (norm.norm_cellulose or 0) +
        (norm.norm_crushed_slate or 0) +
        (norm.norm_asbozurit or 0) +
        (norm.norm_fiberglass or 0)
    )

def get_last_produced_weight_kg(db: Session, line_identifier: str, before_date_str: str = None) -> float:
    try:
        q = db.query(models.Shift).filter(models.Shift.line.like(f"%{line_identifier}%"))
        if before_date_str:
            q = q.filter(models.Shift.date <= before_date_str)
        shifts = q.order_by(models.Shift.date.desc(), models.Shift.id.desc()).limit(100).all()
        if not shifts:
            shifts = db.query(models.Shift).filter(models.Shift.line.like(f"%{line_identifier}%")).order_by(models.Shift.date.desc(), models.Shift.id.desc()).all()
        for s in shifts:
            if s.lfm_reports:
                for r in reversed(s.lfm_reports):
                    if r.lfm_sheets > 0 and r.product_name:
                        return get_product_finished_weight_kg(db, r.product_name)
    except Exception as e:
        print(f"Error in get_last_produced_weight_kg: {e}")
    return 19.6

def get_shift_plan(db: Session, shift: models.Shift) -> int:
    if shift.plan_sheets is not None and shift.plan_sheets > 0:
        return shift.plan_sheets
    sanitary_downtime = 0
    for dt in shift.downtimes:
        if dt.category == "РЎР°РЅРёС‚Р°СЂРЅС‹Р№ РґРµРЅСЊ":
            sanitary_downtime += dt.duration or 0
    if sanitary_downtime > 0:
        return 0
    if getattr(shift, "date", None) and shift.date.weekday() == 0 and shift.shift_name == "Р”РµРЅСЊ":
        return 0
    return 2700 if shift.shift_name == "Р”РµРЅСЊ" else 3300

@router.get("/api/dashboard/daily_report")
def get_daily_report(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    line: str = None,
    shift_number: int = None,
    master_id: int = None,
    db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"
    range_type_param = request.query_params.get("range_type")

    # Dynamic date range calculation based on frontend params
    sd = None
    ed = None

    if start_date:
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        except:
            raise HTTPException(400, "Invalid start_date format")
        if end_date:
            try:
                ed = datetime.strptime(end_date, "%Y-%m-%d").date()
            except:
                raise HTTPException(400, "Invalid end_date format")
        else:
            ed = sd + timedelta(days=6)
    else:
        # Fallback to query parameters passed by app.js (month, week, day, range_type)
        range_type = request.query_params.get("range_type", "month")
        month = request.query_params.get("month")
        week = request.query_params.get("week")
        day = request.query_params.get("day")

        if range_type == "month" and month:
            try:
                y, m = map(int, month.split('-'))
                num_days = calendar.monthrange(y, m)[1]
                sd = datetime(y, m, 1).date()
                ed = datetime(y, m, num_days).date()
            except Exception as e:
                raise HTTPException(400, f"Invalid month format: {e}")
        elif range_type == "week" and week:
            try:
                if not month:
                    now = datetime.now()
                    y, m = now.year, now.month
                else:
                    y, m = map(int, month.split('-'))
                
                week_num = int(week)
                first_day_of_month = datetime(y, m, 1).date()
                diff = -first_day_of_month.weekday()
                current_monday = first_day_of_month + timedelta(days=diff)
                
                if m == 12:
                    first_day_of_next_month = datetime(y + 1, 1, 1).date()
                else:
                    first_day_of_next_month = datetime(y, m + 1, 1).date()
                
                weeks = []
                while current_monday < first_day_of_next_month:
                    current_sunday = current_monday + timedelta(days=6)
                    weeks.append((current_monday, current_sunday))
                    current_monday += timedelta(days=7)
                
                idx = week_num - 1
                if 0 <= idx < len(weeks):
                    sd, ed = weeks[idx]
                elif len(weeks) > 0:
                    sd, ed = weeks[-1]
                else:
                    num_days = calendar.monthrange(y, m)[1]
                    sd = datetime(y, m, 1).date()
                    ed = datetime(y, m, num_days).date()
            except Exception as e:
                raise HTTPException(400, f"Invalid week or month format: {e}")
        elif range_type == "day" and day:
            try:
                sd = datetime.strptime(day, "%Y-%m-%d").date()
                ed = sd
            except Exception as e:
                raise HTTPException(400, f"Invalid day format: {e}")
        else:
            # Fallback to current month if no dates are provided
            now = datetime.now()
            y, m = now.year, now.month
            num_days = calendar.monthrange(y, m)[1]
            sd = datetime(y, m, 1).date()
            ed = datetime(y, m, num_days).date()

    num_days = (ed - sd).days + 1

    if not range_type_param:
        if num_days >= 28:
            effective_range_type = "month"
        elif num_days == 7:
            effective_range_type = "week"
        elif num_days == 1:
            effective_range_type = "day"
        else:
            effective_range_type = "custom"
    else:
        effective_range_type = range_type_param

    shifts_query = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports),
        selectinload(models.Shift.batches)
    ).filter(
        models.Shift.date >= sd,
        models.Shift.date <= ed
    )
    if master_id is not None:
        shifts_query = shifts_query.filter(models.Shift.master_id == master_id)
    shifts = shifts_query.all()
    
    plan_boards_query = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date >= sd,
        models.MonthlyPlanBoard.date <= ed
    )
    if master_id is not None:
        plan_boards_query = plan_boards_query.filter(models.MonthlyPlanBoard.master_id == master_id)
    plan_boards = plan_boards_query.all()
    
    if shift_number is not None:
        # Initialize plans to 0, because we will populate only matching shifts from pb
        data = {
            "line_1": {str(sd + timedelta(days=i)): {"Р”РµРЅСЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": 0, "plan_tons": 0.0, "first_grade": 0, "defect": 0}, "РќРѕС‡СЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": 0, "plan_tons": 0.0, "first_grade": 0, "defect": 0}} for i in range(num_days)},
            "line_2": {str(sd + timedelta(days=i)): {"Р”РµРЅСЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": 0, "plan_tons": 0.0, "first_grade": 0, "defect": 0}, "РќРѕС‡СЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": 0, "plan_tons": 0.0, "first_grade": 0, "defect": 0}} for i in range(num_days)}
        }
    else:
        # Default initialization with standard norms
        data = {
            "line_1": {str(sd + timedelta(days=i)): {"Р”РµРЅСЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700), "plan_tons": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700) * 19.6 / 1000.0, "first_grade": 0, "defect": 0}, "РќРѕС‡СЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": 3300, "plan_tons": 3300 * 19.6 / 1000.0, "first_grade": 0, "defect": 0}} for i in range(num_days)},
            "line_2": {str(sd + timedelta(days=i)): {"Р”РµРЅСЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700), "plan_tons": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700) * 19.6 / 1000.0, "first_grade": 0, "defect": 0}, "РќРѕС‡СЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": 3300, "plan_tons": 3300 * 19.6 / 1000.0, "first_grade": 0, "defect": 0}} for i in range(num_days)}
        }
    
    pb_map = {}
    for pb in plan_boards:
        pb_map[(pb.date, pb.shift_name, pb.line)] = pb
        
        day_key = str(pb.date)
        line_key = "line_1" if pb.line == "Р›Р¤Рњ-1" else "line_2"
        s_name = pb.shift_name
        if day_key in data[line_key] and s_name in ["Р”РµРЅСЊ", "РќРѕС‡СЊ"]:
            if shift_number is not None and pb.shift_number != shift_number:
                continue
            data[line_key][day_key][s_name]["plan_sheets"] = pb.plan_sheets or 0
            data[line_key][day_key][s_name]["plan_tons"] = (pb.plan_sheets or 0) * 19.6 / 1000.0
            
            # Р—Р°РїРёСЃС‹РІР°РµРј С„Р°РєС‚ РґР»СЏ РІСЃРµС…
            if True:
                data[line_key][day_key][s_name]["sheets"] = pb.fact_sheets or 0
                data[line_key][day_key][s_name]["tons"] = (pb.fact_sheets or 0) * 19.6 / 1000.0
                data[line_key][day_key][s_name]["first_grade"] = pb.first_grade or 0
                data[line_key][day_key][s_name]["defect"] = pb.defect or 0
            
    processed_slots = set()
    accumulate_sheets_slots = set()
    for s in shifts:
        if not s.date: continue
        # РќРµ РїСЂРѕРїСѓСЃРєР°РµРј СЃРјРµРЅС‹ РґСЂСѓРіРёС… РјР°СЃС‚РµСЂРѕРІ
        if False and user_role == "master" and s.master_id != user_id:
            continue
        day_key = str(s.date)
        line_key = "line_1" if "1" in s.line else "line_2"
        s_name = "Р”РµРЅСЊ" if s.shift_name == "Р”РµРЅСЊ" else "РќРѕС‡СЊ"
        
        if day_key not in data[line_key]:
            continue
            
        if shift_number is not None:
            pb_line_name = "Р›Р¤Рњ-1" if "1" in s.line else "Р›Р¤Рњ-2"
            pb_entry = pb_map.get((s.date, s.shift_name, pb_line_name))
            if pb_entry is None or pb_entry.shift_number != shift_number:
                continue
            
        total_w = 0
        total_s = 0
        total_1st = 0
        total_def = 0
        for r in s.lfm_reports:
            w_kg = get_product_finished_weight_kg(db, r.product_name)
            total_w += w_kg * r.lfm_sheets
            total_s += r.lfm_sheets
            
        for b in s.batches:
            total_1st += (b.ds_first_grade or 0)
            total_def += (b.ds_defect or 0)
            
        slot_key = (line_key, day_key, s_name)
        if slot_key not in processed_slots:
            processed_slots.add(slot_key)
            data[line_key][day_key][s_name]["tons"] = 0.0
            if data[line_key][day_key][s_name]["sheets"] == 0 or shift_number is not None:
                accumulate_sheets_slots.add(slot_key)
                data[line_key][day_key][s_name]["sheets"] = 0
            # Р’СЃРµРіРґР° РёРЅРёС†РёР°Р»РёР·РёСЂСѓРµРј Рё РЅР°РєР°РїР»РёРІР°РµРј 1 СЃРѕСЂС‚ Рё Р±СЂР°Рє РЅР°РїСЂСЏРјСѓСЋ РёР· СЂР°РїРѕСЂС‚РѕРІ РјР°СЃС‚РµСЂРѕРІ
            data[line_key][day_key][s_name]["first_grade"] = 0
            data[line_key][day_key][s_name]["defect"] = 0
        
        # РќР°РєР°РїР»РёРІР°РµРј 1 СЃРѕСЂС‚ Рё Р±СЂР°Рє Р”РµСЃС‚Р°РєРµСЂР° РЅР°РїСЂСЏРјСѓСЋ РёР· СЃРјРµРЅРЅРѕРіРѕ СЂР°РїРѕСЂС‚Р°
        data[line_key][day_key][s_name]["first_grade"] += total_1st
        data[line_key][day_key][s_name]["defect"] += total_def

        if total_s > 0:
            data[line_key][day_key][s_name]["tons"] += total_w / 1000.0
            if slot_key in accumulate_sheets_slots:
                data[line_key][day_key][s_name]["sheets"] += total_s
            
    last_known_weight = {}
    for l_k in data:
        line_name_for_q = "1" if l_k == "line_1" else "2"
        last_known_weight[l_k] = get_last_produced_weight_kg(db, line_name_for_q, str(sd))

    for i in range(num_days):
        day_k = str(sd + timedelta(days=i))
        for s_nm in ["Р”РµРЅСЊ", "РќРѕС‡СЊ"]:
            for l_k in data:
                if day_k in data[l_k] and s_nm in data[l_k][day_k]:
                    slot_info = data[l_k][day_k][s_nm]
                    if slot_info["sheets"] > 0 and slot_info["tons"] > 0:
                        avg_w = (slot_info["tons"] * 1000.0) / slot_info["sheets"]
                        slot_info["plan_tons"] = slot_info["plan_sheets"] * avg_w / 1000.0
                        last_known_weight[l_k] = avg_w
                    else:
                        slot_info["plan_tons"] = slot_info["plan_sheets"] * last_known_weight[l_k] / 1000.0
            
    # Now structure response as expected by app.js
    days_list = []
    lines_to_include = []
    if line == "lfm1":
        lines_to_include = ["line_1"]
    elif line == "lfm2":
        lines_to_include = ["line_2"]
    else:
        lines_to_include = ["line_1", "line_2"]

    for i in range(num_days):
        dt = sd + timedelta(days=i)
        date_str = str(dt)
        day_num = dt.day
        month_num = dt.month
        
        for s_name in ["Р”РµРЅСЊ", "РќРѕС‡СЊ"]:
            plan_sheets = 0
            fact_sheets = 0
            plan_tons = 0.0
            fact_tons = 0.0
            first_grade = 0
            defect = 0
            
            for l_key in lines_to_include:
                shift_data = data[l_key][date_str][s_name]
                plan_sheets += shift_data["plan_sheets"]
                fact_sheets += shift_data["sheets"]
                plan_tons += shift_data["plan_tons"]
                fact_tons += shift_data["tons"]
                first_grade += shift_data["first_grade"]
                defect += shift_data["defect"]
            
            suffix = "Р”" if s_name == "Р”РµРЅСЊ" else "Рќ"
            label = f"{day_num:02d}.{month_num:02d} ({suffix})"
            
            days_list.append({
                "date": date_str,
                "label": label,
                "plan_sheets": plan_sheets,
                "fact_sheets": fact_sheets,
                "plan_tons": plan_tons,
                "fact_tons": fact_tons,
                "first_grade": first_grade,
                "defect": defect
            })
        
            unique_shifts = set()
    for s in shifts:
        if line and not (("1" in s.line and line == "lfm1") or ("2" in s.line and line == "lfm2") or line == "all"):
            continue
        
        lfm_sheets = sum((r.lfm_sheets or 0) for r in s.lfm_reports) if getattr(s, 'lfm_reports', None) else 0
        warehouse_gp = sum((b.ds_condition or 0) for b in s.batches) if getattr(s, 'batches', None) else 0
        plan_sheets = s.plan_sheets or 0
        zo_batches = s.zo_batches or 0
        
        if plan_sheets == 0 and lfm_sheets == 0 and warehouse_gp == 0 and zo_batches == 0 and not getattr(s, 'zo_submitted', False):
            continue
            
        unique_shifts.add((s.date, s.shift_name, s.line))
            
    total_shifts = len(unique_shifts)
    total_fact_sheets = sum(d["fact_sheets"] for d in days_list)
    total_fact_tons = sum(d["fact_tons"] for d in days_list)

    if master_id is None and shift_number is None:
        if effective_range_type == "month" or num_days >= 28:
            total_plan_sheets = 160000 * len(lines_to_include)
        elif effective_range_type == "week" and num_days == 7:
            total_plan_sheets = 39000 * len(lines_to_include)
        else:
            total_plan_sheets = sum(d["plan_sheets"] for d in days_list)
    else:
        total_plan_sheets = sum(d["plan_sheets"] for d in days_list)

    if total_fact_sheets > 0:
        avg_period_weight = (total_fact_tons * 1000.0) / total_fact_sheets
        total_plan_tons = round((total_plan_sheets * avg_period_weight) / 1000.0, 2)
    else:
        total_plan_tons = round((total_plan_sheets * 19.6) / 1000.0, 2)

    avg_plan_percent = (total_fact_sheets / total_plan_sheets * 100.0) if total_plan_sheets > 0 else 0.0
    
    total_first_grade = sum(d["first_grade"] for d in days_list)
    total_defect = sum(d["defect"] for d in days_list)
    defect_percent = (total_defect / total_fact_sheets * 100.0) if total_fact_sheets > 0 else 0.0
    first_grade_percent = (total_first_grade / total_fact_sheets * 100.0) if total_fact_sheets > 0 else 0.0
    
    lag_sheets = total_plan_sheets - total_fact_sheets
    lag_tons = round(total_plan_tons - total_fact_tons, 2)
    
    return {
        "total_shifts": total_shifts,
        "total_fact_sheets": total_fact_sheets,
        "total_fact_tons": total_fact_tons,
        "total_plan_sheets": total_plan_sheets,
        "total_plan_tons": total_plan_tons,
        "lag_sheets": lag_sheets,
        "lag_tons": lag_tons,
        "total_first_grade": total_first_grade,
        "first_grade_percent": first_grade_percent,
        "total_defect": total_defect,
        "avg_plan_percent": avg_plan_percent,
        "defect_percent": defect_percent,
        "days": days_list
    }

@router.get("/api/dashboard/export_daily_report")
def export_daily_report(request: Request, start_date: str, line: str = None, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
    except:
        raise HTTPException(400, "Invalid date format")
        
    num_days = 14
    ed = sd + timedelta(days=num_days - 1)
    
    shifts = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports)
    ).filter(
        models.Shift.date >= sd,
        models.Shift.date <= ed
    ).all()
    
    plan_boards = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date >= sd,
        models.MonthlyPlanBoard.date <= ed
    ).all()
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    lines_to_export = [("Р›РёРЅРёСЏ 1", "Р›Р¤Рњ-1"), ("Р›РёРЅРёСЏ 2", "Р›Р¤Рњ-2")]
    if line == 'lfm1':
        lines_to_export = [("Р›РёРЅРёСЏ 1", "Р›Р¤Рњ-1")]
    elif line == 'lfm2':
        lines_to_export = [("Р›РёРЅРёСЏ 2", "Р›Р¤Рњ-2")]
        
    for line_id, line_label in lines_to_export:
        ws = wb.create_sheet(title=line_label)
        ws.append(["Р”Р°С‚Р°", "РЎРјРµРЅР°", "РџР»Р°РЅ (Р›РёСЃС‚С‹)", "Р¤Р°РєС‚ (Р›РёСЃС‚С‹)", "РџР»Р°РЅ (РўРѕРЅРЅС‹)", "Р¤Р°РєС‚ (РўРѕРЅРЅС‹)", "1-Р№ СЃРѕСЂС‚", "Р‘СЂР°Рє"])
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        
        day_data = {str(sd + timedelta(days=i)): {
            "Р”РµРЅСЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700), "plan_tons": (0 if (sd + timedelta(days=i)).weekday() == 0 else 2700) * 19.6 / 1000.0, "first_grade": 0, "defect": 0}, 
            "РќРѕС‡СЊ": {"sheets": 0, "tons": 0.0, "plan_sheets": 3300, "plan_tons": 3300 * 19.6 / 1000.0, "first_grade": 0, "defect": 0}
        } for i in range(num_days)}
        
        for pb in plan_boards:
            if pb.line != line_label: continue
            day_key = str(pb.date)
            s_name = pb.shift_name
            if day_key in day_data and s_name in ["Р”РµРЅСЊ", "РќРѕС‡СЊ"]:
                day_data[day_key][s_name]["plan_sheets"] = pb.plan_sheets or 0
                day_data[day_key][s_name]["plan_tons"] = (pb.plan_sheets or 0) * 19.6 / 1000.0
                
                # Р¤Р°РєС‚ Р·Р°РїРёСЃС‹РІР°РµРј С‚РѕР»СЊРєРѕ РґР»СЏ С‚РµРєСѓС‰РµРіРѕ РјР°СЃС‚РµСЂР° (РёР»Рё РµСЃР»Рё СЂРѕР»СЊ РЅРµ master)
                if user_role != "master" or pb.master_id == user_id:
                    day_data[day_key][s_name]["sheets"] = pb.fact_sheets or 0
                    day_data[day_key][s_name]["tons"] = (pb.fact_sheets or 0) * 19.6 / 1000.0
                    day_data[day_key][s_name]["first_grade"] = pb.first_grade or 0
                    day_data[day_key][s_name]["defect"] = pb.defect or 0
        
        processed_slots = set()
        accumulate_sheets_slots = set()
        for s in shifts:
            if not s.date or s.line != line_id: continue
            # РџСЂРѕРїСѓСЃРєР°РµРј СЃРјРµРЅС‹ РґСЂСѓРіРёС… РјР°СЃС‚РµСЂРѕРІ РґР»СЏ СЂРѕР»Рё master
            if False and user_role == "master" and s.master_id != user_id:
                continue
            day_key = str(s.date)
            if day_key not in day_data: continue
            
            s_name = "Р”РµРЅСЊ" if s.shift_name == "Р”РµРЅСЊ" else "РќРѕС‡СЊ"
            
            total_w = 0
            total_s = 0
            for r in s.lfm_reports:
                w_kg = get_product_finished_weight_kg(db, r.product_name)
                total_w += w_kg * r.lfm_sheets
                total_s += r.lfm_sheets
                
            if total_s > 0:
                slot_key = (day_key, s_name)
                if slot_key not in processed_slots:
                    processed_slots.add(slot_key)
                    day_data[day_key][s_name]["tons"] = 0.0
                    if day_data[day_key][s_name]["sheets"] == 0:
                        accumulate_sheets_slots.add(slot_key)
                        day_data[day_key][s_name]["sheets"] = 0
                
                day_data[day_key][s_name]["tons"] += total_w / 1000.0
                if slot_key in accumulate_sheets_slots:
                    day_data[day_key][s_name]["sheets"] += total_s
                
        last_w = get_last_produced_weight_kg(db, "1" if line_id == "lfm1" else "2", str(sd))
        for i in range(num_days):
            day_k = str(sd + timedelta(days=i))
            if day_k in day_data:
                for s_nm in ["Р”РµРЅСЊ", "РќРѕС‡СЊ"]:
                    slot_info = day_data[day_k][s_nm]
                    if slot_info["sheets"] > 0 and slot_info["tons"] > 0:
                        avg_w = (slot_info["tons"] * 1000.0) / slot_info["sheets"]
                        slot_info["plan_tons"] = slot_info["plan_sheets"] * avg_w / 1000.0
                        last_w = avg_w
                    else:
                        slot_info["plan_tons"] = slot_info["plan_sheets"] * last_w / 1000.0
                
        row_idx = 2
        for i in range(num_days):
            d_str = str(sd + timedelta(days=i))
            ws.append([d_str, "Р”РµРЅСЊ", day_data[d_str]["Р”РµРЅСЊ"]["plan_sheets"], day_data[d_str]["Р”РµРЅСЊ"]["sheets"], round(day_data[d_str]["Р”РµРЅСЊ"]["plan_tons"], 2), round(day_data[d_str]["Р”РµРЅСЊ"]["tons"], 2), day_data[d_str]["Р”РµРЅСЊ"]["first_grade"], day_data[d_str]["Р”РµРЅСЊ"]["defect"]])
            ws.append([d_str, "РќРѕС‡СЊ", day_data[d_str]["РќРѕС‡СЊ"]["plan_sheets"], day_data[d_str]["РќРѕС‡СЊ"]["sheets"], round(day_data[d_str]["РќРѕС‡СЊ"]["plan_tons"], 2), round(day_data[d_str]["РќРѕС‡СЊ"]["tons"], 2), day_data[d_str]["РќРѕС‡СЊ"]["first_grade"], day_data[d_str]["РќРѕС‡СЊ"]["defect"]])
            row_idx += 2
            
        chart_sheets = BarChart()
        chart_sheets.type = "col"
        chart_sheets.style = 10
        chart_sheets.title = f"Р’С‹СЂР°Р±РѕС‚РєР° {line_label} (Р›РёСЃС‚С‹)"
        chart_sheets.y_axis.title = 'РљРѕР»РёС‡РµСЃС‚РІРѕ (Р›РёСЃС‚С‹)'
        chart_sheets.x_axis.title = 'Р”Р°С‚Р° / РЎРјРµРЅР°'
        
        data_sheets = Reference(ws, min_col=3, min_row=1, max_row=row_idx-1, max_col=4)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row_idx-1, max_col=2)
        
        chart_sheets.add_data(data_sheets, titles_from_data=True)
        chart_sheets.set_categories(cats)
        chart_sheets.width = 20
        
        ws.add_chart(chart_sheets, "H2")
        
        chart_tons = BarChart()
        chart_tons.type = "col"
        chart_tons.style = 10
        chart_tons.title = f"Р’С‹СЂР°Р±РѕС‚РєР° {line_label} (РўРѕРЅРЅС‹)"
        chart_tons.y_axis.title = 'Р’РµСЃ (РўРѕРЅРЅС‹)'
        chart_tons.x_axis.title = 'Р”Р°С‚Р° / РЎРјРµРЅР°'
        
        data_tons = Reference(ws, min_col=5, min_row=1, max_row=row_idx-1, max_col=6)
        
        chart_tons.add_data(data_tons, titles_from_data=True)
        chart_tons.set_categories(cats)
        chart_tons.width = 20
        
        ws.add_chart(chart_tons, "H18")
        
    out = io.BytesIO()
    wb.save(out)
    
    filename = f"report_{start_date}_{line or 'all'}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return Response(content=out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@router.post("/api/admin/fix_plan_boards")
def fix_plan_boards(request: Request, db: Session = Depends(get_db)):
    user_role = request.session.get("user_role")
    if user_role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    try:
        boards = db.query(models.MonthlyPlanBoard).filter(models.MonthlyPlanBoard.plan_sheets == 0).all()
        updated_count = 0
        for pb in boards:
            date_val = pb.date
            is_monday = False
            if isinstance(date_val, str):
                try:
                    dt_obj = datetime.strptime(date_val, "%Y-%m-%d").date()
                    is_monday = dt_obj.weekday() == 0
                except:
                    pass
            else:
                try:
                    is_monday = date_val.weekday() == 0
                except:
                    pass
                    
            if is_monday and pb.shift_name == "Р”РµРЅСЊ":
                continue
                
            correct_plan = 2700 if pb.shift_name == "Р”РµРЅСЊ" else 3300
            pb.plan_sheets = correct_plan
            updated_count += 1
            
            log_entry = models.AuditLog(
                timestamp=datetime.utcnow(),
                user_name="System Admin",
                action="UPDATE",
                target_table="monthly_plan_board",
                target_id=pb.id,
                details=f"РСЃРїСЂР°РІР»РµРЅРёРµ РЅСѓР»РµРІРѕРіРѕ РїР»Р°РЅР°. РЈСЃС‚Р°РЅРѕРІР»РµРЅ РїР»Р°РЅ {correct_plan}."
            )
            db.add(log_entry)
            
        db.commit()
        return {"success": True, "updated_count": updated_count}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/dashboard/shift_board")
def get_shift_board(month: str, db: Session = Depends(get_db)):
    try:
        y, m = map(int, month.split('-'))
        num_days = calendar.monthrange(y, m)[1]
    except:
        raise HTTPException(400, "Invalid month format")
        
    month_start = datetime(y, m, 1).date()
    month_end = datetime(y, m, num_days).date()
    shifts = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports)
    ).filter(
        models.Shift.date >= month_start,
        models.Shift.date <= month_end
    ).order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
    board = {}
    for s in shifts:
        if not s.date: continue
        master = s.master_name or "РќРµРёР·РІРµСЃС‚РЅС‹Р№ РјР°СЃС‚РµСЂ"
        if master not in board:
            board[master] = []
            
        total_s = 0
        total_w = 0
        for r in s.lfm_reports:
            w_kg = get_product_finished_weight_kg(db, r.product_name)
            total_s += r.lfm_sheets
            total_w += r.lfm_sheets * w_kg
            
        plan_sheets = get_shift_plan(db, s)
        plan_tons = (plan_sheets * 19.6) / 1000.0
        if total_s > 0:
            avg_w = total_w / total_s
            plan_tons = (plan_sheets * avg_w) / 1000.0
            
        board[master].append({
            "shift_id": s.id,
            "date": str(s.date),
            "shift_name": s.shift_name,
            "line": s.line,
            "plan_sheets": plan_sheets,
            "fact_sheets": total_s,
            "plan_tons": round(plan_tons, 2),
            "fact_tons": round(total_w / 1000.0, 2),
            "closed": s.status == "closed"
        })
        
    return board

@router.get("/api/dashboard/export_shift")
def export_shift(shift_id: int = None, db: Session = Depends(get_db)):
    file_bytes = excel_exporter.generate_flat_report(db)
    filename = "РЎРІРѕРґРЅС‹Р№_РѕС‚С‡РµС‚_Tectum.xlsx"
    from urllib.parse import quote
    safe_filename = quote(filename)
    headers = {
        'Content-Disposition': f'attachment; filename="{safe_filename}"; filename*=UTF-8\'\'{safe_filename}'
    }
    return Response(
        content=file_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@router.post("/api/dashboard/sync_sharepoint")
def sync_sharepoint_manually(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="Р’С‹ РЅРµ Р°РІС‚РѕСЂРёР·РѕРІР°РЅС‹")
        
    if user_role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Р”РѕСЃС‚СѓРї Р·Р°РїСЂРµС‰РµРЅ. РўРѕР»СЊРєРѕ РјР°СЃС‚РµСЂ СЃРјРµРЅС‹ РёР»Рё Р°РґРјРёРЅРёСЃС‚СЂР°С‚РѕСЂ РјРѕРіСѓС‚ Р·Р°РїСѓСЃРєР°С‚СЊ СЃРёРЅС…СЂРѕРЅРёР·Р°С†РёСЋ.")
        
    try:
        file_bytes = excel_exporter.generate_flat_report(db)
        filename = "РЎРІРѕРґРЅС‹Р№_РѕС‚С‡РµС‚_Tectum.xlsx"
        
        # Save locally to static folder as well
        local_path = os.path.join("static", "РЎРІРѕРґРЅС‹Р№_РѕС‚С‡РµС‚_Tectum.xlsx")
        try:
            with open(local_path, "wb") as f:
                f.write(file_bytes)
        except Exception as local_err:
            print(f"Error saving local excel file: {local_err}")
            
        web_url = m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder="Reports")
        
        # Log to AuditLog
        db.add(models.AuditLog(
            user_name=request.session.get("user_email") or f"user_{user_id}",
            action="UPDATE",
            target_table="shifts",
            target_id=0,
            details=f"Р СѓС‡РЅР°СЏ СЃРёРЅС…СЂРѕРЅРёР·Р°С†РёСЏ РѕС‚С‡РµС‚Р° СЃ SharePoint РІС‹РїРѕР»РЅРµРЅР° СѓСЃРїРµС€РЅРѕ. РЎСЃС‹Р»РєР°: {web_url}"
        ))
        db.commit()
        return {"message": "РЎРёРЅС…СЂРѕРЅРёР·Р°С†РёСЏ РІС‹РїРѕР»РЅРµРЅР° СѓСЃРїРµС€РЅРѕ", "url": web_url}
    except Exception as e:
        error_msg = str(e)
        if "423" in error_msg or "Locked" in error_msg:
            raise HTTPException(status_code=423, detail="Р¤Р°Р№Р» РѕС‚С‡РµС‚Р° РІСЃРµ РµС‰Рµ Р·Р°Р±Р»РѕРєРёСЂРѕРІР°РЅ РІ SharePoint (РєС‚Рѕ-С‚Рѕ РѕС‚РєСЂС‹Р» РµРіРѕ РІ Excel Online). Р—Р°РєСЂРѕР№С‚Рµ С„Р°Р№Р» Рё РїРѕРїСЂРѕР±СѓР№С‚Рµ СЃРЅРѕРІР°.")
        else:
            raise HTTPException(status_code=500, detail=f"РћС€РёР±РєР° СЃРёРЅС…СЂРѕРЅРёР·Р°С†РёРё: {error_msg}")

@router.post("/api/dashboard/sync_google_sheets_manual")
def sync_google_sheets_manual(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="Р’С‹ РЅРµ Р°РІС‚РѕСЂРёР·РѕРІР°РЅС‹")
        
    if user_role not in ["master", "admin"]:
        raise HTTPException(status_code=403, detail="Р”РѕСЃС‚СѓРї Р·Р°РїСЂРµС‰РµРЅ. РўРѕР»СЊРєРѕ РјР°СЃС‚РµСЂР° РёР»Рё Р°РґРјРёРЅРёСЃС‚СЂР°С‚РѕСЂС‹ РјРѕРіСѓС‚ Р·Р°РїСѓСЃРєР°С‚СЊ РІС‹РіСЂСѓР·РєСѓ.")
        
    try:
        google_sheets_integration.sync_report_to_google_sheets(db)
        
        # Log to AuditLog
        db.add(models.AuditLog(
            user_name=request.session.get("user_email") or f"user_{user_id}",
            action="UPDATE",
            target_table="shifts",
            target_id=0,
            details="Р’С‹РїРѕР»РЅРµРЅР° СЂСѓС‡РЅР°СЏ РІС‹РіСЂСѓР·РєР° СЃРІРѕРґРЅРѕРіРѕ РѕС‚С‡РµС‚Р° РІ Google РўР°Р±Р»РёС†С‹."
        ))
        db.commit()
        return {"message": "Р’С‹РіСЂСѓР·РєР° РІ Google РўР°Р±Р»РёС†С‹ РІС‹РїРѕР»РЅРµРЅР° СѓСЃРїРµС€РЅРѕ!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"РћС€РёР±РєР° РІС‹РіСЂСѓР·РєРё РІ Google: {str(e)}")

@router.get("/api/dashboard/view_archive")
def view_archive(db: Session = Depends(get_db)):
    try:
        url = m365_integration.get_file_web_url("РЎРІРѕРґРЅС‹Р№_РѕС‚С‡РµС‚_Tectum.xlsx", folder="Reports")
        return RedirectResponse(url=url)
    except Exception as e:
        print("РћС€РёР±РєР° РїРѕР»СѓС‡РµРЅРёСЏ СЃСЃС‹Р»РєРё РёР· SharePoint, РїСЂРѕР±СѓРµРј СЃРіРµРЅРµСЂРёСЂРѕРІР°С‚СЊ Рё Р·Р°РіСЂСѓР·РёС‚СЊ РѕС‚С‡РµС‚:", e)
        try:
            file_bytes = excel_exporter.generate_flat_report(db)
            filename = "РЎРІРѕРґРЅС‹Р№_РѕС‚С‡РµС‚_Tectum.xlsx"
            url = m365_integration.upload_file_to_sharepoint(file_bytes, filename, folder="Reports")
            return RedirectResponse(url=url)
        except Exception as upload_err:
            import traceback
            traceback.print_exc()
            raise HTTPException(
                status_code=500,
                detail=f"РќРµ СѓРґР°Р»РѕСЃСЊ РѕС‚РєСЂС‹С‚СЊ СЃРІРѕРґРЅС‹Р№ РѕС‚С‡РµС‚ РІ SharePoint. РћС€РёР±РєР° Р°РІС‚РѕР·Р°РіСЂСѓР·РєРё: {upload_err}. РСЃС…РѕРґРЅР°СЏ РѕС€РёР±РєР°: {e}"
            )

@router.get("/api/dashboard/export_week")
def export_week(request: Request, start_date: str, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
    except:
        raise HTTPException(400, "Invalid date format, use YYYY-MM-DD")
        
    ed = sd + timedelta(days=6)
    
    query = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports),
        selectinload(models.Shift.receipts),
        selectinload(models.Shift.downtimes)
    ).filter(
        models.Shift.date >= sd,
        models.Shift.date <= ed
    )
    if False and user_role == "master" and user_id:
        query = query.filter(models.Shift.master_id == user_id)
    shifts = query.order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"РќРµРґРµР»СЏ {sd} - {ed}"
    
    ws.append([f"РћС‚С‡РµС‚ Р·Р° РЅРµРґРµР»СЋ СЃ {sd} РїРѕ {ed}"])
    ws.append(["Р”Р°С‚Р°", "РЎРјРµРЅР°", "РњР°СЃС‚РµСЂ", "Р›РёРЅРёСЏ", "РџР»Р°РЅ (Р›РёСЃС‚С‹)", "Р¤Р°РєС‚ (Р›РёСЃС‚С‹)", "РџР»Р°РЅ (РўРѕРЅРЅС‹)", "Р¤Р°РєС‚ (РўРѕРЅРЅС‹)"])
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    plan_boards = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date >= sd,
        models.MonthlyPlanBoard.date <= ed
    ).all()
    pb_dict = {(pb.date, pb.shift_name, pb.line): pb for pb in plan_boards}

    active_lines = set([s.line.replace("Р›РёРЅРёСЏ ", "Р›Р¤Рњ-") for s in shifts if s.line] + [pb.line for pb in plan_boards])
    if not active_lines:
        active_lines = {"Р›Р¤Рњ-2"}

    for l_key in active_lines:
        last_w = get_last_produced_weight_kg(db, "1" if "1" in l_key else "2", str(sd)) / 1000.0
        for i in range(7):
            d = sd + timedelta(days=i)
            for s_name in ["Р”РµРЅСЊ", "РќРѕС‡СЊ"]:
                plan_sheets = 0 if d.weekday() == 0 and s_name == "Р”РµРЅСЊ" else (2700 if s_name == "Р”РµРЅСЊ" else 3300)
                
                pb = pb_dict.get((d, s_name, l_key))
                if pb and pb.plan_sheets is not None:
                    plan_sheets = pb.plan_sheets
                    
                slot_shifts = [shift for shift in shifts if shift.date == d and shift.shift_name == s_name and (shift.line.replace("Р›РёРЅРёСЏ ", "Р›Р¤Рњ-") if shift.line else "Р›Р¤Рњ-1") == l_key]
                s = slot_shifts[0] if slot_shifts else None
                
                show_fact = (user_role != "master" or (pb and pb.master_id == user_id) or (s and s.master_id == user_id))
                total_sheets = pb.fact_sheets if (pb and show_fact) else 0
                
                if s and show_fact:
                    sum_lfm_sheets = sum(r.lfm_sheets for sh in slot_shifts for r in sh.lfm_reports)
                    sum_lfm_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) / 1000.0 for sh in slot_shifts for r in sh.lfm_reports)
                    if sum_lfm_sheets > 0:
                        avg_w = (sum_lfm_tons / sum_lfm_sheets)
                        last_w = avg_w
                    else:
                        avg_w = last_w
                    if total_sheets == 0 and sum_lfm_sheets > 0:
                        total_sheets = sum_lfm_sheets
                    master_name = s.master.name if s.master else "Рќ/Р”"
                else:
                    avg_w = last_w
                    master_name = "Рќ/Р”" if show_fact else "РЎРјРµРЅР° РґСЂ. РјР°СЃС‚РµСЂР°"
                    total_sheets = pb.fact_sheets if (pb and show_fact) else 0
                    
                plan_tons = plan_sheets * avg_w
                total_tons = sum_lfm_tons if (s and show_fact and sum_lfm_sheets > 0) else (total_sheets * avg_w)
                
                ws.append([str(d), s_name, master_name, l_key, plan_sheets, total_sheets, round(plan_tons, 2), round(total_tons, 2)])
    out = io.BytesIO()
    wb.save(out)
    
    filename = f"week_{sd}.xlsx"
    return Response(content=out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={'Content-Disposition': f'attachment; filename="{filename}"'})

@router.get("/api/dashboard/weekly")
def get_weekly_json(request: Request, start_date: str, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    user_role = request.session.get("user_role") or "admin"

    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
    except:
        raise HTTPException(400, "Invalid date format, use YYYY-MM-DD")
        
    ed = sd + timedelta(days=6)
    
    query = db.query(models.Shift).options(
        selectinload(models.Shift.lfm_reports),
        selectinload(models.Shift.receipts),
        selectinload(models.Shift.downtimes)
    ).filter(
        models.Shift.date >= sd,
        models.Shift.date <= ed
    )
    if False and user_role == "master" and user_id:
        query = query.filter(models.Shift.master_id == user_id)
    shifts = query.order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
    plan_boards = db.query(models.MonthlyPlanBoard).filter(
        models.MonthlyPlanBoard.date >= sd,
        models.MonthlyPlanBoard.date <= ed
    ).all()
    pb_dict = {(pb.date, pb.shift_name, pb.line): pb for pb in plan_boards}
    
    active_lines = set([s.line.replace("Р›РёРЅРёСЏ ", "Р›Р¤Рњ-") for s in shifts if s.line] + [pb.line for pb in plan_boards])
    if not active_lines:
        active_lines = {"Р›Р¤Рњ-2"}
        
    data = []
    
    for l_key in active_lines:
        last_w = get_last_produced_weight_kg(db, "1" if "1" in l_key else "2", str(sd)) / 1000.0
        for i in range(7):
            d = sd + timedelta(days=i)
            day_str = str(d)
            for s_name in ["Р”РµРЅСЊ", "РќРѕС‡СЊ"]:
                plan_sheets = 0 if d.weekday() == 0 and s_name == "Р”РµРЅСЊ" else (2700 if s_name == "Р”РµРЅСЊ" else 3300)
                
                pb = pb_dict.get((d, s_name, l_key))
                if pb and pb.plan_sheets is not None:
                    plan_sheets = pb.plan_sheets
                    
                slot_shifts = [shift for shift in shifts if shift.date == d and shift.shift_name == s_name and (shift.line.replace("Р›РёРЅРёСЏ ", "Р›Р¤Рњ-") if shift.line else "Р›Р¤Рњ-1") == l_key]
                s = slot_shifts[0] if slot_shifts else None
                
                show_fact = (user_role != "master" or (pb and pb.master_id == user_id) or (s and s.master_id == user_id))
                total_sheets = pb.fact_sheets if (pb and show_fact) else 0
                
                if s and show_fact:
                    sum_lfm_sheets = sum(r.lfm_sheets for sh in slot_shifts for r in sh.lfm_reports)
                    sum_lfm_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) / 1000.0 for sh in slot_shifts for r in sh.lfm_reports)
                    if sum_lfm_sheets > 0:
                        avg_w = (sum_lfm_tons / sum_lfm_sheets)
                        last_w = avg_w
                    else:
                        avg_w = last_w
                    if total_sheets == 0 and sum_lfm_sheets > 0:
                        total_sheets = sum_lfm_sheets
                        
                    if pb and (pb.first_grade or pb.defect):
                        ds_first = pb.first_grade
                        ds_defect = pb.defect
                    else:
                        ds_first = sum(b.ds_first_grade for sh in slot_shifts for b in sh.batches)
                        ds_defect = sum(b.ds_defect for sh in slot_shifts for b in sh.batches)
                        
                    qcd_first = sum(b.ds_first_grade for sh in slot_shifts for b in sh.batches)
                    qcd_defect = sum(b.ds_defect for sh in slot_shifts for b in sh.batches)
                    
                    sanitary_note = ""
                    for dt in s.downtimes:
                        if dt.category == "РЎР°РЅРёС‚Р°СЂРЅС‹Р№ РґРµРЅСЊ":
                            sanitary_note = "РЎР°РЅРёС‚Р°СЂРЅС‹Р№ РґРµРЅСЊ"
                            if dt.duration:
                                sanitary_note += f" ({dt.duration} РјРёРЅ)"
                            break
                    master_name = s.master.name if s.master else "Рќ/Р”"
                    shift_id = s.id
                else:
                    avg_w = last_w
                    ds_first = pb.first_grade if (pb and show_fact) else 0
                    ds_defect = pb.defect if (pb and show_fact) else 0
                    qcd_first = 0
                    qcd_defect = 0
                    sanitary_note = "РЎР°РЅРёС‚Р°СЂРЅС‹Р№ РґРµРЅСЊ (РїР»Р°РЅ 0)" if d.weekday() == 0 and s_name == "Р”РµРЅСЊ" else ("РќРµС‚ РґР°РЅРЅС‹С…" if show_fact else "РЎРјРµРЅР° РґСЂСѓРіРѕРіРѕ РјР°СЃС‚РµСЂР°")
                    master_name = "Рќ/Р”"
                    shift_id = None
                    
                plan_tons = plan_sheets * avg_w
                total_tons = sum_lfm_tons if (s and show_fact and sum_lfm_sheets > 0) else (total_sheets * avg_w)
                
                data.append({
                    "id": shift_id,
                    "date": day_str,
                    "shift_name": s_name,
                    "master": master_name,
                    "line": l_key,
                    "plan_sheets": plan_sheets,
                    "fact_sheets": total_sheets,
                    "plan_tons": round(plan_tons, 2),
                    "fact_tons": round(total_tons, 2),
                    "ds_first_grade": ds_first,
                    "ds_defect": ds_defect,
                    "qcd_first_grade": qcd_first,
                    "qcd_defect": qcd_defect,
                    "note": sanitary_note
                })
        
    return {
        "start_date": str(sd),
        "end_date": str(ed),
        "data": data
    }


@router.get("/api/shifts/{shift_id}/materials_report", response_model=schemas.RawMaterialReport)
def get_materials_report(shift_id: int, db: Session = Depends(get_db)):
    shift = db.query(models.Shift).get(shift_id)
    if not shift:
        raise HTTPException(404, "РЎРјРµРЅР° РЅРµ РЅР°Р№РґРµРЅР°")
    
    # 1. РЎС‡РёС‚Р°РµРј РїСЂРѕРёР·РІРµРґРµРЅРЅСѓСЋ РїСЂРѕРґСѓРєС†РёСЋ (Р¤РѕСЂРјРѕРІРєР°)
    lfm_reports = db.query(models.LFMReport).filter(models.LFMReport.shift_id == shift_id).all()
    product_counts = {}
    for r in lfm_reports:
        product_counts[r.product_name] = product_counts.get(r.product_name, 0) + r.lfm_sheets
        
    # 2. РџРѕР»СѓС‡Р°РµРј РЅРѕСЂРјС‹ РґР»СЏ СЌС‚РёС… РїСЂРѕРґСѓРєС‚РѕРІ Рё СЃС‡РёС‚Р°РµРј С‚РµРѕСЂРёСЋ
    theoretical = {
        "chrysotile_4_20": 0.0, "chrysotile_5_65": 0.0, "chrysotile_6_40": 0.0,
        "cement": 0.0, "cellulose": 0.0, "crushed_slate": 0.0,
        "asbozurit": 0.0, "fiberglass": 0.0
    }
    
    for prod_name, sheets in product_counts.items():
        norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == prod_name).first()
        if norm:
            theoretical["chrysotile_4_20"] += sheets * norm.norm_chrysotile_4_20
            theoretical["chrysotile_5_65"] += sheets * norm.norm_chrysotile_5_65
            theoretical["chrysotile_6_40"] += sheets * norm.norm_chrysotile_6_40
            theoretical["cement"] += sheets * norm.norm_cement
            theoretical["cellulose"] += sheets * norm.norm_cellulose
            theoretical["crushed_slate"] += sheets * norm.norm_crushed_slate
            theoretical["asbozurit"] += sheets * norm.norm_asbozurit
            theoretical["fiberglass"] += sheets * norm.norm_fiberglass

    # 3. Р¤РѕСЂРјРёСЂСѓРµРј РґРµС‚Р°Р»СЊРЅС‹Р№ РѕС‚С‡РµС‚ (Р¤Р°РєС‚ РёР· ZO - РўРµРѕСЂРёСЏ)
    details = []
    total_dev = 0.0
    
    mapping = [
        ("РҐСЂРёР·РѕС‚РёР» 4-20", shift.zo_chrysotile_4_20, theoretical["chrysotile_4_20"]),
        ("РҐСЂРёР·РѕС‚РёР» 5-65", shift.zo_chrysotile_5_65, theoretical["chrysotile_5_65"]),
        ("РҐСЂРёР·РѕС‚РёР» 6-40", shift.zo_chrysotile_6_40, theoretical["chrysotile_6_40"]),
        ("Р¦РµРјРµРЅС‚", shift.zo_cement, theoretical["cement"]),
        ("Р¦РµР»Р»СЋР»РѕР·Р°", shift.zo_cellulose, theoretical["cellulose"]),
        ("Р”СЂРѕР±Р»РµРЅС‹Р№ С€РёС„РµСЂ", shift.zo_crushed_slate, theoretical["crushed_slate"]),
        ("РђСЃР±РѕР·СѓСЂРёС‚", shift.zo_asbozurit, theoretical["asbozurit"]),
        ("РЎС‚РµРєР»РѕРІРѕР»РѕРєРЅРѕ", shift.zo_fiberglass, theoretical["fiberglass"]),
        ("Р›Р°РїСЂРѕР»", shift.zo_laprol, 0.0),
        ("РђСЃР±РѕРєР°СЂС‚РѕРЅ", shift.zo_asbocarton, 0.0)
    ]

    
    total_sheets = sum(product_counts.values())
    
    for mat_name, actual, theory in mapping:
        actual_val = actual or 0.0
        theory_val = theory or 0.0
        dev = actual_val - theory_val
        total_dev += dev
        
        unit_actual = actual_val / total_sheets if total_sheets > 0 else 0.0
        unit_theory = theory_val / total_sheets if total_sheets > 0 else 0.0
        unit_dev = dev / total_sheets if total_sheets > 0 else 0.0
        
        details.append({
            "material": mat_name,
            "actual": round(actual_val, 2),
            "theoretical": round(theory_val, 2),
            "deviation": round(dev, 2),
            "unit_actual": round(unit_actual, 4),
            "unit_theoretical": round(unit_theory, 4),
            "unit_deviation": round(unit_dev, 4)
        })
        
    return {
        "shift_id": shift_id,
        "total_deviation_kg": round(total_dev, 2),
        "details": details
    }


