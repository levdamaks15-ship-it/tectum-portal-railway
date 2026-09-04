import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from sqlalchemy.orm import Session
import models

def get_product_finished_weight_kg(db: Session, product_name: str) -> float:
    norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == product_name).first()
    if not norm or not norm.weight_kg:
        return 19.6
    return norm.weight_kg

def get_shift_plan(db: Session, shift) -> int:
    if shift.plan_sheets is not None and shift.plan_sheets > 0:
        return shift.plan_sheets
    sanitary_downtime = 0
    for dt in shift.downtimes:
        if dt.category == "Санитарный день":
            sanitary_downtime += dt.duration or 0
    if sanitary_downtime > 0:
        return 0
    if getattr(shift, "date", None) and shift.date.weekday() == 0 and shift.shift_name == "День":
        return 0
    return 2700 if shift.shift_name == "День" else 3300

def generate_flat_report(db: Session) -> bytes:
    shifts = db.query(models.Shift).filter(models.Shift.status == "closed").order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводный отчет"
    
    headers = [
        "Дата", "№ партии", "Линия", "Смена", "Мастер", "Наименование продукта", "Назначение (Экспорт)",
        "Количество замесов", "Формовка (листы)", "Формовка (тонны)",
        "Кондиция (на склад)", "1-сорт", "Брак", "Сбросы наката",
        "Слив асб. (кг)", "Слив цем. (кг)",
        "Расход Хризотила 4-20 (кг)", "Расход Хризотила 5-65 (кг)", "Расход Хризотила 6-40 (кг)", "Расход Хризотила общ. (кг)",
        "Расход Цемента С1 (кг)", "Расход Цемента С2 (кг)", "Расход Цемента С3 (кг)", "Расход Цемента С4 (кг)", "Расход Цемента общ. (кг)",
        "Расход Асбокартона (кг)", "Расход Лапрола (кг)", "Расход Целлюлозы (кг)", "Расход Стекловолокна (кг)",
        "Расход Дробленого шифера (кг)", "Расход Асбозурита (кг)",
        "Отклонение Хризотила 4-20 (%)", "Отклонение Хризотила 5-65 (%)", "Отклонение Хризотила 6-40 (%)", "Отклонение Хризотила общ. (%)",
        "Отклонение Цемента общ. (%)", "Отклонение Асбокартона (%)", "Отклонение Лапрола (%)", "Отклонение Целлюлозы (%)",
        "Отклонение Стекловолокна (%)", "Отклонение Дробленого шифера (%)", "Отклонение Асбозурита (%)"
    ]
    
    # Header styling (navy blue theme)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    ws.append(headers)
    
    # Format header row
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin
        
    # Content rows style
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft pastel red
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft pastel green
    
    for s in shifts:
        # Проверяем, есть ли плановые или фактические показатели производства в смене
        plan_sheets_check = s.plan_sheets or 0
        formovka_sheets_check = sum(r.lfm_sheets for r in s.lfm_reports)
        warehouse_gp_check = sum(b.qcd_condition for b in s.batches)
        zo_batches_check = s.zo_batches or 0
        
        if plan_sheets_check == 0 and formovka_sheets_check == 0 and warehouse_gp_check == 0 and zo_batches_check == 0 and not s.zo_submitted:
            continue

        # Date
        date_str = s.date.strftime("%d.%m.%Y") if s.date else ""
        
        # Batch numbers
        batch_numbers = ", ".join(b.batch_number for b in s.batches if b.batch_number)
        
        # Products
        product_names = ", ".join(set(r.product_name for r in s.lfm_reports if r.product_name))
        
        # Formovka sheets and tons
        formovka_sheets = sum(r.lfm_sheets for r in s.lfm_reports)
        formovka_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) for r in s.lfm_reports) / 1000.0
        
        # Quality (QCD / СКК)
        qcd_condition = sum(b.qcd_condition for b in s.batches)
        qcd_first = sum(b.qcd_first_grade for b in s.batches)
        qcd_defect = sum(b.qcd_defect for b in s.batches)
        
        # Wind resets
        wind_resets = sum(r.lfm_wind_resets for r in s.lfm_reports)
        
        # Raw materials theoretical norms sum
        theory = {
            "chrysotile_4_20": 0.0, "chrysotile_5_65": 0.0, "chrysotile_6_40": 0.0,
            "cement": 0.0, "cellulose": 0.0, "crushed_slate": 0.0,
            "asbozurit": 0.0, "fiberglass": 0.0, "asbocarton": 0.0, "laprol": 0.0
        }
        for r in s.lfm_reports:
            norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == r.product_name).first()
            if norm:
                theory["chrysotile_4_20"] += r.lfm_sheets * (norm.norm_chrysotile_4_20 or 0.0)
                theory["chrysotile_5_65"] += r.lfm_sheets * (norm.norm_chrysotile_5_65 or 0.0)
                theory["chrysotile_6_40"] += r.lfm_sheets * (norm.norm_chrysotile_6_40 or 0.0)
                theory["cement"] += r.lfm_sheets * (norm.norm_cement or 0.0)
                theory["cellulose"] += r.lfm_sheets * (norm.norm_cellulose or 0.0)
                theory["crushed_slate"] += r.lfm_sheets * (norm.norm_crushed_slate or 0.0)
                theory["asbozurit"] += r.lfm_sheets * (norm.norm_asbozurit or 0.0)
                theory["fiberglass"] += r.lfm_sheets * (norm.norm_fiberglass or 0.0)
                
        # Factual consumption
        fact = {
            "chrysotile_4_20": s.zo_chrysotile_4_20 or 0.0,
            "chrysotile_5_65": s.zo_chrysotile_5_65 or 0.0,
            "chrysotile_6_40": s.zo_chrysotile_6_40 or 0.0,
            "cement_silo1": s.zo_cement_silo1 or 0.0,
            "cement_silo2": s.zo_cement_silo2 or 0.0,
            "cement_silo3": s.zo_cement_silo3 or 0.0,
            "cement_silo4": s.zo_cement_silo4 or 0.0,
            "asbocarton": s.zo_asbocarton or 0.0,
            "laprol": s.zo_laprol or 0.0,
            "cellulose": s.zo_cellulose or 0.0,
            "fiberglass": s.zo_fiberglass or 0.0,
            "crushed_slate": s.zo_crushed_slate or 0.0,
            "asbozurit": s.zo_asbozurit or 0.0,
        }
        
        total_fact_asbestos = fact["chrysotile_4_20"] + fact["chrysotile_5_65"] + fact["chrysotile_6_40"]
        total_theo_asbestos = theory["chrysotile_4_20"] + theory["chrysotile_5_65"] + theory["chrysotile_6_40"]
        
        total_fact_cement = fact["cement_silo1"] + fact["cement_silo2"] + fact["cement_silo3"] + fact["cement_silo4"]
        theory_cement = theory["cement"]
        
        # Percentage deviation helper
        def get_pct_deviation(fact_val, theo_val):
            if theo_val <= 0:
                if fact_val == 0:
                    return 0.0
                return 100.0
            return ((fact_val - theo_val) / theo_val) * 100.0
            
        export_type = s.export_type or "Эталон"
        
        row_data = [
            date_str,
            batch_numbers,
            s.line or "",
            s.shift_name or "",
            s.master.name if s.master else "",
            product_names,
            export_type,
            s.zo_batches or 0,
            formovka_sheets,
            round(formovka_tons, 3),
            qcd_condition,
            qcd_first,
            qcd_defect,
            wind_resets,
            s.zo_asb_drain or 0.0,
            s.zo_cem_drain or 0.0,
            fact["chrysotile_4_20"],
            fact["chrysotile_5_65"],
            fact["chrysotile_6_40"],
            total_fact_asbestos,
            fact["cement_silo1"],
            fact["cement_silo2"],
            fact["cement_silo3"],
            fact["cement_silo4"],
            total_fact_cement,
            fact["asbocarton"],
            fact["laprol"],
            fact["cellulose"],
            fact["fiberglass"],
            fact["crushed_slate"],
            fact["asbozurit"],
            # Deviations (in %!)
            get_pct_deviation(fact["chrysotile_4_20"], theory["chrysotile_4_20"]),
            get_pct_deviation(fact["chrysotile_5_65"], theory["chrysotile_5_65"]),
            get_pct_deviation(fact["chrysotile_6_40"], theory["chrysotile_6_40"]),
            get_pct_deviation(total_fact_asbestos, total_theo_asbestos),
            get_pct_deviation(total_fact_cement, theory_cement),
            get_pct_deviation(fact["asbocarton"], theory["asbocarton"]),
            get_pct_deviation(fact["laprol"], theory["laprol"]),
            get_pct_deviation(fact["cellulose"], theory["cellulose"]),
            get_pct_deviation(fact["fiberglass"], theory["fiberglass"]),
            get_pct_deviation(fact["crushed_slate"], theory["crushed_slate"]),
            get_pct_deviation(fact["asbozurit"], theory["asbozurit"])
        ]
        
        ws.append(row_data)
        
        curr_row = ws.max_row
        
        # Style cells
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.border = border_thin
            
            # Alignments
            if col_idx in [1, 2, 3, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                
            # Defect cell (col 13)
            if col_idx == 13:
                val = cell.value
                if val == 0:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
                    
            # Deviation cols (cols 32 to 42)
            if col_idx in range(32, 43):
                val = cell.value
                if val is not None and isinstance(val, (int, float)):
                    # Format as percentage string with sign
                    sign = "+" if val > 0 else ""
                    cell.value = f"{sign}{val:.2f}%"
                    # Deviation color rules: > 0.1% red, otherwise green
                    if val > 0.1:
                        cell.fill = red_fill
                    else:
                        cell.fill = green_fill
                        
    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    # Convert range to Excel Table (Formatted Table with Auto-Filter)
    if ws.max_row > 1:
        from openpyxl.worksheet.table import Table, TableStyleInfo
        last_col_letter = get_column_letter(len(headers))
        ref_range = f"A1:{last_col_letter}{ws.max_row}"
        table = Table(displayName="TectumSummaryTable", ref=ref_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def create_initial_directories_xlsx(db: Session) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    ws_norm = wb.create_sheet(title="Нормативы")
    ws_norm.views.sheetView[0].showGridLines = True
    
    norm_headers = [
        "Продукция", "Вес готового листа (кг)", 
        "Норма Хризотил 4-20", "Норма Хризотил 5-65", "Норма Хризотил 6-40",
        "Норма Цемент", "Норма Целлюлоза", "Норма Дробленый шифер", 
        "Норма Асбозурит", "Норма Стекловолокно"
    ]
    ws_norm.append(norm_headers)
    
    norms = db.query(models.ProductNorm).all()
    for n in norms:
        ws_norm.append([
            n.product_name, n.weight_kg or 19.6,
            n.norm_chrysotile_4_20 or 0.0, n.norm_chrysotile_5_65 or 0.0, n.norm_chrysotile_6_40 or 0.0,
            n.norm_cement or 0.0, n.norm_cellulose or 0.0, n.norm_crushed_slate or 0.0,
            n.norm_asbozurit or 0.0, n.norm_fiberglass or 0.0
        ])
        
    ws_dir = wb.create_sheet(title="Справочник простоев")
    ws_dir.views.sheetView[0].showGridLines = True
    
    dir_headers = ["Участок", "Оборудование", "Неисправность/Причина", "Категория", "Комментарий"]
    ws_dir.append(dir_headers)
    
    entries = db.query(models.DowntimeDirectory).all()
    for e in entries:
        ws_dir.append([
            e.department, e.node, e.breakdown, e.category or "Механические", e.comment or ""
        ])
        
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for sheet in wb.worksheets:
        sheet.row_dimensions[1].height = 25
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def sync_directories_from_excel_bytes(file_bytes: bytes, db: Session):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    
    if "Нормативы" in wb.sheetnames:
        ws_norm = wb["Нормативы"]
        header_row = [cell.value for cell in ws_norm[1]]
        if header_row and "Продукция" in header_row:
            db.query(models.ProductNorm).delete()
            for row_idx in range(2, ws_norm.max_row + 1):
                p_name = ws_norm.cell(row=row_idx, column=1).value
                if not p_name: continue
                p_name = str(p_name).strip()
                if p_name == "Шифер 8 волн пиленый":
                    p_name = "Шифер 8 волн рифленый"
                
                weight = float(ws_norm.cell(row=row_idx, column=2).value or 0.0)
                n_c4 = float(ws_norm.cell(row=row_idx, column=3).value or 0.0)
                n_c5 = float(ws_norm.cell(row=row_idx, column=4).value or 0.0)
                n_c6 = float(ws_norm.cell(row=row_idx, column=5).value or 0.0)
                n_cem = float(ws_norm.cell(row=row_idx, column=6).value or 0.0)
                n_cel = float(ws_norm.cell(row=row_idx, column=7).value or 0.0)
                n_sl = float(ws_norm.cell(row=row_idx, column=8).value or 0.0)
                n_asb = float(ws_norm.cell(row=row_idx, column=9).value or 0.0)
                n_fib = float(ws_norm.cell(row=row_idx, column=10).value or 0.0)
                
                db.add(models.ProductNorm(
                    product_name=p_name,
                    weight_kg=weight,
                    norm_chrysotile_4_20=n_c4,
                    norm_chrysotile_5_65=n_c5,
                    norm_chrysotile_6_40=n_c6,
                    norm_cement=n_cem,
                    norm_cellulose=n_cel,
                    norm_crushed_slate=n_sl,
                    norm_asbozurit=n_asb,
                    norm_fiberglass=n_fib
                ))
            db.commit()
            print("Successfully synced Norms from Excel.")
            
    if "Справочник простоев" in wb.sheetnames:
        ws_dir = wb["Справочник простоев"]
        header_row = [cell.value for cell in ws_dir[1]]
        if header_row and "Участок" in header_row:
            db.query(models.DowntimeDirectory).delete()
            for row_idx in range(2, ws_dir.max_row + 1):
                dept = ws_dir.cell(row=row_idx, column=1).value
                node = ws_dir.cell(row=row_idx, column=2).value
                bd = ws_dir.cell(row=row_idx, column=3).value
                if not dept or not bd: continue
                
                dept = str(dept).strip()
                node = str(node).strip() if node else "Общее"
                bd = str(bd).strip()
                cat = str(ws_dir.cell(row=row_idx, column=4).value or "Механические").strip()
                comm = str(ws_dir.cell(row=row_idx, column=5).value or "").strip()
                
                db.add(models.DowntimeDirectory(
                    department=dept,
                    node=node,
                    breakdown=bd,
                    category=cat,
                    comment=comm
                ))
            db.commit()
            print("Successfully synced Downtimes Directory from Excel.")


def generate_full_backup_excel(db: Session) -> bytes:
    """
    Генерирует единую комплексную книгу Excel (.xlsx), содержащую 4 полных листа:
    1. «Сводный отчет» (43 колонки: производство, расходы сырья, силосы, отклонения)
    2. «Приход сырья» (18 колонок: дата, смена, линия, мастер, все виды сырья)
    3. «Простои» (9 колонок: дата, смена, линия, мастер, описание, время, остановка оборудования)
    4. «Переборка» (19 колонок: партии, 7 дефектов своей смены, 7 дефектов прошлой смены)
    В точности соответствует колонкам и структуре Google Таблиц.
    """
    from sqlalchemy import func
    from datetime import date, timedelta
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Удаляем дефолтный пустой лист
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # -------------------------------------------------------------
    # ЛИСТ 1: Сводный отчет
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Сводный отчет")
    ws1.views.sheetView[0].showGridLines = True
    
    headers1 = [
        "Дата", "№ партии", "Линия", "Смена", "Мастер", "Наименование продукта", "Назначение (Экспорт)",
        "Количество замесов", "Формовка (листы)", "Формовка (тонны)",
        "Кондиция (на склад)", "1-сорт", "Брак", "Сбросы наката",
        "Слив асб. (кг)", "Слив цем. (кг)",
        "Расход Хризотила 4-20 (кг)", "Расход Хризотила 5-65 (кг)", "Расход Хризотила 6-40 (кг)", "Расход Хризотила общ. (кг)",
        "Расход Цемента С1 (кг)", "Расход Цемента С2 (кг)", "Расход Цемента С3 (кг)", "Расход Цемента С4 (кг)", "Расход Цемента общ. (кг)",
        "Расход Асбокартона (кг)", "Расход Лапрола (кг)", "Расход Целлюлозы (кг)", "Расход Стекловолокна (кг)",
        "Расход Дробленого шифера (кг)", "Расход Асбозурита (кг)",
        "Отклонение Хризотила 4-20 (%)", "Отклонение Хризотила 5-65 (%)", "Отклонение Хризотила 6-40 (%)", "Отклонение Хризотила общ. (%)",
        "Отклонение Цемента общ. (%)", "Отклонение Асбокартона (%)", "Отклонение Лапрола (%)", "Отклонение Целлюлозы (%)",
        "Отклонение Стекловолокна (%)", "Отклонение Дробленого шифера (%)", "Отклонение Асбозурита (%)"
    ]
    ws1.append(headers1)
    ws1.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin

    # Включаем как закрытые, так и все существующие смены для полноценного бэкапа
    shifts = db.query(models.Shift).order_by(
        models.Shift.date.asc(),
        models.Shift.line.asc(),
        models.Shift.shift_name.asc(),
        models.Shift.batch_number.asc(),
        models.Shift.id.asc()
    ).all()
    
    for s in shifts:
        date_str = s.date.strftime("%d.%m.%Y") if s.date else ""
        batch_numbers = ", ".join(b.batch_number for b in s.batches if b.batch_number)
        if not batch_numbers and s.batch_number:
            batch_numbers = s.batch_number
            
        product_names = ", ".join(set(r.product_name for r in s.lfm_reports if r.product_name))
        if not product_names and s.product_name:
            product_names = s.product_name
            
        formovka_sheets = sum(r.lfm_sheets for r in s.lfm_reports)
        formovka_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) for r in s.lfm_reports) / 1000.0
        
        qcd_condition = sum(b.qcd_condition for b in s.batches)
        qcd_first = sum(b.qcd_first_grade for b in s.batches)
        qcd_defect = sum(b.qcd_defect for b in s.batches)
        wind_resets = sum(r.lfm_wind_resets for r in s.lfm_reports)
        
        theory = {
            "chrysotile_4_20": 0.0, "chrysotile_5_65": 0.0, "chrysotile_6_40": 0.0,
            "cement": 0.0, "cellulose": 0.0, "crushed_slate": 0.0,
            "asbozurit": 0.0, "fiberglass": 0.0, "asbocarton": 0.0, "laprol": 0.0
        }
        for r in s.lfm_reports:
            norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == r.product_name).first()
            if norm:
                theory["chrysotile_4_20"] += r.lfm_sheets * (norm.norm_chrysotile_4_20 or 0.0)
                theory["chrysotile_5_65"] += r.lfm_sheets * (norm.norm_chrysotile_5_65 or 0.0)
                theory["chrysotile_6_40"] += r.lfm_sheets * (norm.norm_chrysotile_6_40 or 0.0)
                theory["cement"] += r.lfm_sheets * (norm.norm_cement or 0.0)
                theory["cellulose"] += r.lfm_sheets * (norm.norm_cellulose or 0.0)
                theory["crushed_slate"] += r.lfm_sheets * (norm.norm_crushed_slate or 0.0)
                theory["asbozurit"] += r.lfm_sheets * (norm.norm_asbozurit or 0.0)
                theory["fiberglass"] += r.lfm_sheets * (norm.norm_fiberglass or 0.0)
                
        fact = {
            "chrysotile_4_20": s.zo_chrysotile_4_20 or 0.0,
            "chrysotile_5_65": s.zo_chrysotile_5_65 or 0.0,
            "chrysotile_6_40": s.zo_chrysotile_6_40 or 0.0,
            "cement_silo1": s.zo_cement_silo1 or 0.0,
            "cement_silo2": s.zo_cement_silo2 or 0.0,
            "cement_silo3": s.zo_cement_silo3 or 0.0,
            "cement_silo4": s.zo_cement_silo4 or 0.0,
            "asbocarton": s.zo_asbocarton or 0.0,
            "laprol": s.zo_laprol or 0.0,
            "cellulose": s.zo_cellulose or 0.0,
            "fiberglass": s.zo_fiberglass or 0.0,
            "crushed_slate": s.zo_crushed_slate or 0.0,
            "asbozurit": s.zo_asbozurit or 0.0,
        }
        
        total_fact_asbestos = fact["chrysotile_4_20"] + fact["chrysotile_5_65"] + fact["chrysotile_6_40"]
        total_theo_asbestos = theory["chrysotile_4_20"] + theory["chrysotile_5_65"] + theory["chrysotile_6_40"]
        total_fact_cement = fact["cement_silo1"] + fact["cement_silo2"] + fact["cement_silo3"] + fact["cement_silo4"]
        theory_cement = theory["cement"]
        export_type = s.export_type or "Эталон"
        
        def _get_pct(act, theo):
            if not theo or theo <= 0: return 0.0
            return round(((act - theo) / theo) * 100.0, 2)
            
        row_data = [
            date_str,
            batch_numbers,
            s.line or "",
            s.shift_name or "",
            s.master.name if s.master else "",
            product_names,
            export_type,
            s.zo_batches or 0,
            formovka_sheets,
            round(formovka_tons, 3),
            qcd_condition,
            qcd_first,
            qcd_defect,
            wind_resets,
            s.zo_asb_drain or 0.0,
            s.zo_cem_drain or 0.0,
            fact["chrysotile_4_20"],
            fact["chrysotile_5_65"],
            fact["chrysotile_6_40"],
            total_fact_asbestos,
            fact["cement_silo1"],
            fact["cement_silo2"],
            fact["cement_silo3"],
            fact["cement_silo4"],
            total_fact_cement,
            fact["asbocarton"],
            fact["laprol"],
            fact["cellulose"],
            fact["fiberglass"],
            fact["crushed_slate"],
            fact["asbozurit"],
            _get_pct(fact["chrysotile_4_20"], theory["chrysotile_4_20"]),
            _get_pct(fact["chrysotile_5_65"], theory["chrysotile_5_65"]),
            _get_pct(fact["chrysotile_6_40"], theory["chrysotile_6_40"]),
            _get_pct(total_fact_asbestos, total_theo_asbestos),
            _get_pct(total_fact_cement, theory_cement),
            _get_pct(fact["asbocarton"], theory["asbocarton"]),
            _get_pct(fact["laprol"], theory["laprol"]),
            _get_pct(fact["cellulose"], theory["cellulose"]),
            _get_pct(fact["fiberglass"], theory["fiberglass"]),
            _get_pct(fact["crushed_slate"], theory["crushed_slate"]),
            _get_pct(fact["asbozurit"], theory["asbozurit"])
        ]
        ws1.append(row_data)
        curr_row = ws1.max_row
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=curr_row, column=col_idx)
            cell.border = border_thin
            if col_idx in range(1, 8):
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                
    for col in ws1.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    # -------------------------------------------------------------
    # ЛИСТ 2: Приход сырья
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Приход сырья")
    ws2.views.sheetView[0].showGridLines = True
    headers2 = [
        "Дата", "Смена", "Линия", "Мастер",
        "Хризотил 4-20 (кг)", "Хризотил 5-65 (кг)", "Хризотил 6-40 (кг)",
        "Цемент С1 (кг)", "Цемент С2 (кг)", "Цемент С3 (кг)", "Цемент С4 (кг)", 
        "Целлюлоза (кг)", "Дробленый шифер (кг)",
        "Асбозурит (кг)", "Асбокартон (кг)", "Паллеты (шт)",
        "Стекловолокно (кг)", "Лапрол (кг)"
    ]
    ws2.append(headers2)
    ws2.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin
        
    receipts = db.query(models.RawMaterialReceipt).outerjoin(models.Shift).order_by(
        func.coalesce(models.RawMaterialReceipt.date, models.Shift.date).asc(),
        models.RawMaterialReceipt.id.asc()
    ).all()
    
    for r in receipts:
        r_date = r.date or (r.shift.date if r.shift else None)
        date_str = r_date.strftime("%d.%m.%Y") if hasattr(r_date, 'strftime') else (str(r_date) if r_date else "")
        shift_name = r.shift_name or (r.shift.shift_name if r.shift else "")
        line = r.line or (r.shift.line if r.shift else "")
        master_name = r.master.name if r.master else (r.shift.master.name if (r.shift and r.shift.master) else "")
        
        row = [
            date_str,
            shift_name,
            line,
            master_name,
            r.chrysotile_4_20 or 0.0,
            r.chrysotile_5_65 or 0.0,
            r.chrysotile_6_40 or 0.0,
            r.cement_silo1 or 0.0,
            r.cement_silo2 or 0.0,
            r.cement_silo3 or 0.0,
            r.cement_silo4 or 0.0,
            r.cellulose or 0.0,
            r.crushed_slate or 0.0,
            r.asbozurit or 0.0,
            r.asbocarton or 0.0,
            r.pallets or 0.0,
            r.fiberglass or 0.0,
            r.laprol or 0.0
        ]
        ws2.append(row)
        curr_row = ws2.max_row
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=curr_row, column=col_idx)
            cell.border = border_thin
            if col_idx in [1, 2, 3, 4]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                
    for col in ws2.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # -------------------------------------------------------------
    # ЛИСТ 3: Простои
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Простои")
    ws3.views.sheetView[0].showGridLines = True
    ws3.freeze_panes = "H2"  # Закрепление 1 строки и столбцов A-G (от даты до длительности включительно)
    headers3 = [
        "Дата", "Смена", "Линия", "Мастер",
        "Начало", "Конец", "Длительность (мин)",
        "Описание простоя",
        "Остановка оборудования"
    ]
    ws3.append(headers3)
    ws3.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin
        
    downtimes = db.query(models.Downtime).outerjoin(models.Shift).order_by(
        func.coalesce(models.Downtime.date, models.Shift.date).asc(),
        models.Downtime.id.asc()
    ).all()
    
    for d in downtimes:
        d_date = d.date or (d.shift.date if d.shift else None)
        date_str = d_date.strftime("%d.%m.%Y") if hasattr(d_date, 'strftime') else (str(d_date) if d_date else "")
        shift_name_val = d.shift_name or (d.shift.shift_name if d.shift else "")
        line_val = d.line or (d.shift.line if d.shift else "")
        master_val = d.master.name if d.master else (d.shift.master.name if (d.shift and d.shift.master) else "")
        desc_text = (d.description or d.comment or "").strip()
        
        row = [
            date_str,
            shift_name_val,
            line_val,
            master_val,
            d.start_time or "",
            d.end_time or "",
            d.duration or 0,
            desc_text,
            "Да" if d.is_equipment_downtime else "Нет"
        ]
        ws3.append(row)
        curr_row = ws3.max_row
        for col_idx in range(1, len(headers3) + 1):
            cell = ws3.cell(row=curr_row, column=col_idx)
            cell.border = border_thin
            if col_idx in [1, 2, 3, 4, 8]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")
                
    for col in ws3.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # -------------------------------------------------------------
    # ЛИСТ 4: Переборка
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Переборка")
    ws4.views.sheetView[0].showGridLines = True
    ws4.freeze_panes = "G2"  # Закрепление 1 строки и столбцов A:F (№ партии ... Продукт)
    headers4 = [
        "№ партии", "Дата", "День нед.", "Время смены", "Мастер ЛФМ", "Продукт", "Формовка, шт",
        "Смена", "1 сорт, шт", "Брак, шт", "% брака", "Детализация брака",
        "Прошлая смена", "Мастер (прошлая)", "1 сорт (прошлая), шт", "Брак (прошлая), шт", "Детализация брака (прошлая)",
        "Всего 1 сорт, шт", "Всего брак, шт"
    ]
    ws4.append(headers4)
    ws4.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers4) + 1):
        cell = ws4.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin
        
    all_schedules = {s.date_str: s for s in db.query(models.ShiftScheduleEntry).all()}
    shifts_lookup = {(s.date.strftime("%Y-%m-%d") if s.date else "", s.shift_name, s.line): s for s in db.query(models.Shift).all()}
    
    from sqlalchemy import case, cast, Integer
    batches = db.query(models.Batch).join(models.Shift).order_by(
        models.Shift.date.asc(),
        case(
            (models.Shift.shift_name.ilike('%день%'), 1),
            (models.Shift.shift_name.ilike('%ночь%'), 2),
            else_=3
        ).asc(),
        cast(models.Batch.batch_number, Integer).asc(),
        models.Batch.id.asc()
    ).all()
    
    days_map = {0: "Пн", 1: "Вт", 2: "Ср", 3: "Чт", 4: "Пт", 5: "Сб", 6: "Вс"}
    for b in batches:
        shift_date = b.shift.date if b.shift else None
        date_str_fmt = shift_date.strftime("%d.%m.%Y") if shift_date else ""
        day_of_week = days_map.get(shift_date.weekday(), "") if shift_date else ""
        shift_name = b.shift.shift_name if b.shift else ""
        shift_line = b.shift.line if b.shift else ""
        master_name = b.shift.master.name if (b.shift and b.shift.master) else ""
        product_name = b.product_name or (b.shift.product_name if b.shift else "")
        
        sched = all_schedules.get(date_str_fmt)
        shift_group = ""
        prev_shift_group = ""
        prev_master_name = ""
        
        if shift_name == "День":
            shift_group = sched.day_shift_group if sched else ""
            if shift_date:
                prev_date = shift_date - timedelta(days=1)
                prev_date_fmt = prev_date.strftime("%d.%m.%Y")
                prev_date_iso = prev_date.strftime("%Y-%m-%d")
                prev_sched = all_schedules.get(prev_date_fmt)
                prev_shift_group = prev_sched.night_shift_group if prev_sched else ""
                prev_shift_obj = shifts_lookup.get((prev_date_iso, "Ночь", shift_line))
                if prev_shift_obj and prev_shift_obj.master:
                    prev_master_name = prev_shift_obj.master.name
        else:
            shift_group = sched.night_shift_group if sched else ""
            if shift_date:
                curr_date_iso = shift_date.strftime("%Y-%m-%d")
                prev_shift_obj = shifts_lookup.get((curr_date_iso, "День", shift_line))
                if prev_shift_obj and prev_shift_obj.master:
                    prev_master_name = prev_shift_obj.master.name
                    
        ds_first = b.ds_first_grade or 0
        ds_def = b.ds_defect or 0
        prev_f = b.prev_first_grade or 0
        prev_d = b.prev_defect or 0
        
        lfm_sheets = 0
        if b.shift and b.shift.lfm_reports:
            lfm_sheets = sum(r.lfm_sheets for r in b.shift.lfm_reports)
        elif b.stacked_stacks:
            lfm_sheets = b.stacked_stacks
            
        pct_defect = round((ds_def / lfm_sheets) * 100.0, 2) if lfm_sheets > 0 else 0.0
        
        def_parts = []
        if b.ds_defect_scratch: def_parts.append(f"Сдир ({b.ds_defect_scratch})")
        if b.ds_defect_bad_cut: def_parts.append(f"Плохой рез ({b.ds_defect_bad_cut})")
        if b.ds_defect_stick_top: def_parts.append(f"Налип сверху ({b.ds_defect_stick_top})")
        if b.ds_defect_broken: def_parts.append(f"Сломан ({b.ds_defect_broken})")
        if b.ds_defect_fell_box: def_parts.append(f"Упал коробки ({b.ds_defect_fell_box})")
        if b.ds_defect_thickness: def_parts.append(f"Не соотв. толщины ({b.ds_defect_thickness})")
        if b.ds_defect_edge: def_parts.append(f"Кромка ({b.ds_defect_edge})")
        note_defect = ", ".join(def_parts)

        prev_parts = []
        if b.prev_defect_scratch: prev_parts.append(f"Сдир ({b.prev_defect_scratch})")
        if b.prev_defect_bad_cut: prev_parts.append(f"Плохой рез ({b.prev_defect_bad_cut})")
        if b.prev_defect_stick_top: prev_parts.append(f"Налип сверху ({b.prev_defect_stick_top})")
        if b.prev_defect_broken: prev_parts.append(f"Сломан ({b.prev_defect_broken})")
        if b.prev_defect_fell_box: prev_parts.append(f"Упал коробки ({b.prev_defect_fell_box})")
        if b.prev_defect_thickness: prev_parts.append(f"Не соотв. толщины ({b.prev_defect_thickness})")
        if b.prev_defect_edge: prev_parts.append(f"Кромка ({b.prev_defect_edge})")
        prev_note = ", ".join(prev_parts)
        
        row = [
            b.batch_number or "",
            date_str_fmt,
            day_of_week,
            shift_name,
            master_name,
            product_name,
            lfm_sheets,
            shift_group,
            ds_first,
            ds_def,
            f"{pct_defect}%",
            note_defect,
            prev_shift_group,
            prev_master_name,
            prev_f,
            prev_d,
            prev_note,
            ds_first + prev_f,
            ds_def + prev_d
        ]
        ws4.append(row)
        curr_row = ws4.max_row
        for col_idx in range(1, len(headers4) + 1):
            cell = ws4.cell(row=curr_row, column=col_idx)
            cell.border = border_thin
            if col_idx in [1, 2, 3, 4, 5, 6, 8, 12, 13, 14, 17]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                
    for col in ws4.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws4.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def restore_from_backup_excel(file_bytes: bytes, db: Session, user_name: str = "Администратор") -> dict:
    """
    Восстанавливает данные из полного файла бэкапа Excel.
    Обновляет или добавляет:
    - Приходы сырья (из листа «Приход сырья»)
    - Простои оборудования (из листа «Простои»)
    - Расход и выработку смен (из листа «Сводный отчет»)
    """
    import openpyxl
    from datetime import datetime
    
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {
        "status": "success",
        "updated_receipts": 0,
        "created_receipts": 0,
        "updated_downtimes": 0,
        "created_downtimes": 0,
        "updated_shifts": 0
    }
    
    # 1. Приход сырья
    if "Приход сырья" in wb.sheetnames:
        ws = wb["Приход сырья"]
        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            if not date_val: continue
            
            p_date = None
            if isinstance(date_val, datetime):
                p_date = date_val.date()
            elif isinstance(date_val, str):
                try:
                    p_date = datetime.strptime(date_val.strip(), "%d.%m.%Y").date()
                except Exception:
                    try:
                        p_date = datetime.strptime(date_val.strip(), "%Y-%m-%d").date()
                    except Exception:
                        continue
            if not p_date: continue
            
            s_name = str(ws.cell(row=row_idx, column=2).value or "").strip()
            line = str(ws.cell(row=row_idx, column=3).value or "").strip()
            m_name = str(ws.cell(row=row_idx, column=4).value or "").strip()
            
            master_id = None
            if m_name:
                m_obj = db.query(models.Master).filter(models.Master.name == m_name).first()
                if m_obj: master_id = m_obj.id
                
            c4 = float(ws.cell(row=row_idx, column=5).value or 0.0)
            c5 = float(ws.cell(row=row_idx, column=6).value or 0.0)
            c6 = float(ws.cell(row=row_idx, column=7).value or 0.0)
            cs1 = float(ws.cell(row=row_idx, column=8).value or 0.0)
            cs2 = float(ws.cell(row=row_idx, column=9).value or 0.0)
            cs3 = float(ws.cell(row=row_idx, column=10).value or 0.0)
            cs4 = float(ws.cell(row=row_idx, column=11).value or 0.0)
            cell = float(ws.cell(row=row_idx, column=12).value or 0.0)
            sl = float(ws.cell(row=row_idx, column=13).value or 0.0)
            asb = float(ws.cell(row=row_idx, column=14).value or 0.0)
            asbc = float(ws.cell(row=row_idx, column=15).value or 0.0)
            pal = float(ws.cell(row=row_idx, column=16).value or 0.0)
            fib = float(ws.cell(row=row_idx, column=17).value or 0.0)
            lap = float(ws.cell(row=row_idx, column=18).value or 0.0)
            
            # Ищем существующую запись
            existing = db.query(models.RawMaterialReceipt).filter(
                models.RawMaterialReceipt.date == p_date,
                models.RawMaterialReceipt.shift_name == s_name,
                models.RawMaterialReceipt.line == line
            ).first()
            
            if existing:
                existing.chrysotile_4_20 = c4
                existing.chrysotile_5_65 = c5
                existing.chrysotile_6_40 = c6
                existing.cement_silo1 = cs1
                existing.cement_silo2 = cs2
                existing.cement_silo3 = cs3
                existing.cement_silo4 = cs4
                existing.cellulose = cell
                existing.crushed_slate = sl
                existing.asbozurit = asb
                existing.asbocarton = asbc
                existing.pallets = pal
                existing.fiberglass = fib
                existing.laprol = lap
                if master_id: existing.master_id = master_id
                result["updated_receipts"] += 1
            else:
                new_r = models.RawMaterialReceipt(
                    date=p_date,
                    shift_name=s_name,
                    line=line,
                    master_id=master_id,
                    chrysotile_4_20=c4,
                    chrysotile_5_65=c5,
                    chrysotile_6_40=c6,
                    cement_silo1=cs1,
                    cement_silo2=cs2,
                    cement_silo3=cs3,
                    cement_silo4=cs4,
                    cellulose=cell,
                    crushed_slate=sl,
                    asbozurit=asb,
                    asbocarton=asbc,
                    pallets=pal,
                    fiberglass=fib,
                    laprol=lap
                )
                db.add(new_r)
                result["created_receipts"] += 1
                
        db.commit()

    # 2. Простои
    if "Простои" in wb.sheetnames:
        ws = wb["Простои"]
        # Определяем порядок колонок по шапке (поддержка старого и нового формата)
        header_row = [str(ws.cell(row=1, column=c).value or "").strip().lower() for c in range(1, 10)]
        is_new_layout = ("начало" in header_row[4] if len(header_row) > 4 else False) or ("время начала" in header_row[4] if len(header_row) > 4 else False)

        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            if not date_val: continue
            
            p_date = None
            if isinstance(date_val, datetime):
                p_date = date_val.date()
            elif isinstance(date_val, str):
                try:
                    p_date = datetime.strptime(date_val.strip(), "%d.%m.%Y").date()
                except Exception:
                    continue
            if not p_date: continue
            
            s_name = str(ws.cell(row=row_idx, column=2).value or "").strip()
            line = str(ws.cell(row=row_idx, column=3).value or "").strip()
            m_name = str(ws.cell(row=row_idx, column=4).value or "").strip()
            
            if is_new_layout:
                # Новый порядок: Дата, Смена, Линия, Мастер, Начало, Конец, Длительность, Описание, Остановка оборудования
                start_t = str(ws.cell(row=row_idx, column=5).value or "").strip()
                end_t = str(ws.cell(row=row_idx, column=6).value or "").strip()
                dur = int(float(ws.cell(row=row_idx, column=7).value or 0))
                desc = str(ws.cell(row=row_idx, column=8).value or "").strip()
                is_eq = (str(ws.cell(row=row_idx, column=9).value or "").strip().lower() == "да")
            else:
                # Прежний порядок: Дата, Смена, Линия, Мастер, Описание, Начало, Конец, Длительность, Остановка оборудования
                desc = str(ws.cell(row=row_idx, column=5).value or "").strip()
                start_t = str(ws.cell(row=row_idx, column=6).value or "").strip()
                end_t = str(ws.cell(row=row_idx, column=7).value or "").strip()
                dur = int(float(ws.cell(row=row_idx, column=8).value or 0))
                is_eq = (str(ws.cell(row=row_idx, column=9).value or "").strip().lower() == "да")
            
            master_id = None
            if m_name:
                m_obj = db.query(models.Master).filter(models.Master.name == m_name).first()
                if m_obj: master_id = m_obj.id
                
            existing_dt = db.query(models.Downtime).filter(
                models.Downtime.date == p_date,
                models.Downtime.shift_name == s_name,
                models.Downtime.line == line,
                models.Downtime.start_time == start_t
            ).first()
            
            if existing_dt:
                existing_dt.description = desc
                existing_dt.end_time = end_t
                existing_dt.duration = dur
                existing_dt.is_equipment_downtime = is_eq
                if master_id: existing_dt.master_id = master_id
                result["updated_downtimes"] += 1
            else:
                new_dt = models.Downtime(
                    date=p_date,
                    shift_name=s_name,
                    line=line,
                    master_id=master_id,
                    description=desc,
                    start_time=start_t,
                    end_time=end_t,
                    duration=dur,
                    is_equipment_downtime=is_eq
                )
                db.add(new_dt)
                result["created_downtimes"] += 1
                
        db.commit()

    db.add(models.AuditLog(
        user_name=user_name,
        action="RESTORE",
        target_table="shifts/receipts/downtimes",
        target_id=0,
        details=f"Восстановление из Excel бэкапа: приходов +{result['created_receipts']}/~{result['updated_receipts']}, простоев +{result['created_downtimes']}/~{result['updated_downtimes']}"
    ))
    db.commit()
    return result

