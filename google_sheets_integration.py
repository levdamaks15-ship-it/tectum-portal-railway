import os
import json
from datetime import datetime, timedelta, date
from collections import defaultdict
from google.oauth2 import service_account
from googleapiclient.discovery import build
from sqlalchemy.orm import Session
import models
from dotenv import load_dotenv

load_dotenv()

SPREADSHEET_ID = os.getenv("GOOGLE_SPREADSHEET_ID")
CREDENTIALS_PATH = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "google_credentials.json")

def get_sheets_service():
    # 1. Сначала пробуем загрузить из переменной окружения (для Railway/Render)
    creds_json = os.getenv("GOOGLE_CREDENTIALS_JSON")
    if creds_json:
        try:
            info = json.loads(creds_json)
            if "private_key" in info:
                info["private_key"] = info["private_key"].replace("\\n", "\n")
            creds = service_account.Credentials.from_service_account_info(
                info,
                scopes=["https://www.googleapis.com/auth/spreadsheets"]
            )
            return build("sheets", "v4", credentials=creds)
        except Exception as env_err:
            print(f"Ошибка парсинга GOOGLE_CREDENTIALS_JSON из переменных окружения: {env_err}")

    # 2. Если переменной нет, считываем локальный файл
    if not os.path.exists(CREDENTIALS_PATH):
        raise FileNotFoundError(f"Файл ключа Google не найден по пути: {CREDENTIALS_PATH}")
    
    with open(CREDENTIALS_PATH, "r", encoding="utf-8") as f:
        info = json.load(f)
    
    if "private_key" in info:
        info["private_key"] = info["private_key"].replace("\\n", "\n")
        
    creds = service_account.Credentials.from_service_account_info(
        info,
        scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    return build("sheets", "v4", credentials=creds)

def get_product_finished_weight_kg(db: Session, product_name: str) -> float:
    norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == product_name).first()
    if not norm or not norm.weight_kg:
        return 19.6
    return norm.weight_kg

def get_pct_deviation(fact_val: float, theo_val: float) -> float:
    if theo_val <= 0:
        if fact_val == 0:
            return 0.0
        return 100.0
    return ((fact_val - theo_val) / theo_val) * 100.0

# Единый стандарт оформления таблиц (как на листе «Сводный отчет»)
PASTEL_GREEN_HEADER = {
    "backgroundColor": {"red": 217 / 255.0, "green": 234 / 255.0, "blue": 211 / 255.0},  # #d9ead3 пастельно-зеленый
    "textFormat": {
        "bold": True,
        "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 0.0},
        "fontFamily": "Arial",
        "fontSize": 10
    },
    "horizontalAlignment": "CENTER",
    "verticalAlignment": "MIDDLE",
    "wrapStrategy": "CLIP"
}

STANDARD_BORDER_STYLE = {
    "top": {"style": "SOLID", "width": 1, "color": {"red": 0.8, "green": 0.8, "blue": 0.8}},
    "bottom": {"style": "SOLID", "width": 1, "color": {"red": 0.8, "green": 0.8, "blue": 0.8}},
    "left": {"style": "SOLID", "width": 1, "color": {"red": 0.8, "green": 0.8, "blue": 0.8}},
    "right": {"style": "SOLID", "width": 1, "color": {"red": 0.8, "green": 0.8, "blue": 0.8}},
    "innerHorizontal": {"style": "SOLID", "width": 1, "color": {"red": 0.85, "green": 0.85, "blue": 0.85}},
    "innerVertical": {"style": "SOLID", "width": 1, "color": {"red": 0.85, "green": 0.85, "blue": 0.85}}
}

def sync_report_to_google_sheets(db: Session):
    """
    Генерирует сводную таблицу рапортов смен аналогично Excel-отчету
    и выгружает ее в Google Таблицу по SPREADSHEET_ID с точным воссозданием форматирования.
    """
    if not SPREADSHEET_ID or SPREADSHEET_ID.startswith("1_mock"):
        print("Синхронизация с Google Таблицами пропущена: не задан реальный GOOGLE_SPREADSHEET_ID в .env")
        return
    
    service = get_sheets_service()
    
    # 1. Извлекаем данные из БД (все смены без фильтра по статусу "closed")
    shifts = db.query(models.Shift).order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.Shift.id.asc()).all()
    
    # Записываем отладочный лог в AuditLog
    db.add(models.AuditLog(
        user_name="System Debug Sheets",
        action="INFO",
        target_table="shifts",
        target_id=0,
        details=f"Синхронизация Google Sheets: найдено {len(shifts)} смен в базе данных."
    ))
    db.commit()
    
    headers = [
        "Дата", "№ партии", "Линия", "Смена", "Мастер", "Наименование продукта", "Назначение (Экспорт)",
        "Количество замесов", "Формовка (листы)", "Формовка (тонны)",
        "Кондиция (на склад)", "1-сорт", "Брак", "Сбросы наката",
        "Слив асб. (кг)", "Слив цем. (кг)",
        "Расход Хризотила 4-20 (кг)", "Расход Хризотила 5-65 (кг)", "Расход Хризотила 6-40 (кг)", "Расход Хризотила общ. (кг)",
        "Расход Цемента С1 (кг)", "Расход Цемента С2 (кг)", "Расход Цемента С3 (кг)", "Расход Цемента С4 (кг)", "Расход Цемента общ. (кг)",
        "Расход Асбокартона (кг)", "Расход Лапрола (кг)", "Расход Целлюлозы (кг)", "Расход Стекловолокна (кг)",
        "Расход Дробленого шифера (кг)", "Расход Асбозурита (кг)",
        "Отклонение Хризотила 4-20 (%)", "Отклонение Хризотила 5-65 (%)", "Отклонение Хризотила 6-40 (%)", "Отклонение Хризотила общ. (%)",
        "Отклонение Цемента общ. (%)", "Отклонение Асбокартона (%)", "Отклонение Лапрола (%)", "Отклонение Целлюлозы (%)",
        "Отклонение Стекловолокна (%)", "Отклонение Дробленого шифера (%)", "Отклонение Асбозурита (%)"
    ]
    
    # Создаем массив строк для Google Sheets
    rows_data = []
    
    # Записываем шапку
    rows_data.append(headers)
    
    for s in shifts:
        # Проверяем, есть ли плановые или фактические показатели производства в смене
        plan_sheets_check = s.plan_sheets or 0
        formovka_sheets_check = sum(r.lfm_sheets for r in s.lfm_reports)
        warehouse_gp_check = sum(b.qcd_condition for b in s.batches)
        zo_batches_check = s.zo_batches or 0
        
        if plan_sheets_check == 0 and formovka_sheets_check == 0 and warehouse_gp_check == 0 and zo_batches_check == 0 and not s.zo_submitted:
            continue
            
        date_str = s.date.strftime("%d.%m.%Y") if s.date else ""
        batch_numbers = ", ".join(b.batch_number for b in s.batches if b.batch_number)
        product_names = ", ".join(set(r.product_name for r in s.lfm_reports if r.product_name))
        formovka_sheets = formovka_sheets_check
        formovka_tons = sum(r.lfm_sheets * get_product_finished_weight_kg(db, r.product_name) for r in s.lfm_reports) / 1000.0
        
        qcd_condition = warehouse_gp_check
        qcd_first = sum(b.qcd_first_grade for b in s.batches)
        qcd_defect = sum(b.qcd_defect for b in s.batches)
        wind_resets = sum(r.lfm_wind_resets for r in s.lfm_reports)
        
        theory = {
            "chrysotile_4_20": 0.0, "chrysotile_5_65": 0.0, "chrysotile_6_40": 0.0,
            "cement": 0.0, "cellulose": 0.0, "crushed_slate": 0.0,
            "asbozurit": 0.0, "fiberglass": 0.0
        }
        for r in s.lfm_reports:
            norm = db.query(models.ProductNorm).filter(models.ProductNorm.product_name == r.product_name).first()
            if norm:
                theory["chrysotile_4_20"] += r.lfm_sheets * (norm.norm_chrysotile_4_20 or 0.0)
                theory["chrysotile_5_65"] += r.lfm_sheets * (norm.norm_chrysotile_5_65 or 0.0)
                theory["chrysotile_6_40"] += r.lfm_sheets * (norm.norm_chrysotile_6_40 or 0.0)
                theory["cement"] += r.lfm_sheets * (norm.norm_cement or 0.0)
                theory["cellulose"] += r.lfm_sheets * (norm.norm_cellulose or 0.0)
                theory["crushed_slate"] += r.lfm_sheets * (norm.norm_crushed_slate or 0.0)
                theory["asbozurit"] += r.lfm_sheets * (norm.norm_asbozurit or 0.0)
                theory["fiberglass"] += r.lfm_sheets * (norm.norm_fiberglass or 0.0)
                
        fact = {
            "chrysotile_4_20": s.zo_chrysotile_4_20 or 0.0,
            "chrysotile_5_65": s.zo_chrysotile_5_65 or 0.0,
            "chrysotile_6_40": s.zo_chrysotile_6_40 or 0.0,
            "cement_silo1": s.zo_cement_silo1 or 0.0,
            "cement_silo2": s.zo_cement_silo2 or 0.0,
            "cement_silo3": s.zo_cement_silo3 or 0.0,
            "cement_silo4": s.zo_cement_silo4 or 0.0,
            "asbocarton": s.zo_asbocarton or 0.0,
            "laprol": s.zo_laprol or 0.0,
            "cellulose": s.zo_cellulose or 0.0,
            "fiberglass": s.zo_fiberglass or 0.0,
            "crushed_slate": s.zo_crushed_slate or 0.0,
            "asbozurit": s.zo_asbozurit or 0.0,
        }
        
        total_fact_asbestos = fact["chrysotile_4_20"] + fact["chrysotile_5_65"] + fact["chrysotile_6_40"]
        total_theo_asbestos = theory["chrysotile_4_20"] + theory["chrysotile_5_65"] + theory["chrysotile_6_40"]
        total_fact_cement = fact["cement_silo1"] + fact["cement_silo2"] + fact["cement_silo3"] + fact["cement_silo4"]
        theory_cement = theory["cement"]
        export_type = s.export_type or "Эталон"
        
        row_data = [
            date_str,
            batch_numbers,
            s.line or "",
            s.shift_name or "",
            s.master.name if s.master else "",
            product_names,
            export_type,
            s.zo_batches or 0,
            formovka_sheets,
            round(formovka_tons, 3),
            qcd_condition,
            qcd_first,
            qcd_defect,
            wind_resets,
            s.zo_asb_drain or 0.0,
            s.zo_cem_drain or 0.0,
            fact["chrysotile_4_20"],
            fact["chrysotile_5_65"],
            fact["chrysotile_6_40"],
            total_fact_asbestos,
            fact["cement_silo1"],
            fact["cement_silo2"],
            fact["cement_silo3"],
            fact["cement_silo4"],
            total_fact_cement,
            fact["asbocarton"],
            fact["laprol"],
            fact["cellulose"],
            fact["fiberglass"],
            fact["crushed_slate"],
            fact["asbozurit"],
            # Deviations (in %!)
            get_pct_deviation(fact["chrysotile_4_20"], theory["chrysotile_4_20"]) / 100.0,
            get_pct_deviation(fact["chrysotile_5_65"], theory["chrysotile_5_65"]) / 100.0,
            get_pct_deviation(fact["chrysotile_6_40"], theory["chrysotile_6_40"]) / 100.0,
            get_pct_deviation(total_fact_asbestos, total_theo_asbestos) / 100.0,
            get_pct_deviation(total_fact_cement, theory_cement) / 100.0,
            get_pct_deviation(fact["asbocarton"], 0.0) / 100.0, # no theory for carton/laprol
            get_pct_deviation(fact["laprol"], 0.0) / 100.0,
            get_pct_deviation(fact["cellulose"], theory["cellulose"]) / 100.0,
            get_pct_deviation(fact["fiberglass"], theory["fiberglass"]) / 100.0,
            get_pct_deviation(fact["crushed_slate"], theory["crushed_slate"]) / 100.0,
            get_pct_deviation(fact["asbozurit"], theory["asbozurit"]) / 100.0
        ]
        rows_data.append(row_data)

    # 2. Выгружаем данные на лист "Сводный отчет"
    sheet_name = "Сводный отчет"
    
    # Проверим, существует ли лист, если нет - создадим
    spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheets_titles = [sh["properties"]["title"] for sh in spreadsheet["sheets"]]
    
    if sheet_name not in sheets_titles:
        body = {
            "requests": [{
                "addSheet": {
                    "properties": {
                        "title": sheet_name,
                        "gridProperties": {
                            "rowCount": 1000,
                            "columnCount": 45
                        }
                    }
                }
            }]
        }
        service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
        # Обновим информацию о листах
        spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        
    sheet_id = next(sh["properties"]["sheetId"] for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    
    # 1. Получаем текущие данные на листе, чтобы узнать, какие строки уже есть
    result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1:AO1000"
    ).execute()
    existing_rows = result.get("values", [])
    
    # Полная очистка диапазона перед записью новых данных, чтобы избежать наложения и сдвига колонок
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1:AZ2000"
    ).execute()
    
    # Очищаем старые правила условного форматирования для этого листа
    sheet_meta = next(sh for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    existing_rules = sheet_meta.get("conditionalFormats", [])
    if existing_rules:
        clear_requests = []
        for idx in range(len(existing_rules) - 1, -1, -1):
            clear_requests.append({
                "deleteConditionalFormatRule": {
                    "sheetId": sheet_id,
                    "index": idx
                }
            })
        service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body={"requests": clear_requests}).execute()
        
    # Записываем шапку и все строки данных разом
    service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1",
        valueInputOption="USER_ENTERED",
        body={"values": rows_data}
    ).execute()
    
    # 3. Обновляем форматирование, автофильтр и закрепление шапки
    sheet_meta = next(sh for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    has_basic_filter = "basicFilter" in sheet_meta

    total_rows = len(rows_data)
    requests = []

    # Закрепление верхней строки заголовка
    requests.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "gridProperties": {
                    "frozenRowCount": 1
                }
            },
            "fields": "gridProperties.frozenRowCount"
        }
    })

    # Фиксированная компактная высота шапки (28px)
    requests.append({
        "updateDimensionProperties": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": 0,
                "endIndex": 1
            },
            "properties": {
                "pixelSize": 28
            },
            "fields": "pixelSize"
        }
    })

    # Стилизация шапки (пастельно-зеленый фон, жирный текст)
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 0,
                "endRowIndex": 1,
                "startColumnIndex": 0,
                "endColumnIndex": len(headers)
            },
            "cell": {
                "userEnteredFormat": PASTEL_GREEN_HEADER
            },
            "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy)"
        }
    })

    # Границы для всей таблицы
    requests.append({
        "updateBorders": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 0,
                "endRowIndex": max(total_rows, 2),
                "startColumnIndex": 0,
                "endColumnIndex": len(headers)
            },
            **STANDARD_BORDER_STYLE
        }
    })

    if has_basic_filter:
        requests.append({
            "clearBasicFilter": {
                "sheetId": sheet_id
            }
        })

    # Всегда устанавливаем автофильтр на точный размер актуальных данных
    requests.append({
        "setBasicFilter": {
            "filter": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                }
            }
        }
    })

    # Форматирование колонки 0 (Дата) как ДАТА (dd.MM.yyyy)
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": total_rows,
                "startColumnIndex": 0,
                "endColumnIndex": 1
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "DATE",
                        "pattern": "dd.MM.yyyy"
                    },
                    "horizontalAlignment": "CENTER"
                }
            },
            "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
        }
    })

    # Выравнивание колонок 1-6 (Смена, Линия, Партия, Мастер, Продукт, Назначение)
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": total_rows,
                "startColumnIndex": 1,
                "endColumnIndex": 4
            },
            "cell": {
                "userEnteredFormat": {
                    "horizontalAlignment": "CENTER"
                }
            },
            "fields": "userEnteredFormat.horizontalAlignment"
        }
    })

    # Числовой формат для колонок данных (4-30: обычные числа)
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": total_rows,
                "startColumnIndex": 4,
                "endColumnIndex": 31
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "NUMBER",
                        "pattern": "#,##0"
                    },
                    "horizontalAlignment": "RIGHT"
                }
            },
            "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
        }
    })

    # Форматирование отклонений (колонки AF-AP, индексы 31-42) как проценты (+0.00% / -0.00%)
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": total_rows,
                "startColumnIndex": 31,
                "endColumnIndex": 42
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "PERCENT",
                        "pattern": "+0.00%;-0.00%;0.00%"
                    },
                    "horizontalAlignment": "RIGHT"
                }
            },
            "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
        }
    })

    # Автоподбор ширины столбцов
    requests.append({
        "autoResizeDimensions": {
            "dimensions": {
                "sheetId": sheet_id,
                "dimension": "COLUMNS",
                "startIndex": 0,
                "endIndex": len(headers)
            }
        }
    })
    
    if requests:
        body = {"requests": requests}
        service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
    print("Синхронизация отчета с Google Таблицами выполнена успешно.")


def export_norms_to_google_sheets(db: Session):
    """
    Создает или обновляет лист 'Нормативы' в Google Таблице, выгружая текущие нормативы
    """
    if not SPREADSHEET_ID or SPREADSHEET_ID.startswith("1_mock"):
        return
        
    service = get_sheets_service()
    sheet_name = "Нормативы"
    
    # 1. Проверяем существование листа
    spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheets_titles = [sh["properties"]["title"] for sh in spreadsheet["sheets"]]
    
    if sheet_name not in sheets_titles:
        body = {
            "requests": [{
                "addSheet": {
                    "properties": {
                        "title": sheet_name,
                        "gridProperties": {
                            "rowCount": 50,
                            "columnCount": 11
                        }
                    }
                }
            }]
        }
        service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
        spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        
    sheet_id = next(sh["properties"]["sheetId"] for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    
    headers = [
        "Продукция", "Вес готового листа (кг)", 
        "Норма Хризотил 4-20", "Норма Хризотил 5-65", "Норма Хризотил 6-40",
        "Норма Цемент", "Норма Целлюлоза", "Норма Дробленый шифер", 
        "Норма Асбозурит", "Норма Стекловолокно"
    ]
    
    norms = db.query(models.ProductNorm).all()
    rows_data = [headers]
    for n in norms:
        rows_data.append([
            n.product_name,
            n.weight_kg or 19.6,
            n.norm_chrysotile_4_20 or 0.0,
            n.norm_chrysotile_5_65 or 0.0,
            n.norm_chrysotile_6_40 or 0.0,
            n.norm_cement or 0.0,
            n.norm_cellulose or 0.0,
            n.norm_crushed_slate or 0.0,
            n.norm_asbozurit or 0.0,
            n.norm_fiberglass or 0.0
        ])
        
    # Очищаем старые значения
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1:K50"
    ).execute()
    
    # Записываем новые значения
    service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1",
        valueInputOption="USER_ENTERED",
        body={"values": rows_data}
    ).execute()
    
    total_rows = len(rows_data)
    sheet_meta = next(sh for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    has_basic_filter = "basicFilter" in sheet_meta

    requests = [
        # Закрепление шапки
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "frozenRowCount": 1
                    }
                },
                "fields": "gridProperties.frozenRowCount"
            }
        },
        # Фиксированная компактная высота шапки (28px)
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 0,
                    "endIndex": 1
                },
                "properties": {
                    "pixelSize": 28
                },
                "fields": "pixelSize"
            }
        },
        # Стилизация шапки (пастельно-зеленый фон)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                },
                "cell": {
                    "userEnteredFormat": PASTEL_GREEN_HEADER
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy)"
            }
        },
        # Границы для всей таблицы
        {
            "updateBorders": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": max(total_rows, 2),
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                },
                **STANDARD_BORDER_STYLE
            }
        },
        # Выравнивание названия продукта (колонка 0)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": 1
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "LEFT"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Форматирование веса и норм (колонки 1-9) как числа
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 1,
                    "endColumnIndex": len(headers)
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "NUMBER",
                            "pattern": "#,##0.00"
                        },
                        "horizontalAlignment": "RIGHT"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        }
    ]

    if has_basic_filter:
        requests.append({
            "clearBasicFilter": {
                "sheetId": sheet_id
            }
        })

    requests.append({
        "setBasicFilter": {
            "filter": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                }
            }
        }
    })

    requests.append({
        "autoResizeDimensions": {
            "dimensions": {
                "sheetId": sheet_id,
                "dimension": "COLUMNS",
                "startIndex": 0,
                "endIndex": len(headers)
            }
        }
    })
    service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body={"requests": requests}).execute()
    print("Нормативы успешно экспортированы в Google Таблицу.")


def sync_norms_from_google_sheets(db: Session):
    """
    Считывает измененные нормативы с листа 'Нормативы' Google Таблицы и записывает их в БД.
    """
    if not SPREADSHEET_ID or SPREADSHEET_ID.startswith("1_mock"):
        raise ValueError("GOOGLE_SPREADSHEET_ID не настроен")
        
    service = get_sheets_service()
    sheet_name = "Нормативы"
    
    result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1:K50"
    ).execute()
    
    rows = result.get("values", [])
    if not rows:
        raise ValueError("Лист 'Нормативы' пустой или не найден")
        
    header = rows[0]
    if "Продукция" not in header:
        raise ValueError("Неверный формат шапки листа 'Нормативы'")
        
    # Вспомогательная функция для безопасного парсинга float
    def safe_float(val):
        if val is None:
            return 0.0
        val_str = str(val).strip().replace(" ", "").replace(",", ".")
        if not val_str:
            return 0.0
        try:
            return float(val_str)
        except ValueError:
            return 0.0
            
    # Начинаем синхронизацию
    # Сначала удаляем старые нормы, чтобы перезаписать
    db.query(models.ProductNorm).delete()
    
    for row in rows[1:]:
        if not row or not row[0]:
            continue
            
        p_name = str(row[0]).strip()
        # Трансляция пиленого шифера в рифленый на лету
        if p_name == "Шифер 8 волн пиленый":
            p_name = "Шифер 8 волн рифленый"
            
        weight = safe_float(row[1] if len(row) > 1 else 19.6)
        n_c4 = safe_float(row[2] if len(row) > 2 else 0.0)
        n_c5 = safe_float(row[3] if len(row) > 3 else 0.0)
        n_c6 = safe_float(row[4] if len(row) > 4 else 0.0)
        n_cem = safe_float(row[5] if len(row) > 5 else 0.0)
        n_cel = safe_float(row[6] if len(row) > 6 else 0.0)
        n_sl = safe_float(row[7] if len(row) > 7 else 0.0)
        n_asb = safe_float(row[8] if len(row) > 8 else 0.0)
        n_fib = safe_float(row[9] if len(row) > 9 else 0.0)
        
        db.add(models.ProductNorm(
            product_name=p_name,
            weight_kg=weight,
            norm_chrysotile_4_20=n_c4,
            norm_chrysotile_5_65=n_c5,
            norm_chrysotile_6_40=n_c6,
            norm_cement=n_cem,
            norm_cellulose=n_cel,
            norm_crushed_slate=n_sl,
            norm_asbozurit=n_asb,
            norm_fiberglass=n_fib
        ))
        
    db.commit()
    print("Нормативы успешно обновлены из Google Таблицы.")


def export_receipt_to_google_sheets(db: Session):
    """
    Создает или обновляет лист 'Приход сырья' в Google Таблице,
    выгружая данные прихода сырья из всех смен (история накопления).
    Вызывается при сохранении сменного рапорта мастера.
    """
    if not SPREADSHEET_ID or SPREADSHEET_ID.startswith("1_mock"):
        print("Экспорт прихода сырья в Google Таблицы пропущен: не задан реальный GOOGLE_SPREADSHEET_ID в .env")
        return

    service = get_sheets_service()
    sheet_name = "Приход сырья"

    # 1. Проверяем существование листа, если нет — создаем
    spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheets_titles = [sh["properties"]["title"] for sh in spreadsheet["sheets"]]

    if sheet_name not in sheets_titles:
        body = {
            "requests": [{
                "addSheet": {
                    "properties": {
                        "title": sheet_name,
                        "gridProperties": {
                            "rowCount": 1000,
                            "columnCount": 18
                        }
                    }
                }
            }]
        }
        service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
        spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()

    sheet_id = next(sh["properties"]["sheetId"] for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)

    # 2. Формируем заголовки
    headers = [
        "Дата", "Смена", "Линия", "Мастер",
        "Хризотил 4-20 (кг)", "Хризотил 5-65 (кг)", "Хризотил 6-40 (кг)",
        "Цемент С1 (кг)", "Цемент С2 (кг)", "Цемент С3 (кг)", "Цемент С4 (кг)", 
        "Целлюлоза (кг)", "Дробленый шифер (кг)",
        "Асбозурит (кг)", "Асбокартон (кг)", "Паллеты (шт)",
        "Стекловолокно (кг)", "Лапрол (кг)"
    ]

    # 3. Собираем данные из БД — все записи прихода сырья
    # Выгружаем каждую запись прихода как отдельную строку
    receipts = db.query(models.RawMaterialReceipt).join(models.Shift).order_by(models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Shift.batch_number.asc(), models.RawMaterialReceipt.id.asc()).all()

    rows_data = []
    rows_data.append(headers)

    for r in receipts:
        date_str = r.shift.date.strftime("%d.%m.%Y") if r.shift and r.shift.date else ""
        shift_name = r.shift.shift_name or ""
        line = r.shift.line or ""
        master_name = r.master.name if r.master else (r.shift.master.name if r.shift and r.shift.master else "")
        
        row = [
            date_str,
            shift_name,
            line,
            master_name,
            r.chrysotile_4_20 or 0.0,
            r.chrysotile_5_65 or 0.0,
            r.chrysotile_6_40 or 0.0,
            r.cement_silo1 or 0.0,
            r.cement_silo2 or 0.0,
            r.cement_silo3 or 0.0,
            r.cement_silo4 or 0.0,
            r.cellulose or 0.0,
            r.crushed_slate or 0.0,
            r.asbozurit or 0.0,
            r.asbocarton or 0.0,
            r.pallets or 0.0,
            r.fiberglass or 0.0,
            r.laprol or 0.0,
        ]
        rows_data.append(row)

    # 4. Полностью перезаписываем лист (очищаем и пишем заново)
    total_rows = len(rows_data)

    # Очищаем старые данные
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'"
    ).execute()

    # Записываем данные
    if rows_data:
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{sheet_name}'!A1",
            valueInputOption="USER_ENTERED",
            body={"values": rows_data}
        ).execute()

    # 5. Форматирование
    sheet_meta = next(sh for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    has_basic_filter = "basicFilter" in sheet_meta

    requests = [
        # Фиксация строки заголовка
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "frozenRowCount": 1
                    }
                },
                "fields": "gridProperties.frozenRowCount"
            }
        },
        # Фиксированная компактная высота шапки (28px)
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 0,
                    "endIndex": 1
                },
                "properties": {
                    "pixelSize": 28
                },
                "fields": "pixelSize"
            }
        },
        # Стилизация шапки (пастельно-зеленый фон)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                },
                "cell": {
                    "userEnteredFormat": PASTEL_GREEN_HEADER
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy)"
            }
        },
        # Границы для всей таблицы
        {
            "updateBorders": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": max(total_rows, 2),
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                },
                **STANDARD_BORDER_STYLE
            }
        },
        # Форматирование колонки 0 (Дата) как ДАТА (dd.MM.yyyy, CENTER)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": max(total_rows, 2),
                    "startColumnIndex": 0,
                    "endColumnIndex": 1
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "DATE",
                            "pattern": "dd.MM.yyyy"
                        },
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        },
        # Выравнивание колонок 1-2 (Смена, Линия) по центру
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": max(total_rows, 2),
                    "startColumnIndex": 1,
                    "endColumnIndex": 3
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Выравнивание колонки 3 (Мастер) по левому краю
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": max(total_rows, 2),
                    "startColumnIndex": 3,
                    "endColumnIndex": 4
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "LEFT"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Форматирование числовых колонок (4-17) как целые числа (#,##0, RIGHT)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": max(total_rows, 2),
                    "startColumnIndex": 4,
                    "endColumnIndex": len(headers)
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "NUMBER",
                            "pattern": "#,##0"
                        },
                        "horizontalAlignment": "RIGHT"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        }
    ]

    if has_basic_filter:
        requests.append({
            "clearBasicFilter": {
                "sheetId": sheet_id
            }
        })

    requests.append({
        "setBasicFilter": {
            "filter": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                }
            }
        }
    })

    requests.append({
        "autoResizeDimensions": {
            "dimensions": {
                "sheetId": sheet_id,
                "dimension": "COLUMNS",
                "startIndex": 0,
                "endIndex": len(headers)
            }
        }
    })

    service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body={"requests": requests}).execute()
    print(f"Экспорт прихода сырья в Google Таблицы выполнен успешно. Выгружено {len(rows_data) - 1} смен.")


def export_downtimes_to_google_sheets(db: Session):
    """
    Создает или обновляет лист 'Простои' в Google Таблице,
    выгружая данные о простоях из всех смен с автофильтрами и форматированием как в сводном отчете.
    """
    if not SPREADSHEET_ID or SPREADSHEET_ID.startswith("1_mock"):
        print("Экспорт простоев в Google Таблицы пропущен: не задан реальный GOOGLE_SPREADSHEET_ID в .env")
        return

    service = get_sheets_service()
    sheet_name = "Простои"

    # 1. Проверяем существование листа, если нет — создаем
    spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheets_titles = [sh["properties"]["title"] for sh in spreadsheet["sheets"]]

    if sheet_name not in sheets_titles:
        body = {
            "requests": [{
                "addSheet": {
                    "properties": {
                        "title": sheet_name,
                        "gridProperties": {
                            "rowCount": 2000,
                            "columnCount": 14
                        }
                    }
                }
            }]
        }
        service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
        spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()

    sheet_id = next(sh["properties"]["sheetId"] for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)

    # 2. Формируем заголовки
    headers = [
        "Дата", "Смена", "Линия", "Мастер",
        "Узел / Оборудование", "Категория", "Простой / Описание",
        "Время начала", "Время окончания",
        "Длительность (мин)", "Остановка оборудования"
    ]

    # 3. Собираем данные из БД — все простои с информацией о смене
    downtimes = db.query(models.Downtime).join(models.Shift).order_by(
        models.Shift.date.asc(), models.Shift.line.asc(), models.Shift.shift_name.asc(), models.Downtime.start_time.asc(), models.Downtime.id.asc()
    ).all()

    rows_data = []
    rows_data.append(headers)

    for d in downtimes:
        shift = d.shift
        date_str = shift.date.strftime("%d.%m.%Y") if (shift and shift.date) else ""
        shift_name_val = shift.shift_name if shift else ""
        line_val = shift.line if shift else ""
        master_val = shift.master.name if (shift and shift.master) else ""
        node_val = (d.node or "").strip()
        category_val = (d.category or "").strip()
        desc_text = (d.description or d.comment or "").strip()

        row = [
            date_str,
            shift_name_val,
            line_val,
            master_val,
            node_val,
            category_val,
            desc_text,
            d.start_time or "",
            d.end_time or "",
            d.duration or 0,
            "Да" if d.is_equipment_downtime else "Нет"
        ]
        rows_data.append(row)

    # 4. Полностью перезаписываем лист (очищаем и пишем заново)
    total_rows = len(rows_data)

    # Очищаем старые данные
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'"
    ).execute()

    # Записываем данные
    if rows_data:
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{sheet_name}'!A1",
            valueInputOption="USER_ENTERED",
            body={"values": rows_data}
        ).execute()

    # 5. Проверяем наличие существующего автофильтра
    sheet_meta = next(sh for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    has_basic_filter = "basicFilter" in sheet_meta

    requests = []

    # Закрепляем первую строку (шапку)
    requests.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "gridProperties": {
                    "frozenRowCount": 1
                }
            },
            "fields": "gridProperties.frozenRowCount"
        }
    })

    # Фиксированная компактная высота шапки (28px)
    requests.append({
        "updateDimensionProperties": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "ROWS",
                "startIndex": 0,
                "endIndex": 1
            },
            "properties": {
                "pixelSize": 28
            },
            "fields": "pixelSize"
        }
    })

    # Сбрасываем старый автофильтр, если был
    if has_basic_filter:
        requests.append({
            "clearBasicFilter": {
                "sheetId": sheet_id
            }
        })

    # Устанавливаем актуальный автофильтр на весь диапазон данных
    requests.append({
        "setBasicFilter": {
            "filter": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": max(total_rows, 1),
                    "startColumnIndex": 0,
                    "endColumnIndex": len(headers)
                }
            }
        }
    })

    # Стилизация заголовка (пастельно-зеленый фон, жирный текст)
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 0,
                "endRowIndex": 1,
                "startColumnIndex": 0,
                "endColumnIndex": len(headers)
            },
            "cell": {
                "userEnteredFormat": PASTEL_GREEN_HEADER
            },
            "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy)"
        }
    })

    # Сетка границ для всех строк
    requests.append({
        "updateBorders": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 0,
                "endRowIndex": max(total_rows, 2),
                "startColumnIndex": 0,
                "endColumnIndex": len(headers)
            },
            **STANDARD_BORDER_STYLE
        }
    })

    # Форматирование колонки 0 (Дата) как ДАТА (dd.MM.yyyy, CENTER)
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": max(total_rows, 2),
                "startColumnIndex": 0,
                "endColumnIndex": 1
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "DATE",
                        "pattern": "dd.MM.yyyy"
                    },
                    "horizontalAlignment": "CENTER"
                }
            },
            "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
        }
    })

    # Выравнивание колонок 1-2 (Смена, Линия) по центру
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": max(total_rows, 2),
                "startColumnIndex": 1,
                "endColumnIndex": 3
            },
            "cell": {
                "userEnteredFormat": {
                    "horizontalAlignment": "CENTER"
                }
            },
            "fields": "userEnteredFormat.horizontalAlignment"
        }
    })

    # Выравнивание колонок 3-6 (Мастер, Узел, Категория, Описание) по левому краю
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": max(total_rows, 2),
                "startColumnIndex": 3,
                "endColumnIndex": 7
            },
            "cell": {
                "userEnteredFormat": {
                    "horizontalAlignment": "LEFT"
                }
            },
            "fields": "userEnteredFormat.horizontalAlignment"
        }
    })

    # Выравнивание колонок 7-8 (Время начала, Время окончания) по центру
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": max(total_rows, 2),
                "startColumnIndex": 7,
                "endColumnIndex": 9
            },
            "cell": {
                "userEnteredFormat": {
                    "horizontalAlignment": "CENTER"
                }
            },
            "fields": "userEnteredFormat.horizontalAlignment"
        }
    })

    # Форматирование колонки 9 (Длительность) как ЧИСЛО (#,##0, RIGHT)
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": max(total_rows, 2),
                "startColumnIndex": 9,
                "endColumnIndex": 10
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "NUMBER",
                        "pattern": "#,##0"
                    },
                    "horizontalAlignment": "RIGHT"
                }
            },
            "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
        }
    })

    # Выравнивание колонки 10 (Остановка оборудования) по центру
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": max(total_rows, 2),
                "startColumnIndex": 10,
                "endColumnIndex": 11
            },
            "cell": {
                "userEnteredFormat": {
                    "horizontalAlignment": "CENTER"
                }
            },
            "fields": "userEnteredFormat.horizontalAlignment"
        }
    })

    # Авто-размер ширины колонок
    requests.append({
        "autoResizeDimensions": {
            "dimensions": {
                "sheetId": sheet_id,
                "dimension": "COLUMNS",
                "startIndex": 0,
                "endIndex": len(headers)
            }
        }
    })

    service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body={"requests": requests}).execute()
    print(f"Экспорт простоев в Google Таблицы выполнен успешно. Выгружено {len(rows_data) - 1} записей с автофильтрами.")




from datetime import timedelta

def get_iso_week_key(d):
    # Returns (year, week) tuple for grouping
    if not d:
        return (0, 0)
    return d.isocalendar()[:2]

def sync_qcd_reports_to_google_sheets(db: Session):
    if not SPREADSHEET_ID or SPREADSHEET_ID.startswith("1_mock"):
        return
        
    service = get_sheets_service()
    sheet_name = "Переборка"
    
    # 1. Проверяем существование листа
    spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheets_titles = [sh["properties"]["title"] for sh in spreadsheet["sheets"]]
    
    if sheet_name not in sheets_titles:
        body = {
            "requests": [{
                "addSheet": {
                    "properties": {
                        "title": sheet_name,
                        "gridProperties": {
                            "rowCount": 2000,
                            "columnCount": 25,
                            "frozenRowCount": 1
                        }
                    }
                }
            }]
        }
        service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body=body).execute()
        spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        
    sheet_id = next(sh["properties"]["sheetId"] for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    
    headers = [
        "№ партии", "Дата", "День нед.", "Время смены", "Мастер ЛФМ", "Продукт", "Формовка, шт",
        "Смена", "1 сорт, шт", "Брак, шт", "% брака", "Детализация брака",
        "Прошлая смена", "Мастер (прошлая)", "1 сорт (прошлая), шт", "Брак (прошлая), шт", "Детализация брака (прошлая)",
        "Всего 1 сорт, шт", "Всего брак, шт"
    ]
    
    from datetime import date
    # Фильтруем записи с 1 сентября 2026 года (начало нового формата ведения)
    start_filter_date = date(2026, 9, 1)
    batches = (
        db.query(models.Batch)
        .join(models.Shift)
        .filter(models.Shift.date >= start_filter_date)
        .order_by(models.Shift.date.asc(), models.Batch.batch_number.asc())
        .all()
    )
    
    # Кэш записей графика сменности по датам
    all_schedules = {e.date_str: e for e in db.query(models.ShiftScheduleEntry).all()}
    
    # Кэш всех смен в БД для поиска мастера сдавшей смены
    all_shifts_list = db.query(models.Shift).all()
    shifts_lookup = {}
    for s in all_shifts_list:
        if s.date:
            key = (s.date.strftime("%Y-%m-%d"), s.shift_name, s.line)
            shifts_lookup[key] = s
    
    rows_data = [headers]
    
    # Отсортированный список дат для поиска сдавшей смены
    sorted_sched_dates = sorted(all_schedules.keys(), key=lambda d: datetime.strptime(d, "%d.%m.%Y"))

    for b in batches:
        shift_date = b.shift.date if b.shift else None
        shift_name = b.shift.shift_name if b.shift else ""
        shift_line = b.shift.line if b.shift else "Линия 1"
        product_name = b.product_name or "Неизвестный продукт"
        master_name = b.shift.master.name if (b.shift and b.shift.master) else ""
        
        date_str_fmt = shift_date.strftime("%d.%m.%Y") if shift_date else ""
        sched = all_schedules.get(date_str_fmt)
        day_of_week = sched.day_of_week if sched else (shift_date.strftime("%a") if shift_date else "")
        
        is_day = (shift_name == "День")
        shift_group = (sched.day_shift_group if is_day else sched.night_shift_group) if sched else ""
        
        # Определяем сдающую (предшествующую) бригаду и мастера из БД/графика
        prev_shift_group = ""
        prev_master_name = ""
        if is_day:
            if shift_date:
                prev_d_obj = shift_date - timedelta(days=1)
                prev_date_fmt = prev_d_obj.strftime("%d.%m.%Y")
                prev_date_iso = prev_d_obj.strftime("%Y-%m-%d")
                prev_sched = all_schedules.get(prev_date_fmt)
                prev_shift_group = prev_sched.night_shift_group if prev_sched else ""
                
                prev_shift_obj = shifts_lookup.get((prev_date_iso, "Ночь", shift_line))
                if prev_shift_obj and prev_shift_obj.master:
                    prev_master_name = prev_shift_obj.master.name
        else:
            prev_shift_group = sched.day_shift_group if sched else ""
            if shift_date:
                curr_date_iso = shift_date.strftime("%Y-%m-%d")
                prev_shift_obj = shifts_lookup.get((curr_date_iso, "День", shift_line))
                if prev_shift_obj and prev_shift_obj.master:
                    prev_master_name = prev_shift_obj.master.name
            
        ds_first = b.ds_first_grade or 0
        ds_def = b.ds_defect or 0
        
        prev_f = b.prev_first_grade or 0
        prev_d = b.prev_defect or 0
        
        lfm_sheets = 0
        if b.shift and b.shift.lfm_reports:
            lfm_sheets = sum(r.lfm_sheets for r in b.shift.lfm_reports)
        elif b.stacked_stacks:
            lfm_sheets = b.stacked_stacks
            
        total_sheets = lfm_sheets
        pct_defect = (ds_def / total_sheets) if total_sheets > 0 else 0
        
        # 7 видов брака текущей смены
        def_parts = []
        if b.ds_defect_scratch: def_parts.append(f"Сдир ({b.ds_defect_scratch})")
        if b.ds_defect_bad_cut: def_parts.append(f"Плохой рез ({b.ds_defect_bad_cut})")
        if b.ds_defect_stick_top: def_parts.append(f"Налип сверху ({b.ds_defect_stick_top})")
        if b.ds_defect_broken: def_parts.append(f"Сломан ({b.ds_defect_broken})")
        if b.ds_defect_fell_box: def_parts.append(f"Упал коробки ({b.ds_defect_fell_box})")
        if b.ds_defect_thickness: def_parts.append(f"Не соотв. толщины ({b.ds_defect_thickness})")
        if b.ds_defect_edge: def_parts.append(f"Кромка ({b.ds_defect_edge})")
        note_defect = ", ".join(def_parts)

        # 7 видов брака сдавшей смены
        prev_parts = []
        if b.prev_defect_scratch: prev_parts.append(f"Сдир ({b.prev_defect_scratch})")
        if b.prev_defect_bad_cut: prev_parts.append(f"Плохой рез ({b.prev_defect_bad_cut})")
        if b.prev_defect_stick_top: prev_parts.append(f"Налип сверху ({b.prev_defect_stick_top})")
        if b.prev_defect_broken: prev_parts.append(f"Сломан ({b.prev_defect_broken})")
        if b.prev_defect_fell_box: prev_parts.append(f"Упал коробки ({b.prev_defect_fell_box})")
        if b.prev_defect_thickness: prev_parts.append(f"Не соотв. толщины ({b.prev_defect_thickness})")
        if b.prev_defect_edge: prev_parts.append(f"Кромка ({b.prev_defect_edge})")
        prev_note = ", ".join(prev_parts)
        
        total_first_all = ds_first + prev_f
        total_def_all = ds_def + prev_d
        
        rows_data.append([
            b.batch_number or "",
            date_str_fmt,
            day_of_week or "",
            shift_name,
            master_name,
            product_name,
            total_sheets,
            shift_group or "",
            ds_first,
            ds_def,
            pct_defect,
            note_defect,
            prev_shift_group or "",
            prev_master_name or "",
            prev_f,
            prev_d,
            prev_note,
            total_first_all,
            total_def_all
        ])
        
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{sheet_name}'!A1:Z3000"
    ).execute()
    
    if len(rows_data) > 0:
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"'{sheet_name}'!A1",
            valueInputOption="USER_ENTERED",
            body={"values": rows_data}
        ).execute()
        
    total_rows = max(len(rows_data), 2)
    total_cols = len(headers)
    
    sheet_meta = next(sh for sh in spreadsheet["sheets"] if sh["properties"]["title"] == sheet_name)
    has_basic_filter = "basicFilter" in sheet_meta

    requests = [
        # Фиксация строки заголовка
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "frozenRowCount": 1
                    }
                },
                "fields": "gridProperties.frozenRowCount"
            }
        },
        # Фиксированная компактная высота шапки (28px)
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 0,
                    "endIndex": 1
                },
                "properties": {
                    "pixelSize": 28
                },
                "fields": "pixelSize"
            }
        },
        # Стилизация шапки (пастельно-зеленый фон)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": total_cols
                },
                "cell": {
                    "userEnteredFormat": PASTEL_GREEN_HEADER
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy)"
            }
        },
        # Общие границы таблицы
        {
            "updateBorders": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": total_cols
                },
                **STANDARD_BORDER_STYLE
            }
        },
        # Форматирование колонки 1 (Дата) как ДАТА (dd.MM.yyyy, CENTER)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 1,
                    "endColumnIndex": 2
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "DATE",
                            "pattern": "dd.MM.yyyy"
                        },
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        },
        # Выравнивание текстовых колонок (Партия, День недели, Смена, Бригада) по центру
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": 1
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 2,
                    "endColumnIndex": 4
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Выравнивание Мастера и Продукта (колонки 4, 5) по левому краю
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 4,
                    "endColumnIndex": 6
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "LEFT"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Формат чисел (Формовка: колонка 6)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 6,
                    "endColumnIndex": 7
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "NUMBER", "pattern": "#,##0"},
                        "horizontalAlignment": "RIGHT"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        },
        # Выравнивание Бригады (колонка 7) по центру
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 7,
                    "endColumnIndex": 8
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Формат чисел (1 сорт и Брак текущей бригады: колонки 8, 9)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 8,
                    "endColumnIndex": 10
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "NUMBER", "pattern": "#,##0"},
                        "horizontalAlignment": "RIGHT"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        },
        # Формат процентов (% брака: колонка 10)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 10,
                    "endColumnIndex": 11
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "PERCENT", "pattern": "0.00%"},
                        "horizontalAlignment": "RIGHT"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        },
        # Выравнивание примечания дефектов (колонка 11) по левому краю
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 11,
                    "endColumnIndex": 12
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "LEFT"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Выравнивание сдавшей бригады и мастера (колонки 12, 13)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 12,
                    "endColumnIndex": 13
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 13,
                    "endColumnIndex": 14
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "LEFT"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Формат чисел (1 сорт и Брак сдавшей бригады: колонки 14, 15)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 14,
                    "endColumnIndex": 16
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "NUMBER", "pattern": "#,##0"},
                        "horizontalAlignment": "RIGHT"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        },
        # Выравнивание примечания дефектов сдавшей бригады (колонка 16)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 16,
                    "endColumnIndex": 17
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "LEFT"
                    }
                },
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        },
        # Формат чисел (Итоги Всего 1 сорт и Всего брак: колонки 17, 18)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 17,
                    "endColumnIndex": 19
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "NUMBER", "pattern": "#,##0"},
                        "horizontalAlignment": "RIGHT"
                    }
                },
                "fields": "userEnteredFormat.numberFormat,userEnteredFormat.horizontalAlignment"
            }
        }
    ]

    if has_basic_filter:
        requests.append({
            "clearBasicFilter": {
                "sheetId": sheet_id
            }
        })

    requests.append({
        "setBasicFilter": {
            "filter": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": total_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": total_cols
                }
            }
        }
    })

    requests.append({
        "autoResizeDimensions": {
            "dimensions": {
                "sheetId": sheet_id,
                "dimension": "COLUMNS",
                "startIndex": 0,
                "endIndex": total_cols
            }
        }
    })
    
    service.spreadsheets().batchUpdate(spreadsheetId=SPREADSHEET_ID, body={"requests": requests}).execute()
    print("Лист 'Переборка' (чистая непрерывная таблица с бригадами) успешно экспортирован.")


EMPLOYEES_SPREADSHEET_ID = "1QyDBTkU_y-E_pxgOp-l1J5ejvEwcDhBvt7s5GVOv5I8"
SCHEDULE_SPREADSHEET_ID = "1WOp9ME0ThkQn8Uf7uZ4HZ03PaNtCn2Y65PpuVT0pkME"
CHECKLISTS_SPREADSHEET_ID = os.getenv("CHECKLISTS_SPREADSHEET_ID") or "17AqmwqceblrAzlHfSztc80cmGJJ71B_godpYoK0QHrc"

def get_department_by_position(pos: str, shift_g: str = "") -> str:
    """Определяет утвержденный производственный участок по должности."""
    p = pos.lower().strip()
    if "мастер" in p:
        return "Цех ХЦИ"
    if "машин" in p or "помощник" in p:
        return "ЛФМ"
    if "дестакер" in p:
        return "Дестакер"
    if "стакер" in p:
        return "Стакер"
    if "целлюлоз" in p or "рекуператор" in p:
        return "Рекуператоры и целлюлоза"
    if "зо" in p or "заготовительн" in p:
        return "Заготовительное отделение (ЗО)"
    if "дозиров" in p:
        return "Участок дозировки"
    if "дежурный слесарь" in p or "дежурный электрик" in p or "слесар" in p or "электрик" in p:
        return "Сменные ремслужбы"
    if "лаборант" in p or "отк" in p or "скк" in p:
        return "СКК"
    if "котельн" in p:
        return "Котельная"
    if "насос" in p:
        return "Зона ремонта насосов"
    if "цилиндр" in p:
        return "Участок ремонта цилиндров"
    if "мастерская" in p:
        return "Слесарная мастерская"
    return "Цех ХЦИ"

def sync_employees_from_google_sheets(db: Session):
    """Импортирует или обновляет список сотрудников из Google Таблицы."""
    import urllib.request, csv, io
    
    url = f"https://docs.google.com/spreadsheets/d/{EMPLOYEES_SPREADSHEET_ID}/export?format=csv&gid=652222344"
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req) as resp:
        content = resp.read().decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        
    if not rows or len(rows) < 2:
        return {"status": "error", "message": "Пустой ответ от Google Таблицы"}
        
    count = 0
    # Header: ['№', 'Смена', 'Должность', 'ФИО']
    for r in rows[1:]:
        if len(r) >= 4 and r[3].strip():
            num_val = int(r[0].strip()) if r[0].strip().isdigit() else None
            shift_g = r[1].strip()
            pos = r[2].strip()
            name_val = r[3].strip()
            dept = get_department_by_position(pos, shift_g)
            
            emp = db.query(models.ChecklistEmployee).filter(
                models.ChecklistEmployee.name == name_val,
                models.ChecklistEmployee.shift_group == shift_g,
                models.ChecklistEmployee.position == pos
            ).first()
            
            if not emp:
                emp = models.ChecklistEmployee(
                    num=num_val,
                    shift_group=shift_g,
                    position=pos,
                    name=name_val,
                    department=dept,
                    is_active=True
                )
                db.add(emp)
                count += 1
            else:
                emp.num = num_val
                emp.department = dept
                emp.is_active = True
                
    db.commit()
    return {"status": "ok", "synced_count": count, "total_rows": len(rows)-1}

def sync_schedule_from_excel_or_google(db: Session):
    """Импортирует или обновляет график сменности из локального файла Excel (или Google Таблицы при отсутствии)."""
    import os, datetime
    
    # 1. Проверяем локальный эталонный Excel файл в 'График и табеля'
    excel_path = os.path.join(os.getcwd(), "График и табеля", "График_сменности_2026_до_конца_года.xlsx")
    if os.path.exists(excel_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.worksheets[0]
            
            month_configs = [
                ('Июль', 9, 7, 31),
                ('Август', 16, 8, 31),
                ('Сентябрь', 23, 9, 30),
                ('Октябрь', 30, 10, 31),
                ('Ноябрь', 37, 11, 30),
                ('Декабрь', 44, 12, 31)
            ]
            dow_ru = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
            
            existing = {e.date_str: e for e in db.query(models.ShiftScheduleEntry).all()}
            count = 0
            
            for m_name, start_r, m_num, days_in_m in month_configs:
                for d in range(1, days_in_m + 1):
                    dt = datetime.date(2026, m_num, d)
                    d_str = dt.strftime('%d.%m.%Y')
                    dow = dow_ru[dt.weekday()]
                    
                    s1 = str(ws.cell(start_r + 2, 1 + d).value or '').strip()
                    s2 = str(ws.cell(start_r + 3, 1 + d).value or '').strip()
                    s3 = str(ws.cell(start_r + 4, 1 + d).value or '').strip()
                    s4 = str(ws.cell(start_r + 5, 1 + d).value or '').strip()
                    
                    day_g = ''
                    night_g = ''
                    for c_idx, s_val in [(1, s1), (2, s2), (3, s3), (4, s4)]:
                        if 'Д' in s_val: day_g = f'Смена {c_idx}'
                        if 'Н' in s_val: night_g = f'Смена {c_idx}'
                        
                    entry = existing.get(d_str)
                    if not entry:
                        entry = models.ShiftScheduleEntry(
                            date_str=d_str,
                            day_of_week=dow,
                            shift1_status=s1,
                            shift2_status=s2,
                            shift3_status=s3,
                            shift4_status=s4,
                            day_shift_group=day_g,
                            night_shift_group=night_g
                        )
                        db.add(entry)
                        count += 1
                    else:
                        entry.day_of_week = dow
                        entry.shift1_status = s1
                        entry.shift2_status = s2
                        entry.shift3_status = s3
                        entry.shift4_status = s4
                        entry.day_shift_group = day_g
                        entry.night_shift_group = night_g
                        count += 1
            db.commit()
            return {"status": "ok", "source": "excel", "synced_count": count}
        except Exception as e:
            db.rollback()
            print(f"[ScheduleSync] Ошибка загрузки из Excel: {e}")
            
    # 2. Fallback: Google Sheets CSV
    import urllib.request, csv, io
    url = f"https://docs.google.com/spreadsheets/d/{SCHEDULE_SPREADSHEET_ID}/export?format=csv&gid=1540648819"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            content = resp.read().decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            
        if not rows or len(rows) < 2:
            return {"status": "error", "message": "Пустой ответ от Google Таблицы графика"}
            
        count = 0
        for r in rows[1:]:
            if len(r) >= 8 and r[0].strip():
                d_str = r[0].strip()
                dow = r[1].strip()
                s1 = r[2].strip()
                s2 = r[3].strip()
                s3 = r[4].strip()
                s4 = r[5].strip()
                day_g = r[6].strip()
                night_g = r[7].strip()
                
                entry = db.query(models.ShiftScheduleEntry).filter(models.ShiftScheduleEntry.date_str == d_str).first()
                if not entry:
                    entry = models.ShiftScheduleEntry(
                        date_str=d_str,
                        day_of_week=dow,
                        shift1_status=s1,
                        shift2_status=s2,
                        shift3_status=s3,
                        shift4_status=s4,
                        day_shift_group=day_g,
                        night_shift_group=night_g
                    )
                    db.add(entry)
                    count += 1
                else:
                    entry.day_of_week = dow
                    entry.shift1_status = s1
                    entry.shift2_status = s2
                    entry.shift3_status = s3
                    entry.shift4_status = s4
                    entry.day_shift_group = day_g
                    entry.night_shift_group = night_g
        db.commit()
        return {"status": "ok", "source": "google_sheets", "synced_count": count}
    except Exception as e:
        db.rollback()
        return {"status": "error", "message": str(e)}

# Backward compatibility alias
sync_schedule_from_google_sheets = sync_schedule_from_excel_or_google

def get_or_create_sheet(service, spreadsheet_id: str, sheet_title: str):
    """Возвращает sheet_id существующего листа или создает новый лист с таким названием."""
    spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    for sh in spreadsheet.get("sheets", []):
        if sh["properties"]["title"] == sheet_title:
            return sh["properties"]["sheetId"]
            
    # Если лист не найден, создаем
    body = {
        "requests": [{
            "addSheet": {
                "properties": {
                    "title": sheet_title,
                    "gridProperties": {
                        "rowCount": 1000,
                        "columnCount": 30
                    }
                }
            }
        }]
    }
    res = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
    new_sheet_id = res["replies"][0]["addSheet"]["properties"]["sheetId"]
    return new_sheet_id

def export_checklists_to_google_sheets(db: Session):
    """
    Выгружает заполненные чек-листы в отдельную Google Таблицу с разделением по листам:
    1. 'Сводный журнал' (все чек-листы)
    2. 'Чек-листы Мастеров' (мастера ХЦИ)
    3. 'Чек-листы Рабочих' (сменные рабочие по участкам)
    4. 'Инспекции ИТР' (дневные проверки и аудит)
    """
    target_id = CHECKLISTS_SPREADSHEET_ID
    if not target_id or target_id.startswith("1_mock"):
        print(f"Skipping checklist export: target_id={target_id}")
        return
        
    try:
        service = get_sheets_service()
        submissions = db.query(models.ChecklistSubmission).order_by(models.ChecklistSubmission.created_at.asc(), models.ChecklistSubmission.id.asc()).all()
        
        headers = [
            "ID", "Дата", "Время сохранения", "Смена (День/Ночь)", "Бригада",
            "Тип чек-листа", "Участок", "Принимающий / Проверяющий", "Сдающий",
            "Статус", "Кол-во замечаний", "Выявленные замечания / Пункты с дефектами", "Общие примечания"
        ]

        def build_row(sub):
            defects_list = []
            try:
                items = json.loads(sub.items_data or "[]")
                for it in items:
                    if it.get("status") == "fail":
                        cm = f" ({it.get('comment')})" if it.get('comment') else ""
                        defects_list.append(f"❌ {it.get('title')}{cm}")
            except Exception:
                pass
                
            defects_str = "\n".join(defects_list) if defects_list else "— Нет замечаний (Норма)"
            status_display = "Замечания" if sub.remarks_count > 0 else "Норма"
            created_time_str = sub.created_at.strftime("%H:%M:%S") if sub.created_at else ""
            
            return [
                str(sub.id),
                sub.date_str or "",
                created_time_str,
                sub.shift_name or "",
                sub.shift_group or "",
                sub.template_title or sub.template_code or "",
                sub.department or "Общий",
                sub.inspector_name or "",
                sub.submitter_name or "",
                status_display,
                str(sub.remarks_count),
                defects_str,
                sub.notes or ""
            ]

        # Группировка записей по листам
        all_rows = [headers] + [build_row(s) for s in submissions]
        
        master_subs = [s for s in submissions if s.template_code == "master_shift" or "мастер" in (s.template_title or "").lower()]
        master_rows = [headers] + [build_row(s) for s in master_subs]

        worker_subs = [s for s in submissions if s.template_code == "worker_shift_handover" or "рабоч" in (s.template_title or "").lower()]
        worker_rows = [headers] + [build_row(s) for s in worker_subs]

        day_subs = [s for s in submissions if s.template_code == "day_inspection" or "дневн" in (s.template_title or "").lower() or "итр" in (s.template_title or "").lower()]
        day_rows = [headers] + [build_row(s) for s in day_subs]

        sheets_to_sync = [
            {"title": "Чек-листы (Премирование)", "rows": all_rows, "color": {"red": 0.12, "green": 0.35, "blue": 0.65}},
            {"title": "Чек-листы Мастеров", "rows": master_rows, "color": {"red": 0.18, "green": 0.45, "blue": 0.25}},
            {"title": "Чек-листы Рабочих", "rows": worker_rows, "color": {"red": 0.85, "green": 0.45, "blue": 0.12}},
            {"title": "Инспекции ИТР", "rows": day_rows, "color": {"red": 0.45, "green": 0.22, "blue": 0.65}}
        ]

        for item in sheets_to_sync:
            stitle = item["title"]
            srows = item["rows"]
            header_color = item["color"]
            
            sheet_id = get_or_create_sheet(service, target_id, stitle)
            
            # Очищаем и записываем
            service.spreadsheets().values().clear(
                spreadsheetId=target_id,
                range=f"'{stitle}'!A1:Z{max(len(srows)+10, 100)}"
            ).execute()
            
            service.spreadsheets().values().update(
                spreadsheetId=target_id,
                range=f"'{stitle}'!A1",
                valueInputOption="USER_ENTERED",
                body={"values": srows}
            ).execute()
            
            # Форматирование
            requests = [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": 0,
                            "endColumnIndex": len(headers)
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": header_color,
                                "textFormat": {
                                    "bold": True,
                                    "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                                    "fontSize": 11,
                                    "fontFamily": "Calibri"
                                },
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE"
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)"
                    }
                },
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": len(srows),
                            "startColumnIndex": 0,
                            "endColumnIndex": len(headers)
                        },
                        "top": {"style": "SOLID"},
                        "bottom": {"style": "SOLID"},
                        "left": {"style": "SOLID"},
                        "right": {"style": "SOLID"},
                        "innerHorizontal": {"style": "SOLID"},
                        "innerVertical": {"style": "SOLID"}
                    }
                },
                {
                    "autoResizeDimensions": {
                        "dimensions": {
                            "sheetId": sheet_id,
                            "dimension": "COLUMNS",
                            "startIndex": 0,
                            "endIndex": len(headers)
                        }
                    }
                }
            ]
            service.spreadsheets().batchUpdate(spreadsheetId=target_id, body={"requests": requests}).execute()

        print(f"Экспорт чек-листов по листам в Google Sheets завершен. Всего выгружено {len(submissions)} записей.")
    except Exception as e:
        print(f"Ошибка экспорта чек-листов в Google Sheets: {e}")
        raise e

